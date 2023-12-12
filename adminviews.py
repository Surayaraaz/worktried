import json
from django.db import connection
import requests
from django.contrib import messages
from django.contrib.auth.models import User
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render,redirect
from django.urls import reverse
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.core.mail import send_mail
from django.conf import settings
from django.db.models import Sum
from django.views.decorators.csrf import csrf_exempt
from ehrms.forms import AddEmployForm,EditEmployForm
from ehrms.models import CustomUser,empdocs,admin_ad_salary,admin_project_create,details,Reimbursement,MonthlyTotal,documents_setup,ad_salary,salary_struct,admin_add_form1,integrations,companylogo,employ_add_form,checkout,checkin,admin_home_drop,salary_struc,set_salary_structure,set_payroll_date,duration_year,duration_months,Progress,company_details_first,pvt_pub,sole_proprietorship,partnership,trust_ngo,admin_drop,adminnav,AdminHod,LeaveReportEmploy,employ_payslip,Employs,NotificationEmploy,employ_tax_form,Admin_Reimbursement,types,admin_doc,typeofd,TDS,P_tax,ESIC,PF, Company,company_details,company_logo,masterctc,admin_tax_details,reimbursementsetup1,reimbursementsetup,documents_setup1
from django.core.paginator import Paginator

from ehrms.models import Monitoring,MonitoringDetails,Screenshots,SystemStatus
from .models import Companys

import datetime

from django.shortcuts import render, redirect
from django.contrib import messages
import requests,random
from django.core.mail import send_mail
from django.http import JsonResponse, HttpResponseNotAllowed





def view_app_web(request): 

    if request.method == 'POST':
        admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
        admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
        data = Companys.objects.filter(usernumber=request.user.id).first()
        data1=data.id
        user=CustomUser.objects.filter(id=request.user.id).first()
        userid1=user.id

        # s=adminnav.objects.filter(is_projectmanager = 1)
        s=adminnav.objects.all()

        projectm=admin_project_create.objects.filter(admin_id= userid1)
        h=HR.objects.all()
        employs_all=Employs.objects.all()
        projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
        


        
        e_id = request.POST['e_id']
        m_date = request.POST['date_log']
        m_date_f1 = datetime.strptime(m_date, '%Y-%m-%d')
        m_date_f2 = datetime.strftime(m_date_f1, '%Y-%m-%d')
        moni_details = Monitoring.objects.filter(employee_id=e_id, m_log_ts__startswith=m_date_f2).exclude(m_title="").values()
        return render(request, 'admin-template/ViewMoniLogs.html', {"msg": moni_details,'admin_drops':admin_drops,'admin_home_drops':admin_home_drops,'data':data,'data1':data1,'s':s,'projectm':projectm,'employs_all':employs_all,'projects_drops':projects_drops,'user':user,'h':h})
    else:
   
        emp_details = Employs.objects.all()
        return render(request, 'admin-template/SelectMoniEmp.html', {"msg": emp_details})


def depth_view_app_web(request):
    if request.method == 'POST':
        e_id = request.POST['e_id']
        md_date = request.POST['date_log']
        md_date_f1 = datetime.strptime(md_date, '%Y-%m-%d')
        md_date_f2 = datetime.strftime(md_date_f1, '%Y-%#m-%#d')
        depth_moni_details = MonitoringDetails.objects.filter( employee_id=e_id).exclude(md_title="").values()
        return render(request, 'admin-template/ViewDepthMoniLogs.html', {"msg": depth_moni_details})
    else:

        emp_details = Employs.objects.all()
        return render(request, 'admin-template/SelectDepthMoniEmp.html', {"msg": emp_details})
from .models import Screenshots
def ss_monitoring(request):
    if request.method == 'POST':
        e_id = request.POST['e_id']
        ss_date = request.POST['date_log']
        ss_date_f1 = datetime.strptime(ss_date, '%Y-%m-%d')
        ss_date_f2 = datetime.strftime(ss_date_f1, '%Y-%#m-%#d')
        ss_moni_details = Screenshots.objects.filter( employee_id=e_id, timestamp__startswith=ss_date_f2).values()
        return render(request, 'admin-template/ViewSSMoniLogs.html', {"msg": ss_moni_details})
    else:
        emp_details = Employs.objects.all()
        return render(request, 'admin-template/SelectSSMoniEmp.html', {"msg": emp_details})
    
def power_monitoring(request):
    if request.method == 'POST':
        e_id = request.POST['e_id']
        pm_date = request.POST['date_log']
        pm_date_f1 = datetime.strptime(pm_date, '%Y-%m-%d')
        pm_date_f2 = datetime.strftime(pm_date_f1, '%Y-%#m-%#d')
        ss_power_details = SystemStatus.objects.filter( employee_id=e_id, timestamp__startswith=pm_date_f2).values()
        return render(request, 'admin-template/ViewPowerMoniLogs.html', {"msg": ss_power_details})
    else:
        emp_details = Employs.objects.all()
        return render(request, 'admin-template/SelectPowerMoniEmp.html', {"msg": emp_details})




def admin_home(request):
    # next_pay_roll_date 
    current_date = date.today()
    # Assuming payroll date is in the format of 'dd' (e.g., '01', '15', '30')
    payrolldate = set_payroll_date.objects.first().payrolldate  # Fetch the payrolldate from the database
    try:
        day = int(payrolldate)
    except ValueError:
        day = 1  # Default to 1 if the payrolldate is not a valid integer
    
    if current_date.day <= day:
        next_month = current_date.replace(month=current_date.month)
    else:
        next_month = current_date.replace(month=current_date.month )


    next_month = next_month.replace(day=day)
    # Format the date as 'dd,month,yyyy'
    next_month_payroll_date_str = next_month.strftime("%d/%b/%Y")
    user=CustomUser.objects.filter(id=request.user.id).first()
    userid1=user.id
    employ_count1=Employs.objects.all().count()
    # data = AdminHod.objects.filter(admin=request.user.id,options=0)
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()


    s=adminnav.objects.filter(is_projectmanager = 1)
    projectm=admin_project_create.objects.filter(admin_id= userid1)
    h=HR.objects.all()
    employs_all=Employs.objects.all()
    admin_drops=admin_drop.objects.filter(parent_category=None,show = 0).order_by('id')
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    noti=NotificationEmploy.objects.all()
    progress_obj, _ = Progress.objects.get_or_create(pk=1, defaults={'value': 0, 'progress_form1':0, 'progress_form2':0, 'progress_form3':0, 'progress_form4':0, 'progress_form5':0, 'progress_form6':0})
    today_date = datetime.now().date()

    # Count the number of check-ins for today
    # Count the number of check-ins for today
    b = checkin.objects.filter(date=today_date,companyid=data,is_employee="1").count()   
    c = Employs.objects.filter(companyid=data1).count()
    d = c - b

    print(f"Today's date: {today_date}")
    print(f"Check-ins for today: {b}")
    print(f"Total employees: {c}")
    print(f"Non check-in employees: {d}")

    leave_applicants_count = LeaveReportEmploy.objects.filter(employ_id__companyid=data1).count()
    tomorrow_date = date.today() + timedelta(days=1)
    approved_leave_count = LeaveReportEmploy.objects.filter(leave_status=1).count()
    pending_leave_count = LeaveReportEmploy.objects.filter(leave_status=0).count()

    # Calculate the total leave applications
    total_leave_count = approved_leave_count + pending_leave_count

    # Get the count of members who have applied for leave permission tomorrow
    tomorrow_leave_applicants_count = LeaveReportEmploy.objects.filter(leave_date=tomorrow_date).count()
    # checked_in_employee_ids = checkin.objects.filter(date=today_date).values_list('empid', flat=True)

    # # Get the total number of employees who have not checked in today
    # d = Employs.objects.exclude(empid__in=checked_in_employee_ids).count()
    num_progress_submitted = sum([int(progress_obj.progress_form1), int(progress_obj.progress_form2),int(progress_obj.progress_form3),int(progress_obj.progress_form4),int(progress_obj.progress_form5),int(progress_obj.progress_form6)])
    form_status = sum([int(progress_obj.value), int(progress_obj.value1),int(progress_obj.value2),int(progress_obj.value3),int(progress_obj.value4),int(progress_obj.value5)])
    if form_status > 0:
        progress_s = num_progress_submitted
        message=form_status
    else:
        progress_s = 0
        message=0
    users=company_details_first.objects.first()
    users1=set_payroll_date.objects.first()
    users2=set_salary_structure.objects.first()
    da1=Employs.objects.filter(admin=request.user.id).first()
    if da1:
       da2=da1.id
       project_details = Project.objects.filter(o_id=da2).prefetch_related('teammember_set') # Adjust this query to filter projects as needed
    else:
        da2=None
        project_details=None
    project_cm=TeamMember.objects.filter(is_team_lead=1)
    if project_details:
      team_members = TeamMember.objects.filter(project__in=project_details)
      team_member_scores = [(team_member, team_member.calculate_total_score_in_project()) for team_member in team_members]
      team_member_scores = sorted(team_member_scores, key=lambda x: x[1], reverse=True)
    else :
       team_members=None
       team_member_scores=None

    
    if request.method == "POST":
        for i in projectm:
            project_id = i.id
            status = request.POST.get('status-' + str(project_id))
            up = admin_project_create.objects.get(id=project_id)
            up.status = status
            up.save()
        return redirect('/admin_home')



    return render(request,"admin-template/admin-dashboard.html",{'h':h,'a':a,
        'project_cm':project_cm,'da1':da1,'da2':da2,'project_details':project_details,"projects_drops":projects_drops,'next_month_payroll_date':next_month_payroll_date_str,'h':h,'data':data,'user':user,'message':message,'users':users,'users1':users1,'users2':users2,'progress_s':progress_s,'noti':noti,'admin_home_drops':admin_home_drops,'admin_drops':admin_drops,"employ_count1":employ_count1,'s':s,'b':b,'c':c,'d':d,'leave_applicants_count':leave_applicants_count,'tomorrow_leave_applicants_count':tomorrow_leave_applicants_count,'approved_leave_count': approved_leave_count,'pending_leave_count': pending_leave_count,'total_leave_count': total_leave_count,'userid1':userid1,'projectm':projectm})




def checkin_countdata(request):

    # Get the current date
    today_date = date.today()
    compid=Companys.objects.filter(usernumber=request.user.id).first()

    # Fetch checkin data for the current date, including the 'status' field
    checkin_data = checkin.objects.filter(date=today_date,companyid=compid).values('empid', 'status')

    # Get the empids and status for employees who checked in today
    empid_status_data = {record['empid']: record['status'] for record in checkin_data}

    # Fetch employee details for employees who checked in today
    employees_data = Employs.objects.filter(email__in=empid_status_data.keys())

    # Update employee data with status
    for employee in employees_data:
        employee.status = empid_status_data.get(employee.email, 'N/A')

    items_per_page = 10  
    paginator = Paginator(employees_data, items_per_page)
    page = request.GET.get('page')

    try:
        employees_data = paginator.page(page)
    except PageNotAnInteger:
        employees_data = paginator.page(1)
    except EmptyPage:
        employees_data = paginator.page(paginator.num_pages)
        
    user=CustomUser.objects.filter(id=request.user.id).first()
    userid1=user.id
    employ_count1=Employs.objects.all().count()
    # data = AdminHod.objects.filter(admin=request.user.id,options=0)
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    s=adminnav.objects.filter(is_projectmanager = 1)
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    a=company_details.objects.filter(companyid=data1).first()

    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    da1=Employs.objects.filter(admin=request.user.id).first()
    if da1:
      da2=da1.id
    else:
        da2=None
    projectm=admin_project_create.objects.filter(admin_id= userid1)
    h=HR.objects.all()
    employs_all=Employs.objects.all()

    return render(request, "admin-template/checkin_count.html", {
        'employee_data': employees_data,
        'today_date': today_date,
        'user':user,
        'da1':da1,
        'da2':da2,
        'admin_drops':admin_drops,
        'a':a,
        'admin_home_drops':admin_home_drops,
        's':s,
        'h':h,
        'projectm':projectm,
        'projects_drops':projects_drops,
        'data':data,
        'employs_all':employs_all,

    })


def working(request):
    if request.method=="POST":
        starting_time=request.POST['starting_time']
        ending_time=request.POST['ending_time']
        cutoff_time=request.POST['cutoff_time']
        shift_name = request.POST['shift_name']        
        k=working_shifts(starting_time=starting_time,ending_time=ending_time,cutoff_time=cutoff_time,shift_name=shift_name)
        k.save();
    return render(request,'admin-template/admin_shifts.html')
   

def update_payroll_setup(request):
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = AdminHod.objects.filter(admin=request.user.id).first()
    # da2 = da1.id
    # admin = AdminHod.objects.get(id=1)
    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    # data = AdminHod.objects.filter(id=request.user.id)
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')

    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()




    existing_payroll_data = set_payroll_date.objects.first()  # Retrieve the existing data

    if request.method == "POST":
        auto_run_payroll = request.POST.get('auto_run_payroll')
        advance_salary_request = request.POST.get('advance_salary_request')
        payrolldate = request.POST.get('payrolldate')
        payroll = request.POST.get('payroll')
        existing_payroll_data=set_payroll_date.objects.first()
        existing_payroll_data.auto_run_payroll = auto_run_payroll
        existing_payroll_data.advance_salary_request = advance_salary_request
        existing_payroll_data.payrolldate = payrolldate
        existing_payroll_data.payroll = payroll

        if payroll == "on":
            existing_payroll_data.payroll = 'Yes'
        else:
            existing_payroll_data.payroll = 'No'

        if auto_run_payroll == 'on':
            existing_payroll_data.auto_run_payroll = 'Yes'
        else:
            existing_payroll_data.auto_run_payroll = 'No'

        if advance_salary_request == 'on':
            existing_payroll_data.advance_salary_request = 'Enabled'
        else:
            existing_payroll_data.advance_salary_request = 'Disabled'

        existing_payroll_data.save()
        messages.success(request,"your details were sucessfully submitted")

        return redirect('/setting/')
    
    return render(request, 'admin-template/update_payroll_setup.html', {
        'admin_drops': admin_drops,
        'existing_payroll_data': existing_payroll_data,  # Pass the existing data to the template
        'user':user,
        'da1':da1,
        'a':a,
        'admin_home_drops':admin_home_drops,
        's':s,
        'h':h,
        'projectm':projectm,
        'projects_drops':projects_drops,
        'data':data,
        'employs_all':employs_all

    })

def display_employee_records(request):
    compid=Companys.objects.filter(usernumber=request.user.id).first()
    today_date = datetime.now().date()
    checked_in_employee_ids = checkin.objects.filter(date=today_date,companyid=compid).values_list('empid', flat=True)
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')

    not_checkin_employees = Employs.objects.exclude(email__in=checked_in_employee_ids)



    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = Employs.objects.filter(admin=request.user.id).first()
    if da1:
      da2 = da1.id
    else:
        da2=None
    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops = admin_home_drop.objects.filter(parent_category=None).order_by('id')
    # data = AdminHod.objects.filter(id=request.user.id)
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id

    projects_drops = project_drop.objects.filter(parent_category=None).order_by('id')
    a=company_details.objects.filter(companyid=data1).first()

    data_to_display = not_checkin_employees.filter(companyid=compid)
    records_count = data_to_display.count()
    items_per_page = 10  # Adjust the number of items per page as needed
    page = request.GET.get('page')  
    paginator = Paginator(data_to_display, items_per_page)  

    try:
        data_to_display = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver the first page
        data_to_display = paginator.page(1)
    except EmptyPage:
        # If page is out of range, deliver the last page of results
        data_to_display = paginator.page(paginator.num_pages)
    

    return render(request, 'admin-template/list.html', {'data_to_display': data_to_display,
                                                         'admin_drops': admin_drops,
                                                         'user': user,
                                                         'da1': da1,
                                                         'da2': da2,
                                                         'a':a,
                                                         'admin_home_drops': admin_home_drops,
                                                         's': s,
                                                         'h': h,
                                                         'projectm': projectm,
                                                         'projects_drops': projects_drops,
                                                         'data': data,
                                                         'employs_all': employs_all,
                                                         'records_count': records_count})

# def display_employee_records(request):
#     today_date = datetime.now().date()
#     checked_in_employee_ids = checkin.objects.filter(date=today_date).values_list('empid', flat=True)
#     not_checkin_employees = Employs.objects.exclude(empid__in=checked_in_employee_ids)
    
#     records_count = not_checkin_employees.count()
    
#     records_count = not_checkin_employees.count()
#     admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
#     employ_records = Employs.objects.filter(status=0)
#     data_to_display = employ_records.values('first_name', 'contactno', 'empid')

#     # Count the number of records
#     records_count = employ_records.count()

#     user = CustomUser.objects.filter(id=request.user.id).first()
#     userid1 = user.id
#     da1 = AdminHod.objects.filter(admin=request.user.id).first()
#     da2 = da1.id
#     admin = AdminHod.objects.get(id=1)
#     s = adminnav.objects.all()
#     projectm = admin_project_create.objects.filter(admin_id=userid1)
#     h = HR.objects.all()
#     employs_all = Employs.objects.all()
#     admin_home_drops = admin_home_drop.objects.filter(parent_category=None).order_by('id')
#     data = AdminHod.objects.filter(id=request.user.id)
#     projects_drops = project_drop.objects.filter(parent_category=None).order_by('id')

#     return render(request, 'admin-template/list.html', {'data_to_display': data_to_display,
#                                                          'admin_drops': admin_drops,
#                                                          'user': user,
#                                                          'da1': da1,
#                                                          'da2': da2,
#                                                          'admin': admin,
#                                                          'admin_home_drops': admin_home_drops,
#                                                          's': s,
#                                                          'h': h,
#                                                          'projectm': projectm,
#                                                          'projects_drops': projects_drops,
#                                                          'data': data,
#                                                          'employs_all': employs_all,
#                                                          'records_count': records_count})





def admin_basic_details_mm(request):
    s=adminnav.objects.all()
    if request.method == "POST":
     
        Organisationtype=request.POST["Organisationtype"]
        companypan=request.POST["companypan"]
        companyname=request.POST["companyname"]
        companyGSTIN=request.POST["companyGSTIN"]
        brandname=request.POST["brandname"]
        registeraddress=request.POST["registeraddress"]
        address=request.POST["address"]
        State=request.POST["State"]
        pincode=request.POST["pincode"]
        payrolldate=request.POST["payrolldate"]
        auto_run_payroll=request.POST["auto_run_payroll"]
        advance_salary_request=request.POST["advance_salary_request"]
        FBP_allowances=request.POST["FBP_allowances"]
        progress =  20 
        value_message=3
        val=4
          # Example progress value for form 1
        progress_obj, _ = Progress.objects.get_or_create(pk=1, defaults={'value': 0})
       
        progress_obj.value = value_message
        progress_obj.progress_form1 = progress
        
        progress_obj.save()
        user_obj=admin_basic_details(Organisationtype=Organisationtype,companypan=companypan,companyname=companyname,companyGSTIN=companyGSTIN,brandname=brandname,registeraddress=registeraddress,address=address,State=State,pincode=pincode,payrolldate=payrolldate,auto_run_payroll=auto_run_payroll,advance_salary_request=advance_salary_request,FBP_allowances=FBP_allowances)
        instance=admin_home_drop.objects.get(id=1)
        instance.progress_value=val
        instance.save()
        
        user_obj.save()
    return render(request,"hod_template/admin_basic_details.html",{'s':s})
    

def admin_apply_leave(request):
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)
    staff_obj = AdminHod.objects.get(admin=request.user.id)
    leave_data=LeaveReportAdmin.objects.filter(admin_id=staff_obj)
    return render(request,"admin-template/admin_apply_leave.html",{"leave_data":leave_data,'s':s})

def admin_apply_leave_save(request):
    if request.method!="POST":
        return HttpResponseRedirect(reverse("admin_apply_leave"))
    else:
        employ_obj=AdminHod.objects.get(admin=request.user.id)
        leave_date=request.POST.get("leave_date")
        leave_type=request.POST.get("leave_type")
        to_date=request.POST.get("to_date")
        leave_msg=request.POST.get("leave_msg")

        
        # try:
        leave_report=LeaveReportAdmin(admin_id=employ_obj.id,leave_type=leave_type,leave_date=leave_date,to_date=to_date,leave_message=leave_msg,leave_status=0)
        leave_report.save()
        messages.success(request, "Successfully Applied for Leave")
        return HttpResponseRedirect(reverse("admin_apply_leave"))
        # except:
        #     messages.error(request, "Failed To Apply for Leave")
        #     return HttpResponseRedirect(reverse("admin_apply_leave"))

def admin_approve(request,leave_id):
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    leave=LeaveReportAdmin.objects.get(id=leave_id)
    leave.leave_status=1
    leave.save()
    return HttpResponseRedirect(reverse("admin_leave_view"))

def admin_disapprove(request,leave_id):
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    leave=LeaveReportAdmin.objects.get(id=leave_id)
    leave.leave_status=2
    leave.save()
    return HttpResponseRedirect(reverse("admin_leave_view"))


def admin_leave_view(request):
    leaves=LeaveReportAdmin.objects.all()
    s=adminnav.objects.all()
    return render(request,"admin-template/admin_leave_view.html",{"leaves":leaves,'s':s})


def admin_basic_detailss(request):
    objs=AdminHod.objects.get(admin=request.user.id)
    users=admin_basic_details.objects.get(id=request.user.id)

    return render(request,"admin-template/admin_basic_details.html",{"users":users})

def admin_basic_details_save(request):
    if request.method!="POST":
        return HttpResponseRedirect(reverse("admin_basic_detailss"))
    else:    
        Organisationtype=request.POST.get("Organisationtype")
        companypan=request.POST.get("companypan")
        companyname=request.POST.get("companyname")
        companyGSTIN=request.POST.get("companyGSTIN")
        brandname=request.POST.get("brandname")
        registeraddress=request.POST.get("registeraddress")
        address=request.POST.get("address")
        State=request.POST.get("State")
        pincode=request.POST.get("pincode")
        payrolldate=request.POST.get("payrolldate")
        auto_run_payroll=request.POST.get("auto_run_payroll")
        advance_salary_request=request.POST.get("advance_salary_request")
        FBP_allowances=request.POST.get("FBP_allowances")
        
        
        try:
        
            user_subobj=admin_basic_details.objects.get(id=request.user.id)
            user_subobj.Organisationtype=Organisationtype
            user_subobj.companypan=companypan
            user_subobj.companyname=companyname
            user_subobj.companyGSTIN=companyGSTIN
            user_subobj.brandname=brandname
            user_subobj.registeraddress=registeraddress
            user_subobj.address=address
            user_subobj.State=State
            user_subobj.pincode=pincode
            user_subobj.payrolldate=payrolldate
            user_subobj.auto_run_payroll=auto_run_payroll
            user_subobj.advance_salary_request=advance_salary_request
            user_subobj.FBP_allowances=FBP_allowances
            user_subobj.save()
            return HttpResponseRedirect(reverse("admin_basic_detailss"))
        except:
            
            return HttpResponseRedirect(reverse("admin_basic_detailss"))



@csrf_exempt
def check_email_exist(request):
    email=request.POST.get("email")
    user_obj=CustomUser.objects.filter(email=email).exists()
    if user_obj:
        return HttpResponse(True)
    else:
        return HttpResponse(False)

@csrf_exempt
def check_username_exist(request):
    username=request.POST.get("username")
    user_obj=CustomUser.objects.filter(username=username).exists()
    if user_obj:
        return HttpResponse(True)
    else:
        return HttpResponse(False)

@csrf_exempt
def check_empid_exist(request):
    empid=request.POST.get("empid")
    user_obj=CustomUser.objects.filter(empid=empid).exists()
    if user_obj:
        return HttpResponse(True)
    else:
        return HttpResponse(False)

def admin_details(request):
    s=adminnav.objects.all()
    user=AdminHod.objects.get(admin=request.user.id)
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    if request.method=="POST":
        firstname1=request.POST["firstname1"]
        lastname1=request.POST["lastname1"]
        
        pan=request.POST["pan"]
        ifsecode=request.POST["ifsecode"]
        acno=request.POST["acno"]
        beneficiaryname=request.POST["beneficiaryname"]
        phno=request.POST["phno"]  

        profilepic=request.FILES.get("profilepic")
       
        dob=request.POST["dob"]
        address1=request.POST["address1"]
        heq=request.POST["heq"]
        aadharno=request.POST["aadharno"] 
        bloodgroup=request.POST["bloodgroup"] 
        student_obj=admin_add_form1.objects.get(admin_id=user)
        try:
            emp_instance,_=admin_add_form1.objects.get_or_create(admin_id=user)
            emp_instance.firstname1=firstname1
            emp_instance.lastname1=lastname1
            emp_instance.pan=pan
            emp_instance.ifsecode=ifsecode
            emp_instance.acno=acno
            emp_instance.beneficiaryname=beneficiaryname
            emp_instance.phno=phno
            emp_instance.dob=dob
            emp_instance.address1=address1
            emp_instance.heq=heq
            emp_instance.aadharno=aadharno
            emp_instance.bloodgroup=bloodgroup
            emp_instance.profilepic=profilepic
            emp_instance.save()
        except admin_add_form1.DoesNotExist:
            k=admin_add_form1(admin=user,firstname1=firstname1,lastname1=lastname1,email2=student_obj.email,pan=pan,ifsecode=ifsecode,acno=acno,beneficiaryname=beneficiaryname,phno=phno,gender1=1,dob=dob,address1=address1,heq=heq,aadharno=aadharno,bloodgroup=bloodgroup,profilepic=profilepic)
            k.save();
        # return HttpResponseRedirect(reverse("/"))
        return HttpResponseRedirect(reverse("admin_profile"))
    return render(request,"admin-template/addadmindetails.html",{'s':s,'admin_drops':admin_drops})  


def admin_profile(request):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    # admin = AdminHod.objects.get(admin=request.user.id) 
    user_obj = Companys.objects.filter(usernumber=request.user.id)  # Define user_obj here
    # datas = admin_add_form1.objects.filter(admin=user_obj)

    user = CustomUser.objects.filter(id=request.user.id).first()
    da1 = AdminHod.objects.filter(admin=request.user.id).first()
    s = adminnav.objects.all()
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()

   
    return render(request, "admin-template/admin_account.html", {'a':a,'admin_drops':admin_drops,'user':user,'da1':da1,'admin_home_drops':admin_home_drops,'s':s,'h':h,'projects_drops':projects_drops,'data':data,'employs_all':employs_all,'user_obj':user_obj})




def admin_profile_save(request):
     if request.method!="POST":
        return HttpResponseRedirect(reverse("admin_profile"))
     else:
        first_name=request.POST.get("first_name")
        last_name=request.POST.get("last_name")
        password=request.POST.get("password")
        location = request.POST.get("location")


        
        try:
            customuser=CustomUser.objects.get(id=request.user.id)
            customuser.first_name=first_name
            customuser.last_name=last_name
            if password!=None and password!="":
                customuser.set_password(password)
            customuser.save()

            employ=AdminHod.objects.get(admin=customuser)
            employ.first_name=first_name
            employ.last_name=last_name
            employ.location = location  

            employ.save()
            messages.success(request, "Successfully Updated Profile")
            return HttpResponseRedirect(reverse("admin_profile"))
        except:
            messages.error(request, "Failed to Update Profile")
            return HttpResponseRedirect(reverse("admin_profile"))

def admin_upload_photo(request):
    user = CustomUser.objects.filter(id=request.user.id).first()

    # userid1 = user.id
    da1 = Employs.objects.filter(admin=request.user.id).first()
    if da1:
      da2 = da1.id
    else:
        da2=None
    s = adminnav.objects.all()
    # projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    # employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    # projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()

    if request.method == "POST":
        adminprofilepic = request.FILES.get("adminprofilepic")
        if adminprofilepic:
            prof=Companys.objects.get(usernumber=request.user.id)
            prof.profilepic=adminprofilepic
            prof.save()
            return redirect('/admin_home')
        else:
            messages.error(request, "Profile picture upload failed.")

    return render(request, "admin-template/adminprofilepic.html", {'a':a,'admin_drops':admin_drops,'user':user,'da1':da1,'da2':da2,'admin_home_drops':admin_home_drops,'s':s,'h':h,'data':data})


# def admin_upload_photo(request):
#     user_obj = AdminHod.objects.get(admin=request.user.id)
#     admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
#     # data = AdminHod.objects.get(admin=request.user.id)

#     user = CustomUser.objects.filter(id=request.user.id).first()
#     userid1 = user.id
#     da1 = AdminHod.objects.filter(admin=request.user.id).first()
#     da2 = da1.id
#     admin = AdminHod.objects.get(id=1)
#     s = adminnav.objects.all()
#     projectm = admin_project_create.objects.filter(admin_id=userid1)
#     h = HR.objects.all()
#     employs_all = Employs.objects.all()
#     admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
#     data = AdminHod.objects.filter(id=request.user.id)
#     projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')


    

#     if request.method == "POST":
#         adminprofilepic = request.FILES.get("adminprofilepic")
#         if adminprofilepic:
           
#             data.adminprofilepic = adminprofilepic
            
#                 # Create a new instance if it doesn't exist
#             data,_ = AdminHod.objects.get_or_create(admin=request.user.id)
#             data.adminprofilepic=adminprofilepic

#             data.save()
#             return redirect('/admin_home')
#         else:
#             messages.error(request, "Profile picture upload failed.")

#     return render(request, "admin-template/adminprofilepic.html", {'data': data,'admin_drops':admin_drops,'user':user,'da1':da1,'da2':da2,'admin':admin,'admin_home_drops':admin_home_drops,'s':s,'h':h,'projectm':projectm,'projects_drops':projects_drops,'data':data,'employs_all':employs_all})




def admindelete_profile_pic(request):
    if request.method == "POST":
        user_obj = AdminHod.objects.get(admin=request.user.id)
        data = AdminHod.objects.get(admin=request.user.id)

        # Delete the profile picture and save the object
        if data.adminprofilepic:
            data.adminprofilepic.delete()
            data.save()
            messages.success(request, "Profile picture removed successfully.")
        else:
            messages.warning(request, "No profile picture to remove.")

    return redirect('admin_home')



# def admindelete_profile_pic(request):
#     user_obj = AdminHod.objects.get(admin=request.user.id)
#     try:
#         data = AdminHod.objects.get(admin=user_obj.id)
#         data.adminprofilepic.delete()  # Delete the profile picture file
#         data.adminprofilepic = None  # Clear the profilepic field
#         data.save()
#         messages.success(request, "Profile picture removed successfully.")
#     except AdminHod.DoesNotExist:
#         messages.error(request, "User data not found. Please add user data first.")
    
#     return redirect("admin_upload_photo")



def add_admin(request):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    s=adminnav.objects.all()
    return render(request,"admin-template/admin_profile.html",{'s':s,'admin_drops':admin_drops})


def add_admin_save(request):
    if request.method!="POST":
        return HttpResponse("Method Not Allowed")
    else:
        
        if request.method=="POST":
            first_name=request.POST["first_name"]
            last_name=request.POST["last_name"]
            email=request.POST["email"]
            empid=request.POST["empid"]
            phnos=request.POST["phnos"]
            address=request.POST["address"]
            dateofbirth=request.POST["dateofbirth"]
            dateofjoining=request.POST["dateofjoining"]
            department=request.POST["department"]
            manager=request.POST["manager"]
            location=request.POST["location"]
            package=request.POST['package']
            pincode=request.POST["pincode"]
            designation=request.POST["designation"]
            status=request.POST["status"]
           
            bloodgroup=request.POST["bloodgroup"]
            admin_profilepic=request.FILES.get('admin_profilepic')
            gender=request.POST["gender"]
            

            # try:
            user=CustomUser.objects.create_user(username=first_name,password=last_name,email=email,user_type=1)
            user.adminhod.firstname=first_name
            user.adminhod.lastname=last_name
            user.adminhod.email=email
            user.adminhod.empid=empid
            user.adminhod.phnos=phnos
            user.adminhod.address=address
            user.adminhod.dateofbirth=dateofbirth
            user.adminhod.dateofjoining=dateofjoining
            user.adminhod.Department=department
            user.adminhod.manager=manager
            user.adminhod.location=location
            user.adminhod.package=package
            user.adminhod.pincode=pincode
            user.adminhod.designation=designation
            user.adminhod.status=status
           
            user.adminhod.bloodgroup=bloodgroup
            user.adminhod.gender=gender
            user.adminhod.adminprofilepic=admin_profilepic
            user.save()
            html_content = render_to_string("email_template.html",{'title':'test email','first_name':first_name,'last_name':last_name,'empid':empid})
            text_content=strip_tags(html_content)
            subject="WELCOME TO DEVELOPTRESS"
            email = EmailMultiAlternatives(
                subject,
                text_content,
                settings.EMAIL_HOST_USER,
                [email],
                )
            email.attach_alternative(html_content,"text/html")
            email.fail_silently = True
            email.send()
            messages.success(request,"Successfully Added Admin")
            return redirect("/add_admin")
        return render(request,"data is inserted")

from django.core.exceptions import ObjectDoesNotExist, MultipleObjectsReturned  # Import MultipleObjectsReturned
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.shortcuts import render, redirect
from django.contrib import messages
from django.conf import settings
from .models import CustomUser, Progress, admin_home_drop, adminnav, Employs, HR, list, admin_drop


def add_employee(request):
    li=list.objects.all()

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = AdminHod.objects.filter(admin=request.user.id).first()
    if da1:
       da2 = da1.id
    else:
        da2=None
    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    companyid=Companys.objects.filter(usernumber=request.user.id).first()

    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()

    existing_user_email = None
    existing_user_username = None

    if request.method == "POST":
        email = request.POST.get("email")
        username = request.POST.get("username")

        if email:
            try:
                existing_user_email = CustomUser.objects.get(email=email)
                messages.info(request, 'Email already exists')
                return redirect("/add_employee")  # Redirect here to avoid processing the rest of the view
            except ObjectDoesNotExist:
                existing_user_email = None
            except MultipleObjectsReturned:
                messages.error(request, ' Email already exists.')
                return redirect("/add_employee")  # Redirect here to avoid processing the rest of the view

        if username:
            try:
                existing_user_username = CustomUser.objects.get(username=username)
                messages.error(request, 'Username already exists')
                return redirect("/add_employee")  # Redirect here to avoid processing the rest of the view
            except ObjectDoesNotExist:
                existing_user_username = None
            except MultipleObjectsReturned:
                messages.error(request, 'Both username and email already exist. Contact the administrator.')
                return redirect("/add_employee")  # Redirect here to avoid processing the rest of the view

    if request.method == "POST":
        first_name = request.POST["first_name"]
        last_name = request.POST["last_name"]
        username = request.POST["username"]
        email = request.POST["email"]
        web_mail = request.POST["web_mail"]
        password = request.POST["password"]
        address = request.POST["address"]
        empid = request.POST["empid"]
        designation = request.POST.get("designation")
        location = request.POST["location"]
        package = request.POST["package"]
        pincode = request.POST["pincode"]
        contactno = request.POST["contactno"]
        bloodgroup = request.POST.get("bloodgroup")
        dateofjoining = request.POST['dateofjoining']
        sex = request.POST["gender"]
        role = request.POST.get('role') # Added field for role
    
        if role == "Employ":
            user = CustomUser.objects.create_user(username=username, password=password, email=email, user_type=2)
            user.employs.first_name = first_name
            user.employs.last_name = last_name
            user.employs.email = email
            user.employs.password = password
            user.employs.address = address
            user.employs.empid = empid
            user.employs.web_mail = web_mail
            user.employs.designation = designation
            user.employs.location = location
            user.employs.package = package
           
            user.employs.pincode = pincode
            user.employs.contactno = contactno
            user.employs.bloodgroup = bloodgroup
            user.employs.dateofjoining = dateofjoining
            user.employs.role = role  # Set the role field for HR
            user.employs.hroptions=0
            user.employs.companyid=companyid
            user.save()
            messages.success(request, "Successfully Added Employee")
                
        elif role == "HR":
            user = CustomUser.objects.create_user(username=username, password=password, email=email, user_type=2)
            user.employs.first_name = first_name
            user.employs.last_name = last_name
            user.employs.email = email
            user.employs.password = password
            user.employs.address = address
            user.employs.empid = empid
            user.employs.web_mail = web_mail
            user.employs.designation = designation
            user.employs.location = location
            user.employs.package = package
          
            user.employs.pincode = pincode
            user.employs.contactno = contactno
            user.employs.bloodgroup = bloodgroup
            user.employs.dateofjoining = dateofjoining
            user.employs.role = role  # Set the role field for HR
            user.employs.hroptions=1
            user.employs.companyid=companyid
            user.save()
            messages.success(request, "Successfully Added HR")
        else:
            user = CustomUser.objects.create_user(username=username, password=password, email=email, user_type=2)
            user.employs.first_name = first_name
            user.employs.last_name = last_name
            user.employs.email = email
            user.employs.password = password
            user.employs.address = address
            user.employs.empid = empid
            user.employs.web_mail = web_mail
            user.employs.designation = designation
            user.employs.location = location
            
            user.employs.package = package
            user.employs.pincode = pincode
            user.employs.contactno = contactno
            user.employs.bloodgroup = bloodgroup
            user.employs.dateofjoining = dateofjoining
            user.employs.role = role  # Set the role field for HR
            user.employs.hroptions=0
            user.employs.projectmanagerop=1
            user.employs.companyid=companyid
            user.save()
            messages.success(request, "Successfully Added ProjectManager")
            

       

       
        html_content = render_to_string("email_template.html", {'title': 'test email', 'first_name': first_name,
                                                                'last_name': last_name, 'empid': empid})
        text_content = strip_tags(html_content)
        subject = "WELCOME TO DEVELOPTRESS"
        email = EmailMultiAlternatives(
            subject,
            text_content,
            settings.EMAIL_HOST_USER,
            [email],
        )
        email.attach_alternative(html_content, "text/html")
        email.fail_silently = True
        email.send()
        progress = 20
        val = 4
        value_message = 1
        progress_obj, _ = Progress.objects.get_or_create(pk=1, defaults={'value': 0})

        progress_obj.value3 = value_message
        progress_obj.progress_form4 = progress

        progress_obj.save()
        instance = admin_home_drop.objects.get(id=5)
        instance.progress_value = val
        instance.save()
        
        return redirect("/add_employee")

    # If the request method is not POST, render the form
    s = adminnav.objects.all()
    form = Employs.objects.all()
    a1=working_shifts.objects.all()
    hrform = HR.objects.all()
    role = list.objects.all()
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    return render(request, "admin-template/add_employ_template.html",
                  {'role': role, 'hrform': hrform, "form": form,'a':a, 's': s, 'admin_drops': admin_drops,'li':li,'a1':a1,'admin_drops':admin_drops,'user':user,'da1':da1,'da2':da2,'admin_home_drops':admin_home_drops,'s':s,'h':h,'projectm':projectm,'projects_drops':projects_drops,'employs_all':employs_all,'data':data})


def manage_employ(request):
    employs=Employs.objects.all()
    s=adminnav.objects.all()
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    return render(request,"admin-template/manage_employ_template.html",{"employs":employs,'s':s,'admin_drops':admin_drops})


def edit_employ(request,student_id):
    request.session['student_id']=student_id
    student=Employs.objects.get(admin=student_id)
    form=EditEmployForm()
    form.fields['email'].initial=student.admin.email
    form.fields['first_name'].initial=student.admin.first_name
    form.fields['last_name'].initial=student.admin.last_name
    form.fields['username'].initial=student.admin.username
    form.fields['address'].initial=student.address
    form.fields['empid'].initial=student.empid
    form.fields['Manager'].initial=student.Manager
    form.fields['role'].initial=student.role
    form.fields['location'].initial=student.location
    form.fields['package'].initial=student.package
    form.fields['pincode'].initial=student.pincode
    form.fields['contactno'].initial=student.contactno
    form.fields['sex'].initial=student.gender
  
    return render(request,"admin-template/edit_employ_template.html",{"form":form,"id":student_id,"username":student.admin.username})

def edit_employ_save(request):
    if request.method!="POST":
        return HttpResponse("<h2>Method Not Allowed</h2>")
    else:
        student_id=request.session.get("student_id")
        if student_id==None:
            return HttpResponseRedirect(reverse("manage_student"))

        form=EditEmployForm(request.POST,request.FILES)
        if form.is_valid():
            first_name = form.cleaned_data["first_name"]
            last_name = form.cleaned_data["last_name"]
            username = form.cleaned_data["username"]
            email = form.cleaned_data["email"]
            address = form.cleaned_data["address"]
            empid = form.cleaned_data["empid"]
            Manager = form.cleaned_data["Manager"]
            role = form.cleaned_data["role"]
            location = form.cleaned_data["location"]
            package = form.cleaned_data["package"]
            pincode = form.cleaned_data["pincode"]
            contactno = form.cleaned_data["contactno"]
            sex = form.cleaned_data["sex"]

            if request.FILES.get('profile_pic',False):
                profile_pic=request.FILES['profile_pic']
                fs=FileSystemStorage()
                filename=fs.save(profile_pic.name,profile_pic)
                profile_pic_url=fs.url(filename)
            else:
                profile_pic_url=None


            try:
                user=CustomUser.objects.get(id=student_id)
                user.first_name=first_name
                user.last_name=last_name
                user.username=username
                user.email=email
                user.save()

                student=Employs.objects.get(admin=student_id)
                student.address=address
                student.empid=empid
                student.role=role
                student.location=location
                student.package=package
                student.pincode=pincode
                student.contactno=contactno
                student.gender=sex
              
                if profile_pic_url!=None:
                    student.profile_pic=profile_pic_url
                student.save()
                del request.session['student_id']
                messages.success(request,"Successfully Edited employ")
                return HttpResponseRedirect(reverse("edit_employ",kwargs={"student_id":student_id}))
            except:
                messages.error(request,"Failed to Edit employ")
                return HttpResponseRedirect(reverse("edit_employ",kwargs={"student_id":student_id}))
        else:
            form=EditEmployForm(request.POST)
            student=Employs.objects.get(admin=student_id)
            return render(request,"admin-template/edit_employ_template.html",{"form":form,"id":student_id,"username":student.admin.username})

def employ_leave_view(request):

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = Employs.objects.filter(admin=request.user.id).first()
    if da1:
      da2 = da1.id
    else:
        da2=None
    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    a=company_details.objects.filter(companyid=data1).first()



    leaves=LeaveReportEmploy.objects.filter(employ_id__companyid=data1)
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    leave_applicants_count = LeaveReportEmploy.objects.filter(employ_id__companyid=data1).count()
    tomorrow_date = date.today() + timedelta(days=1)
    approved_leave_count = LeaveReportEmploy.objects.filter(employ_id__companyid=data1, leave_status=1).count()
    pending_leave_count = LeaveReportEmploy.objects.filter(employ_id__companyid=data1, leave_status=0).count()
    total_leave_count = approved_leave_count + pending_leave_count
    tomorrow_leave_applicants_count = LeaveReportEmploy.objects.filter(employ_id__companyid=data1, leave_date=tomorrow_date).count()

    items_per_page = 10  # Adjust the number of items per page as needed

    paginator = Paginator(leaves, items_per_page)
    page = request.GET.get('page')

    try:
        leaves = paginator.page(page)
    except PageNotAnInteger:
        leaves = paginator.page(1)
    except EmptyPage:
       leaves = paginator.page(paginator.num_pages)

    return render(request,"admin-template/employ_leave_view.html",{"leaves":leaves,'admin_drops':admin_drops,'a':a, 'leave_applicants_count':leave_applicants_count,'total_leave_count':total_leave_count,'tomorrow_leave_applicants_count':tomorrow_leave_applicants_count,'approved_leave_count':approved_leave_count,'pending_leave_count':pending_leave_count,'user':user,'da1':da1,'da2':da2,'admin_home_drops':admin_home_drops,'s':s,'h':h,'projectm':projectm,'projects_drops':projects_drops,'data':data,'employs_all':employs_all})





def admin_payslip_view(request):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = Employs.objects.filter(admin=request.user.id).first()
    if da1:
      da2 = da1.id
    else:
        da2=None
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')


    # s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)
    s=adminnav.objects.all()
    employ = Employs.objects.all()
    leaves = payslip_request.objects.filter(student_id__companyid=data1)

    total_count = leaves.count()
    approved_count = leaves.filter(status=1).count()
    pending_count = leaves.filter(status=0).count()
    reject_count = leaves.filter(status=2).count()
    items_per_page = 10  
    paginator = Paginator(leaves, items_per_page)
    page = request.GET.get('page')

    try:
        leaves = paginator.page(page)
    except PageNotAnInteger:
        leaves = paginator.page(1)
    except EmptyPage:
        leaves = paginator.page(paginator.num_pages)

    context = {
        'approved_count': approved_count,
        'pending_count': pending_count,
        'reject_count': reject_count,
        'total_count': total_count,
        'leaves': leaves,
        'admin_drops': admin_drops,
        'employ': employ,
        's':s,
        'projects_drops':projects_drops,
        'a':a,
        'user':user,
        'da1':da1,
        'da2':da2,
        'data':data,
        'admin_home_drops':admin_home_drops,
        'employs_all':employs_all,
        'h':h,
    
        'projectm':projectm,
    }

    return render(request, "admin-template/payslip_request_status.html", context)


# def employ_leave_view(request):
#     leaves=LeaveReportEmploy.objects.all()
    
#     return render(request,"admin-template/employ_leave_view.html",{"leaves":leaves})

# def employ_approve_leave(request,leave_id):
#     leave=LeaveReportEmploy.objects.get(id=leave_id)
#     leave.leave_status=1
#     leave.save()
#     return HttpResponseRedirect(reverse("employ_leave_view"))
#     leave=LeaveReportEmploy.objects.get(id=leave_id)
#     leave.leave_status=2
#     leave.save()
#     return HttpResponseRedirect(reverse("employ_leave_view"))

def reimbursement_status_view(request):
    s=adminnav.objects.all()
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    employ=Employs.objects.all()
    leaves=Reimbursement.objects.all()
    return render(request,"admin-template/reimbursement_status_view.html",{"admin_drops":admin_drops,"leaves":leaves,'s':s,"employ":employ})

def reimbursement_approve_status(request,leave_id):
    leave=Reimbursement.objects.get(id=leave_id)
    leave.reimbursement_status=1
    leave.save()
    return HttpResponseRedirect(reverse("reimbursement_status_view"))

def reimbursement_disapprove_status(request,leave_id):
    leave=Reimbursement.objects.get(id=leave_id)
    leave.reimbursement_status=2
    leave.save()
    return HttpResponseRedirect(reverse("reimbursement_status_view"))

def advance_salary_status(request):
    s=adminnav.objects.all()
    employ=Employs.objects.all()
    leaves=ad_salary.objects.all()
    return render(request,"admin-template/advance_salary_status.html",{'s':s,'employ':employ,'leaves':leaves})

def advance_salary_status_approve(request,leave_id):
    leave=ad_salary.objects.get(id=leave_id)
    leave.request_status=1
    leave.save()
    return HttpResponseRedirect(reverse("admin_advancesalary_apply_view"))

def advance_salary_status_disapprove(request,leave_id):
    leave=ad_salary.objects.get(id=leave_id)
    leave.request_status=2
    leave.save()
    return HttpResponseRedirect(reverse("admin_advancesalary_apply_view"))


def admin_send_notification_employ(request):
    employ=Employs.objects.all()
    s=adminnav.objects.all()
    return render(request,"admin-template/employ_notification.html",{"employ":employ,"s":s})

@csrf_exempt
def send_employ_notification(request):
    id=request.POST.get("id")
    message=request.POST.get("message")
    employ=Employs.objects.get(admin=id)
    token=Employs.fcm_token
    url="https://fcm.googleapis.com/fcm/send"
    body={
        "notification":{
            "title":"Developtreeshrms",
            "body":message,
            "click_action": "https://studentmanagementsystem22.herokuapp.com/employ_all_notification",
            "icon": "http://studentmanagementsystem22.herokuapp.com/static/dist/img/user2-160x160.jpg"
        },
        "to":token
    }
    headers={"Content-Type":"application/json","Authorization":"key=SERVER_KEY_HERE"}
    data=requests.post(url,data=json.dumps(body),headers=headers)
    notification=NotificationEmploy(employ_id=employ,message=message)
    notification.save()
    print(data.text)
    return HttpResponse("True")

def reports(request):
    s=adminnav.objects.all()
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()

    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    return render(request,"admin-template/reports.html",{'a':a,'data':data,'s':s,'admin_drops':admin_drops,'admin_home_drops':admin_home_drops})

# def admin_reimbursement_apply_view(request):
#     projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')

#     user = AdminHod.objects.get(admin=request.user.id)

#     leave_data=Admin_Reimbursement.objects.filter(admin_id=user)
#     s=adminnav.objects.all()
#     k3=reimbursementsetup1.objects.all()

#     admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
#     admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
#     return render(request,"admin-template/reimbursement_apply_view.html",{'projects_drops':projects_drops,'s':s,'k3':k3,'admin_drops':admin_drops,'admin_home_drops':admin_home_drops,'leave_data':leave_data})


def admin_reimbursement_apply_view(request):
    
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)
    da1 = Employs.objects.filter(admin=request.user.id).first()

    if da1:
       da2 = da1.id
    else:
        da2=None
    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id

    # s = adminnav.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()

    # da1=AdminHod.objects.filter(admin=request.user.id).first()
    s=adminnav.objects.filter(is_projectmanager = 1)
    projectm=admin_project_create.objects.filter(admin_id= userid1)
    h=HR.objects.all()
   

    employ=empdocs.objects.all()
    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id

    # projectm = admin_project_create.objects.filter(admin_id=userid1)
    # h = HR.objects.all()
    # employs_all = Employs.objects.all()
  
    # data = AdminHod.objects.filter(id=request.user.id)
    r=documents_setup1.objects.all()
    # projectm = admin_project_create.objects.filter(admin_id=userid1)
    # h = HR.objects.all()
    # employs_all = Employs.objects.all()
    # data = AdminHod.objects.filter(id=request.user.id)
    staff_obj = Employs.objects.all()


    k = types.objects.all()

    st4 = request.POST.get("ss")
    st3 = request.POST.get("vk")
    st = request.POST.get("d1")
    st1 = request.POST.get("d2")

    leave_data = Reimbursement.objects.filter(employ_id__companyid=data1)

    # Apply filters based on user input
    if st4 and st4 != '----Select----':  # Check if a valid status is selected
        leave_data = leave_data.filter(reimbursement_status=st4)
    if st3:  # If "Select Type" is selected
        leave_data = leave_data.filter(typea__icontains=st3)
    if st and st1:
        leave_data = leave_data.filter(date__range=[st, st1])

    total_approved = leave_data.filter(reimbursement_status=1).aggregate(Sum('amount'))['amount__sum'] or 0
    total_pending = leave_data.filter(reimbursement_status=0).aggregate(Sum('amount'))['amount__sum'] or 0
    items_per_page = 10  # Adjust the number of items per page as needed

    paginator = Paginator(leave_data, items_per_page)
    page = request.GET.get('page')

    try:
        leave_data = paginator.page(page)
    except PageNotAnInteger:
        leave_data = paginator.page(1)
    except EmptyPage:
       leave_data = paginator.page(paginator.num_pages)
    return render(request, "admin-template/reimbursement_apply_view.html", {
        'leave_data': leave_data,
        's': s,
        'k': k,
        'total': total_approved,
        'total1': total_pending,
        'staff_obj':staff_obj,
        'admin_drops':admin_drops,
        'admin_home_drops':admin_home_drops,
        'user':user,
        'a':a,
        'data':data,
        'employs_all':employs_all,
        'h':h,
        'da2':da2, 
        'projectm':projectm,
        'employ':employ,


    })


def admin_reimbursement_apply_view_save(request):
    if request.method!="POST":
        return HttpResponseRedirect(reverse("admin_reimbursement_apply_view"))
    else:
        typea=request.POST.get("typea")
        date=request.POST.get("date")
        detail=request.POST.get("detail")
        amount=request.POST.get("amount")
        image=request.POST.get("image")

        student_obj=AdminHod.objects.get(admin=request.user.id)
        try:
            leave_report=Admin_Reimbursement(admin_id=student_obj,typea=typea,date=date,detail=detail,amount=amount,image=image,reimbursement_status=0)
            leave_report.save()
            messages.success(request, "Successfully Applied for Reimbursement")
            return HttpResponseRedirect(reverse("admin_reimbursement_apply_view"))
        except:
            messages.error(request, "Failed To Apply for Reimbursement")
            return HttpResponseRedirect(reverse("admin_reimbursement_apply_view"))


from django.shortcuts import render, redirect, get_object_or_404
from .models import reimbursementsetup1
from django.contrib import messages

def delete_reimbursementsetup1(request, reimbursement_id):
    # Use get_object_or_404 to get the reimbursement entry by its ID or return a 404 response if not found
    reimbursement = get_object_or_404(reimbursementsetup1, pk=reimbursement_id)

    # Delete the reimbursement entry
    reimbursement.delete()
    
    messages.success(request, "Your Reimbursement was successfully removed")  # Optional: Display a success message

    # Redirect to a page after the delete operation (e.g., a list of reimbursements)
    return redirect('/reimbursement_update') 


def update_document_setup1(request):

    k = documents_setup1.objects.all()
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = AdminHod.objects.filter(admin=request.user.id).first()
  
    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')

    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()


    if request.method == "POST":
        for r in k:
            document_type = r.document_type
            compulsory = 'Yes' if request.POST.get(f'compulsory_{r.id}') == 'on' else 'No'
            Enabled = 'Yes' if request.POST.get(f'Enabled_{r.id}') == 'on' else 'No'
            r.compulsory = compulsory
            r.Enabled = Enabled
            r.save()

        document_type = request.POST.get('document_type')
        compulsory = 'Yes' if 'compulsory' in request.POST else 'No'
        Enabled = 'Yes' if 'Enabled' in request.POST else 'No'

        k = documents_setup1( document_type=document_type, compulsory=compulsory, Enabled=Enabled)
        k.save()

        messages.success(request,"Your data was successfully added")
        return redirect('/setting/')

    return render(request, 'admin-template/document_setup_update1.html', {'a':a,'admin_drops':admin_drops,'admin_home_drops':admin_home_drops,'k': k,'user':user,'da1':da1,'admin_home_drops':admin_home_drops,'s':s,'h':h,'projectm':projectm,'projects_drops':projects_drops,'data':data,'employs_all':employs_all})


from django.shortcuts import get_object_or_404

def delete_document_setup1(request, document_id):
    # Use get_object_or_404 to get the object by its ID or return a 404 response if not found
    document = get_object_or_404(documents_setup1, pk=document_id)

    # Delete the document
    document.delete()
    messages.success(request,"Your data was successfully deleted")


    # Redirect to a page after the delete operation (e.g., a list of documents)
    return redirect('/update_document_setup1/')


def admin_reg(request):
    staff_obj = AdminHod.objects.get(admin=request.user.id)
    leave_data=Admin_Reimbursement.objects.filter(admin_id=staff_obj) 

    s=adminnav.objects.all()
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    k=types.objects.all()
    
    total=Admin_Reimbursement.objects.filter(reimbursement_status=1).aggregate(Sum('amount'))['amount__sum']
    total1=Admin_Reimbursement.objects.filter(reimbursement_status=0).aggregate(Sum('amount'))['amount__sum']
 
    st=request.POST.get("d1")
    st1=request.POST.get("d2")
    st3=request.POST.get("vk")
    st4=request.POST.get("ss")
    
    if(st4 and st and st3):
       leave_data=Admin_Reimbursement.objects.filter(reimbursement_status__icontains=st4).filter(typea__icontains=st3).filter(date__range=[st,st1]).values();
       return render(request,"admin-template/adminreim.html",{'leave_data':leave_data,'s':s,'k':k,'admin_drops':admin_drops,'admin_home_drops':admin_home_drops,'total':total,'total1':total1});
    elif(st4 and st3):
       leave_data=Admin_Reimbursement.objects.filter(reimbursement_status__icontains=st4).filter(typea__icontains=st3).values();
       return render(request,"admin-template/adminreim.html",{'leave_data':leave_data,'s':s,'k':k,'admin_drops':admin_drops,'admin_home_drops':admin_home_drops,'total':total,'total1':total1});
    
    elif(st4):
       leave_data=Admin_Reimbursement.objects.filter(reimbursement_status__icontains=st4).values();
       return render(request,"admin-template/adminreim.html",{'leave_data':leave_data,'s':s,'k':k,'admin_drops':admin_drops,'admin_home_drops':admin_home_drops,'total':total,'total1':total1});
    elif(st3 and st):
       leave_data=Admin_Reimbursement.objects.filter(typea__icontains=st3).filter(date__range=[st,st1]).values();
       return render(request,"admin-template/adminreim.html",{'leave_data':leave_data,'s':s,'k':k,'admin_drops':admin_drops,'admin_home_drops':admin_home_drops,'total':total,'total1':total1});
    elif(st3):
       leave_data=Admin_Reimbursement.objects.filter(typea__icontains=st3).values();
       return render(request,"admin-template/adminreim.html",{'leave_data':leave_data,'s':s,'k':k,'admin_drops':admin_drops,'admin_home_drops':admin_home_drops,'total':total,'total1':total1})
    elif(st):
        leave_data=Admin_Reimbursement.objects.filter(date__range=[st,st1]).values();
        return render(request,"admin-template/adminreim.html",{'leave_data':leave_data,'s':s,'k':k,'admin_drops':admin_drops,'admin_home_drops':admin_home_drops,'total':total,'total1':total1})

    else:
       return render(request,"admin-template/adminreim.html",{'leave_data':leave_data,'s':s,'k':k,'admin_drops':admin_drops,'admin_home_drops':admin_home_drops,'total':total,'total1':total1})
   
def reim_delete(request,id):
    k=Admin_Reimbursement.objects.get(id=id)
    k.delete()
    return redirect("/admin_reg")

def reim_edit(request,id):
    ks=Admin_Reimbursement.objects.get(id=id)
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    k=types.objects.all()
    return render(request,"admin-template/reim_update.html",{'ks':ks,'k':k,'admin_drops':admin_drops})

def reim_update(request,id):
    if request.method=="POST":
        reimbursement_status=request.POST["reimbursement_status"]
        date=request.POST["date"]
        detail=request.POST["detail"]
        typea=request.POST["typea"]
        amount=request.POST["amount"]
        image=request.FILES.get('image')
        k=Admin_Reimbursement.objects.get(id=id);
        k.reimbursement_status=reimbursement_status
        k.date=date
        k.detail=detail
        k.typea=typea
        k.amount=amount
        k.image=image
        k.save();
        return redirect("/admin_reg")
    return render(request,"admin-template/reim_update.html")



from django.shortcuts import render
from .models import payslip_request, Employs
# def admin_payslip_view(request):
#     admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
#     s=adminnav.objects.all()
#     employ=Employs.objects.all()
#     leaves=payslip_request.objects.all()
#     return render(request,"admin-template/payslip_request_status.html",{"admin_drops":admin_drops,"leaves":leaves,'s':s,'employ':employ})



def admin_payslip_approve_status(request,leave_id):
    leave=payslip_request.objects.get(id=leave_id)
    leave.status=1
    leave.save()
    return HttpResponseRedirect(reverse("admin_payslip_view"))

from django.shortcuts import render, HttpResponseRedirect, reverse

def admin_payslip_disapprove_status(request, leave_id):

    a=payslip_request.objects.all();
    remarks= request.POST.get('remarks')

    if request.method == 'POST':
        leave = payslip_request.objects.get(id=leave_id)
        leave.status = 2  
        leave.remarks=remarks
        leave.save()
        return HttpResponseRedirect(reverse("admin_payslip_view"))
    else:
       
        pass

# def admin_chart_ac_invite(request):
#     if request.method=='POST':
        
#         chart_ac_name=request.POST["chart_ac_name"]
#         chart_ac_email=request.POST["chart_ac_email"]
#         chart_ac_no=request.POST["chart_ac_no"]
#         try:
#             k=admin_chart_ac(chart_ac_name=chart_ac_name,chart_ac_email=chart_ac_email,chart_ac_no=chart_ac_no)
#             k.save()
#             html_content = render_to_string("charted_ac_invitaion.html",{'title':'test email','chart_ac_name':chart_ac_name})
#             text_content=strip_tags(html_content)
#             subject="WELCOME TO DEVELOPTRESS"
#             email = EmailMultiAlternatives(
#                 subject,
#                 text_content,
#                 settings.EMAIL_HOST_USER,
#                 [chart_ac_email],
#                 )
#             email.attach_alternative(html_content,"text/html")
#             email.fail_silently = True
#             email.send()
#             messages.success(request,"Successfully Added ")
#             return HttpResponseRedirect(reverse('admin_home'))
        
#         except:
#             messages.error(request,"Failed to Add ")
#             return HttpResponseRedirect(reverse('admin_home'))

# from django.http import HttpResponseBadRequest
# def doc_hod(request):
#     projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
#     staff_obj = AdminHod.objects.get(id=request.user.id)
#     leave_data=admin_doc.objects.filter(admin_id=staff_obj)
#     admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
#     admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
#     s=adminnav.objects.all()
#     r=documents_setup1.objects.all()
#     if request.method=="POST":
#         documenttype1=request.POST["documenttype1"]
#         imagefile1=request.FILES.get('imagefile1')
#         description=request.POST["description"]
#         employ_obj=AdminHod.objects.get(id=request.user.id)
#         k=admin_doc(admin=employ_obj,documenttype1=documenttype1,imagefile1=imagefile1,description=description)
#         k.save()
#         max_size = 5 * 1024 *1024
#         if imagefile1.size > max_size:
#             return HttpResponseBadRequest("File Size Exceeds Maximum Limit 5MB")
#         return redirect("/doc_hod")
#     return render(request,"admin-template/admin_documents_uploaded_view.html",{'projects_drops':projects_drops,'r':r,'s':s,'leave_data':leave_data,'admin_drops':admin_drops,'admin_home_drops':admin_home_drops})

from django.http import HttpResponseBadRequest
def doc_hod(request):
    employ=empdocs.objects.all()
    a=company_details.objects.filter(companyid=data1).first()

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = AdminHod.objects.filter(admin=request.user.id).first()
    da2 = da1.id
    admin = AdminHod.objects.get(id=1)
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    s=adminnav.objects.all()
    employs_all = Employs.objects.all()
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    data = AdminHod.objects.filter(id=request.user.id)
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')


    return render(request,"admin-template/admin_documents_uploaded_view.html",{'employ':employ,'admin_drops': admin_drops,'user':user,'a':a, 'da1':da1,'da2':da2,'data':data, 'projectm':projectm, 'admin_home_drops':admin_home_drops,'projects_drops':projects_drops,'employs_all':employs_all,'s':s,'h':h,'admin':admin})

def doc_delete(request,id):
    leave_data=admin_doc.objects.get(id=id)
    leave_data.delete()
    return redirect('/doc_hod')

def employ_delete(request,id):
    leave_data=empdocs.objects.get(id=id)
    leave_data.delete()
    return redirect('/')


def admin_tax_slip(request):
    user = CustomUser.objects.get(id=request.user.id)
    s=adminnav.objects.all()
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    taxs=admin_tax_details.objects.get(admin_id=request.user.id)
    return render(request,"admin-template/admin_tax_retrive_table.html",{'taxs':taxs,'s':s,'user':user,'admin_drops':admin_drops})

def admin_tax(request):
    kol=Employs.objects.all();
    admin_email=CustomUser.objects.filter(user_type=1);
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    s=adminnav.objects.all();
    if request.method=="POST":
        email=request.POST.get("email");
        email1=request.POST.get("email1");
        Current_Monthly_Rent=request.POST["Current_Monthly_Rent"];
        Name_of_landlord=request.POST["Name_of_landlord"];
        PAN_of_landlord=request.POST["PAN_of_landlord"];
        Address_of_landlord=request.POST["Address_of_landlord"];
        Section_80C=request.POST["Section_80C"];
        Section_80CCD=request.POST["Section_80CCD"];
        Section_80D=request.POST["Section_80D"];
        Section_80DD=request.POST["Section_80DD"];
        Section_80E=request.POST["Section_80E"];
        Section_80EEB=request.POST["Section_80EEB"];
        Section_80G=request.POST["Section_80G"];
        Section_80U=request.POST["Section_80U"];
        Section_80DDB=request.POST["Section_80DDB"];
        Section_80TTA=request.POST["Section_80TTA"];
        Section_80TTB=request.POST["Section_80TTB"];

        Annual_interest_payable=request.POST["Annual_interest_payable"];
        Additional_benefit_under_Section=request.POST["Additional_benefit_under_Section"];
        Name_of_lender=request.POST["Name_of_lender"];
        PAN_of_lender=request.POST["PAN_of_lender"];
        Address_of_lender=request.POST["Address_of_lender"];
        Section_80EEA=request.POST["Section_80EEA"];
        Amount=request.POST["Amount"];
        Origin=request.POST["Origin"];
        Destination=request.POST["Destination"];
        std_obj=AdminHod.objects.get(admin=request.user.id)
        if email and email1:
            pass
             
        # Both email and email1 are submitted. Handle the error case if needed.
                           
        elif email:
            k=employ_tax_form(email=email,employ_id=email,Current_Monthly_Rent=Current_Monthly_Rent,Name_of_landlord=Name_of_landlord,PAN_of_landlord=PAN_of_landlord,Address_of_landlord=Address_of_landlord,Section_80C=Section_80C,Section_80CCD=Section_80CCD,Section_80D=Section_80D,Section_80DD=Section_80DD,Section_80E=Section_80E,Section_80EEB=Section_80EEB,Section_80G=Section_80G,Section_80U=Section_80U,Section_80DDB=Section_80DDB,Section_80TTA=Section_80TTA,Section_80TTB=Section_80TTB,Annual_interest_payable=Annual_interest_payable,Additional_benefit_under_Section=Additional_benefit_under_Section,Name_of_lender=Name_of_lender,PAN_of_lender=PAN_of_lender,Address_of_lender=Address_of_lender,Section_80EEA=Section_80EEA,Amount=Amount,Origin=Origin,Destination=Destination)
            k.save();
        elif email1:
            try:
            # admin_hod_instance = AdminHOD.objects.get(id=email1)
                obj=admin_tax_details(admin_id_id=email1,Current_Monthly_Rent=Current_Monthly_Rent,Name_of_landlord=Name_of_landlord,PAN_of_landlord=PAN_of_landlord,Address_of_landlord=Address_of_landlord,Section_80C=Section_80C,Section_80CCD=Section_80CCD,Section_80D=Section_80D,Section_80DD=Section_80DD,Section_80E=Section_80E,Section_80EEB=Section_80EEB,Section_80G=Section_80G,Section_80U=Section_80U,Section_80DDB=Section_80DDB,Section_80TTA=Section_80TTA,Section_80TTB=Section_80TTB,Annual_interest_payable=Annual_interest_payable,Additional_benefit_under_Section=Additional_benefit_under_Section,Name_of_lender=Name_of_lender,PAN_of_lender=PAN_of_lender,Address_of_lender=Address_of_lender,Section_80EEA=Section_80EEA,Amount=Amount,Origin=Origin,Destination=Destination)
                obj.save()
            except AdminHod.DoesNotExist:
                # Handle the case when the specified AdminHOD instance does not exist.
                pass
            
        else:
            pass
        
    return render(request,"admin-template/admin_tax_form.html",{'s':s,'kol':kol,'admin_email':admin_email,'admin_drops':admin_drops})

def edit_admin_home_rent(request,id):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    s=adminnav.objects.all()
    k = admin_tax_details.objects.get(id=id);
    return render(request,"admin-template/update_admin_home_rent.html",{'k':k,'s':s,'admin_drops':admin_drops })

def update_admin_home_rent(request,id):
    if (request.method=="POST"):
        Current_Monthly_Rent=request.POST.get("Current_Monthly_Rent")
        Name_of_landlord=request.POST.get("Name_of_landlord")
        PAN_of_landlord=request.POST.get("PAN_of_landlord")
        Address_of_landlord=request.POST.get("Address_of_landlord")
        home_rent_proof=request.FILES.get("home_rent_proof")
        from_month=request.POST.get("from_month")
        to_month=request.POST.get("to_month")
        
        k=admin_tax_details.objects.get(id=id);
        
        k.Current_Monthly_Rent=Current_Monthly_Rent
        k.Name_of_landlord=Name_of_landlord
        k.PAN_of_landlord=PAN_of_landlord
        k.Address_of_landlord=Address_of_landlord
        k.home_rent_proof=home_rent_proof
        k.from_month=from_month
        k.to_month=to_month
        k.save();
        return redirect(reverse("admin_tax_slip"))
    return render(request,"admin-template/update_admin_home_rent.html")

def edit_admin_tax_sections(request,id):
    s=adminnav.objects.all()
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    k = admin_tax_details.objects.get(id=id);
    return render(request,"admin-template/update_admin_tax_sections.html",{'k':k,'s':s,'admin_drops':admin_drops})

def update_admin_tax_sections(request,id):
    if (request.method=="POST"):
        Section_80C=request.POST.get("Section_80C")
        Section_80CCD=request.POST.get("Section_80CCD")
        Section_80D=request.POST.get("Section_80D")
        Section_80DD=request.POST.get("Section_80DD")
        Section_80E=request.POST.get("Section_80E")
        Section_80EEB=request.POST.get("Section_80EEB")
        Section_80G=request.POST.get("Section_80G")
        Section_80U=request.POST.get("Section_80U")
        Section_80DDB=request.POST.get("Section_80DDB")
        Section_80TTA=request.POST.get("Section_80TTA")
        Section_80TTB=request.POST.get("Section_80TTB")
        Section_80C_proof=request.FILES.get("Section_80C_proof")
        Section_80D_proof=request.FILES.get("Section_80D_proof")
        Section_80CCD1BS_proof=request.FILES.get("Section_80CCD1BS_proof")
        Section_80DD_proof=request.FILES.get("Section_80DD_proof")
        Section_80E_proof=request.FILES.get("Section_80E_proof")
        Section_80G_proof=request.FILES.get("Section_80G_proof")
        Section_80U_proof=request.FILES.get("Section_80U_proof")
        Section_80EEB_proof=request.FILES.get("Section_80EEB_proof")
        Section_80DDB_proof=request.FILES.get("Section_80DDB_proof")
        Section_80TTA_proof=request.FILES.get("Section_80TTA_proof")
        Section_80TTB_proof=request.FILES.get("Section_80TTB_proof")

        k=admin_tax_details.objects.get(id=id);
        k.Section_80C=Section_80C
        k.Section_80CCD=Section_80CCD
        k.Section_80D=Section_80D
        k.Section_80DD=Section_80DD
        k.Section_80E=Section_80E
        k.Section_80EEB=Section_80EEB
        k.Section_80G=Section_80G
        k.Section_80U=Section_80U
        k.Section_80DDB=Section_80DDB
        k.Section_80TTA=Section_80TTA
        k.Section_80TTB=Section_80TTB
        k.Section_80C_proof=Section_80C_proof
        k.Section_80D_proof=Section_80D_proof
        k.Section_80CCD1BS_proof=Section_80CCD1BS_proof
        k.Section_80DD_proof=Section_80DD_proof
        k.Section_80E_proof=Section_80E_proof
        k.Section_80G_proof=Section_80G_proof
        k.Section_80U_proof=Section_80U_proof
        k.Section_80EEB_proof=Section_80EEB_proof
        k.Section_80DDB_proof=Section_80DDB_proof
        k.Section_80TTA_proof=Section_80TTA_proof
        k.Section_80TTB_proof=Section_80TTB_proof
        k.save();
        return redirect(reverse("admin_tax_slip"))
    return render(request,"admin-template/update_admin_tax_sections.html")

def edit_admin_home_loan(request,id):
    s=adminnav.objects.all()
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    k = admin_tax_details.objects.get(id=id);
    return render(request,"admin-template/update_admin_home_loan.html",{'k':k,'s':s,'admin_drops':admin_drops})

def update_admin_home_loan(request,id):
    if (request.method=="POST"):
        Annual_interest_payable=request.POST.get("Annual_interest_payable")
        Additional_benefit_under_Section=request.POST.get("Additional_benefit_under_Section")
        Name_of_lender=request.POST.get("Name_of_lender")
        PAN_of_lender=request.POST.get("PAN_of_lender")
        Address_of_lender=request.POST.get("Address_of_lender")
        Section_80EEA=request.POST.get("Section_80EEA")
        House_Property_proof=request.FILES.get("House_Property_proof")
        Section_80EE_proof=request.FILES.get("Section_80EE_proof")
        Section_80EEA_proof=request.FILES.get("Section_80EEA_proof")

        k=admin_tax_details.objects.get(id=id);
        k.Annual_interest_payable=Annual_interest_payable
        k.Additional_benefit_under_Section=Additional_benefit_under_Section
        k.Name_of_lender=Name_of_lender
        k.PAN_of_lender=PAN_of_lender
        k.Address_of_lender=Address_of_lender
        k.Section_80EEA=Section_80EEA
        k.House_Property_proof=House_Property_proof
        k.Section_80EE_proof=Section_80EE_proof
        k.Section_80EEA_proof=Section_80EEA_proof
        k.save();
        return redirect(reverse("admin_tax_slip"))
    return render(request,"admin-template/update_admin_home_loan.html")

def edit_admin_travel_allowance(request,id):
    s=adminnav.objects.all()
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    k = admin_tax_details.objects.get(id=id);
    return render(request,"admin-template/update_admin_travel_allowance.html",{'k':k,'s':s,'admin_drops':admin_drops})

def update_admin_travel_allowance(request,id):
    if (request.method=="POST"):
        Amount=request.POST.get("Amount")
        Origin=request.POST.get("Origin")
        Destination=request.POST.get("Destination")
        TravelStartDate=request.POST.get("TravelStartDate")
        k=admin_tax_details.objects.get(id=id);
        k.Amount=Amount
        k.Origin=Origin
        k.Destination=Destination
        k.TravelStartDate=TravelStartDate
        k.save();
        return redirect(reverse("admin_tax_slip"))
    return render(request,"admin-template/update_admin_travel_allowance.html")

from django.shortcuts import render
from datetime import datetime
from django.db.models import Sum

from .models import company_details_first,adminMonthlyTotal
from .utils import get_current_month_holidays_count

def count_business_days(year, month):
    total_days = calendar.monthrange(year, month)[1]
    business_days = 0
    
    for day in range(1, total_days + 1):
        if datetime(year, month, day).weekday() < 5:  # Monday to Friday (0 to 4)
            business_days += 1
    
    return business_days

def get_month_name(month_number):
    return calendar.month_abbr[month_number]

def admin_payslip(request):
    
    today = datetime.now()
    
    current_year = today.year
    current_month = today.month
    company_detail=company_details_first.objects.all()
    objs=AdminHod.objects.get(admin=request.user.id)
    details=admin_add_form1.objects.filter(admin=objs)
    # Filter data for the current month
    # data_for_month = CheckData.objects.filter(
    #     created_at__year=current_year,
    #     created_at__month=current_month
    # )
    data_for_months = checkout.objects.filter(
        date__year=current_year,
        date__month=current_month,
        empid=objs.email
    )
    india_holidays_count = get_current_month_holidays_count('india')
    user_obj=AdminHod.objects.get(id=request.user.id)
    package = int(user_obj.package)
    tax= int(package * 0.66667)
    
    b_tax=int(tax * 0.5)
    h_tax=int(b_tax * 0.5)
    s_tax=int(b_tax * 0.3)
    l_tax=int(b_tax * 0.2)
    standard_deductions= 50000
    net_taxable_income=tax - standard_deductions
    quotient = int(package) / 12
    main_salary=quotient
    # Multiply the remainder by 0.5%
    month_salary = main_salary * 0.5
    value=(main_salary * 0.25) +  (main_salary * 0.15) + (main_salary * 0.10)
    rent=int(main_salary * 0.25)
    travel=int(main_salary * 0.15)
    leave_travel=int(main_salary * 0.10)
    # Get the number of days in the current month
    business_days_in_current_month = count_business_days(current_year, current_month)

    # Adjust the amount based on the number of days in the current month
    quotients = month_salary / business_days_in_current_month 
    try:
        monthly_total_result = data_for_months.aggregate(total=Sum('date_value'))
        monthly_total = monthly_total_result['total'] + india_holidays_count
    except TypeError:
        monthly_total = 0
    totalss = int(monthly_total * quotients)
    
    total= int(totalss + value )
    month_and_year = f"{get_month_name(current_month)} {current_year}"
    if monthly_total is None:
        monthly_total = 0
    monthly_total_obj, created = adminMonthlyTotal.objects.update_or_create(
        year=current_year,
        month=current_month,
        student_id=user_obj,
        month_and_year = month_and_year,
        defaults={'total': totalss}
    )
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    data=adminMonthlyTotal.objects.get(student_id=user_obj)
    return render(request, 'admin-template/admin_payslip.html', {'admin_drops':admin_drops,'standard_deductions':standard_deductions,'net_taxable_income':net_taxable_income,'tax':tax,'b_tax':b_tax,'h_tax':h_tax,'s_tax':s_tax,'l_tax':l_tax,'travel':travel,'leave_travel':leave_travel,'rent':rent,'details':details,'user_obj':user_obj,'data':data,'company_detail':company_detail,'total':total,'monthly_total': monthly_total_obj.total,'totalss':totalss})

def admin_payroll_table(request):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    user=CustomUser.objects.get(id=request.user.id)
    user_obj= AdminHod.objects.get(admin=user)
    email_id=request.session.get('e_mail')
    p=employ_payslip.objects.filter(email=email_id).values()
    return render(request,"admin-template/admin_table.html",{'p':p,'user_obj':user_obj,'admin_drops':admin_drops})




def admin_send_payslip_to_employeemail(request):
    from_email = settings.EMAIL_HOST_USER
    email=request.session.get('email_2')
    e=Employs.objects.filter(email=email).values();
    recipient_list =[email]
    html_content = render_to_string("admin-template/admin_email_template.html",{'title':'test email','e':e})
    text_content=strip_tags(html_content)
    subject="YOUR PAYSLIP"
    email = EmailMultiAlternatives(
        subject,
        text_content,
        from_email,
        recipient_list,
        )
    email.attach_alternative(html_content,"text/html")
    email.fail_silently = True
    email.send() 
    return redirect('/admin_send_payslip_mail',{'e':e})


def admin_send_payslip_mail(request):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    m=MonthlyTotal.objects.all()
    s=adminnav.objects.all()
    return render(request,'admin-template/adminsearch_payslip.html', {'s':s,'m':m,'admin_drops':admin_drops})
   

def admin_payslip_table(request):
    user=CustomUser.objects.get(id=request.user.id)
    employ=Employs.objects.get(admin=user)
    email=request.session.get('email_2')
    e=Employs.objects.filter(email=email).values()
    return render(request,"admin-template/adminpaysliptable.html",{'e':e,'employ':employ})

from django.shortcuts import render, get_object_or_404,redirect
from .models import Employee_runpay,loca,depart
from .forms import dataForm
from django.db import connection


def homerunpay(request):
    # k=Employee.objects.all()
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    locations=loca.objects.all()
    departments=depart.objects.all()
    employees = Employee_runpay.objects.all()
    paid = 0 
    finalized = 0
    skipped = 0
    total = 0
    

    for employee in employees:
        total += 1
        if employee.monthly_ctc <= employee.addition + employee.deduction:
            skipped += 1
        # Modify the condition to compare monthly_ctc with gross_pay
        elif employee.monthly_ctc == employee.gross_pay:
            finalized += 1
        paid += employee.monthly_ctc

    context = {
        
        'paid': paid,
        'finalized': finalized,
        'skipped': skipped,
        'total': total
        
    }
    

    return render(request,"admin-template/homerunpay.html",{'admin_drops':admin_drops,'employees':employees,'departments':departments,'locations':locations,'context':context})

def calculate_values(request):
    employees = Employee_runpay.objects.all()
    paid = 0 
    finalized = 0
    skipped = 0
    total = 0
    

    for employee in employees:
        total += 1
        if employee.monthly_ctc <= employee.addition + employee.deduction:
            skipped += 1
        # Modify the condition to compare monthly_ctc with gross_pay
        elif employee.monthly_ctc == employee.gross_pay:
            finalized += 1
        paid += employee.monthly_ctc

    context = {
        
        'paid': paid,
        'finalized': finalized,
        'skipped': skipped,
        'total': total
        
    }

    return render(request, 'admin-template/calculate.html', context)

def employee_edit(request):
    # employees= get_object_or_404(Employee, id=id)
    if request.method == 'POST':
        form = dataForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('employee_list')
    else:
        form = dataForm()
    return render(request, 'admin-template/data_form.html', {'form': form})

def runpay_edit(request,id):
    post=Employee_runpay.objects.get(id=id)
    return render(request,"admin-template/runpayroll.html",{'post':post})

def runpay_update(request,id):
    if request.method=="POST":
        name=request.POST.get('name');
        monthly_ctc=request.POST.get('monthly_ctc');
        addition=request.POST.get('addition');
        deduction=request.POST.get('deduction');
        reimbursement=request.POST.get('reimbursement');
        remarks=request.POST.get('remarks');
        gross_pay=request.POST.get('gross_pay');
        post=Employee_runpay.objects.get(id=id);
        post.name=name;
        post.monthly_ctc=monthly_ctc;
        post.addition=addition;
        post.deduction=deduction;
        post.reimbursement=reimbursement;
        post.remarks=remarks;
        post.gross_pay=gross_pay;
        post.save()
        return redirect('home')
    return render(request,"admin-template/runpayroll.html")



# def enter_company_details(request):
#     users=company_details_first.objects.get()
#     admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
#     if(request.method=="POST"):
#         Organisationtype=request.POST.get("Organisationtype")
#         companypan=request.POST.get("companypan")
#         companyname=request.POST.get("companyname")
#         companyGSTIN=request.POST.get("companyGSTIN")
#         brandname=request.POST.get("brandname")
#         registeraddress=request.POST.get("registeraddress")
#         address=request.POST.get("address")
#         State=request.POST.get("State")
#         pincode=request.POST.get("pincode")
#         progress = 20
#         value_message=1
#         val=4

#         # Example progress value for form 1
#         progress_obj, _ = Progress.objects.get_or_create(pk=1, defaults={'value': 0})
    
#         progress_obj.value = value_message
#         progress_obj.progress_form1 = progress
        
        
#         progress_obj.save()
#         user_obj=company_details_first(pk=1,Organisationtype=Organisationtype,companypan=companypan,companyname=companyname,companyGSTIN=companyGSTIN,brandname=brandname,registeraddress=registeraddress,address=address,State=State,pincode=pincode)
#         instance=admin_home_drop.objects.get(id=1)
#         instance.progress_value=val
        
#         instance.save()
#         instance1=admin_home_drop.objects.get(id=2)
#         instance1.check_value=val
#         instance1.save()
#         user_obj.save()
#     return render(request,"admin-template/company_details.html",{'admin_drops':admin_drops})


def enter_company_details_mm(request):
   
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')

    users=company_details_first.objects.get()
    if(request.method=="POST"):
        Organisationtype=request.POST.get("Organisationtype")
        companypan=request.POST.get("companypan")
        companyname=request.POST.get("companyname")
        companyGSTIN=request.POST.get("companyGSTIN")
        brandname=request.POST.get("brandname")
        registeraddress=request.POST.get("registeraddress")
        address=request.POST.get("address")
        State=request.POST.get("State")
        pincode=request.POST.get("pincode")

        progress = 20
        value_message=1
        val=4

        # Example progress value for form 1
        progress_obj, _ = Progress.objects.get_or_create(pk=1, defaults={'value': 0})
    
        progress_obj.value = value_message
        progress_obj.progress_form1 = progress
        
        
       
        user_obj=company_details_first(pk=1,Organisationtype=Organisationtype,companypan=companypan,companyname=companyname,companyGSTIN=companyGSTIN,brandname=brandname,registeraddress=registeraddress,address=address,State=State,pincode=pincode)
        instance=admin_home_drop.objects.get(id=1)
        instance.progress_value=val
        progress_obj.save()
        instance.save()
        instance1=admin_home_drop.objects.get(id=2)
        instance1.check_value=val
        instance1.save()
        user_obj.save()
    return render(request,"admin-template/company_details.html",{ 'users':users,  'admin_drops':admin_drops,'admin_home_drops':admin_home_drops})



def enter_company_details(request):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id

    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    data = Companys.objects.filter(usernumber=request.user.id).first()
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    a=company_details.objects.filter(companyid=data).first()



    # objs=AdminHod.objects.get(admin=request.user.id)
    users=company_details_first.objects.first()
    messages.success(request,"your Details was added successfully ")


    return render(request,"admin-template/company_details.html",{"users":users,'a':a, 'admin_drops':admin_drops,'admin_home_drops':admin_home_drops,'user':user,'admin_home_drops':admin_home_drops,'s':s,'h':h,'projectm':projectm,'projects_drops':projects_drops,'data':data,'employs_all':employs_all

})



def enter_company_details_save(request):
    if request.method!="POST":
        return HttpResponseRedirect(reverse("enter_company_details"))
    else:    
        Organisationtype=request.POST.get("Organisationtype")
        companypan=request.POST.get("companypan")
        companyname=request.POST.get("companyname")
        companyGSTIN=request.POST.get("companyGSTIN")
        brandname=request.POST.get("brandname")
        registeraddress=request.POST.get("registeraddress")
        address=request.POST.get("address")
        State=request.POST.get("State")
        pincode=request.POST.get("pincode")
          
        try:
        
            user_subobj, _=company_details_first.objects.get_or_create(admin_id=objs)
            user_subobj.Organisationtype=Organisationtype
            user_subobj.companypan=companypan
            user_subobj.companyname=companyname
            user_subobj.companyGSTIN=companyGSTIN
            user_subobj.brandname=brandname
            user_subobj.registeraddress=registeraddress
            user_subobj.address=address
            user_subobj.State=State
            user_subobj.pincode=pincode  
            user_subobj.save()
            return HttpResponseRedirect(reverse("enter_company_details"))
        except:
            
            return HttpResponseRedirect(reverse("enter_company_details"))




def company_details_edit(request):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    com=company_details_first.objects.first()
    return render(request,"admin-template/company_details_update.html",{'com':com,'admin_drops':admin_drops})

def company_details_update(request):
    com=company_details_first.objects.first()

    if(request.method=="POST"):
        Organisationtype=request.POST.get('Organisationtype')
        companypan=request.POST.get('companypan')
        companyname=request.POST.get('companyname')
        companyGSTIN=request.POST.get('companyGSTIN')
        brandname=request.POST.get('brandname')
        registeraddress=request.POST.get('registeraddress')
        address=request.POST.get('address')
        State=request.POST.get('State')
        pincode=request.POST.get('pincode')
        com=company_details_first.objects.first()
        com.Organisationtype=Organisationtype;
        com.companypan=companypan;
        com.companyname=companyname;
        com.companyGSTIN=companyGSTIN;
        com.brandname=brandname;
        com.registeraddress=registeraddress;
        com.address=address;
        com.State=State;
        com.pincode=pincode;
        com.save()
        return redirect("/admin_home")
    return render (request,"admin-template/company_details_update.html",{'com':com})




# def set_payroll(request):
    
#     admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
#     if(request.method=="POST"):
#         payrolldate=request.POST.get("payrolldate")
#         auto_run_payroll=request.POST.get("auto_run_payroll")
#         advance_salary_request=request.POST.get("advance_salary_request")
#         progress = 15 
#         value_message=1
#         val=4
#         # Example progress value for form 1
#         progress_obj, _ = Progress.objects.get_or_create(pk=1, defaults={'value': 0})
    
#         progress_obj.value1 = value_message
#         progress_obj.progress_form2 = progress
        
#         progress_obj.save()
#         user_obj=set_payroll_date(pk=1,payrolldate=payrolldate,auto_run_payroll=auto_run_payroll,advance_salary_request=advance_salary_request)
#         instance=admin_home_drop.objects.get(id=1)
#         instance.progress_value=val
#         instance.save()
        
#         user_obj.save()
#     return render(request,"admin-template/set_payroll.html",{'admin_drops':admin_drops})


def set_payroll(request):
    
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    
    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    data = Companys.objects.filter(usernumber=request.user.id).first()
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    a=company_details.objects.filter(companyid=data).first()




    if(request.method=="POST"):

        payrolldate=request.POST.get("payrolldate")
        payroll=request.POST.get("payroll")
        if payroll == "on":
            payroll = "Yes"
        else:
            payroll = "No"

        auto_run_payroll=request.POST.get("auto_run_payroll")
        if auto_run_payroll == "on":
            auto_run_payroll = "Yes"
        else:
            auto_run_payroll = "No"

        advance_salary_request=request.POST.get("advance_salary_request")
        
        if advance_salary_request == "on":
            advance_salary_request = "enabled"
        else:
            advance_salary_request = "disabled"

        progress = 15 
        value_message=1
        val=4
        # Example progress value for form 1
        progress_obj, _ = Progress.objects.get_or_create(pk=1, defaults={'value': 0})
    
        progress_obj.value1 = value_message
        progress_obj.progress_form2 = progress
        
        progress_obj.save()
        user_obj=set_payroll_date(pk=1,payrolldate=payrolldate,payroll=payroll,auto_run_payroll=auto_run_payroll,advance_salary_request=advance_salary_request)
        instance=admin_home_drop.objects.get(id=1)
        instance.progress_value=val
        instance.save()
        
        user_obj.save()
        messages.success(request,"Your data was successfully saved")

        return redirect('setting')
    return render(request,"admin-template/set_payroll.html",{'admin_drops':admin_drops,'user':user,'admin_home_drops':admin_home_drops,'a':a, 's':s,'h':h,'projectm':projectm,'projects_drops':projects_drops,'data':data,'employs_all':employs_all})





def edit_payroll_setup(request):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    post=set_payroll_date.objects.first();
    return render(request,"admin-template/update_payroll_setup.html",{'post':post,'admin_drops':admin_drops})

# def update_payroll_setup(request):
#     if (request.method =="POST"):
       
#         payrolldate=request.POST.get('payrolldate');
#         auto_run_payroll=request.POST.get('auto_run_payroll');
#         advance_salary_request=request.POST.get('advance_salary_request');
#         default_salary=request.POST.get('default_salary')
#         post=set_payroll_date.objects.first()
        
#         post.payrolldate=payrolldate;
#         post.auto_run_payroll=auto_run_payroll;
#         post.advance_salary_request=advance_salary_request;
#         post.default_salary=default_salary;
#         post.save()
#         return redirect('/admin_home')
#     return render(request,'admin-template/update_payroll_setup.html',{'post': post})



# def default_sal(request):
#     k=salary_struc.objects.all()
   
#     admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
#     if(request.method=="POST"):
#         default_salary=request.POST.get("default_salary")
#         FBP_allowances=request.POST.get("FBP_allowances")
#         progress = 15
#         value_message=1
#         val=2
#         # Example progress value for form 1
#         progress_obj, _ = Progress.objects.get_or_create(pk=1, defaults={'value': 0})
    
#         progress_obj.value2 = value_message
#         progress_obj.progress_form3 = progress
        
#         progress_obj.save()
#         user_obj=set_salary_structure(pk=1,default_salary=default_salary,FBP_allowances=FBP_allowances)
#         instance=admin_home_drop.objects.get(id=1)
#         instance.progress_value=val
#         instance.save()
        
#         user_obj.save()
#     return render(request,"admin-template/default_sal.html",{'k':k,'admin_drops':admin_drops})


from django.shortcuts import render, redirect
from .models import Progress, set_salary_structure, salary_struc, admin_drop, admin_home_drop

def default_sal(request):
    k = salary_struc.objects.all()
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    data = Companys.objects.filter(usernumber=request.user.id).first()
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    a=company_details.objects.filter(companyid=data).first()


    user_obj=None

    if request.method == "POST":
        # Check if the "default_salary" checkbox is checked
        default_salary = request.POST.get("default_salary") == "on"

        # Check if the "FBP_allowances" checkbox is checked
        FBP_allowances = request.POST.get("FBP_allowances") == "on"

        # Example progress values
        progress = 15
        value_message = 1
        val = 2

        # Update the Progress model
        progress_obj, _ = Progress.objects.get_or_create(pk=1, defaults={'value': 0})
        progress_obj.value2 = value_message
        progress_obj.progress_form3 = progress
        progress_obj.save()

        # Update the set_salary_structure object or create it if it doesn't exist
        user_obj, _ = set_salary_structure.objects.get_or_create(pk=1)
        user_obj.default_salary = default_salary
        user_obj.FBP_allowances = FBP_allowances
        user_obj.save()

        # Update the status field in the salary_struc model
        salary_struc_objs = salary_struc.objects.all()
        for obj in salary_struc_objs:
            obj.status = int(default_salary)  # Convert bool to int (1 or 0)
            obj.save()
        # Update the progress_value field in admin_home_drop
        instance = admin_home_drop.objects.get(id=1)
        instance.progress_value = val
        instance.save()
        messages.success(request,"Your details were successfully saved")

    return render(request, "admin-template/default_sal.html", {'k': k, 'a':a, 'admin_drops': admin_drops, 'user_obj': user_obj,'user':user,'admin_home_drops':admin_home_drops,'s':s,'h':h,'projectm':projectm,'projects_drops':projects_drops,'data':data,'employs_all':employs_all})






# def esic(request):
#     a=ESIC.objects.all()
#     s=adminnav.objects.all()
#     admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
#     if request.method=="POST":
#         esi_status=request.POST.get('esi_status')
#         esi_payment=request.POST.get('esi_payment')
#         username2=request.POST.get('username2')
#         password3=request.POST.get('password3')
#         esi_settings=request.POST.get('esi_settings')
#         esi_settings1=request.POST.get('esi_settings1')

#         progress = 20 
#         value_message=1
#         val=4
#         k=ESIC(pk=1,esi_settings1=esi_settings1,esi_status=esi_status,esi_payment=esi_payment,username2=username2,password3=password3,esi_settings=esi_settings)
#         progress_obj, _ = Progress.objects.get_or_create(pk=1, defaults={'value': 0})
#         progress_obj.value4 = value_message
#         progress_obj.progress_form5 = progress
#         progress_obj.save()
#         instance=admin_home_drop.objects.get(id=6)
#         instance.progress_value=val
#         instance.save()

#         k.save()
#         return redirect("/company_detail")
#     return render(request,"admin-template/admin_esic.html",{'a':a,'s':s,'admin_drops':admin_drops})




def pvt_pub_form(request):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    if request.method=="POST":
        certificate=request.FILES.get('certificate')
        cmpny_pancard=request.FILES.get('cmpny_pancard')
        cheque=request.FILES.get('cheque')
        owner_pancard=request.FILES.get('owner_pancard')
        id_proof=request.FILES.get('id_proof')
        gst=request.FILES.get('gst')
        progress = 15 
        value_message=1
        val=4
        DATA="ENABLE"
        admin_obj=AdminHod.objects.get(admin=request.user.id)
        k=pvt_pub(admin_id=admin_obj,certificate=certificate,cmpny_pancard=cmpny_pancard,cheque=cheque,owner_pancard=owner_pancard,id_proof=id_proof,gst=gst) 
        k.save()
        objs,_=Company.objects.get_or_create(id=1)
        objs.KYCSTATUS= DATA
        objs.save()
        progress_obj, _ = Progress.objects.get_or_create(pk=1, defaults={'value': 0})
        progress_obj.value5 = value_message
        progress_obj.progress_form6 = progress
        progress_obj.save()
        instance=admin_home_drop.objects.get(id=10)
        instance.progress_value=val
        instance.save() 
    return render(request,"admin-template/admin_kyc.html",{'admin_drops':admin_drops}) 




def sole_proprietorship_form(request):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    if request.method=="POST":
        gst_certificate=request.FILES.get('gst_certificate')
        cancelled_cheque=request.FILES.get('cancelled_cheque')
        owner_pancard1=request.FILES.get('owner_pancard1')
        idproof=request.FILES.get('idproof')
        GST=request.FILES.get('GST')
        progress = 15 
        value_message=1
        val=4
        DATA="ENABLE"
        admin_obj=AdminHod.objects.get(admin=request.user.id)
        k=sole_proprietorship(admin_id=admin_obj,gst_certificate=gst_certificate,cancelled_cheque=cancelled_cheque,owner_pancard1=owner_pancard1,idproof=idproof,GST=GST) 
        k.save()
        objs,_=Company.objects.get_or_create(id=1)
        objs.KYCSTATUS= DATA
        objs.save()
        progress_obj, _ = Progress.objects.get_or_create(pk=1, defaults={'value': 0})
        progress_obj.value5 = value_message
        progress_obj.progress_form6 = progress
        progress_obj.save()
        instance=admin_home_drop.objects.get(id=10)
        instance.progress_value=val
        instance.save() 
    return render(request,"admin-template/admin_kyc.html",{'admin_drops':admin_drops})  






def partnership_form(request):
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    if request.method=="POST":
        shop_license=request.FILES.get('shop_license')
        cmpny_pan=request.FILES.get('cmpny_pan')
        cheque1=request.FILES.get('cheque1')
        owner_pan=request.FILES.get('owner_pan')
        id_proof1=request.FILES.get('id_proof1')
        gst1=request.FILES.get('gst1')
        progress = 15 
        value_message=1
        val=4
        DATA="ENABLE"
        admin_obj=AdminHod.objects.get(admin=request.user.id)
        k=partnership(admin_id=admin_obj,shop_license=shop_license,cmpny_pan=cmpny_pan,cheque1=cheque1,owner_pan=owner_pan,id_proof1=id_proof1,gst1=gst1) 
        k.save() 
        objs,_=Company.objects.get_or_create(id=1)
        objs.KYCSTATUS= DATA
        objs.save()
        progress_obj, _ = Progress.objects.get_or_create(pk=1, defaults={'value': 0})
        progress_obj.value5 = value_message
        progress_obj.progress_form6 = progress
        progress_obj.save()
        instance=admin_home_drop.objects.get(id=10)
        instance.progress_value=val
        instance.save()  
    return render(request,"admin-template/admin_kyc.html",{'admin_drops':admin_drops}) 





def trust_ngo_form(request):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    if request.method=="POST":
        shop_license1=request.FILES.get('shop_license1')
        cmpny_pan1=request.FILES.get('cmpny_pan1')
        cheque2=request.FILES.get('cheque2')
        owner_pan1=request.FILES.get('owner_pan1')
        id_proof2=request.FILES.get('id_proof2')
        gst2=request.FILES.get('gst2')
        progress = 15 
        value_message=1
        val=4
        DATA="ENABLE"
        admin_obj=AdminHod.objects.get(admin=request.user.id)        
        k=trust_ngo(admin_id=admin_obj,shop_license1=shop_license1,cmpny_pan1=cmpny_pan1,cheque2=cheque2,owner_pan1=owner_pan1,id_proof2=id_proof2,gst2=gst2) 
        k.save()  
        objs,_=Company.objects.get_or_create(id=1)
        objs.KYCSTATUS= DATA
        objs.save()
        progress_obj, _ = Progress.objects.get_or_create(pk=1, defaults={'value': 0})
        progress_obj.value5 = value_message
        progress_obj.progress_form6 = progress
        progress_obj.save()
        instance=admin_home_drop.objects.get(id=10)
        instance.progress_value=val
        instance.save() 
    return render(request,"admin-template/admin_kyc.html",{'admin_drops':admin_drops}) 


def kyc_update(request):
    k8 = Company.objects.first()
    
    if request.method=="POST":
        new_field_d_data = request.POST.get('field_d_data')
        Company.objects.update(KYCSTATUS=new_field_d_data)
        
        return redirect('/company_detail')
    return render(request,'admin-template/kyc_update.html', {'k8': k8})

def card(request):
    today_date = datetime.now().date()

    b = checkin.objects.filter(date=today_date).count()
    c=Employs.objects.count()
    d=c-b
    leave_applicants_count = LeaveReportEmploy.objects.count()
    tomorrow_date = date.today() + timedelta(days=1)
    approved_leave_count = LeaveReportEmploy.objects.filter(leave_status=1).count()
    pending_leave_count = LeaveReportEmploy.objects.filter(leave_status=0).count()

    # Calculate the total leave applications
    total_leave_count = approved_leave_count + pending_leave_count

    # Get the count of members who have applied for leave permission tomorrow
    tomorrow_leave_applicants_count = LeaveReportEmploy.objects.filter(leave_date=tomorrow_date).count()
    # checked_in_employee_ids = checkin.objects.filter(date=today_date).values_list('empid', flat=True)
    return render(request, "admin-template/e.html",{'b':b,'c':c,'d':d,'leave_applicants_count':leave_applicants_count,'tomorrow_leave_applicants_count':tomorrow_leave_applicants_count,'approved_leave_count': approved_leave_count,'pending_leave_count': pending_leave_count,'total_leave_count': total_leave_count})


# def tds(request):
#     s=adminnav.objects.all()
#     admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
#     if request.method=="POST":
#         tds_payment=request.POST['tds_payment']
#         verify_tan=request.POST['verify_tan']
#         tds_filling_setup=request.POST['tds_filling_setup']
#         filling_form=request.POST['filling_form']
#         userid=request.POST['userid']
#         password=request.POST['password']
#         username=request.POST['username']
#         password1=request.POST['password1']
#         k=TDS(tds_payment=tds_payment,verify_tan=verify_tan,tds_filling_setup=tds_filling_setup,filling_form=filling_form,userid=userid,password=password,username=username,password1=password1)
#         k.save()
#         return redirect("/pr_tax")
#     return render(request,"admin-template/tds.html",{'s':s,'admin_drops':admin_drops})




# def pr_tax(request):
#     a=P_tax.objects.all()
#     s=adminnav.objects.all()
#     admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
#     if request.method=='POST':
#         professional_tax=request.POST.get('professional_tax')
#         pt_payment=request.POST.get('pt_payment')
#         username3=request.POST.get('username3')
#         password4=request.POST.get('password4')
#         pt_setup=request.POST.get('pt_setup')
#         pt_setup1=request.POST.get('pt_setup1')
#         pt_setup2=request.POST.get('pt_setup2')
#         pt_setup3=request.POST.get('pt_setup3')
#         pt_setup4=request.POST.get('pt_setup4')
#         pt_setup5=request.POST.get('pt_setup5')
#         k=P_tax(pk=1,pt_setup5=pt_setup5,pt_setup4=pt_setup4,pt_setup3=pt_setup3,pt_setup1=pt_setup1,pt_setup2=pt_setup2,password4=password4,pt_setup=pt_setup,username3=username3,pt_payment=pt_payment,professional_tax=professional_tax)
#         k.save()
#         instance=details.objects.create(form1=professional_tax)
#         return redirect("/pf")
#     return render(request,"admin-template/p_tax.html",{'a':a,'s':s,'admin_drops':admin_drops})


# def pf(request):
#     b=PF.objects.all()
#     s=adminnav.objects.all()
#     admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
#     if request.method=='POST':
#         pf_status=request.POST.get('pf_status')
#         pf_payment=request.POST.get('pf_payment')
#         username1=request.POST.get('username1')
#         password2=request.POST.get('password2')
#         pf_setup=request.POST.get('pf_setup')
#         pf_setup1=request.POST.get('pf_setup1')
#         pf_setup2=request.POST.get('pf_setup2')
#         pf_setup3=request.POST.get('pf_setup3')
#         pf_setup4=request.POST.get('pf_setup4')

#         k=PF(pk=1,pf_setup1=pf_setup1,pf_setup2=pf_setup2,pf_setup3=pf_setup3,pf_setup4=pf_setup4,pf_status=pf_status,pf_payment=pf_payment,username1=username1,password2=password2,pf_setup=pf_setup)
#         k.save()

#         instance=details.objects.create(form2=pf_status)
        
#         return redirect("/esic")
#     return render(request,"admin-template/pf.html",{'b':b,'s':s,'admin_drops':admin_drops})



def company_form(request):
    s=adminnav.objects.all()
    k = salary_struc.objects.all()
    if request.method=="POST":
        Organisationtype=request.POST['Organisationtype']
        companypan=request.POST['companypan']
        companyGSTIN=request.POST['companyGSTIN']
        brandname=request.POST['brandname']
        registeraddress=request.POST['registeraddress']
        address=request.POST['address']
        companyname=request.POST['companyname']
        logo=request.FILES.get("media/")
        State=request.POST['State']
        pincode=request.POST['pincode']   
        n=company_details(Organisationtype=Organisationtype,companypan=companypan,companyname=companyname,companyGSTIN=companyGSTIN,brandname=brandname,registeraddress=registeraddress,address=address,logo=logo,State=State,pincode=pincode)
        n.save();
        return redirect("/admin_home")
    return render(request,"admin-template/companyform.html",{'s':s,'k':k})

# def company_detail(request):
#     a=companylogo.objects.all()
#     s=adminnav.objects.all()
#     n=company_details_first.objects.first()
#     k11=P_tax.objects.first()
#     k21=PF.objects.first()
#     k31=ESIC.objects.first()
#     k41=P_tax.objects.all()
#     k8=Company.objects.first()
#     company = Company.objects.first()
    
#     admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
   
#     return render(request, 'admin-template/company_detail.html', {'company': company,'n':n,'s':s,'admin_drops':admin_drops,'k11':k11,'k21':k21,'k31':k31,'k41':k41,'k8':k8,'a':a})

def company_detail(request):
    # a=companylogo.objects.all()
    s=adminnav.objects.all()
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()

    n=company_details.objects.filter(companyid=data1).first()
    k=TDS.objects.all()
    k1=P_tax.objects.all()
    k2=PF.objects.all()
    k3=ESIC.objects.all()
    k4=P_tax.objects.all()
    k8=Company.objects.first()
    post1=TDS.objects.first()


    company = Company.objects.first()
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
   
    return render(request, 'admin-template/company_detail.html', {'data':data,'post1':post1,'company': company,'n':n,'k8':k8,'s':s,'admin_drops':admin_drops,'k1':k1,'k2':k2,'k3':k3,'a':a ,'k':k,'k4':k4})



def company_form_edit(request,id):
    a=companylogo.objects.all()
    s=adminnav.objects.all()
    com=company_details_first.objects.get(id=id)
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    return render(request,"admin-template/form_update.html",{'com':com,'s':s,'admin_drops':admin_drops,'a':a})

def company_details_update1(request):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    s=adminnav.objects.all()
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()
    comp=company_details.objects.all()
    if(request.method=="POST"):
        companyname=request.POST.get('companyname')
        registeraddress=request.POST.get('registeraddress')
        State=request.POST.get('State')
        pincode=request.POST.get('pincode')
        companyid=data
        compnaycheck=company_details.objects.filter(companyid=data).first()
        if compnaycheck:
            com_instance=company_details.objects.get(companyid=data)
            com_instance.companyname=companyname;
            com_instance.registeraddress=registeraddress;
            com_instance.State=State;
            com_instance.pincode=pincode;
            com_instance.save()
        else:
            comdet=company_details(companyname=companyname,registeraddress=registeraddress,State=State,pincode=pincode,companyid=companyid)
            comdet.save()
        messages.success(request,"Your details were successfully submitted")
        return redirect("/company_detail")
    return render (request,"admin-template/company_details_update1.html",{'a':a, 'admin_drops':admin_drops,'data':data,'s':s})



def compney_logos(request):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    s=adminnav.objects.all()
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()
    if request.method=="POST":
        logo=request.FILES.get("logo")
        if a:    
            k=company_details.objects.get(companyid=data1)
            k.logo=logo
            k.save()
        else:
            k=company_details(logo=logo,companyid=data)
            k.save()
    return render(request,"admin-template/compney_logos.html",{'a':a,'admin_drops':admin_drops,'s':s,'data':data})

def compney_logos1(request,):
    comp=Companys.objects.filter(usernumber=request.user.id).first()
    a=company_details.objects.get(companyid=comp)
    a.logo.delete()
    return redirect("/compney_logos")


from django.shortcuts import render, redirect
from .models import P_tax, PF, ESIC, admin_drop  # Import your models

def update_data_view(request):
    # Retrieving data from the database for each model
    k1 = P_tax.objects.first()
    k2 = PF.objects.first()
    k3 = ESIC.objects.first()
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    s=adminnav.objects.all()

    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()

    # Default values for the dropdowns
    field_a_data = k1.professional_tax if k1 else None
    field_b_data = k2.pf_status if k2 else None
    field_c_data = k3.esi_status if k3 else None

    # Checking if the form was submitted
    if request.method == 'POST':
        # Extracting data from the form
        new_field_a_data = request.POST.get('field_a_data')
        new_field_b_data = request.POST.get('field_b_data')
        new_field_c_data = request.POST.get('field_c_data')

        # Updating data for P_tax, PF, and ESIC
        if k1:
            k1.professional_tax = new_field_a_data
            k1.save()
        if k2:
            k2.pf_status = new_field_b_data
            k2.save()
        if k3:
            k3.esi_status = new_field_c_data
            k3.save()

        messages.success(request,"Your details were successfully submitted")

        # Redirecting to the same page after the update
        return redirect('/company_detail')

    return render(request, 'admin-template/compliance.html', {
        'k1': k1,
        'k2': k2,
        'k3': k3,
        'admin_drops': admin_drops,
        'field_a_data': field_a_data,
        'field_b_data': field_b_data,
        'field_c_data': field_c_data,
        's':s,
        'data':data,
        'a':a
    })




def company_update4(request):
    k=TDS.objects.first()
    p=PF.objects.first()
    e=ESIC.objects.first()
    f=P_tax.objects.first()

    lu=AdminHod.objects.first()
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    s=adminnav.objects.all()

    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()

    if request.method=="POST":
        userid=request.POST.get('userid');
        password=request.POST.get('password');
        username=request.POST.get('username');
        password1=request.POST.get('password1');
        k=TDS.objects.first();
        k.userid=userid;
        k.password=password;
        k.username=username;
        k.password1=password1;
        k.save();
        username1=request.POST.get('username1');
        password2=request.POST.get('password2');
        p=PF.objects.first();
        p.username1=username1;
        p.password2=password2;
        p.save();
        username2=request.POST.get('username2');
        password3=request.POST.get('password3');
        e=ESIC.objects.first();
        e.username2=username2;
        e.password3=password3;
        e.save()
        username3=request.POST.get('username3');
        password4=request.POST.get('password4');
        f=P_tax.objects.first();
        f.username3=username3;
        f.password4=password4;
        f.save()

        location=request.POST.get('location');
        lu=AdminHod.objects.first()
        lu.location=location;
        lu.save()

        messages.success(request,"Your details were successfully submitted")

        return redirect('/company_detail/')
    return render(request,"admin-template/company_edit4.html",{'a':a,'k':k,'p':p,'e':e,'f':f,'lu':lu,'a':a, 'admin_drops':admin_drops,'s':s,'data':data})



def salary(request):
    employ=Employs.objects.all()
    leaves=ad_salary.objects.all()
    pending=ad_salary.objects.raw('SELECT count(amount) as sum from ehrms_ad_salary where request_status=0;')
    sum=ad_salary.objects.raw('SELECT sum(amount) as sum from ehrms_ad_salary where request_status=1;')
    emp_req=ad_salary.objects.all()
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    if request.method=="POST":
        amount=request.POST['amount']
        emi=request.POST['emi']
        reason=request.POST['reason']
    
        k=admin_ad_salary(amount=amount,emi=emi,reason=reason)
        k.save()
        return redirect("/admin_home")
    return render(request,"admin-template/adminadvancesalary.html",{'admin_drops':admin_drops,'emp_req':emp_req,'employ':employ})  

# def admin_advancesalary_apply_view(request):
#     user = CustomUser.objects.filter(id=request.user.id).first()
#     userid1 = user.id
#     da1 = AdminHod.objects.filter(admin=request.user.id).first()
#     da2 = da1.id
#     admin = AdminHod.objects.get(id=1)
#     projectm = admin_project_create.objects.filter(admin_id=userid1)
#     h = HR.objects.all()
#     s=adminnav.objects.all()
#     employs_all = Employs.objects.all()
#     projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
#     data = AdminHod.objects.filter(id=request.user.id)
#     admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
#     admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')


#     employ=Employs.objects.all()
#     leaves=ad_salary.objects.all()
#     data = CustomUser.objects.filter(id=request.user.id,user_type=1)
#     pending=ad_salary.objects.raw('SELECT count(amount) as sum from ehrms_ad_salary where request_status=0;')
#     sum=ad_salary.objects.raw('SELECT sum(amount) as sum from ehrms_ad_salary where request_status=1;')
#     emp_req=ad_salary.objects.all()
#     admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
#     items_per_page = 10  # Adjust the number of items per page as needed

#     paginator = Paginator(leaves, items_per_page)
#     page = request.GET.get('page')

#     try:
#         leaves = paginator.page(page)
#     except PageNotAnInteger:
#         leaves = paginator.page(1)
#     except EmptyPage:
#         leaves = paginator.page(paginator.num_pages)
    
#     return render(request,"admin-template/adminadvancesalary.html",{'data':data,'admin_drops':admin_drops,'emp_req':emp_req,'employ':employ,'admin_drops': admin_drops,'user':user,'da1':da1,'da2':da2,'data':data,'admin_home_drops':admin_home_drops,'projects_drops':projects_drops,'employs_all':employs_all,'s':s,'h':h,'admin':admin})



def admin_advancesalary_apply_view(request):

    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    employs = Employs.objects.all()
    leaves = ad_salary.objects.all()
    data = CustomUser.objects.filter(id=request.user.id, user_type=1)
    pending = ad_salary.objects.filter(request_status=0).count()
    total_amount = ad_salary.objects.filter(request_status=1).aggregate(Sum('amount'))
    emp_req = ad_salary.objects.filter(employ_id__companyid=data1)
    s=adminnav.objects.all()
    a=company_details.objects.filter(companyid=data1).first()

    items_per_page = 10  # Adjust the number of items per page as needed

    paginator = Paginator(leaves, items_per_page)
    page = request.GET.get('page')

    try:
        leaves = paginator.page(page)
    except PageNotAnInteger:
        leaves = paginator.page(1)
    except EmptyPage:
        leaves = paginator.page(paginator.num_pages)

    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    
    return render(request, "admin-template/adminadvancesalary.html", {
        'data': data,
        'admin_drops': admin_drops,
        'emp_req': emp_req,
        'employs': employs,
        'leaves': leaves,
        'pending_count': pending,
        's':s,
        'a':a,
        'total_amount': total_amount['amount__sum'] if total_amount['amount__sum'] else 0
    })


def admin_advancesalary_apply_view_save(request):
    if request.method!="POST":
        return HttpResponseRedirect(reverse("admin_advancesalary_apply_view"))
    else:
        amount=request.POST.get("amount")
        emi=request.POST.get("emi")
        reason=request.POST.get("reason")
        amount=request.POST.get("amount")

        student_obj=AdminHod.objects.get(admin=request.user.id)
        try:
            leave_report=admin_ad_salary(admin_id=student_obj,amount=amount,emi=emi,reason=reason,request_status=0)
            leave_report.save()
            messages.success(request, "Successfully Applied for advancesalary")
            return HttpResponseRedirect(reverse("admin_advancesalary_apply_view"))
        except:
            messages.error(request, "Failed To Apply for advancesalary")
            return HttpResponseRedirect(reverse("admin_advancesalary_apply_view"))


def otp(request):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    k1=Employs.objects.all()
    employs_count=Employs.objects.raw("select count(first_name) from ehrms_employs;")
    if request.method=="POST":
        amount=request.POST['amount']
        paymenttype=request.POST['paymenttype']
        k=otp(amount=amount,paymenttype=paymenttype)
        k.save()
        return redirect("/admin_home")
    return render(request,"admin-template/employ_otp.html",{'k1':k1,'employs_count':employs_count,'admin_drops':admin_drops})


def company_edit1(request):
    a=companylogo.objects.all()
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    post1=Company.objects.first()
    return render(request,"admin-template/company_edit1.html",{'post1':post1,'admin_drops':admin_drops,'a':a})


def company_update1(request):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')

    s=adminnav.objects.all()

    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()
    tdscom=TDS.objects.filter(companyid=data).first()
    com_instance=company_details.objects.get(companyid=data)
    post1=TDS.objects.get(companyid=data)


    if request.method=="POST":
        companypan=request.POST.get('companypan');
        companyGSTIN=request.POST.get('companyGSTIN');
        verify_tan=request.POST.get('verify_tan');

        # com_instance=company_details.objects.first();
        # com_instance.companypan=companypan;
        # com_instance.companyGSTIN=companyGSTIN;
        # com_instance.save();
        # verify_tan=request.POST.get('verify_tan');
        # post1.verify_tan=verify_tan;
        # post1.save();
        if a:
            if com_instance is not None:
                com_instance.companypan = companypan
                com_instance.companyGSTIN = companyGSTIN
                com_instance.save()
        else:
            com_instance=company_details(companypan=companypan,companyGSTIN=companyGSTIN)
            com_instance.save()
        if tdscom:
            # Ensure post1 is not None before accessing its attributes
            if post1 is not None:
                verify_tan = request.POST.get('verify_tan')
                post1.verify_tan = verify_tan
                post1.save()
        else:
            post1=TDS(verify_tan=verify_tan,companyid=data1)
            post1.save()

        messages.success(request,"Your details were successfully submitted")

        return redirect('company_detail')
    return render(request,"admin-template/company_edit1.html",{'post1':post1,'a':a,'com_instance':com_instance,'admin_drops':admin_drops,'admin_home_drops':admin_home_drops,'s':s,'data':data})



def company_update2(request,id):
    if request.method=="POST":
        KYCSTATUS=request.POST.get('KYCSTATUS');
        post=Company.objects.get(id=id);
        post.KYCSTATUS=KYCSTATUS;
        post.save();
        return redirect('company_detail')
    return render(request,"admin-template/company_edit2.html")



# def masterctcs(request):

#     s = adminnav.objects.all()
#     a = companylogo.objects.all()
#     admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
#     user = CustomUser.objects.filter(id=request.user.id).first()
#     userid1 = user.id
#     da1 = AdminHod.objects.filter(admin=request.user.id).first()
#     da2 = da1.id
#     admin = AdminHod.objects.get(id=1)
#     projectm = admin_project_create.objects.filter(admin_id=userid1)
#     h = HR.objects.all()
#     s=adminnav.objects.all()
#     employs_all = Employs.objects.all()
#     projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
#     data = Companys.objects.filter(usernumber=request.user.id).first()
#     data1=data.id
#     admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
#     admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')


#     employ=Employs.objects.all()
#     cursor = connection.cursor()
#     cursor.execute("SELECT ehrms_employs.empid, ehrms_customuser.first_name, ehrms_customuser.email, ehrms_employs.role, ehrms_employs.location, ehrms_employs.status, ehrms_employs.dateofjoining, ehrms_employs.dateofbirth, ehrms_employs.package FROM ehrms_employs INNER JOIN ehrms_customuser ON ehrms_employs.id=ehrms_customuser.id;")

#     all_data = cursor.fetchall()
#     items_per_page = 10  # Adjust the number of items per page as needed

#     paginator = Paginator(all_data, items_per_page)
#     page = request.GET.get('page')

#     try:
#         rs = paginator.page(page)
#     except PageNotAnInteger:
#         rs = paginator.page(1)
#     except EmptyPage:
#         rs = paginator.page(paginator.num_pages)

#     return render(request, 'admin-template/masterctc.html', {'rs': rs, 's': s, 'admin_drops': admin_drops, 'a': a, 'data':data,'admin_drops':admin_drops,'employ':employ,'admin_drops': admin_drops,'user':user,'da1':da1,'da2':da2,'data':data,'admin_home_drops':admin_home_drops,'projects_drops':projects_drops,'employs_all':employs_all,'s':s,'h':h,'admin':admin})


def masterctcs(request):
    s = adminnav.objects.all()
    # a = companylogo.objects.all()
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1 = data.id
    a=company_details.objects.filter(companyid=data1).first()
    rst = Employs.objects.filter(companyid=data1)

    # Pagination
    page = request.GET.get('page', 1)
    items_per_page = 1
    paginator = Paginator(rst, items_per_page)

    try:
        rst = paginator.page(page)
    except PageNotAnInteger:
        rst = paginator.page(1)
    except EmptyPage:
        rst = paginator.page(paginator.num_pages)

    return render(request, 'admin-template/masterctc.html', {'data': data, 'rst': rst, 's': s, 'admin_drops': admin_drops, 'a': a})

def all_employees_reimbursement_report(request):
    
    staff_obj = Employs.objects.all()
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id

    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    a=company_details.objects.filter(companyid=data1).first()




    s=adminnav.objects.all()
    k = types.objects.all()

    st4 = request.POST.get("ss")
    st3 = request.POST.get("vk")
    st = request.POST.get("d1")
    st1 = request.POST.get("d2")

    leave_data = Reimbursement.objects.filter(employ_id__companyid=data1)

    # Apply filters based on user input
    if st4 and st4 != '----Select----':  # Check if a valid status is selected
        leave_data = leave_data.filter(reimbursement_status=st4)
    if st3:  # If "Select Type" is selected
        leave_data = leave_data.filter(typea__icontains=st3)
    if st and st1:
        leave_data = leave_data.filter(date__range=[st, st1])

    total_approved = leave_data.filter(reimbursement_status=1).aggregate(Sum('amount'))['amount__sum'] or 0
    total_pending = leave_data.filter(reimbursement_status=0).aggregate(Sum('amount'))['amount__sum'] or 0
    items_per_page = 10  # Adjust the number of items per page as needed

    paginator = Paginator(leave_data, items_per_page)
    page = request.GET.get('page')

    try:
        leave_data = paginator.page(page)
    except PageNotAnInteger:
        leave_data = paginator.page(1)
    except EmptyPage:
        leave_data = paginator.page(paginator.num_pages)
    return render(request, "admin-template/all_employees_reimbursement_report.html", {
        'leave_data': leave_data,
        's': s,
        'k': k,
        'total': total_approved,
        'total1': total_pending,
        'staff_obj':staff_obj,
        'admin_drops':admin_drops,
        'user':user,
        'a':a,
        'admin_home_drops':admin_home_drops,
        'h':h,
        'projectm':projectm,
        'projects_drops':projects_drops,
        'data':data,
        'employs_all':employs_all

    })

def reimbursements_approve(request,leave_id):
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    leave=Reimbursement.objects.get(id=leave_id)
    leave.reimbursement_status=1
    leave.save()
    if leave:
        messages.success(request,"Reimbursement as been approve")
   
    
    return redirect(request.META.get('HTTP_REFERER', '/'))
from django.shortcuts import get_object_or_404, HttpResponseRedirect
from django.contrib import messages
from django.urls import reverse

def reimbursements_disapprove(request, leave_id):
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    
    try:
        leave = Reimbursement.objects.get(id=leave_id)
        leave.reimbursement_status = 2
        leave.save()
        messages.success(request, "Reimbursement has been disapproved")
        
        # Redirect to admin_reimbursement_apply_view and pass a message
        return HttpResponseRedirect(reverse('admin_reimbursement_apply_view'))
    except Reimbursement.DoesNotExist:
        messages.error(request, "Reimbursement not found")
        return HttpResponseRedirect(reverse('admin_reimbursement_apply_view'))  # Redirect to the reimbursement view with an error message

# def reimbursements_disapprove(request,leave_id):
#     admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
#     leave=Reimbursement.objects.get(id=leave_id)
#     leave.reimbursement_status=2
#     leave.save()
#     if leave:
#         messages.success(request,"Reimbursement as been disapprove")
#     return HttpResponseRedirect(reverse("all_employees_reimbursement_report", args=[leave.employ_id_id]))


# def register1(request):
#     s=adminnav.objects.all()
#     admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
#     user = CustomUser.objects.filter(id=request.user.id).first()
#     userid1 = user.id
#     da1 = AdminHod.objects.filter(admin=request.user.id).first()
#     da2 = da1.id
#     admin = AdminHod.objects.get(id=1)
#     s = adminnav.objects.all()
#     projectm = admin_project_create.objects.filter(admin_id=userid1)
#     h = HR.objects.all()
#     employs_all = Employs.objects.all()
#     admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
#     admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
#     data = AdminHod.objects.filter(id=request.user.id)
#     projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
#     query = request.GET.get('query')
#     x=connection.cursor()
#     x.execute("SELECT ehrms_employs.id, ehrms_employs.empid, ehrms_employs.first_name, ehrms_employs.last_name, ehrms_employs.email, ehrms_employ_add_form.phno, ehrms_employs.gender, ehrms_employ_add_form.pan, ehrms_employs.dateofjoining, ehrms_employs.role FROM ehrms_employs INNER JOIN ehrms_employ_add_form ON ehrms_employs.email = ehrms_employ_add_form.email2;")
#     rs=x.fetchall()
#     items_per_page = 10  # Adjust the number of items per page as needed

#     paginator = Paginator(rs, items_per_page)
#     page = request.GET.get('page')

#     try:
#         res = paginator.page(page)
#     except PageNotAnInteger:
#         res = paginator.page(1)
#     except EmptyPage:
#         res = paginator.page(paginator.num_pages)

    # paginator = Paginator(rs, 10)  # Show 10 items per page
    # page_number = request.GET.get('page')
    # res = paginator.get_page(page_number)
    # return render(request, 'admin-template/register_data.html', {'res': res,'s':s,'admin_drops':admin_drops,'query':query, 'admin_drops':admin_drops,
    #     'user':user,
    #     'da1':da1,
    #     'da2':da2,
    #     'admin':admin,
    #     'admin_home_drops':admin_home_drops,
    #     'h':h,
    #     'projectm':projectm,
    #     'projects_drops':projects_drops,
    #     'data':data,
    #     'employs_all':employs_all})


def register1(request):
    s = adminnav.objects.all()
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    query = request.GET.get('query')
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1 = data.id
    rts = employ_add_form.objects.filter(student_id__companyid__usernumber=request.user.id)
    a=company_details.objects.filter(companyid=data1).first()

    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(rts, 2)  # Show 10 items per page
    try:
        rts = paginator.page(page)
    except PageNotAnInteger:
        rts = paginator.page(1)
    except EmptyPage:
        rts = paginator.page(paginator.num_pages)

    return render(request, 'admin-template/register_data.html', {'rts': rts, 's': s, 'a':a, 'admin_drops': admin_drops, 'query': query,'data':data})

def audit(request):
    s=adminnav.objects.all()
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    query = request.GET.get('query')
    # rs = Register_data.objects.select_related('Students_id','employ_form_id','admin').all()
    x=connection.cursor()
    x.execute("SELECT ehrms_checkin.date,ehrms_checkin.time,ehrms_employs.first_name,ehrms_employs.last_name FROM ehrms_employs INNER JOIN ehrms_checkin ON ehrms_checkin.empid=ehrms_employs.email;")
    rs=x.fetchall()
    return render(request,'admin-template/audit_report.html',{'rs':rs,'s':s,'admin_drops':admin_drops})


def reimreport(request):
    r=Reimbursement.objects.filter(reimbursement_status=1)
    employ=Employs.objects.all()
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    return render(request,"admin-template/reportrmb.html",{'admin_drops':admin_drops,'r':r,'employ':employ})

# def documentreport(request):
#     admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
#     a=companylogo.objects.all()
#     x=connection.cursor()
#     x.execute("SELECT ehrms_employs.first_name,ehrms_empdocs.date,ehrms_empdocs.documenttype1,ehrms_empdocs.description,ehrms_empdocs.imagefile FROM ehrms_employs INNER JOIN ehrms_empdocs on ehrms_employs.id=ehrms_empdocs.employ_id_id;")
#     doc=x.fetchall()
#     return render (request,"admin-template/docreport.html",{'doc':doc,'admin_drops':admin_drops,'a':a})

def documentreport(request):
 
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()

    employ=empdocs.objects.filter(employ_id__companyid=data1)

    items_per_page = 10  # Adjust the number of items per page as needed

    paginator = Paginator(employ, items_per_page)
    page = request.GET.get('page')

    try:
        employ = paginator.page(page)
    except PageNotAnInteger:
        employ = paginator.page(1)
    except EmptyPage:
        employ = paginator.page(paginator.num_pages)


    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id

    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')





    return render (request,"admin-template/docreport.html",{'employ':employ, 'a':a, 'admin_drops':admin_drops,'user':user,'admin_home_drops':admin_home_drops,'s':s,'h':h,'projectm':projectm,'projects_drops':projects_drops,'data':data,'employs_all':employs_all})

 
import zipfile
from django.http import HttpResponse
from io import BytesIO
from PIL import Image
from .models import Employs, empdocs
import base64

def image_to_base64(image_path):
    with open(image_path, "rb") as image_file:
        image_data = image_file.read()
        base64_encoded = base64.b64encode(image_data).decode("utf-8")
    return base64_encoded

def generate_zip(request):
    # Create a BytesIO buffer to write the ZIP file to
    zip_buffer = BytesIO()

    # Create a ZIP file object
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
        # Query your database to retrieve the 'Employs' documents
        employs_documents = Employs.objects.all()

        # Loop through the retrieved 'Employs' documents and add their content to the ZIP file
        for employ in employs_documents:
            # Create an HTML file for each employee
            html_content = f"<html><body><h2>Employee: {employ.first_name}</h2>"

            empdocs_documents = empdocs.objects.filter(employ_id=employ)

            # Loop through the related 'empdocs' documents for this 'Employs' record 
            for empdoc in empdocs_documents:
                document_content = f"<p>Document Type: {empdoc.documenttype}</p>"

                # Get the image URL for this empdoc
                image_path = empdoc.imagefile.path
                image_filename = empdoc.imagefile.name.split("/")[-1]  # Get the image file name

                # Convert the image to base64
                image_base64 = image_to_base64(image_path)

                # Embed the base64-encoded image in the HTML
                image_tag = f'<img src="data:image/png;base64,{image_base64}" alt="{image_filename}">'
                document_content += image_tag

                # Add the document content to the HTML file
                html_content += document_content

            # Close the HTML tags for this employee
            html_content += "</body></html>"

            # Add the HTML content to the ZIP file
            html_name = f'documents/{employ.first_name}_info.html'
            zipf.writestr(html_name, html_content)

    # Close and seek to the beginning of the buffer
    zip_buffer.seek(0)

    # Create an HttpResponse with the ZIP file as content
    response = HttpResponse(zip_buffer.read(), content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="document_files.zip"'

    return response


from django.core.mail import send_mail
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Employs, empdocs, employ_add_form
from django.http import Http404

from django.core.exceptions import ObjectDoesNotExist

def all_employees_missing_info(request):
    
    
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id 
    a=company_details.objects.filter(companyid=data1).first()

    s = adminnav.objects.all()
   
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.filter(companyid=data1)
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    


    # a=companylogo.objects.all()
    employees_list=Employs.objects.all()

    items_per_page = 10  # Adjust the number of items per page as needed
    paginator = Paginator(employees_list, items_per_page)
    page = request.GET.get('page')

    try:
        employees = paginator.page(page)
    except PageNotAnInteger:
        employees = paginator.page(1)
    except EmptyPage:
        employees = paginator.page(paginator.num_pages)
    if request.method == 'POST':
        selected_employee_ids = request.POST.getlist('selected_employees')

        # Loop through the selected employee IDs and send emails
        for employee_id in selected_employee_ids:
            try:
                employee = Employs.objects.get(id=employee_id)
                employee_email = employee.email
                try:
                    employ_add_form_info = employ_add_form.objects.get(student_id=employee)
                except employ_add_form.DoesNotExist:
                    employ_add_form_info = None

                empdocs_info = empdocs.objects.filter(employ_id=employee)

                # Compose the email message with missing information
                missing_info = []
                if employee.missing_info():
                    missing_info.append(", ".join(employee.missing_info().keys()))
                if employ_add_form_info and employ_add_form_info.missing_info():
                    missing_info.append(", ".join(employ_add_form_info.missing_info().keys()))
                for empdoc in empdocs_info:
                    if empdoc.missing_info():
                        missing_info.append(", ".join(empdoc.missing_info().keys()))
                        
                if missing_info:
                    subject = f"Mr/Miss {employee.first_name}.{employee.last_name} Please Update Your Missing Information."
                    message = f"Dear {employee.first_name} .{employee.last_name},\n\n"
                    message += f"The following information is missing or incomplete:\n"
                    message += "\n".join(missing_info)
                    # message += "\n\nPlease Login to DevelopTrees HRMS to update your information as soon as possible."
                    message += "\n\nPlease click the following link to update your information as soon as possible.\n"
                    # Generate the profile URL using the employee's ID
                    # message += "\n"
                    # profile_url = reverse('employ_profile', kwargs={'employee_id': employee_id})
                    # message += request.build_absolute_uri(profile_url)

                    from_email = settings.EMAIL_HOST_USER  # Replace with your email address
                    recipient_list = [employee.email]  # Send email to the employee's email address

                    send_mail(subject, message, from_email, recipient_list)

            except Employs.DoesNotExist:
                pass  # Handle the case where an employee does not exist
    
        return redirect('all_employees_missing_info')

    employees = Employs.objects.filter(companyid=data1)
    employee_data = []
    
    for employee in employees:
        try:
            employ_add_form_info = employ_add_form.objects.get(student_id=employee)
        except ObjectDoesNotExist:
            employ_add_form_info = None

        employee_info = {
            'employs': employee,
            'empdocs_info': empdocs.objects.filter(employ_id=employee),
            'employ_add_form_info': employ_add_form_info,
        }
        employee_data.append(employee_info)
    employees_list = Employs.objects.filter(companyid=data1)
    
    items_per_page = 10  # Adjust the number of items per page to 10
    paginator = Paginator(employees_list, items_per_page)  #

    page = request.GET.get('page')
    try:
        employees = paginator.page(page)
    except PageNotAnInteger:
        employees = paginator.page(1)
    except EmptyPage:
        employees = paginator.page(paginator.num_pages)
    context = {
        'employee_data': employee_data,
        'admin_drops':admin_drops,
        'a':a,
        'user':user,
        'admin_home_drops':admin_home_drops,
        's':s,
        'h':h,
        'projectm':projectm,
        'projects_drops':projects_drops,
        'data':data,
        'employs_all':employs_all,
        'employees_data':employees

    }

    return render(request, 'admin-template/all_employees_missing_info1.html', context)



def doct_allemp_email(request):
    e = CustomUser.objects.filter(is_active="1") 
    subject = 'Important Notification'
    email_template = "doct_activeemp_email.html"  # Update with the correct path

    recipient_list = []  # Initialize an empty list for recipients

    for i in e:
        message = render_to_string(email_template, {'i': i})
        recipient_list.append(i.email)  # Add email to the recipient list

    if recipient_list:
        # Send the email to all recipients at once
        send_mail(subject, '', 'bethapudianil1999@gmail.com', recipient_list, html_message=message)
        return HttpResponse("Email sent successfully.")
    else:
        return HttpResponse("No employees found to send emails to.")
    
def doct_inactiveemp_email(request):
    e = CustomUser.objects.filter(is_active="0") 
    subject = 'Important Notification'
    email_template = "doct_inactive_email.html"  # Update with the correct path

    recipient_list = []  # Initialize an empty list for recipients

    for i in e:
        message = render_to_string(email_template, {'i': i})
        recipient_list.append(i.email)  # Add email to the recipient list

    if recipient_list:
        # Send the email to all recipients at once
        send_mail(subject, '', 'saipathivada1234@gmail.com', recipient_list, html_message=message)
        return HttpResponse("Emails sent successfully to all inactive employees.")
    else:
        return HttpResponse("No employees found to send emails to.")    
    


from django.shortcuts import render
from .models import Reimbursement, Employs
from datetime import datetime
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

def yearmonth_uploaded(request):
    # a = companylogo.objects.all()
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id

    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()
    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops = admin_home_drop.objects.filter(parent_category=None).order_by('id')
    # data = AdminHod.objects.filter(id=request.user.id)
    projects_drops = project_drop.objects.filter(parent_category=None).order_by('id')
    # data2 = Companys.objects.filter(usernumber=request.user.id).first()
    # data1 = data2.id
    rst = Employs.objects.filter(companyid=data1)
    ad = ad_salary.objects.first()
    b = Reimbursement.objects.first()

    # Assuming 'rst' is a queryset of Employs objects, modify the following line
    # to filter Reimbursement objects based on the employ_id relationship.
    reimbursments = Reimbursement.objects.filter(employ_id__in=rst)

    page = request.GET.get('page', 1)
    items_per_page = 1  # Adjust the number of items per page as needed
    paginator = Paginator(reimbursments, items_per_page)

    try:
        reimbursments = paginator.page(page)
    except PageNotAnInteger:
        reimbursments = paginator.page(1)
    except EmptyPage:
        reimbursments = paginator.page(paginator.num_pages)

    return render(request, "admin-template/salaryreport.html", {
        'reimbursments': reimbursments,
        'a': a,
        'b': b,
        'admin_drops': admin_drops,
        'ad': ad,
        'user': user,   
        'admin_home_drops': admin_home_drops,
        's': s,
        'h': h,
        'projectm': projectm,
        'projects_drops': projects_drops,
        'data': data,
        'employs_all': employs_all,
        'paginator': paginator,
        'rst':rst
    })

from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.http import HttpResponse  # Import HttpResponse
from .models import Employs  # Make sure to import your Employs model

def allemp_email(request):
    e = CustomUser.objects.filter(is_active="1") 
    subject = 'Important Notification'
    email_template = "activeemp_email.html"  # Update with the correct path

    recipient_list = []  # Initialize an empty list for recipients

    for i in e:
        message = render_to_string(email_template, {'i': i})
        recipient_list.append(i.email)  # Add email to the recipient list

    if recipient_list:
        # Send the email to all recipients at once
        send_mail(subject, '', 'saipathivada1234@gmail.com', recipient_list, html_message=message)
        return HttpResponse("Emails sent successfully to all employees.")
    else:
        return HttpResponse("No employees found to send emails to.")
    
def inactiveemp_email(request):
    e = CustomUser.objects.filter(is_active="0") 
    subject = 'Important Notification'
    email_template = "inactive_email.html"  # Update with the correct path

    recipient_list = []  # Initialize an empty list for recipients

    for i in a:
        message = render_to_string(email_template, {'i': i})
        recipient_list.append(i.email)  # Add email to the recipient list

    if recipient_list:
        # Send the email to all recipients at once
        send_mail(subject, '', 'saipathivada1234@gmail.com', recipient_list, html_message=message)
        return HttpResponse("Emails sent successfully to all inactive employees.")
    else:
        return HttpResponse("No employees found to send emails to.")    
    



def admin_tax_table(request):
    employ=Employs.objects.all()
    tax=admin_tax_form.objects.all()
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    a=companylogo.objects.all()
    leave_date=LeaveReportEmploy.objects.filter()
    return render(request,"admin-template/admin_tax_table.html",{'employ':employ,'tax':tax,'leave_date':leave_date,'admin_drops':admin_drops,'a':a})

def total(request):
    employ=Employs.objects.all()
    a=companylogo.objects.all()
    return render(request,"people.html",{'employ':employ,'a':a})

def edit_people_admin(request,std):
    request.session['employ_id']=std
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = AdminHod.objects.filter(admin=request.user.id).first()
    
    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')


    # a=companylogo.objects.all()
    employ=Employs.objects.get(admin=std)
    return render(request,"admin-template/example.html",{'employ':employ,'admin_drops':admin_drops,'a':a,'user':user,'da1':da1,'admin_home_drops':admin_home_drops,'s':s,'h':h,'projectm':projectm,'projects_drops':projects_drops,'data':data,'employs_all':employs_all})





def people_count(request):
    
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()

    projects_drops = project_drop.objects.filter(parent_category=None).order_by('id')
    s = adminnav.objects.all()
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    # a = companylogo.objects.all()
    employs = Employs.objects.filter(companyid=data1)
    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = AdminHod.objects.filter(admin=request.user.id).first()
    if da1:
      da2 = da1.id
    else:
        da2=None
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')



    emp = CustomUser.objects.filter(employs__companyid=data1, is_active=True).count()  # Count active employees
    dis = CustomUser.objects.filter(employs__companyid=data1, is_active=False).count()  # Count inactive employees
    total_employees_count = emp + dis
    
    items_per_page = 10  # Adjust the number of items per page as needed

    # Retrieve all employees for pagination
    employees = Employs.objects.filter(companyid=data1)

    paginator = Paginator(employees, items_per_page)
    page_number = request.GET.get('page')

    try:
        employees = paginator.page(page_number)
    except PageNotAnInteger:
        employees = paginator.page(1)
    except EmptyPage:
        employees = paginator.page(paginator.num_pages)
    
    return render(request, "admin-template/peoples.html", {
        's': s,
        'admin_drops': admin_drops,
        'employees': employees,
        'employs': employs,
        'emp': emp,
        'dis': dis,
        'total_employees_count': total_employees_count,
        'admin_drops': admin_drops,
        'a': a,
        'projects_drops': projects_drops,
        'user':user,
        'da1':da1,
        'da2':da2,
        'admin_home_drops':admin_home_drops,
        'h':h,
        'projectm':projectm,
        'data':data,'data1':data1,'employs_all':employs_all
    })
    
    
    
    
    

def upload_excel(request,message=None, error_message=None):
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = AdminHod.objects.filter(admin=request.user.id).first()
    # da2 = da1.id
    # admin = AdminHod.objects.get(id=1)
    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    # data = AdminHod.objects.filter(id=request.user.id)
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')

    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()

    # a=companylogo.objects.all()
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']

        if excel_file.name.endswith('.xlsx'):
            try:
                df = pd.read_excel(excel_file)

                # Assuming the columns in the Excel file match the fields in Employs model
                for index, row in df.iterrows():
                    user=CustomUser.objects.create_user(username=row['username'],password=row['empid'],first_name=row['first_name'],last_name=row['last_name'],email=row['email'],user_type=2)
                    user.employs.first_name=row['first_name']
                    user.employs.last_name=row['last_name']
                    user.employs.email=row['email']
                    user.employs.address=row['address']
                    user.employs.empid=row['empid']
                    user.employs.role=row['role']
                    user.employs.location=row['location']
                    user.employs.package=row['package']
                    user.employs.pincode=row['pincode']
                    user.employs.contactno=row['contactno']
                    user.employs.gender=row['sex']
                    date_string = str(row['dateofjoining'])
                    date_obj = datetime.strptime(date_string, '%Y-%m-%d %H:%M:%S')  # Adjust the format as per your Excel file
                    user.employs.dateofjoining = date_obj.date()
                    user.save()
                    html_content = render_to_string("email_template.html",{'title':'test email','first_name':row['first_name'],'last_name':row['last_name'],'empid':row['empid']})
                    text_content=strip_tags(html_content)
                    subject="WELCOME TO DEVELOPTRESS"
                    email = EmailMultiAlternatives(
                            subject,
                            text_content,
                            settings.EMAIL_HOST_USER,
                            [row['email']],
                            )
                    email.attach_alternative(html_content,"text/html")
                    email.fail_silently = True
                    email.send()
                    
                message = f"{len(df)} records imported successfully."  
                return render(request, 'admin-template/upload_bulk_data.html',{'message':message})
            except Exception as e:
                # Handle any exceptions that may occur during processing
                # return HttpResponseRedirect(reverse("upload_excel"))
                error_message = f"{str(e)} An error occurred during processing."
        else:
                error_message = 'Please upload a valid Excel file.'
        
    return render(request, 'admin-template/upload_bulk_data.html', {'error_message': error_message, 'admin_drops': admin_drops,'a': a,'user':user,'da1':da1,'admin_home_drops':admin_home_drops,'s':s,'h':h,'projectm':projectm,'a':a,'projects_drops':projects_drops,'data':data,'employs_all':employs_all})





def edit_people_admin(request,std):
    work=working_shifts.objects.all()
    request.session['employ_id']=std
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = AdminHod.objects.filter(admin=request.user.id).first()
    # da2 = da1.id
    # admin = AdminHod.objects.get(id=1)
    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')


    # a=companylogo.objects.all()
    employ=Employs.objects.get(admin=std)
    objss=employ.id 
    users=CustomUser.objects.get(id=std)
    datas=employ_add_form.objects.filter(student_id=objss)
    return render(request,"admin-template/people_employ.html",{'work':work,'a':a,'employ':employ,'admin_drops':admin_drops,'datas':datas,'objss':objss,'users':users, 'user':user,'da1':da1,'admin_home_drops':admin_home_drops,'s':s,'h':h,'projectm':projectm,'projects_drops':projects_drops,'data':data,'employs_all':employs_all})


from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from .models import Employs  # Import your Employee model

def update_shift(request, id):
    if request.method == 'POST':
        employee = get_object_or_404(Employs, id=id)
        new_shift = request.POST.get('working12')

        if new_shift:
            employee.working12 = new_shift
            employee.save()
            return JsonResponse({'success': 'Shift updated successfully'})
        else:
            return JsonResponse({'error': 'Invalid shift selection'})

    return JsonResponse({'error': 'Invalid request'})


def enable_button(request, std):

    request.session['employ_id']=std
    instance = CustomUser.objects.get(id=std)
    instance.is_active = True
    instance.save()
    return redirect('edit_people_admin', std=std)  # Redirect to the detail view

def disable_button(request, std):
    request.session['employ_id']=std
    instance = CustomUser.objects.get(id=std)
    instance.is_active = False
    instance.save()
    return redirect('edit_people_admin',std=std)

def individual_employ_tax_report(request,std):
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = AdminHod.objects.filter(admin=request.user.id).first()
    # da2 = da1.id
    # admin = AdminHod.objects.get(id=1)
    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')


    # a=companylogo.objects.all()
    request.session['employ_id']=std
    employ=Employs.objects.all()
    tax=employ_tax_form.objects.filter(employ=std)
    
    leave_date=LeaveReportEmploy.objects.filter()
    return render(request,"admin-template/individual_employ_taxReport.html",{'a':a,'employ':employ,'admin_drops':admin_drops,'tax':tax,'leave_date':leave_date,'user':user,'da1':da1,'admin_home_drops':admin_home_drops,'s':s,'h':h,'projectm':projectm,'projects_drops':projects_drops,'data':data,'employs_all':employs_all})

def individual_documentreport(request,std):
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = AdminHod.objects.filter(admin=request.user.id).first()
    # da2 = da1.id
    # admin = AdminHod.objects.get(id=1)
    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')



    # a=companylogo.objects.all()
    request.session['employ_id']=std
    employ_id=Employs.objects.all()
    doc=empdocs.objects.filter(employ_id=std)
 
    return render (request,"admin-template/individual_docreport.html",{'a':a,'doc':doc,'employ_id':employ_id,'admin_drops':admin_drops,'user':user,'da1':da1,'admin_home_drops':admin_home_drops,'s':s,'h':h,'projectm':projectm,'projects_drops':projects_drops,'data':data,'employs_all':employs_all})


# def individual_employ_attendance(request,std):
#     admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
#     request.session['employ_id']=std
#     employ_id=Employs.objects.all()
#     attend=checkin.objects.filter(employ_id=std)
 
#     return render (request,"admin-template/individual_attendance_report.html",{'attend':attend,'employ_id':employ_id,'admin_drops':admin_drops})


def individual_reimbursement_view(request,std):
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = AdminHod.objects.filter(admin=request.user.id).first()
    # da2 = da1.id
    # admin = AdminHod.objects.get(id=1)
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')


    # a=companylogo.objects.all()
    request.session['employ_id']=std
    s=adminnav.objects.all()
    employ_id=Employs.objects.all()
    leaves=Reimbursement.objects.filter(employ_id=std)
    items_per_page = 10  # Adjust the number of items per page as needed

    # Retrieve all employees for pagination
    employees = Employs.objects.all()

    paginator = Paginator(employees, items_per_page)
    page_number = request.GET.get('page')

    try:
        leaves = paginator.page(page_number)
    except PageNotAnInteger:
        leaves = paginator.page(1)
    except EmptyPage:
        leaves = paginator.page(paginator.num_pages)
    return render(request,"admin-template/individual_reimbursement.html",{'a':a,"leaves":leaves,'s':s,"employ_id":employ_id,'admin_drops':admin_drops, 'user':user,'da1':da1,'admin_home_drops':admin_home_drops,'h':h,'projectm':projectm,'projects_drops':projects_drops,'data':data,'employs_all':employs_all})


def reimbursement_approve_status(request,leave_id):
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    leave=Reimbursement.objects.get(id=leave_id)
    leave.reimbursement_status=1
    leave.save()
    user=CustomUser.objects.all()
    return HttpResponseRedirect(reverse("individual_reimbursement_view", args=[leave.employ_id_id]))

def reimbursement_disapprove_status(request,leave_id):
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    leave=Reimbursement.objects.get(id=leave_id)
    leave.reimbursement_status=2
    leave.save()
    return HttpResponseRedirect(reverse("individual_reimbursement_view", args=[leave.employ_id_id]))

def individual_leaveReport(request,std):
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = AdminHod.objects.filter(admin=request.user.id).first()
    # da2 = da1.id
    # admin = AdminHod.objects.get(id=1)
    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')


    # a=companylogo.objects.all()
    request.session['employ_id']=std
    employ_id=Employs.objects.all()
    leaves=LeaveReportEmploy.objects.filter(employ_id=std)
    return render(request,"admin-template/individual_employ_leave_view.html",{'a':a,"leaves":leaves,'admin_drops':admin_drops,'employ_id':employ_id,'user':user,'da1':da1,'admin_home_drops':admin_home_drops,'s':s,'h':h,'projectm':projectm,'projects_drops':projects_drops,'data':data,'employs_all':employs_all})


def employ_approve_leave(request,leave_id):
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    leave=LeaveReportEmploy.objects.get(id=leave_id)
    leave.leave_status=1
    leave.save()
    return HttpResponseRedirect(reverse("individual_leaveReport", args=[leave.employ_id]))

def employ_disapprove_leave(request,leave_id):
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    leave=LeaveReportEmploy.objects.get(id=leave_id)
    leave.leave_status=2
    leave.save()
    return HttpResponseRedirect(reverse("individual_leaveReport", args=[leave.employ_id_id]))
def employ_approve(request,leave_id):
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    leave=LeaveReportEmploy.objects.get(id=leave_id)
    leave.leave_status=1
    leave.save()
    return HttpResponseRedirect(reverse("employ_leave_view"))

def employ_disapprove(request,leave_id):
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    leave=LeaveReportEmploy.objects.get(id=leave_id)
    leave.leave_status=2
    leave.save()
    return HttpResponseRedirect(reverse("employ_leave_view"))


def individual_advancesalary_Report(request,std):
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = AdminHod.objects.filter(admin=request.user.id).first()
    # da2 = da1.id
    # admin = AdminHod.objects.get(id=1)
    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')



    # a=companylogo.objects.all()
    request.session['employ_id']=std
    employ_id=Employs.objects.all()
    leaves=ad_salary.objects.filter(employ_id=std)
    return render(request,"admin-template/individual_advancesalary_Report.html",{'a':a,"leaves":leaves,'admin_drops':admin_drops,'employ_id':employ_id,'user':user,'da1':da1,'admin_home_drops':admin_home_drops,'s':s,'h':h,'projectm':projectm,'projects_drops':projects_drops,'data':data,'employs_all':employs_all})


def employ_approve_advancesalary(request,leave_id):
    # admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    leave=ad_salary.objects.get(id=leave_id)
    leave.leave_status=1
    leave.save()
    return HttpResponseRedirect(reverse("individual_advancesalary_Report", args=[leave.employ_id]))

def employ_disapprove_advancesalary(request,leave_id):
    # admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    leave=ad_salary.objects.get(id=leave_id)
    leave.leave_status=2
    leave.save()
    return HttpResponseRedirect(reverse("individual_advancesalary_Report", args=[leave.employ_id_id]))

def pdf_report_create(request):
    
    # user=CustomUser.objects.get(id=request.user.id)
    # student=Employ.objects.get(admin=user)
    employ=Employs.objects.all()
    tax=admin_tax_form.objects.all()
    template_path = 'admin_tax_pdfReport.html'

    context = {'tax': tax,'employ':employ}

    response = HttpResponse(content_type='application/pdf')

    response['Content-Disposition'] = 'filename="products_report.pdf"'

    template = get_template(template_path)

    html = template.render(context)

    # create a pdf
    pisa_status = pisa.CreatePDF(
       html, dest=response)
    # if error then show some funy view
    if pisa_status.err:
       return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response


def my_view(request):
    ptax = p_tax.objects.all()
    pfund = PF.objects.all()
    esicdata = ESIC.objects.all()
    
    return render(request, 'admin-template/company_edit3.html', {'ptax': ptax, 'pfund': pfund, 'esicdata': esicdata})


def update_pt(request,id):
    pt = p_tax.objects.get(pk=id)
    pt.professional_tax = request.POST.get('new_field1_value')
    pt.save()
    return render(request,"admin-template/company_edit3")

def update_pf(request,id):
    pf = PF.objects.get(pk=id)
    pf.pf_status = request.POST.get('new_field2_value')
    pf.save()
    return render(request,"admin-template/company_edit3")

def update_esic(request,id):
    esic = ESIC.objects.get(pk=id)
    esic.esi_status = request.POST.get('new_field3_value')
    esic.save()
    return render(request,"admin-template/company_edit3")




def search(request):
    s = adminnav.objects.all()
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')




    departments = Department.objects.all()
    employees = Employs.objects.filter(companyid=data1)

    search_term = request.GET.get('search_term')
    designation = request.GET.get('department')
    distinct_department_designations = employees.values('designation').distinct()
    selected_date_range = request.GET.get('dateofjoining')

    employees_by_designation = {}

    if designation == "":
        designation_employees = employees.all()
    else:
        designation_employees = employees.filter(designation=designation)

    if selected_date_range == "":
        pass
    else:
        try:
            selected_date_range = int(selected_date_range)
            today = datetime.now().date()
            start_date = today - timedelta(days=selected_date_range * 30)
            designation_employees = designation_employees.filter(dateofjoining__gte=start_date)
        except (ValueError, TypeError):
            # Handle the case where 'selected_date_range' is not a valid number (e.g., it's None or a non-integer value)
            pass

    if search_term:
        designation_employees = designation_employees.filter(first_name__icontains=search_term)

    total_count = designation_employees.count()
    total_package = designation_employees.aggregate(Sum('package'))['package__sum']

    employees_by_designation[designation] = designation_employees

    labels = []
    salaries = []
    items_per_page = 10  # Adjust the number of items per page as needed
    paginator = Paginator(employees, items_per_page)
    page = request.GET.get('page')

    try:
        employees = paginator.page(page)
    except PageNotAnInteger:
        employees = paginator.page(1)
    except EmptyPage:
        employees = paginator.page(paginator.num_pages)


    context = {
        'employees': employees_by_designation,
        'total_count': total_count,
        'total_package': total_package,
        'labels': labels,
        'salaries': salaries,
        's': s,
        'admin_drops': admin_drops,
        'departments': distinct_department_designations,
        'selected_date_range': selected_date_range,
        'user':user,
        'a':a,
        'admin_home_drops':admin_home_drops,
        'h':h,
        'projectm':projectm,
        'projects_drops':projects_drops,
        'data':data,
        'employs_all':employs_all

    }

    return render(request, 'admin-template/variance_report.html', context)






def home_int(request):
    s=adminnav.objects.all()
    int=integrations.objects.all()
    obj=Employs.objects.all().count()
    objs=Employs.objects.all()
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    if request.method=="POST":
        response=request.POST["response"]
        k=app(response=response)
        k.save()
        return redirect("/home_int")
    return render(request,"admin-template/int_template.html",{'int':int,'s':s,'obj':obj,'objs':objs,'admin_drops':admin_drops})


def payment_data(request):
     k = admin_add_form1.objects.first()
    #   return redirect("/admin_profile")
     return render(request, "admin-template/admin_file.html", {'k': k})


def update_pay_info(request,id):
    if (request.method=="POST"):
        pan=request.POST.get("pan")
        ifsecode=request.POST.get("ifsecode")
        acno=request.POST.get("acno")
        beneficiaryname=request.POST.get("beneficiaryname")
       
        k=admin_add_form1.objects.get(id=id);
        
        k.pan=pan
        k.ifsecode=ifsecode
        k.acno=acno
        k.beneficiaryname=beneficiaryname
        k.save();
        return HttpResponseRedirect(reverse("admin_profile"))
    return render(request,"admin-template/update_pay_info.html")


def otherinfo(request):
    k = admin_add_form1.objects.first()

    return render(request,"admin-template/admin_file2.html",{'k':k})


def update_other_info(request,id):
    if (request.method=="POST"):
        phno=request.POST.get("phno")
        gender1=request.POST.get("gender1")
        dob=request.POST.get("dob")
        address1=request.POST.get("address1")
        heq=request.POST.get("heq")
        aadharno=request.POST.get("aadharno")
       
        bloodgroup=request.POST.get("bloodgroup")

        k=admin_add_form1.objects.get(id=id);
        
        k.phno=phno
        k.gender1=gender1
        k.dob=dob
        k.address1=address1
        k.heq=heq
        k.aadharno=aadharno
     
        k.bloodgroup=bloodgroup
        k.save();
        return HttpResponseRedirect(reverse("admin_profile"))
    return render(request,"admin-template/admin_file2.html")



def edit_other_info(request):
  k= employ_add_form.objects.get(id=id);
  return render(request,"admin-template/admin_file2.html",{'k':k})


def bank(request):
    int=integrations.objects.filter(purpose="BANKING")
    if request.method=="POST":
        response=request.POST["response"]
        k=app(response=response)
        k.save()
        return redirect("/bank")
    return render(request,"admin-template/int_template.html",{'int':int})

def hrms(request):
    int=integrations.objects.filter(purpose="HRMS")
    if request.method=="POST":
        response=request.POST["response"]
        k=app(response=response)
        k.save()
        return redirect("/hrms")
    return render(request,"admin-template/int_template.html",{'int':int})

def communication(request):
    int=integrations.objects.filter(purpose="Communication")
    if request.method=="POST":
        response=request.POST["response"]
        k=app(response=response)
        k.save()
        return redirect("/communication")
    return render(request,"admin-template/int_template.html",{'int':int})

def insurance(request):
    int=integrations.objects.filter(purpose="Insurance")
    if request.method=="POST":
        response=request.POST["response"]
        k=app(response=response)
        k.save()
        return redirect("/insurance")
    return render(request,"admin-template/int_template.html",{'int':int})

def attendance(request):
    int=integrations.objects.filter(purpose="Attendance")
    if request.method=="POST":
        response=request.POST["response"]
        k=app(response=response)
        k.save()
        return redirect("/attendance")
    return render(request,"admin-template/int_template.html",{'int':int})

def itr(request):
    int=integrations.objects.filter(purpose="ITR Filing")
    if request.method=="POST":
        response=request.POST["response"]
        k=app(response=response)
        k.save()
        return redirect("/itr")
    return render(request,"admin-template/int_template.html",{'int':int})

def performance(request):
    int=integrations.objects.filter(purpose="PERFORMANCE")
    if request.method=="POST":
        response=request.POST["response"]
        k=app(response=response)
        k.save()
        return redirect("/performance")
    return render(request,"admin-template/int_template.html",{'int':int})

def bgv(request):
    int=integrations.objects.filter(purpose="BACKGROUND VERIFICATION")
    if request.method=="POST":
        response=request.POST["response"]
        k=app(response=response)
        k.save()
        return redirect("/bgv")
    return render(request,"admin-template/int_template.html",{'int':int})


from datetime import datetime, timedelta, date
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, HttpResponseRedirect
from django.views import generic
from django.urls import reverse
from django.utils.safestring import mark_safe
import calendar
import requests
from django.views.generic import TemplateView
from django.template import Context, Template

from .utils import Calendar1,caltable
from django import forms
from .models import editholiday12

def index(request):
    return HttpResponse('hello')

class HomeForm(forms.Form):
    name=forms.CharField(max_length=100)
    email=forms.EmailField()

class CalendarView_1(TemplateView):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    
    template_name = "admin-template/admin_calendar.html"
    

   
    def get_context_data(self, **kwargs):
        context = super(CalendarView_1,self).get_context_data(**kwargs)
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)
        admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')



        user = CustomUser.objects.filter(id=self.request.user.id).first()
        userid1 = user.id
        projectm = admin_project_create.objects.filter(admin_id=userid1)
        h = HR.objects.all()
        s=adminnav.objects.all()
        employs_all = Employs.objects.all()
        data = AdminHod.objects.filter(id=self.request.user.id)
        admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
        data = Companys.objects.filter(usernumber=self.request.user.id).first()
        data1=data.id
        a=company_details.objects.filter(companyid=data1).first()


        d = get_date(self.request.GET.get('month', None))
        cal = Calendar1(d.year, d.month,d.day)
        html_cal = cal.formatmonth(self.request,withyear=True)
        days_list,weekoff_days =caltable(self,d.year,d.month,self.request)
        
        context['calendar'] = mark_safe(html_cal)
        context['prev_month'] = prev_month(d,self.request)
        context['next_month'] = next_month(d,self.request)
        context['admin_drops'] = self.admin_drops
        context['weekoff_days']=weekoff_days
        context['s']= s
        context['user']= user
        context['projectm']= projectm
        context['s']= s
        context['a']= a

        context['h']= h
        context['employs_all']= employs_all
        context['data']= data
        context['admin_home_drops']= admin_home_drops
        context['admin_drops']=admin_drops



        inline_context={
            "days_list":"days_list"
        }
        inline_html_template=Template(f"{days_list}")
        inline_view=inline_html_template.render(Context(inline_context))
        context['inline_view']=inline_view
        return  context
    
    
    
    def post(self, request, *args, **kwargs):
        form = HomeForm(request.POST)
        if form.is_valid():
            name = form.cleaned_data['name']
            email = form.cleaned_data['email']
            # Do something with the name and email
            return HttpResponseRedirect(reverse('calendar'))
        else:
            context = self.get_context_data(**kwargs)
            context['form'] = form

            context['check_in'] = check_in(self.request)
            return redirect('/hod_calendar/')

class CalendarView_2(TemplateView):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')


    
    template_name = "admin-template/admin_calendar.html"
    

   
    def get_context_data(self, **kwargs):
        context = super(CalendarView_2,self).get_context_data(**kwargs)
        d = get_date(self.request.GET.get('month', None))
        cal = Calendar1(d.year, d.month,d.day)
        html_cal = cal.formatmonth(self.request,withyear=True)
        days_list,weekoff_days =caltable(self,d.year,d.month,self.request)
        admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')

        user = CustomUser.objects.filter(id=self.request.user.id).first()
        userid1 = user.id

        s = adminnav.objects.all()
        projectm = admin_project_create.objects.filter(admin_id=userid1)
        h = HR.objects.all()
        employs_all = Employs.objects.all()
        admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
        admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
        data = AdminHod.objects.filter(id=self.request.user.id)
        projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')


       
        
        context['calendar'] = mark_safe(html_cal)
        context['prev_month'] = prev_month(d,self.request)
        context['next_month'] = next_month(d,self.request)
        context['admin_drops'] = self.admin_drops
        context['weekoff_days']=weekoff_days


        context['admin_drops'] = admin_drops
        context['admin_home_drops'] = admin_home_drops
        context['s'] = s
        context['h'] = h
        context['data'] = data
        context['user'] = user

        context['projects_drops'] = projects_drops
        context['projectm'] = projectm
        context['employs_all'] = employs_all


        inline_context={
            "days_list":"days_list"
        }
        inline_html_template=Template(f"{days_list}")
        inline_view=inline_html_template.render(Context(inline_context))
        context['inline_view']=inline_view
        return  context
    
    
    
    def post(self, request, *args, **kwargs):
        form = HomeForm(request.POST)
        if form.is_valid():
            name = form.cleaned_data['name']
            email = form.cleaned_data['email']
            # Do something with the name and email
            return HttpResponseRedirect(reverse('calendar'))
        else:
            context = self.get_context_data(**kwargs)
            context['form'] = form

            context['check_in'] = check_in(self.request)
            return redirect('/hod_calendar/')




def get_date(req_month):
    if req_month:
        year, month = (int(x) for x in req_month.split('-'))
        curr_day = datetime.now().day
        curr_year = datetime.now().year
        curr_month = datetime.now().month
        return date(year, month, day=1)

    return datetime.today()

def prev_month(d,request):
    first = d.replace(day=1)
    prev_month = first - timedelta(days=1)
    month = 'month=' + str(prev_month.year) + '-' + str(prev_month.month)
    request.session['year']=str(prev_month.year);
    return month

def next_month(d,request):
    days_in_month = calendar.monthrange(d.year, d.month)[1]
    last = d.replace(day=days_in_month)
    next_month = last + timedelta(days=1)
    month = 'month=' + str(next_month.year) + '-' + str(next_month.month)
    request.session['month']=str(next_month.month);
    request.session['year']=str(next_month.year);
    return month



def sidenav(request):
    s=adminnav.objects.all()
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    return render(request,'admin-template/admin_calendar.html',{'s':s,'admin_drops':admin_drops})

        
def check_in(request):
    today = datetime.now().date()
    now = datetime.now().time()
    stat = "present"
    email = request.session.get('email_2')

    if request.method == 'POST' and 'v' in request.POST:
        check = checkin.objects.filter(date=today, empid=email).first()
        if not check:
            checkin.objects.create(date=today, time=now, status=stat, empid=email,)
            formatted_time = now.strftime("%I:%M %p")  # Format time as HH:MM AM/PM
            messages.success(request,f'You have successfully checked in at  {formatted_time}.')
            return redirect('/hod_calendar/')
        else:
            messages.warning(request,'You have already checked in today.')
            return redirect('/hod_calendar/')


    if request.method == 'POST' and 'v1' in request.POST:
        checkt = checkout.objects.filter(date=today, empid=email).first()
        if not checkt:
            checkout.objects.create(date=today, time=now, empid=email,date_value=1)
            formatted_time = now.strftime("%I:%M %p")  # Format time as HH:MM AM/PM
            messages.success(request, f'You have successfully checked out at {formatted_time}.')
            return redirect('/hod_calendar/')

        else:
            messages.warning(request, 'You have already checked out today.')
            return redirect('/hod_calendar/')

    return render(request, 'admin-template/admin_calendar.html')


        
from datetime import datetime, timedelta
from django.shortcuts import render
from django.db.models import Count
from .models import checkin, LeaveReportEmploy

def attendance(request):
    rr = checkin.objects.values('empid').annotate(empid_count=Count('empid'))
    rrr = LeaveReportEmploy.objects.values('empid').annotate(empid_count=Count('empid'))

    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if from_date and to_date:
        rr = rr.filter(date__range=[from_date, to_date])
        rrr = rrr.filter(leave_date__range=[from_date, to_date])

    combined_data = {}

    for record in rr:
        combined_data[record['empid']] = {
            'empid': record['empid'],
            'empid_count_rr': record['empid_count'],
            'empid_count_rrr': 0,
        }

    for record in rrr:
        if record['empid'] in combined_data:
            combined_data[record['empid']]['empid_count_rrr'] = record['empid_count']
        else:
            combined_data[record['empid']] = {
                'empid': record['empid'],
                'empid_count_rr': 0,
                'empid_count_rrr': record['empid_count'],
            }

    total_working_days = 0
    if from_date and to_date:
        from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
        to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()

        current_date = from_date_obj
        while current_date <= to_date_obj:
            day_of_week = current_date.weekday()
            if day_of_week < 5:  # Monday to Friday
                total_working_days += 1
            current_date += timedelta(days=1)
        admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    return render(request, 'hod_template/adattendance.html', {'combined_data': combined_data.values(), 'total_working_days': total_working_days,'admin_drops':admin_drops})







#setting views###############################################################

from .models import check
from django.http import HttpResponse
from django.core.exceptions import ValidationError
import datetime
# Create your views here.
def home(request):
    return render(request, "admin-template/FAQ.html")


def reg(request):
    if request.method == "POST":
        # Process checkbox values
        EnabledAttendance = request.POST.get('EnabledAttendance', 'no')
        AllowEmployees = request.POST.get('AllowEmployees', 'no')
        AllowHalfdayleave = request.POST.get('AllowHalfdayleave', 'no')
        Employeemustenter = request.POST.get('Employeemustenter', 'no')
        Showattendence = request.POST.get('Showattendence', 'no')
        automaticallyadd = request.POST.get('automaticallyadd', 'no')
        lossofpay = request.POST.get('lossofpay', 'no')
        usefinancal = request.POST.get('usefinancal', 'no')
        track = request.POST.get('track', 'no')
        
        # Process other form fields
        type = request.POST['type']
        defaultleave = request.POST['defaultleave']
        monthlyincrement = request.POST['monthlyincrement']
        maxleave = request.POST['maxleave']
        carryforward = request.POST['carryforward']
        Defaultshift = request.POST['Defaultshift']
        graceperiod = request.POST['graceperiod']
        fulltime = request.POST['fulltime']
        halftime = request.POST['halftime']
        date = request.POST.get('date')
        description = request.POST['description']
        attendanceenabled = request.POST.get('attendanceenabled', 'no')
        all = request.POST.get('all', 'no')
        andhrapradesh = request.POST.get('andhrapradesh', 'no')
        telangana = request.POST.get('telangana', 'no')
        date2 = request.POST.get('date2')
        description2 = request.POST.get('description2')
        holidays = request.POST['holidays']
        weekend = request.POST['weekend']
        day = request.POST.get('day')
        
        # Validate date
        if date:
            try:
                datetime.strptime(date, '%Y-%m-%d')
            except ValueError:
                return HttpResponse("Invalid date format. It must be in YYYY-MM-DD format.")
        k=check.objects.first()
        # Save data to the database
        k = check(pk=1,
            EnabledAttendance=EnabledAttendance,
            holidays=holidays,
            weekend=weekend,
            AllowEmployees=AllowEmployees,
            all=all,
            andhrapradesh=andhrapradesh,
            telangana=telangana,
            date2=date2,
            description2=description2,
            day=day,
            AllowHalfdayleave=AllowHalfdayleave,
            Employeemustenter=Employeemustenter,
            Showattendence=Showattendence,
            automaticallyadd=automaticallyadd,
            lossofpay=lossofpay,
            usefinancal=usefinancal,
            track=track,
            type=type,
            defaultleave=defaultleave,
            monthlyincrement=monthlyincrement,
            maxleave=maxleave,
            carryforward=carryforward,
            Defaultshift=Defaultshift,
            graceperiod=graceperiod,
            fulltime=fulltime,
            halftime=halftime,
            date=date,
            description=description,
            attendanceenabled=attendanceenabled,
        )
        k.save()
        
        return HttpResponse("Record is inserted")
    
    return render(request, "admin-template/Leave.html")



def edit_leave(request):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    k1=check.objects.all()
   
    return render(request,"admin-template/Leave_update.html",{'k1':k1,'admin_drops':admin_drops})
from.models import publicholidays
from .models import publicholidays,halfldayvreason
from datetime import datetime
from django.shortcuts import render, redirect
from django.http import HttpResponseServerError
from collections import defaultdict
from .utils import get_current_month_holidays

# def update_leave(request):
#     shift_tim=working_shifts.objects.all()
#     indianholidays = get_current_month_holidays("india")
#     sorted_holidays = sorted(indianholidays.items(), key=lambda x: x[0])
#     checkvalue = editholiday12.objects.first()
#     halfreason = halfldayvreason.objects.first()
#     custom_holidays = customholidays.objects.order_by('-date')
#     admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
#     stored_dates = publicholidays.objects.values_list('publicholiday_date', flat=True)  
#     stored_dates = list(stored_dates)  
#     if request.method == "POST":
#         sun = request.POST.get('sun', '')
#         sat1 = request.POST.get('sat1', '')
#         sat2 = request.POST.get('sat2', '')
#         sat3 = request.POST.get('sat3', '')
#         sat4 = request.POST.get('sat4', '')
#         sat5 = request.POST.get('sat5', '')
#         mon = request.POST.get('mon', '')
#         tue = request.POST.get('tue', '')
#         web = request.POST.get('web', '')
#         thu = request.POST.get('thu', '')
#         fri = request.POST.get('fri', '')
#         alsat = request.POST.get('alsat', '')
#         halfdaylev = request.POST.get('halfdaylev', '')
#         reason1 = request.POST.get('reason1', '')
#         starting_time=request.POST.get('starting_time','')
#         ending_time=request.POST.get('ending_time','')
#         shift_name=request.POST.get('shift_name','')
#         cutoff_time=request.POST.get('cutoff_time','')
#         befor_time=request.POST.get('befor_time','')
#         shift_id=request.POST.getlist('shift_id')
#         if starting_time==ending_time:
#             messages.success(request,f'')
#         else:
#             shft=working_shifts(starting_time=starting_time,ending_time=ending_time,shift_name=shift_name,cutoff_time=cutoff_time,befor_time=befor_time)
#             shft.save()
#         k1 = editholiday12.objects.get(id=1)
#         k1.sun = sun
#         k1.sat1 = sat1
#         k1.sat2 = sat2
#         k1.sat3 = sat3
#         k1.sat4 = sat4
#         k1.sat5 = sat5
#         k1.mon=mon
#         k1.tue=tue
#         k1.web=web
#         k1.thu=thu
#         k1.fri=fri
#         k1.alsat=alsat
#         k1.save()
        
#         k2 = halfldayvreason.objects.get(id=1)
#         k2.halfdaylev = halfdaylev
#         k2.reason1 = reason1
#         k2.save()
        
#         dates = request.POST.getlist('date')
#         reasons = request.POST.getlist('reason')
        
#         for date_str, reason in zip(dates, reasons):
#             if date_str or reason:
#                 # Check if either date_str or reason is not empty before creating and saving the object
#                 date = customholidays(date=date_str, reason=reason)
#                 date.save()
        

        
#         selected_holiday_ids = []
#         for key in request.POST.keys():
#             if key.startswith('publicholiday_datecheck_'):
#                 holiday_id = key.split('_')[-1]
#                 selected_holiday_ids.append(int(holiday_id))

#         # Get all existing holiday records from the database
#         existing_holidays = publicholidays.objects.all()

#         # Loop through existing holidays and delete those not in the selected_holiday_ids list
#         for holiday in existing_holidays:
#             if holiday.id not in selected_holiday_ids:
#                 holiday.delete()

#         for key in request.POST.keys():
#             if key.startswith('publicholiday_datecheck_'):
#                 holiday_id = key.split('_')[-1]
#                 holiday_dates = request.POST.get(f'publicholiday_date_{holiday_id}')
#                 festival_name = request.POST.get(f'festival_name_{holiday_id}')
#                 finding=request.POST.get(f'finding_{holiday_id}')
#                 try:
#                     holiday_date = datetime.strptime(holiday_dates, '%d/%m/%Y').date()
#                 except ValueError:
#                     # Handle invalid date format gracefully (e.g., log an error)
#                     continue
                
#                 # Create a new publicholidays object and save it to the database
#                 publicholidays.objects.create(publicholiday_date=holiday_date, festival_name=festival_name,finding=finding)
#         working_shifts.objects.filter(id__in=shift_id).delete()

#         return redirect('/setting/')  # Redirect to a success page

#     # Prepare a list of flags to determine checkbox status in the template

#     return render(request, 'admin-template/Leave_update123.html', {'checkvalue': checkvalue,'shift_tim':shift_tim, 'custom_holidays': custom_holidays, 'halfreason': halfreason, 'indianholidays': sorted_holidays,'stored_dates':stored_dates,'admin_drops':admin_drops })

def custome_holidaydele(request,id):
    custom_holiday=customholidays.objects.get(id=id)
    custom_holiday.delete()
    return redirect("/update_leave")

from django.shortcuts import render
from .forms import YourForm

def your_view(request):
    if request.method == 'POST':
        form = YourForm(request.POST)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.all = 'Yes' if form.cleaned_data['all'] else 'No'
            instance.save()
    else:
        form = YourForm()
    
    context = {'form': form}
    return render(request, 'admin-template/your_template.html', context)


def edit_employee_notification(request):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    post,_=empnotificationupdate.objects.get_or_create(id=1);
    return render(request,"admin-template/update_employee_notification.html",{'post':post,'admin_drops':admin_drops})

def update_employee_notification(request):
    if request.method == "POST":
        send_email_notify = request.POST.get('send_email_notify', '').lower()
        email_notify = request.POST.get('email_notify', '').lower()
        if send_email_notify == 'yes' and email_notify == 'yes':
            result = 'Yes'
        else:
            result = 'No'
        empnotificationupdate.objects.all().delete() 
        post = empnotificationupdate(send_email_notify=send_email_notify, email_notify=email_notify)
        post.save()
        return redirect('/setting')
    try:
        data = empnotificationupdate.objects.latest('id')
    except empnotificationupdate.DoesNotExist:
        data = None
    return render(request, 'admin-template/update_employee_notification.html', {'data': data})

def default_salary_str(request):
    if request.method=="POST":
        default_salary = request.POST.get('default_salary', '').lower()
        default_salary="Yes" if default_salary == "on" else "No"
        salarycomponent=request.POST['salarycomponent']
        percentageofCTC=request.POST['percentageofCTC']
        percentageorfixed=request.POST['percentageorfixed']
        Taxable=request.POST['Taxable']
        k=salary_struc(default_salary=default_salary,salarycomponent=salarycomponent,percentageofCTC=percentageofCTC,percentageorfixed=percentageorfixed,Taxable=Taxable)
        k.save()
        return HttpResponse("data is inserted")
    return render(request, "admin-template/defaultform_salary_str.html")

def default_sal_str(request):
     k=salary_struc.objects.all()
     return render(request, "admin-template/defaultform_sal_str.html",{'k':k})
    
def edit_default_sal_str(request):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    k=salary_struc.objects.all();
    return render(request,"admin-template/defaultform_sal_str.html",{'k':k,'admin_drops':admin_drops})


# def update_default_sal_str(request):
#     if request.method == "POST":
#         new_field_x_data = request.POST.get('field_x_data')
#         default_sal = "Yes" if new_field_x_data == 'on' else "No"
#         k, _ = default_salary.objects.get_or_create(pk=1)
#         k.default_sal = default_sal
#         k.save()
#         return redirect('/setting/')
#     k = default_salary.objects.first()  
#     return render(request, 'admin-template/default_sal_str.html', {'k': k})


# def tds(request):
#     s=adminnav.objects.all()
#     if request.method=="POST":
#         tds_payment=request.POST.get('tds_payment')
#         tds_payment= "Yes" if tds_payment == "on" else "No"

#         tds_filling_setup=request.POST.get('tds_filling_setup')
#         filling_form=request.POST.get('filling_form')
#         userid=request.POST['userid']
#         password=request.POST['password']
#         username=request.POST['username']
#         password1=request.POST['password1']
#         k=TDS(pk=1,tds_payment=tds_payment,tds_filling_setup=tds_filling_setup,filling_form=filling_form,userid=userid,password=password,username=username,password1=password1)
#         k.save()
#         return redirect("/pr_tax")
#     return render(request,"admin-template/tds.html",{'s':s})

# def esic(request):def p
#     s=adminnav.objects.all()

#     if request.method=="POST":
#         esi_status=request.POST.get('esi_status')
#         esi_payment=request.POST.get('esi_payment')
#         esi_payment="yes" if esi_payment == "yes" else "No"

#         username2=request.POST.get('username2')
#         password3=request.POST.get('password3')
#         k=ESIC(pk=1,esi_status=esi_status,esi_payment=esi_payment,username2=username2,password3=password3)
#         k.save()
#         return redirect("/tds")
#     return render(request,"admin-template/esic.html",{'s':s})

def payments_edit(request):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = AdminHod.objects.filter(admin=request.user.id).first()
    
    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')

    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()


    k=TDS.objects.first()
    k1=PF.objects.first()
    k2=ESIC.objects.first()
    k3=P_tax.objects.first()
    messages.success(request,"your details were sucessfully submitted")
    return render(request,"admin-template/payments.html",{'a':a,'k3':k3,'k':k,'k1':k1,'k2':k2,'admin_drops':admin_drops,'user':user,'da1':da1,'admin_home_drops':admin_home_drops,'s':s,'h':h,'projectm':projectm,'projects_drops':projects_drops,'data':data,'employs_all':employs_all})

    

def payments_update(request):

    if request.method == 'POST':
        tds_payment = request.POST.get('tds_payment')
        tds_payment1 = request.POST.get('tds_payment1')
        k=TDS.objects.first()
        k.tds_payment=tds_payment;
        k.tds_payment1=tds_payment1;
        if tds_payment == "on":
            k.tds_payment = 'Yes'
        else:
            k.tds_payment = 'No'
        if tds_payment1 == "on":
            k.tds_payment1 = 'Yes'
        else:
            k.tds_payment1 = 'No'
   
        k.save()
        pf_payment = request.POST.get('pf_payment')
        pf_setup = request.POST.get('pf_setup')
        pf_setup1 = request.POST.get('pf_setup1')
        pf_setup2 = request.POST.get('pf_setup2')
        pf_setup3 = request.POST.get('pf_setup3')
        k1=PF.objects.first()
        k1.pf_payment = pf_payment;
        k1.pf_setup = pf_setup;
        k1.pf_setup1 = pf_setup1;
        k1.pf_setup2 = pf_setup2;
        k1.pf_setup3 = pf_setup3;
        if pf_payment == "on":
            k1.pf_payment = 'Yes'
        else:
            k1.pf_payment = 'No'
        if pf_setup == "on":
            k1.pf_setup = 'Yes'
        else:
            k1.pf_setup = 'No'
        if pf_setup1 == "on":
            k1.pf_setup1 = 'Yes'
        else:
            k1.pf_setup1 = 'No'
        if pf_setup2 == "on":
            k1.pf_setup2 = 'Yes'
        else:
            k1.pf_setup2 = 'No'
        if pf_setup3 == "on":
            k1.pf_setup3 = 'Yes'
        else:
            k1.pf_setup3 = 'No'

        k1.save()

     
        esi_payment = request.POST.get('esi_payment')
        esi_settings = request.POST.get('esi_settings')
        esi_settings1 = request.POST.get('esi_settings1')
        k2=ESIC.objects.first()
        k2.esi_payment = esi_payment;
        k2.esi_settings = esi_settings;
        k2.esi_settings1 = esi_settings1;
        if esi_payment == "on":
            k2.esi_payment = 'Yes'
        else:
            k2.esi_payment = 'No'
        if esi_settings == "on":
            k2.esi_settings = 'Yes'
        else:
            k2.esi_settings = 'No'
        if esi_settings1 == "on":
            k2.esi_settings1 = 'Yes'
        else:
            k2.esi_settings1 = 'No'

        k2.save()

        pt_payment=request.POST.get('pt_payment')
        k3=P_tax.objects.first()
        k3.pt_payment=pt_payment;
        if pt_payment == "on":
            k3.pt_payment = 'Yes'
        else:
            k3.pt_payment = 'No'

        k3.save()

        return redirect('/setting')

    return render(request, 'admin-template/payments.html')


def edit_Integrations_KlaarHQ(request):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    return render(request, 'admin-template/edit_Integrations_KlaarHQ.html',{'admin_drops':admin_drops})


def reimbursementform(request):
    if request.method=="POST":
        document_type=request.POST['document_type']
        Enabled=request.POST.get('Enabled')
        Enabled="Yes" if Enabled == "on" else "No"
        Reimbursements_Enabled=request.POST.get('Reimbursements_Enabled')
        Reimbursements_Enabled= "Yes" if Reimbursements_Enabled == "on" else "No"
        Make_attachments_compulsory=request.POST.get('Make_attachments_compulsory')
        Make_attachments_compulsory= "Yes" if Make_attachments_compulsory == "on" else "No"
        Include_reimbursements_with_payroll=request.POST.get('Include_reimbursements_with_payroll')
        Include_reimbursements_with_payroll= "Yes" if Include_reimbursements_with_payroll == "on" else "No"
        
        k=reimbursement_setup_settings(Reimbursements_Enabled=Reimbursements_Enabled,Make_attachments_compulsory=Make_attachments_compulsory,Include_reimbursements_with_payroll=Include_reimbursements_with_payroll,document_type=document_type,Enabled=Enabled)
        k.save()
        messages.success(request,"data is inserted")
    return render(request,'admin-template/reimbursementedit.html')

def reimbursement_edit(request):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    k3=reimbursement_setup_settings.objects.all()
    
    return render(request,"admin-template/reimbursementedit.html",{'k3':k3,'admin_drops':admin_drops})

# def reimbursement_update(request):
#     if request.method == 'POST':

#         new_field_x_data = request.POST.get('field_x_data')
#         Reimbursements_Enabled = "Yes" if new_field_x_data == 'on' else "No"
#         new_field_y_data = request.POST.get('field_y_data')
#         Make_attachments_compulsory = "Yes" if new_field_y_data == 'on' else "No"
#         new_field_z_data = request.POST.get('field_z_data')
#         Include_reimbursements_with_payroll = "Yes" if new_field_z_data == 'on' else "No"
#         document_type = request.POST.get('document_type')
#         Enabled = request.POST.get('Enabled')
#         Enabled = "Yes" if Enabled == 'on' else "No"
#         k3=reimbursement_setup_settings(
#         Reimbursements_Enabled = Reimbursements_Enabled,
#         Make_attachments_compulsory = Make_attachments_compulsory,
#         Include_reimbursements_with_payroll = Include_reimbursements_with_payroll,
#         document_type=document_type,
#         Enabled=Enabled),
#         k3.save()
#         return redirect('/setting')
#     return render(request, 'admin-template/reimbursementedit.html')

import json

def reimbursement_update(request):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = AdminHod.objects.filter(admin=request.user.id).first()
    
    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    data = AdminHod.objects.filter(id=request.user.id)
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')

    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()

    k3=reimbursementsetup.objects.first()
    k2=reimbursementsetup1.objects.all()

    if request.method == 'POST':
        for r in k2:
            reimbursement_type = r.reimbursement_type
            select_reimbursement = 'Yes' if request.POST.get(f'select_reimbursement_{r.id}') == 'on' else 'No'
            r.select_reimbursement = select_reimbursement
            r.save()
                
        reimbursement_type = request.POST.get('reimbursement_type')
        select_reimbursement = 'Yes' if request.POST.get('select_reimbursement') == 'on' else 'No'
        k2 = reimbursementsetup1(reimbursement_type=reimbursement_type,select_reimbursement=select_reimbursement)
        k2.save()


        Reimbursements_Enabled = request.POST.get('Reimbursements_Enabled')
        Make_attachments_compulsory = request.POST.get('Make_attachments_compulsory')
        Include_reimbursements_with_payroll = request.POST.get('Include_reimbursements_with_payroll')

        k3=reimbursementsetup.objects.first()
        k3.Reimbursements_Enabled = Reimbursements_Enabled;
        k3.Make_attachments_compulsory = Make_attachments_compulsory;
        k3.Include_reimbursements_with_payroll = Include_reimbursements_with_payroll;


        if Make_attachments_compulsory == "on":
            k3.Make_attachments_compulsory = 'Yes'
        else:
            k3.Make_attachments_compulsory = 'No'

        if Include_reimbursements_with_payroll == "on":
            k3.Include_reimbursements_with_payroll = 'Yes'
        else:
            k3.Include_reimbursements_with_payroll = 'No'
       
        if Reimbursements_Enabled == "on":
            k3.Reimbursements_Enabled = 'Yes'
            # Update adminnav model
            adminnav.objects.update_or_create(pk=7,
                name='REIMBURSEMENTS',
                defaults={'is_name_exist': 1},
                icon='fa fa-credit-card',
                url='/admin_reimbursement_apply_view'
            )
        else:
            k3.Reimbursements_Enabled = 'No'
            # Update adminnav model
            adminnav.objects.filter(name='REIMBURSEMENTS').delete()

        if Reimbursements_Enabled == "on":
            k3.Reimbursements_Enabled = 'Yes'
            # Update adminnav model
            employnav.objects.update_or_create(pk=6,
                name='REIMBURSEMENTS',
                defaults={'is_name_exist': 1},
                icon='fa fa-credit-card',
                url='/reimbursement_apply_view'

            )
        else:
            k3.Reimbursements_Enabled = 'No'
            # Update employnav model
            employnav.objects.filter(name='REIMBURSEMENTS').delete()
        k3.save()
        messages.success(request,"Your details were successfully submitted")

        return redirect('/setting')
    return render(request, 'admin-template/reimbursementedit1.html',{'a':a,'k2':k2,'k3':k3,'admin_drops':admin_drops,'user':user,'da1':da1,'admin_home_drops':admin_home_drops,'s':s,'h':h,'projectm':projectm,'projects_drops':projects_drops,'data':data,'employs_all':employs_all})






def reimbursement_delete(request,id):
    k3=reimbursement_setup_settings.objects.get(id=id);
    k3.delete()
    return redirect("/reimbursement_edit")


    
from .models import *

def tdsfillingsetupform(request):
    if request.method=="POST":
        Automated_filling24=request.POST.get('Automated_filling24')
        Automated_filling24= "Yes" if Automated_filling24 == "on" else "No"
        Automated_filling26=request.POST.get('Automated_filling26')
        Automated_filling26= "Yes" if Automated_filling26 == "on" else "No"
        k=tdsfillingsetup(pk=1,Automated_filling24=Automated_filling24,Automated_filling26=Automated_filling26)
        k.save()
        messages.success(request,"data is inserted")
    return render(request,'admin-template/tdsfillingsetup.html')

def tdsfillingsetup_edit(request):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    k4=tdsfillingsetup.objects.first()
  
    return render(request,"admin-template/tdsfillingsetup.html",{'k4':k4,'admin_drops':admin_drops})
   


def tdsfillingsetupupdate(request):
    if request.method=="POST":
        new_field_x_data = request.POST.get('field_x_data')
        Automated_filling24 = "Yes" if new_field_x_data == 'on' else "No"
        new_field_y_data = request.POST.get('field_y_data')
        Automated_filling26 = "Yes" if new_field_y_data == 'on' else "No"
        k4, _=tdsfillingsetup.objects.get_or_create(id=1)
        k4.Automated_filling24 = Automated_filling24
        k4.Automated_filling26 = Automated_filling26
        k4.save()
        return redirect('/setting')
    return render(request,'admin-template/tdsfillingsetup.html')

def employeform(request):
    if request.method=="POST":
        employe_id_prefix=request.POST['employe_id_prefix']
        employe_directory=request.POST.get('employe_directory')
        employe_directory= "Yes" if employe_directory == "on" else "No"
        additionalinfo=request.POST['additionalinfo']
        k=employedata(employe_id_prefix=employe_id_prefix,employe_directory=employe_directory,additionalinfo=additionalinfo)
        k.save()
        messages.success(request,"data is inserted")
    return render(request,'admin-template/employedataedit.html')

# def employedata_edit(request):
#     admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
#     k5, _=employedata.objects.get_or_create(pk=1)
   
#     return render(request,"admin-template/employedataedit.html",{'k5':k5,'admin_drops':admin_drops})

from .models import task1
def employedata_edit(request):
    half=task1.objects.first()
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = AdminHod.objects.filter(admin=request.user.id).first()
   
    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    # data = AdminHod.objects.filter(id=request.user.id)
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')

    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()


    k5, _=employedata.objects.get_or_create(pk=1)
    if request.method =="POST":
        Personalphonenumber1=request.POST.get("Personalphonenumber1");
        PersonalEmailAddress1=request.POST.get("PersonalEmailAddress1");
        FathersName1=request.POST.get("FathersName1");
        FathersDOB1=request.POST.get("FathersDOB1");
        MothersName1=request.POST.get("MothersName1");
        MothersDOB1=request.POST.get("MothersDOB1");
        Childdetails1=request.POST.get("Childdetails1");
        Childdetails2=request.POST.get("Childdetails2");
        TemporaryAddress1=request.POST.get("TemporaryAddress1");
        HighestEducatonalQualification1=request.POST.get("HighestEducatonalQualification1");
        Addharnumber1=request.POST.get("Addharnumber1");
        maritalstatus1=request.POST.get("maritalstatus1");
        workexperiance1=request.POST.get("workexperiance1");
        previousemploye1=request.POST.get("previousemploye1");
        previousdesignation1=request.POST.get("previousdesignation1");
        Marriageannivarsary1=request.POST.get("Marriageannivarsary1");
        emergencycontactname1=request.POST.get("emergencycontactname1");
        emergencycontactnumber1=request.POST.get("emergencycontactnumber1");
        emergencycontactrelation1=request.POST.get("emergencycontactrelation1");
        bloodgroup1=request.POST.get("bloodgroup1");
        nationality=request.POST.get("nationality");
        post=task1.objects.get(id=1);
        post.Personalphonenumber1=Personalphonenumber1;
        post.PersonalEmailAddress1=PersonalEmailAddress1;
        post.FathersName1=FathersName1;
        post.FathersDOB1=FathersDOB1;
        post.MothersName1=MothersName1;
        post.MothersDOB1=MothersDOB1;
        post.Childdetails1=Childdetails1;
        post.Childdetails2=Childdetails2;
        post.TemporaryAddress1=TemporaryAddress1;
        post.HighestEducatonalQualification1=HighestEducatonalQualification1;
        post.Addharnumber1=Addharnumber1;
        post.maritalstatus1=maritalstatus1;
        post.workexperiance1=workexperiance1;
        post.previousemploye1=previousemploye1;
        post.previousdesignation1=previousdesignation1;
        post.Marriageannivarsary1=Marriageannivarsary1;
        post.emergencycontactname1=emergencycontactname1;
        post.emergencycontactnumber1=emergencycontactnumber1;
        post.emergencycontactrelation1=emergencycontactrelation1;
        post.bloodgroup1=bloodgroup1;
        post.nationality=nationality;
        post.save();
    if (request.method =="POST"):
        additionalinfo=request.POST.getlist('additionalinfo') 
        new_field_d_data = request.POST.get('field_d_data')
        employe_directory = "Yes" if new_field_d_data == 'on' else "No"    
        employe_id_prefix=request.POST.get('employe_id_prefix');
        k5, _=employedata.objects.get_or_create(pk=1)
        k5.additionalinfo=additionalinfo;
        k5.employe_directory=employe_directory;
        k5.employe_id_prefix=employe_id_prefix;
        
        k5.save()
        messages.success(request,"your data was sucessfully updated")

    return render(request,"admin-template/employedataedit.html",{'a':a,'k5':k5,'admin_drops':admin_drops,'half':half,'user':user,'da1':da1,'admin_home_drops':admin_home_drops,'s':s,'h':h,'projectm':projectm,'projects_drops':projects_drops,'data':data,'employs_all':employs_all})



# def employedata_update(request): 
#     if (request.method =="POST"):
#         additionalinfo=request.POST.getlist('additionalinfo') 
#         new_field_d_data = request.POST.get('field_d_data')
#         employe_directory = "Yes" if new_field_d_data == 'on' else "No"    
#         employe_id_prefix=request.POST.get('employe_id_prefix');
#         k5, _=employedata.objects.get_or_create(pk=1)
#         k5.additionalinfo=additionalinfo;
#         k5.employe_directory=employe_directory;
#         k5.employe_id_prefix=employe_id_prefix;
        
#         k5.save()
#         return redirect('/setting/')
#     return render(request,'admin-template/employedataedit.html')


def tax_deduction_edit(request):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    post=tax_deduction.objects.first()
    return render(request,"admin-template/tax_deduction.html",{'post':post,'admin_drops':admin_drops})

def tax_deduction_update(request) :
    if (request.method =="POST"): 
        approval=request.POST.get('approval');
        approval = "Yes" if approval == 'on' else "No"
        verify=request.POST.get('verify');
        verify = "Yes" if verify == 'on' else "No"
        taxable=request.POST.get('taxable');
        taxable = "Yes" if taxable == 'on' else "No"
        post, _=tax_deduction.objects.get_or_create(pk=1)
        post.approval=approval;
        post.verify=verify;
        post.taxable=taxable;
        post.save()
        return redirect('/setting/',)
    return render(request,'admin-template/tax_deduction.html')



def tax_deduction_form(request):
    if request.method=="POST":
        approval=request.POST.get('approval')
        approval= "Yes" if approval == "on" else "No"
        verify=request.POST.get('verify')
        verify= "Yes" if verify == "on" else "No"
        taxable=request.POST.get('taxable')
        taxable= "Yes" if taxable == "on" else "No"
        k, _=tax_deduction.objects.get_or_create(pk=1)
        k=tax_deduction(pk=1,approval=approval,verify=verify,taxable=taxable)
        k.save()
        messages.success(request,"data is inserted")
    return render(request,'admin-template/tax_deduction.html')


def resignationsetupform(request):
    if request.method=="POST":
        resignations=request.POST.get('resignations')
        resignations= "Yes" if resignations == "on" else "No"
        post1, _=emp_reg_setup.objects.get_or_create(pk=1)
        post1=emp_reg_setup(pk=1,resignations=resignations)
        post1.save()
    return render(request,'admin-template/update_resignation_setup.html')


def edit_Employee_Resignation_Setup(request):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    post1=emp_reg_setup.objects.first()
    
    return render(request, 'admin-template/update_resignation_setup.html',{'post1':post1,'admin_drops':admin_drops})

def update_Employee_Resignation_Setup(request):
    if (request.method =="POST"):
       
        resignations=request.POST.get('resignations');
        resignations= "Yes" if resignations == "on" else "No"
        post1, _=emp_reg_setup.objects.get_or_create(pk=1)
        
        post1.resignations=resignations;
        post1.save()
        return redirect('/setting/')
    return render(request,'admin-template/update_resignation_setup.html')

def setting(request):
    # a=companylogo.objects.all()
    s=adminnav.objects.all()
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    p=set_payroll_date.objects.get()
    q=salary_struc.objects.all()
    q=set_salary_structure.objects.all()                                                                                                                                                                                                          
    a2=empnotificationupdate.objects.all()
    k=TDS.objects.all()
    k1=PF.objects.all()
    k2=ESIC.objects.all()
    k3=reimbursement_setup_settings.objects.all()
    k9=reimbursementsetup.objects.all()
    data_list1=list(k3.values())
    k4=tdsfillingsetup.objects.first()
    post=emp_reg_setup.objects.all()
    k5=check.objects.all()  
    k6=employedata.objects.first()
    k7=tax_deduction.objects.all()
    k8=documents_setup1.objects.all()
    k10=P_tax.objects.all()

    data_list=list(k8.values())
    return render(request,"settings.html",{'data':data,'k9':k9,'a':a,'s':s,'admin_drops':admin_drops,'a2':a2,'k10':k10,'p':p,'k':k,'k1':k1,'k2':k2,'k3':k3,'post':post,'q':q,'k4':k4,'k5':k5,'k6':k6,'k7':k7,'k8':k8,'data_list':data_list,'data_list1':data_list1})

def document_setup(request):
    if request.method =="POST":
        document_type=request.POST['document_type']
        compulsory=request.POST.get('compulsory')
        compulsory="Yes" if compulsory == "on" else "No"
        Enabled=request.POST.get('Enabled')
        Enabled="Yes" if Enabled == "on" else "No"
        k=documents_setup(document_type=document_type,compulsory=compulsory,Enabled=Enabled)
        k.save()
        return HttpResponse("data is inserted")
    return render(request,"admin-template/document_setup.html")

def delete_document_setup(request,id):
    post=documents_setup.objects.get(id=id)
    post.delete()
    return redirect("/edit_document_setup")

def edit_document_setup(request):
    post=documents_setup.objects.all()
    # return redirect('/update_document_setup')
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    return render(request,"admin-template/document_setup_update.html",{'post':post,'admin_drops':admin_drops})

def update_document_setup(request):
    post=documents_setup.objects.all()
    if request.method == "POST":
        document_type = request.POST.get('document_type')
        compulsory = request.POST.get('compulsory')
        compulsory = "Yes" if compulsory == 'on' else "No"
        Enabled = request.POST.get('Enabled')
        Enabled = "Yes" if Enabled == 'on' else "No"
        post = documents_setup(
        
        document_type = document_type,
        compulsory = compulsory,
        Enabled = Enabled)
        post.save()
        
        return redirect('/setting/')

    return render(request, 'admin-template/document_setup_update.html',{'post':post})

def read_proj(request):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')

    # a=Project.objects.all()

    if request.method == 'GET':
        status = request.GET.get('status')
        o_id = AdminHod.objects.get(admin=request.user.id)
        project_details = Project.objects.filter(o_id_id = o_id).values()
        
        if status == 'completed':
            projects = Project.objects.filter(status='completed')
        elif status == 'ongoing':
            projects = Project.objects.filter(status='ongoing')
        elif status == 'featured':
            projects = Project.objects.filter(status='featured')
        else:
            projects = Project.objects.all()
        completed_count = Project.objects.filter(status='completed').count()
        ongoing_count = Project.objects.filter(status='ongoing').count()
        featured_count = Project.objects.filter(status='featured').count()

    return render(request, 'admin-template/ViewProjects.html', {'projects': projects,"status":status,"msg": project_details,"completed_count":completed_count,"ongoing_count":ongoing_count,"featured_count":featured_count,'admin_drops':admin_drops})





def update_default_sal_str(request):
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = AdminHod.objects.filter(admin=request.user.id).first()
    # da2 = da1.id
    # admin = AdminHod.objects.get(id=1)
    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    # data = AdminHod.objects.filter(id=request.user.id)
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')

    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()



    k = salary_struct.objects.all()
    f = salary_struc.objects.all()
    l = set_salary_structure.objects.first()

    if request.method == "POST":
        # Check if the form was submitted
        default_salary = request.POST.get('default_salary')
        l = set_salary_structure.objects.first()
        l.default_salary = default_salary
        if default_salary == "on":
            l.default_salary = 'Enabled'
        else:
            l.default_salary = 'Disabled'
        l.save()

        # Update the 'percentageofCTC' values in the database
        for i in f:
            # Get the updated value from the POST data
            updated_percentageofCTC = request.POST.get(f'percentageofCTC_{i.id}')
            
            # Update the 'percentageofCTC' field in the database
            i.percentageofCTC = updated_percentageofCTC
            i.save()
        messages.success(request,"your data was sucessfully updated")

        return redirect('/setting')

    return render(request, 'admin-template/defaultform_sal_str.html', {'a':a,'k': k, 'f': f,'l':l, 'admin_drops': admin_drops,'user':user,'da1':da1,'admin_home_drops':admin_home_drops,'s':s,'h':h,'projectm':projectm,'projects_drops':projects_drops,'data':data,'employs_all':employs_all})




from datetime import timedelta

def read_tasks(request):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    o_id = AdminHod.objects.get(admin=request.user.id)
    board_details = HR.objects.filter(o_id_id=o_id).all()
    projects = Project.objects.filter(o_id=o_id).all()
    tasks=Task.objects.all()
    selected_project_id = request.GET.get('project_filter')
    filtered_tasks = Task.objects.all()
    if selected_project_id and selected_project_id != 'any':
        board_details = board_details.filter(HRids__p_id=selected_project_id)
        filtered_tasks = Task.objects.filter(p_id=selected_project_id)
        # filtered_tasks = Task.objects.filter(is_expired=False)
    if request.method == 'GET':
        return render(request, 'admin-template/ViewTasks.html', {'admin_drops':admin_drops,"board_details": board_details, "projects": projects,'filtered_tasks':filtered_tasks,'tasks':tasks})
    else:
        return render(request, 'admin-template/ViewTasks.html', {"projects": projects})

def create_task(request,pid):
    o_id = AdminHod.objects.get(admin=request.user.id)
    boards = HR.objects.filter(o_id_id=o_id).values()
    projects = Project.objects.filter(o_id_id=o_id,id=pid ).values()
    employees = Employs.objects.all()
    context = {"boards": boards, "projects": projects, "employees": employees}
    if request.method == 'POST':
        t_name = request.POST['t_name']
        t_desc = request.POST['t_desc']
        t_deadline_date = request.POST['t_deadline_date']
        t_status = "todo"
        t_priority = request.POST['t_priority']
        b_id = request.POST['b_id']
        p_id = pid
        e_id = request.POST['e_id']
        taskObj = Task.objects.create(t_name=t_name, t_desc=t_desc, t_deadline_date=t_deadline_date,
                                      t_status=t_status, t_priority=t_priority, o_id_id=o_id.admin.id, b_id_id=b_id, p_id_id=p_id, e_id_id=e_id)
        if taskObj:
            empDetails = Employs.objects.filter(id=e_id).values()
            subject = 'DevelopTrees - New Task Created for you'
            message = f'Hi {empDetails[0]["first_name"]} , Your organization as created a new task : {t_name} , description : {t_desc}, priority : {t_priority} and deadline for task is : {t_deadline_date}, Login in your account to get more information. From: DevelopTrees. '
            email_from = settings.EMAIL_HOST_USER
            recipient_list = [empDetails[0]["email"], ]
            send_mail(subject, message, email_from, recipient_list)
            messages.success(request, "Task was created successfully!")
            dynamic_url = reverse ('projectwise_tasks' , args=[pid])
            return HttpResponseRedirect(dynamic_url)
        else:
            messages.error(request, "Some Error was occurred!")
            return HttpResponseRedirect('/create-task')
    return render(request, 'admin-template/CreateTask.html', context)




def update_task(request, pk):
    # try:
    o_id = AdminHod.objects.get(admin=request.user.id)
    tasks = Task.objects.get(id=pk)
    p_id = tasks.p_id
    pid=tasks.p_id.id
    projects_emp_link = TeamMember.objects.filter(project_id=p_id)
     
    if request.method == 'POST':
        t_name = request.POST['t_name']
        t_desc = request.POST['t_desc']
        t_deadline_date = request.POST['t_deadline_date']
        t_priority = request.POST['t_priority']
        e_id = request.POST['e_id']
        employ_instance = Employs.objects.get(id=e_id)
        t_status = request.POST['t_status']
        tasks.t_deadline_date = datetime.strptime(t_deadline_date, '%Y-%m-%dT%H:%M')
        tasks.t_name = t_name
        tasks.t_desc = t_desc
        
        tasks.t_priority = t_priority
        tasks.t_status = t_status
        tasks.e_id = employ_instance
        tasks.save()
        if tasks:
            empDetails = Employs.objects.filter(id=e_id).values()
            subject = 'DevelopTrees - Task Updated for you'
            message = f'Hi {empDetails[0]["first_name"]} , Your organization as updated a task : {t_name} , description : {t_desc}, priority : {t_priority} and deadline for task is : {t_deadline_date}, Login in your account to get more information. From: DevelopTrees. '
            email_from = settings.EMAIL_HOST_USER
            recipient_list = [empDetails[0]["email"], ]
            send_mail(subject, message, email_from, recipient_list)
            messages.success(request, "Task was updated successfully!")
            dynamic_url = reverse('projectwise_tasks', args=[pid])
            return HttpResponseRedirect(dynamic_url)

        else:
            messages.error(request, "Some Error was occurred!2")
            
    else:
        if tasks:
            return render(request, 'admin-template/UpdateTask.html', {"tasks": tasks , "projects_emp_link": projects_emp_link } )
        else:
            messages.error(request, "Some Error was occurred!1")
            
    dynamic_url = reverse('projectwise_tasks', args=[pid])
    return HttpResponseRedirect(dynamic_url)



def delete_task(request, pk):
    try:
        # Fetch the task and its project ID
        task = Task.objects.get(id=pk)
        pid = task.p_id.id  # Assuming Task has a foreign key to Project model

        # Check if the user has the necessary permissions (you might want to customize this logic)
        if request.user == task.o_id.admin:
            task.delete()
            messages.success(request, "Task was deleted successfully!")
        else:
            messages.error(request, "You don't have permission to delete this task!")

    except Task.DoesNotExist:
        messages.error(request, "Task does not exist.")
    except Exception as e:
        messages.error(request, f"An error occurred: {str(e)}")

    dynamic_url = reverse('projectwise_tasks', args=[pid])
    return HttpResponseRedirect(dynamic_url)



import pandas as pd
import plotly.express as px
from plotly.offline import plot
from datetime import timedelta
import json

def projectwise_task(request, pid):
    o_id = AdminHod.objects.get(admin=request.user.id)
    
    
    team_members = TeamMember.objects.filter(project=pid)
    team_member_scores = [(team_member, team_member.calculate_total_score_in_project()) for team_member in team_members]
    team_member_scores = sorted(team_member_scores, key=lambda x: x[1], reverse=True)
    project_details=Project.objects.filter(id=pid)
    task_details = Task.objects.filter(o_id_id=o_id, id=pid).all()
    # s = adminnav.objects.filter(parent_category=None).order_by('id')

    if request.method == 'GET':
        t_status = request.GET.get('t_status', 'total')  # Default to 'total' if not provided

        if t_status == 'total':
            task_details = Task.objects.filter(o_id_id=o_id, p_id=pid).all()
        elif t_status == 'completed':
            task_details = Task.objects.filter(o_id_id=o_id, p_id=pid, t_status='completed')
        elif t_status == 'in-progress':
            # task_details = Task.objects.filter(o_id_id=o_id, p_id=pid, t_status='in-progress').all()
            task_details = Task.objects.filter(o_id_id=o_id, p_id=pid).exclude(t_status='completed')

        

        tasks = task_details
        count_no_of_total_tasks = Task.objects.filter(o_id_id=o_id, p_id_id=pid).count()
        count_no_of_completed_tasks = Task.objects.filter(o_id_id=o_id, p_id_id=pid, t_status="completed").count()
        count_no_of_pending_tasks = count_no_of_total_tasks - count_no_of_completed_tasks
        task_data = []

        for task in tasks:
            now = timezone.now()
            task_score = task.calculate_task_score()
            remaining_time = task.calculate_remaining_time()
            progress_percentage = min(task_score / 10 * 100, 100)
            excess_points = max(task_score - 10, 0)  # Calculate the excess points (if any)
            neg_points = max(-task_score, 0)

            task_data.append({
                'task': task,
                'progress_percentage': progress_percentage,
                'excess_points': excess_points,
                'neg_points': neg_points,
                'task_score': task_score,
                'remaining_time': remaining_time,
            })
            
        
        context = {
        
        "project_details": project_details,
        'task_total': count_no_of_total_tasks,
        'task_completed': count_no_of_completed_tasks,
        'task_pending': count_no_of_pending_tasks,
        'team_members': team_members,
        # 's': s,
        't_status':t_status,
        "task_details": task_details,
        'task_data': task_data,
        
        'team_member_scores': team_member_scores,

        }
    
    return render(request, 'admin-template/ViewProjectwiseTasks.html', context)

def employee_record(request, employee_id):
    employee = Employs.objects.get(id=employee_id)

    # Get employee's overall performance data
    performance_data = employee.calculate_overall_performance()

    # Get the list of projects associated with the employee
    projects = Project.objects.filter(teammember__employee=employee)
    team_members = TeamMember.objects.filter(employee=employee)
    total_score_across_all_projects = sum(member.calculate_total_score_across_all_projects() for member in team_members)
    # Get the list of tasks associated with the employee
    tasks = Task.objects.filter(e_id=employee)

    context = {
        'employee': employee,
        'performance_data': performance_data,
        'projects': projects,
        'tasks': tasks,
        'total_score_across_all_projects':total_score_across_all_projects,
    }


    return render(request,"admin-template/employee_record.html" , context)


    
def delete_user(request,employee_id , project_id):
        # Assuming your model has a field 'id' representing the user's ID
        employee = TeamMember.objects.get(employee_id=employee_id,project_id=project_id)
        employee.delete()
        referer = request.META.get('HTTP_REFERER')
        if referer:
          return redirect(referer)
        else:
          return redirect(reverse('home'))
        
def view_project_wise_employees(request):
    pids = list(Project.objects.filter(o_id_id=request.user.id).values_list('id', flat=True))
    projEmps = {}
    eids = []
    eid_enames = dict(Employs.objects.filter(o_id_id=request.user.id).values_list('id','first_name','last_name'))
    pnames = list(Project.objects.filter(o_id_id=2).values_list('p_name', flat=True))
    enames = []
    for pid in pids:
        e = list(Project_Employee_Linker.objects.filter(o_id_id=request.session['o_id'],p_id=pid).values_list('e_id', flat=True))
        eids.append(e)
        enames.append([eid_enames[i] for i in e])
    projEmps = dict(zip(pnames, enames))
    print(dict(zip(pnames, enames)))
    return render(request, 'admin-template/ViewProjEmps.html', {'projEmps':projEmps})



def org_update_tasks(request, tid):
    t_update_date = datetime.today().strftime('%Y-%m-%d')
    task_details = Task.objects.filter(id=tid).update(t_update_date=t_update_date, t_status="completed")
    if task_details:
        messages.success(request, "Task Mark as completed!")
        return HttpResponseRedirect('/read-task')
    else:
        messages.error(request, "Some error was occurred!")
        return HttpResponseRedirect('/read-task')

def get_team_members(request, project_id):
    team_members = TeamMember.objects.filter(project_id=project_id)
    data = [{'id': member.employee.id, 'name': f"{member.employee.first_name} {member.employee.last_name}"}
            for member in team_members]
    return JsonResponse({'team_members': data})

def get_employee_tasks(request, employee_id):
    try:
        # Query the Task model to retrieve tasks associated with the employee_id
        tasks = tlassigntask.objects.filter(employid=employee_id)

        # Prepare a list of task data
        task_data = [{'task_name': task.task, } for task in tasks]

        # Return the task data as a JSON response
        return JsonResponse({'tasks': task_data})

    except Exception as e:
        return JsonResponse({'error': str(e)})


from django.db.models import Q

def create_proj(request):
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    pname=admin_project_create.objects.filter(admin_id=request.user.id)
    team_lead_employee = None
    selected_team_lead_id = None
    # s=adminnav.objects.filter(parent_category=None).order_by('id')
    team_members = Employs.objects.filter(
        Q(teammember__isnull=True) | Q(teammember__project__status='completed')
    ).distinct()
    if request.method == 'POST':
        p_name = request.POST['p_name']
        p_desc = request.POST['p_desc']
        pr_deadline = request.POST['project_deadline_date']
        # project_manager = request.POST['manager_name']
        # status = request.POST['status']
        o_id = AdminHod.objects.get(admin=request.user.id)

        selected_employee_ids = request.POST.getlist('selected_employees[]')
        project = Project.objects.create(
            p_name=p_name,
            p_desc=p_desc,
            o_id=o_id,
            # project_manager=project_manager,
            project_deadline=pr_deadline,
            # status=status
            
        )

        selected_team_lead_name = request.POST.get('team_lead')
        existing_team_member_ids = TeamMember.objects.filter(project=project).values_list('employee_id', flat=True)
        available_team_members = Employs.objects.exclude(id__in=existing_team_member_ids)
        team_lead_employee = Employs.objects.get(id=selected_team_lead_name)
        
        if team_lead_employee:
            selected_team_lead_id = team_lead_employee.id
        # Check if a TeamMember entry with the same project and employee already exists
        existing_team_member = TeamMember.objects.filter(project=project, employee=team_lead_employee.id, is_team_lead=True).first()
        if not existing_team_member:
            # Create the team lead only if it doesn't exist
            team_lead_member = TeamMember(
                project=project,
                employee=team_lead_employee,
                is_team_lead=True,
            )
            team_lead_member.save()

        # Mark the selected team lead as a team lead in the database
        

        for emp_id in selected_employee_ids:
            try:
                employee = Employs.objects.get(id=emp_id)
                TeamMember.objects.create(project=project, employee=employee, is_team_lead=False)
                # available_team_members = available_team_members.exclude(id=emp_id)
            except IntegrityError:
                existing_entry = TeamMember.objects.get(project=project, employee=employee, is_team_lead=False)
                existing_entry.save()
        
        # Fetch available team members (exclude team leads)
        # team_members = Employ.objects.filter(
        #         Q(teammember__isnull=True) | Q(teammember__project__status='completed')
        #         ).distinct()

        messages.success(request, 'Project created successfully and employees assigned.')
        return redirect('/admin_home')
    
    

    return render(request, 'admin-template/OrgCreateProject.html', {'s':s,'projects_drops':projects_drops,'pname':pname,"team_members": team_members, "selected_team_lead_id": selected_team_lead_id})


    
    

    return render(request, 'admin-template/OrgCreateProject.html', {'s':s,'projects_drops':projects_drops,'pname':pname,"team_members": team_members, "selected_team_lead_id": selected_team_lead_id})



def del_proj(request, pid):
    o_id = AdminHod.objects.get(admin=request.user.id)
    try:
        project_detail = Project.objects.filter(id=pid,o_id_id=o_id).delete()
        if project_detail:
            messages.success(request, "Project was deleted successfully!")
            return HttpResponseRedirect('/read-proj')
        else:
            messages.error(request, "Some Error was occurred!")
            return HttpResponseRedirect('/read-proj')
    except:
        messages.error(request, "Some Error was occurred!")
        return HttpResponseRedirect('/read-proj')

from django.shortcuts import render, redirect
from .models import AdminHod, Project, TeamMember, Employs
from django.db.utils import IntegrityError
from django.contrib import messages



def tds(request):
    a1=TDS.objects.all()

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = AdminHod.objects.filter(admin=request.user.id).first()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    # data = AdminHod.objects.filter(id=request.user.id)
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    data=Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a = company_details.objects.filter(companyid=data1).first()



    s=adminnav.objects.all()
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    if request.method=="POST":
        tds_payment=request.POST.get('tds_payment')
        if tds_payment == "on":
            tds_payment = "Yes"
        else:
            tds_payment = "No"
        tds_payment1=request.POST.get('tds_payment1')
        if tds_payment1 == "on":
            tds_payment1 = "Yes"
        else:
            tds_payment1 = "No"
        verify_tan=request.POST.get('verify_tan')
        tds_filling_setup=request.POST.get('tds_filling_setup')
        if tds_filling_setup == "on":
            tds_filling_setup = "Yes"
        else:
            tds_filling_setup = "No"
        tds_filling_setup1=request.POST.get('tds_filling_setup1')
        if tds_filling_setup1 == "on":
            tds_filling_setup1 = "Yes"
        else:
            tds_filling_setup1 = "No"
        tds_filling_setup2=request.POST.get('tds_filling_setup2')
        if tds_filling_setup2 == "on":
            tds_filling_setup2 = "Yes"
        else:
            tds_filling_setup2 = "No"

        filling_form=request.POST.get('filling_form')
        userid=request.POST.get('userid')
        password=request.POST.get('password')
        username=request.POST.get('username')
        password1=request.POST.get('password1')

        k=TDS(pk=1,tds_payment1=tds_payment1,tds_filling_setup2=tds_filling_setup2,tds_filling_setup1=tds_filling_setup1,tds_payment=tds_payment,verify_tan=verify_tan,tds_filling_setup=tds_filling_setup,filling_form=filling_form,userid=userid,password=password,username=username,password1=password1)
        k.save()
        return redirect("/pr_tax")
    return render(request,"admin-template/tds.html",{'a1':a1,'s':s,'admin_drops':admin_drops,'user':user,'da1':da1,'a':a,'admin_home_drops':admin_home_drops,'h':h,'projectm':projectm,'projects_drops':projects_drops,'data':data,'employs_all':employs_all})





def pr_tax(request):
    a1=P_tax.objects.all()
    s=adminnav.objects.all()

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = AdminHod.objects.filter(admin=request.user.id).first()
    
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()

    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')


    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    if request.method=='POST':
        professional_tax=request.POST.get('professional_tax')
        pt_payment=request.POST.get('pt_payment')
        if pt_payment == "on":
            pt_payment = "Yes"
        else:
            pt_payment = "No"

        username3=request.POST.get('username3')
        password4=request.POST.get('password4')
        k=P_tax(pk=1,password4=password4,username3=username3,pt_payment=pt_payment,professional_tax=professional_tax)
        k.save()
        instance=details.objects.create(form1=professional_tax)
        return redirect("/pf")
    return render(request,"admin-template/p_tax.html",{'a1':a1,'s':s,'admin_drops':admin_drops,'user':user,'da1':da1,'a':a,'admin_home_drops':admin_home_drops,'h':h,'projectm':projectm,'projects_drops':projects_drops,'data':data,'employs_all':employs_all})


def pf(request):
    a1=PF.objects.all()

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = AdminHod.objects.filter(admin=request.user.id).first()
    
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')


    s=adminnav.objects.all()
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    if request.method=='POST':
        pf_status=request.POST.get('pf_status')
        pf_payment=request.POST.get('pf_payment')
        if pf_payment == "on":
            pf_payment = "Yes"
        else:
            pf_payment = "No"

        username1=request.POST.get('username1')
        password2=request.POST.get('password2')
        pf_setup=request.POST.get('pf_setup')
        if pf_setup == "on":
            pf_setup = "Yes"
        else:
            pf_setup = "No"

        pf_setup1=request.POST.get('pf_setup1')
        if pf_setup1 == "on":
            pf_setup1 = "Yes"
        else:
            pf_setup1 = "No"

        pf_setup2=request.POST.get('pf_setup2')
        if pf_setup2 == "on":
            pf_setup2 = "Yes"
        else:
            pf_setup2 = "No"

        pf_setup3=request.POST.get('pf_setup3')
        if pf_setup3 == "on":
            pf_setup3 = "Yes"
        else:
            pf_setup3 = "No"

        k=PF(pk=1,pf_setup1=pf_setup1,pf_setup2=pf_setup2,pf_setup3=pf_setup3,pf_status=pf_status,pf_payment=pf_payment,username1=username1,password2=password2,pf_setup=pf_setup)
        k.save()

        instance=details.objects.create(form2=pf_status)
        
        return redirect("/esic")
    return render(request,"admin-template/pf.html",{'a1':a1,'s':s,'admin_drops':admin_drops,'user':user,'da1':da1,'a':a,'admin_home_drops':admin_home_drops,'h':h,'projectm':projectm,'projects_drops':projects_drops,'data':data,'employs_all':employs_all})

def esic(request):

    a1=ESIC.objects.all()
    s=adminnav.objects.all()
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = AdminHod.objects.filter(admin=request.user.id).first()
    
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')


    if request.method=="POST":
        esi_status=request.POST.get('esi_status')
        esi_payment=request.POST.get('esi_payment')
        if esi_payment == "on":
            esi_payment = "Yes"
        else:
            esi_payment = "No"

        username2=request.POST.get('username2')
        password3=request.POST.get('password3')
        esi_settings=request.POST.get('esi_settings')
        if esi_settings == "on":
            esi_settings = "Yes"
        else:
            esi_settings = "No"

        esi_settings1=request.POST.get('esi_settings1')
        if esi_settings1 == "on":
            esi_settings1 = "Yes"
        else:
            esi_settings1 = "No"

        progress = 20 
        value_message=1
        val=4
        k=ESIC(pk=1,esi_settings1=esi_settings1,esi_status=esi_status,esi_payment=esi_payment,username2=username2,password3=password3,esi_settings=esi_settings)
        progress_obj, _ = Progress.objects.get_or_create(pk=1, defaults={'value': 0})
        progress_obj.value4 = value_message
        progress_obj.progress_form5 = progress
        progress_obj.save()
        instance=admin_home_drop.objects.get(id=6)
        instance.progress_value=val
        instance.save()

        k.save()
        return redirect("/company_detail")
    return render(request,"admin-template/admin_esic.html",{'a1':a1,'s':s,'admin_drops':admin_drops,'user':user,'da1':da1,'a':a,'admin_home_drops':admin_home_drops,'h':h,'projectm':projectm,'projects_drops':projects_drops,'data':data,'employs_all':employs_all})



# def edit_role_permission(request,id):
#     if request.method == "POST":
#         designation=request.POST.get('designation')
#         k=Employs.objects.get(id=id)
#         k.designation=designation
#         k.save()
        
#     return HttpResponseRedirect(reverse("edit_role_permission", args=[k.id]))
    # return render(request, "admin-template/rolepermission.html", args=[k.employ_id])
# return HttpResponseRedirect(reverse("individual_reimbursement_view", args=[leave.employ_id_id]))
def edit_role_permission(request,id):
    s=adminnav.objects.all()
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    data=Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a = company_details.objects.filter(companyid=data1).first()

    if request.method == "POST":
        designation=request.POST.get('designation')
        k=Employs.objects.get(id=id)
        k.designation=designation
        k.save()
        return redirect('/people_count')

    return render(request, "admin-template/rolepermission.html",{'s':s,"admin_drops":admin_drops,"data":data,"a":a})
 


def edit_edit_role_permission(request,std):
    request.session['employ_id']=std
    k=Employs.objects.get(admin=std)
    objss=k.id 
    users=CustomUser.objects.get(id=std)
    datas=employ_add_form.objects.filter(student_id=objss)
    return redirect('/update_role_permission')
    return render(request,"admin-template/rolepermission.html",{'k':k,'objss':objss,'users':users,'datas':datas})


 

def update_role_permission(request,std):
    request.session['employ_id']=std
    k=Employs.objects.get(admin=std)
    objss=k.id 
    users=CustomUser.objects.get(id=std)
    datas=employ_add_form.objects.filter(student_id=objss)
    if request.method == "POST":
        designation = request.POST.get('designation')
        k.designation=designation
        k.save()
        
        return redirect('/edit_people_admin/')

    return render(request, 'admin-template/rolepermission.html',{'k':k,'objss':objss,'users':users,'std':std,'datas':datas})




def update_leave(request):
    shift_tim=working_shifts.objects.all()
    indianholidays = get_current_month_holidays("india")
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = AdminHod.objects.filter(admin=request.user.id).first()
    
    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')

    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()

    sorted_holidays = sorted(indianholidays.items(), key=lambda x: x[0])
    checkvalue = editholiday12.objects.first()
    halfreason = halfldayvreason.objects.first()
    custom_holidays = customholidays.objects.order_by('-date')
    levid1=employlev.objects.filter(leave_id=1).values()
    levid2=employlev.objects.filter(leave_id=2).values()
    levid3=employlev.objects.filter(leave_id=3).values()
    levid4=employlev.objects.filter(leave_id=4).values()
    levid5=employlev.objects.filter(leave_id=5).values()
    levid6=employlev.objects.filter(leave_id=6).values()
    levid7=employlev.objects.filter(leave_id=7).values()
    levid8=employlev.objects.filter(leave_id=8).values()
    levid9=employlev.objects.filter(leave_id=9).values()
    levid10=employlev.objects.filter(leave_id=10).values()
    levid11=employlev.objects.filter(leave_id=11).values()
    levid12=employlev.objects.filter(leave_id=12).values()
    stored_dates = publicholidays.objects.values_list('publicholiday_date', flat=True)  
    stored_dates = list(stored_dates) 
    if request.method == "POST":
        sun = request.POST.get('sun', '')
        sat1 = request.POST.get('sat1', '')
        sat2 = request.POST.get('sat2', '')
        sat3 = request.POST.get('sat3', '')
        sat4 = request.POST.get('sat4', '')
        sat5 = request.POST.get('sat5', '')
        mon = request.POST.get('mon', '')
        tue = request.POST.get('tue', '')
        web = request.POST.get('web', '')
        thu = request.POST.get('thu', '')
        fri = request.POST.get('fri', '')
        alsat = request.POST.get('alsat', '')
        halfdaylev = request.POST.get('halfdaylev', '')
        reason1 = request.POST.get('reason1', '')
        starting_time=request.POST['starting_time']
        ending_time=request.POST['ending_time']
        shift_name=request.POST.get('shift_name','')
        cutoff_time=request.POST.get('cutoff_time','')
        befor_time=request.POST.get('befor_time','')
        shift_id=request.POST.getlist('shift_id')
        if starting_time==ending_time:
            messages.success(request,f'')
        else:
          shft=working_shifts(starting_time=starting_time,ending_time=ending_time,shift_name=shift_name,befor_time=befor_time,cutoff_time=cutoff_time)
          shft.save()
        k1 = editholiday12.objects.get(id=1)
        k1.sun = sun
        k1.sat1 = sat1
        k1.sat2 = sat2
        k1.sat3 = sat3
        k1.sat4 = sat4
        k1.sat5 = sat5
        k1.mon=mon
        k1.tue=tue
        k1.web=web
        k1.thu=thu
        k1.fri=fri
        k1.alsat=alsat
        k1.save()
        k2 = halfldayvreason.objects.get(id=1)
        k2.halfdaylev = halfdaylev
        k2.reason1 = reason1
        k2.save()
        
        dates = request.POST.getlist('date')
        reasons = request.POST.getlist('reason')
        type=request.POST.getlist('type')
        defaultleave=request.POST.getlist('defaultleave')
        leave_id=request.POST.getlist('leave_id')
        
        for date_s, reason in zip(dates, reasons):
            if date_s or reason:
                date = customholidays(date=date_s, reason=reason)
                date.save()
        if len(type) == len(defaultleave) == len(leave_id):
           for t, dl, lid in zip(type, defaultleave, leave_id):
              if t:
                 try:
                    lev, created = employlev.objects.get_or_create(leave_id=lid, defaults={'type': t, 'defaultleave': dl})
                    if not created:
                        lev.type = t
                        lev.defaultleave = dl
                        lev.save()
                 except Exception as e:
                       print(f"An error occurred while saving data: {str(e)}")

        
        selected_holiday_ids = []
        for key in request.POST.keys():
            if key.startswith('publicholiday_datecheck_'):
                holiday_id = key.split('_')[-1]
                selected_holiday_ids.append(int(holiday_id))

        existing_holidays = publicholidays.objects.all()

        for holiday in existing_holidays:
            if holiday.id not in selected_holiday_ids:
                holiday.delete()

        for key in request.POST.keys():
            if key.startswith('publicholiday_datecheck_'):
                holiday_id = key.split('_')[-1]
                holiday_dates = request.POST.get(f'publicholiday_date_{holiday_id}')
                festival_name = request.POST.get(f'festival_name_{holiday_id}')
                finding=request.POST.get(f'finding_{holiday_id}')
                try:
                    holiday_date = datetime.strptime(holiday_dates, '%d/%m/%Y').date()
                except ValueError:
                    continue
                
                publicholidays.objects.create(publicholiday_date=holiday_date, festival_name=festival_name,finding=finding)
        working_shifts.objects.filter(id__in=shift_id).delete()

        return redirect('/setting/')  # Redirect to a success page


    return render(request, 'admin-template/Leave_update123.html', {'a':a,'checkvalue': checkvalue, 'custom_holidays': custom_holidays,'admin_drops':admin_drops, 'halfreason': halfreason, 'indianholidays': sorted_holidays,'stored_dates':stored_dates,'shift_tim':shift_tim,'levid1':levid1,'levid2':levid2,'levid3':levid3,'levid4':levid4,'levid5':levid5,'levid6':levid6,'levid7':levid7,'levid8':levid8,'levid9':levid9,'levid10':levid10,'levid11':levid11,'levid12':levid12,'user':user,'da1':da1,'admin_home_drops':admin_home_drops,'s':s,'h':h,'projectm':projectm,'projects_drops':projects_drops,'data':data,'employs_all':employs_all})
 




from django.shortcuts import render
from datetime import date, timedelta
from .models import Employs, checkin, LeaveReportEmploy, editholiday12, customholidays, publicholidays
def paid(request):
    
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()
    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')

    results = []

    employees = Employs.objects.filter(companyid=data1)

    current_month = datetime.now().month
    current_year = datetime.now().year

    _, num_days = calendar.monthrange(current_year, current_month)
    start_date = datetime(current_year, current_month, 1)
    
    days_of_month = [start_date + timedelta(days=i) for i in range(num_days)]

    checkin_statuses = {}  # A dictionary to store check-in status for each day

    for employee in employees:
        empid = employee.empid
        first_name = employee.first_name

        today = date.today()
        current_month = today.month
        current_year = today.year

        last_day_of_month = date(current_year, current_month, 1)
        last_day_of_month = last_day_of_month.replace(day=28)
        while last_day_of_month.month == current_month:
            last_day_of_month += timedelta(days=1)
        last_day_of_month -= timedelta(days=1)

        days_in_month = (last_day_of_month - date(current_year, current_month, 1)).days + 1

        editholidays = editholiday12.objects.first()
        admin_holidays = sum(getattr(editholidays, day, False) or 0 for day in ['sun', 'sat1', 'sat2', 'sat3', 'sat4', 'sat5', 'alsat', 'mon', 'tue', 'web', 'thu', 'fri'])

        custom_holidays = customholidays.objects.filter(date__month=current_month).count()
        public_holidays = publicholidays.objects.filter(publicholiday_date__month=current_month).count()

        remaining_days = days_in_month - admin_holidays - custom_holidays - public_holidays

        total_checkins = checkin.objects.filter(empid=employee.email, date__month=current_month).count()

        half_leave = LeaveReportEmploy.objects.filter(employ_id=employee, leave_status=1, leave_date__month=current_month, leave_type="Half-Day").count()
        unpaid_leave = LeaveReportEmploy.objects.filter(employ_id=employee, leave_status=1, leave_date__month=current_month, leave_type="Unpaid_leave").count()
        leave = LeaveReportEmploy.objects.filter(employ_id=employee, leave_status=1, leave_date__month=current_month, leave_type="Leave").count()
        open_request = LeaveReportEmploy.objects.filter(employ_id=employee, leave_status=0, leave_date__month=current_month).count()

        checkin_status_for_employee = []

        for day in days_of_month:
            checkin_exists = checkin.objects.filter(
                empid=employee.email,
                date=day
            ).exists()

            checkin_status_for_employee.append('Present' if checkin_exists else '-NA-')

        results.append({
            'empid': empid,
            'first_name': first_name,
            'total_checkins': total_checkins,
            'remaining_days': remaining_days,
            'original_working_days': remaining_days - total_checkins,
            'half_leave': half_leave,
            'unpaid_leave': unpaid_leave,
            'leave': leave,
            'open_request': open_request,
            'checkin_status': checkin_status_for_employee,
            
        })

    items_per_page = 10
    page = request.GET.get('page')
    paginator = Paginator(results, items_per_page)

    try:
        results = paginator.page(page)
    except PageNotAnInteger:
        results = paginator.page(1)
    except EmptyPage:
        results = paginator.page(paginator.num_pages)

    return render(request, "admin-template/paid.html", {
        'days_of_month': days_of_month,
        'employ_id': employees,
        'checkin_statuses': checkin_statuses,
        'results': results,
        'admin_drops': admin_drops,
        'days_of_month': days_of_month,
        'employ_id': employees,
        'checkin_statuses': checkin_statuses,
        'results': results,
        'admin_drops':admin_drops,
        'user':user,
        'a':a,
        'admin_home_drops':admin_home_drops,
        's':s,
        'h':h,
        'projectm':projectm,
        'projects_drops':projects_drops,
        'data':data,
        'employs_all':employs_all,

    })






from datetime import datetime, timedelta
from calendar import monthrange
from django.shortcuts import render
from .models import Employs, checkin


def paid2(request):
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    employees = Employs.objects.filter(companyid=data1)
    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id

    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')


    # Get start_date and end_date from query parameters (if provided)
    start_date_param = request.GET.get('start_date')
    end_date_param = request.GET.get('end_date')

    # Determine today's date
    today = datetime.now()

    # Calculate start_date and end_date based on provided parameters or default to the current month
    if start_date_param and end_date_param:
        start_date = datetime.strptime(start_date_param, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_param, '%Y-%m-%d')
    else:
        current_month = today.month
        current_year = today.year
        _, num_days = monthrange(current_year, current_month)
        start_date = datetime(current_year, current_month, 1)
        end_date = datetime(current_year, current_month, num_days)

    # Create a list of dates in the date range
    date_range = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]

    # Categorize dates into "previous," "present," or "future"
    previous_dates = [date for date in date_range if date < today]
    present_dates = [date for date in date_range if date == today]
    future_dates = [date for date in date_range if date > today]

    # Paginate the employees
    items_per_page = 10
    page = request.GET.get('page')
    paginator = Paginator(employees, items_per_page)

    try:
        employees_page = paginator.page(page)
    except PageNotAnInteger:
        employees_page = paginator.page(1)
    except EmptyPage:
        employees_page = paginator.page(paginator.num_pages)

    # Initialize the checkin_statuses dictionary for the current page
    checkin_statuses = {}
    for employee in employees_page:
        checkin_status_for_employee = []

        for day in date_range:
            checkin_exists = checkin.objects.filter(
                empid=employee.email,
                date=day
            ).exists()
            checkin_status_for_employee.append('Present' if checkin_exists else '-NA-')

        checkin_statuses[employee] = checkin_status_for_employee

    # Determine the category of the date range
    if start_date < today and end_date < today:
        date_category = "Previous"
    elif start_date > today and end_date > today:
        date_category = "Future"
    else:
        date_category = "Present"

    return render(request, "admin-template/paid2.html", {
        'date_category': date_category,
        'date_range': date_range,
        'employ_id': employees,
        'checkin_statuses': checkin_statuses,
        'previous_dates': previous_dates,
        'present_dates': present_dates,
        'future_dates': future_dates,
        'admin_drops': admin_drops,
        'checkin_statuses_page': employees_page,
        'admin_drops':admin_drops,
        'user':user,
 
        'admin_home_drops':admin_home_drops,
        's':s,
        'h':h,
        'projectm':projectm,
        'projects_drops':projects_drops,
        'data':data,
        'employs_all':employs_all,
    })

 


from django.shortcuts import render
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from datetime import datetime, timedelta, date
from .models import Employs, LeaveReportEmploy  # Adjust the import path based on your models

def balance(request):

    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id 
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    s=adminnav.objects.all()
    # Query your data with date filtering
    employees = Employs.objects.filter(companyid=data1)

    # Populate the results list with data
    results = []
    for employee in employees:
        current_month = datetime.now().month
        current_year = datetime.now().year

        last_day_of_month = date(current_year, current_month, 1)
        last_day_of_month = last_day_of_month.replace(day=28)
        while last_day_of_month.month == current_month:
            last_day_of_month += timedelta(days=1)
        last_day_of_month -= timedelta(days=1)

        # Calculate the number of days in the current month
        days_in_month = (last_day_of_month - date(current_year, current_month, 1)).days + 1

        # Calculate the number of days in the current month
        half_leave = LeaveReportEmploy.objects.filter(employ_id=employee, leave_status=1, leave_date__month=current_month, leave_type="Half-Day").count()
        unpaid_leave = LeaveReportEmploy.objects.filter(employ_id=employee, leave_status=1, leave_date__month=current_month, leave_type="Unpaid_leave").count()
        leave = LeaveReportEmploy.objects.filter(employ_id=employee, leave_status=1, leave_date__month=current_month, leave_type="Leave").count()
        open_request = LeaveReportEmploy.objects.filter(employ_id=employee, leave_status=0, leave_date__month=current_month).count()

        results.append({
            'empid': employee.empid,
            'first_name': employee.first_name,
            'half_leave': half_leave,
            'unpaid_leave': unpaid_leave,
            'leave': leave,
            'open_request': open_request,
        })

    items_per_page = 10  # Adjust the number of items per page as needed

    paginator = Paginator(results, items_per_page)
    page = request.GET.get('page')

    try:
        results = paginator.page(page)
    except PageNotAnInteger:
        results = paginator.page(1)
    except EmptyPage:
        results = paginator.page(paginator.num_pages)

    return render(request, 'admin-template/balance.html', {'results': results,            's':s,'data':data,'data1':data1,'admin_drops':admin_drops,'admin_home_drops':admin_home_drops,'projects_drops':projects_drops
})

from datetime import datetime, timedelta
from calendar import monthrange
from django.shortcuts import render
from .models import Employs, checkin
def opeanrequest(request):
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id

    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    employees = Employs.objects.filter(companyid=data1)

    # Get start_date and end_date from query parameters (if provided)
    start_date_param = request.GET.get('start_date')
    end_date_param = request.GET.get('end_date')

    # Determine today's date
    today = datetime.now()

    # Calculate start_date and end_date based on provided parameters or default to the current month
    if start_date_param and end_date_param:
        start_date = datetime.strptime(start_date_param, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_param, '%Y-%m-%d')
    else:
        current_month = today.month
        current_year = today.year
        _, num_days = monthrange(current_year, current_month)
        start_date = datetime(current_year, current_month, 1)
        end_date = datetime(current_year, current_month, num_days)

    # Create a list of dates in the date range
    date_range = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]

    # Categorize dates into "previous," "present," or "future"
    previous_dates = [date for date in date_range if date < today]
    present_dates = [date for date in date_range if date == today]
    future_dates = [date for date in date_range if date > today]

    # Paginate the employees
    items_per_page = 10
    page = request.GET.get('page')
    paginator = Paginator(employees, items_per_page)

    try:
        employees_page = paginator.page(page)
    except PageNotAnInteger:
        employees_page = paginator.page(1)
    except EmptyPage:
        employees_page = paginator.page(paginator.num_pages)

    # Initialize the checkin_statuses dictionary for the current page
    checkin_statuses = {}
    for employee in employees_page:
        checkin_status_for_employee = []

        for day in date_range:
            checkin_exists = checkin.objects.filter(
                empid=employee.email,
                date=day
            ).exists()
            checkin_status_for_employee.append('Present' if checkin_exists else '-NA-')

        checkin_statuses[employee] = checkin_status_for_employee

    # Determine the category of the date range
    if start_date < today and end_date < today:
        date_category = "Previous"
    elif start_date > today and end_date > today:
        date_category = "Future"
    else:
        date_category = "Present"

    return render(request, "admin-template/opeanrequest.html", {
        'date_category': date_category,
        'date_range': date_range,
        'employ_id': employees,
        'checkin_statuses': checkin_statuses,
        'previous_dates': previous_dates,
        'present_dates': present_dates,
        'future_dates': future_dates,
        'admin_drops': admin_drops,
        'checkin_statuses_page': employees_page,
        'user':user,

        'admin_home_drops':admin_home_drops,
        's':s,
        'h':h,
        'projectm':projectm,
        'projects_drops':projects_drops,
        'employs_all':employs_all,'data':data,'data1':data1,

    })




def update_project(request, pid):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')

    # Get the project to be updated
    project = get_object_or_404(Project, id=pid)
    team_leader_info=project.get_team_leader()
    
    # Get available team members excluding team leads
    teams=Employs.objects.filter(Q(teammember__isnull=True) | Q(teammember__project=project)).distinct()
    team_members = TeamMember.objects.filter(project=pid)

    if request.method == 'POST':

       
        # Retrieve updated project details from the form
        p_name = request.POST['p_name']
        p_desc = request.POST['p_desc']
        pr_deadline = request.POST['project_deadline_date']
        project_manager = request.POST['manager_name']
        status = request.POST.get('status')

        # Update the project
        project.p_name = p_name
        project.p_desc = p_desc
        project.project_deadline = pr_deadline
        project.project_manager = project_manager
        project.status = status
        project.save()  
            # Update team lead if changed
        selected_team_lead_id = request.POST.get('team_lead')
        if selected_team_lead_id:
            try:
                team_lead = Employs.objects.get(id=selected_team_lead_id)
                project.teammember_set.update(is_team_lead=False)  # Remove team leader status from existing team leader

                # Create a new TeamMember if it doesn't exist
                team_member, created = TeamMember.objects.get_or_create(project=project, employee=team_lead)
                team_member.is_team_lead = True
                team_member.save()
            except Employs.DoesNotExist:
                # Handle the case where the selected team leader doesn't exist
                messages.error(request, 'Selected team leader does not exist.')
                return redirect('/read-proj')

        # Update selected team members
        selected_employee_ids = request.POST.getlist('selected_employees[]')
        existing_team_member_ids = TeamMember.objects.filter(project=project).values_list('employee_id', flat=True)
        for emp_id in selected_employee_ids:
            emp_id = int(emp_id)
            if emp_id not in existing_team_member_ids:
                employee = get_object_or_404(Employs, id=emp_id)
                TeamMember.objects.create(project=project, employee=employee, is_team_lead=False)

        messages.success(request, 'Project updated successfully and employees assigned.')
        return redirect('/read-proj')

    # Fetch available team members (exclude team leads)
    # selected_team_lead_id = project.team_lead.id if project.team_lead else None
    return render(request, 'admin-template/update_project.html', {"admin_drops":admin_drops,"project": project,"teams":teams, "team_members": team_members,'team_leader_info':team_leader_info})

from django.shortcuts import render, redirect
from .models import admin_project_create  # Import your project model

# def admin_project(request):
#     creater = CustomUser.objects.filter(adminhod__options=1)
#     data = CustomUser.objects.filter(id=request.user.id, user_type=1)
#     admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')

#     user = CustomUser.objects.filter(id=request.user.id).first()
#     userid1 = user.id
#     da1 = AdminHod.objects.filter(admin=request.user.id).first()
#     da2 = da1.id
#     admin = AdminHod.objects.get(id=1)
#     s = adminnav.objects.all()
#     projectm = admin_project_create.objects.filter(admin_id=userid1)
#     h = HR.objects.all()
#     employs_all = Employs.objects.all()
#     admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
#     projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')


#     if request.method == 'POST':
#         project_name = request.POST['project_name']
#         project_dec = request.POST['project_dec']
#         admin_id = request.POST['admin_id']
#         project_manager = CustomUser.objects.get(id=admin_id)  
#         project = admin_project_create(project_name=project_name, project_dec=project_dec, admin_id=project_manager)
#         project.save()

#     projects = admin_project_create.objects.all()

#     return render(request, 'admin-template/admin_create_projects.html', {'admin_drops': admin_drops, 'data': data, 'creater': creater, 'projects': projects,'user':user,'da1':da1,'da2':da2,'admin':admin,'admin_home_drops':admin_home_drops,'s':s,'h':h,'projectm':projectm,'projects_drops':projects_drops,'employs_all':employs_all})

from .models import admin_project_create  # Import your project model

def admin_project(request):

    # creater = CustomUser.objects.filter(adminhod__options=1)
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()

    creater = Employs.objects.filter(projectmanagerop=1,companyid=data1)
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id

   
    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
        # data = AdminHod.objects.filter(id=self.request.user.id)
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')



    if request.method == 'POST':
        project_name = request.POST['project_name']
        project_dec = request.POST['project_dec']
        admin_id = request.POST['admin_id']

        # Create a new project instance and save it to the database
        project = admin_project_create(project_name=project_name, project_dec=project_dec, admin_id=admin_id)
        project.save()

    # Query the list of projects and pass it to the template
    projects = admin_project_create.objects.all()

    return render(request, 'admin-template/admin_create_projects.html', {'admin_drops': admin_drops,'a':a, 'data': data, 'creater': creater, 'projects': projects,'user':user,'admin_home_drops':admin_home_drops,'s':s,'h':h,'projectm':projectm,'projects_drops':projects_drops,'data':data,'employs_all':employs_all})




def ret_admin_project(request):
    data=admin_project.objects.get(admin_id=request.user.id)
    return render(request,"admin-template/retrive_pm.html",{'data':data})


from django.shortcuts import render, redirect
from .models import admin_project_create, taskdata, TeamMember

def display(request):
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)

    pr=AdminHod.objects.filter(admin=request.user.id).first()
    pr1=pr.id
    projects=Project.objects.filter(o_id=pr1)
    tl=TeamMember.objects.filter(is_team_lead=1)

    if request.method == 'POST':
        p_name = request.POST.get('p_name')
        p_url = request.POST.get('p_url')
        team_lead = request.POST.get('team_lead')

        # Create a new project
        p1 = taskdata(
            p_name=p_name,
            p_url=p_url,
            team_lead=team_lead
        )
        p1.save()

        return redirect('/read-proj')

    return render(request, 'admin-template/data.html', {'tl':tl,'projects': projects,'projects_drops':projects_drops,'s':s})



# def projecttask1(request):
#     pr=AdminHod.objects.filter(admin=request.user.id).first()
#     pr1=pr.id
#     projects=Project.objects.filter(o_id=pr1)
#     if request.method == "POST":
#         p_name = request.POST['p_name']
#         l_name = request.POST['l_name']
#         task_names = request.POST.getlist('task_name[]')
        

#         for task_name in task_names:
#             task = projecttask(p_name=p_name, tasks=task_name, l_name=l_name,date=date)
#             task.save()

#     return render(request, 'admin-template/form123.html', {'projects':projects})

from .models import Project, projecttask
from django.shortcuts import render
from django.utils import timezone

def projecttask1(request):
    pr = AdminHod.objects.filter(admin=request.user.id).first()
    pr1 = pr.id
    projects = Project.objects.filter(o_id=pr1)

    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = AdminHod.objects.filter(admin=request.user.id).first()
    da2 = da1.id
    admin = AdminHod.objects.get(id=1)
    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    data = AdminHod.objects.filter(id=request.user.id)
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')


    if request.method == "POST":
        p_id = request.POST['p_name']  # Assuming 'p_name' contains the ID of the project
        l_name = request.POST['l_name']
        task_names = request.POST.getlist('task_name[]')

        # Fetch the Project instance based on the ID obtained
        project_instance = Project.objects.get(pk=p_id)

        for task_name in task_names:
            # Assign the project instance (not just the ID) to the projecttask's p_name field
            task = projecttask(p_name=project_instance, tasks=task_name, l_name=l_name)
            task.save()

    return render(request, 'admin-template/form123.html', {'projects': projects,'user':user,'da1':da1,'da2':da2,'admin':admin,'admin_home_drops':admin_home_drops,'s':s,'h':h,'projectm':projectm,'projects_drops':projects_drops,'data':data,'employs_all':employs_all})






def display1(request):

    o_id = AdminHod.objects.get(admin=request.user.id)
    projects = Project.objects.filter(o_id_id=o_id).values()
    employees = Employs.objects.all()
    context = { "projects": projects, "employees": employees}
    if request.method == 'POST':
        b_id = request.POST['b_id']
        e_id = request.POST['e_id']
        task=request.POST['task']
        performance=request.POST['performance']
        taskObj = Task.objects.create( o_id_id=o_id.admin.id, b_id_id=b_id,  e_id_id=e_id,task=task,performance=performance)
        if taskObj:
            empDetails = Employs.objects.filter(id=e_id).values()
            message = f'Hi {empDetails[0]["first_name"]}'
        else:
            messages.error(request, "Some Error was occurred!")
            return HttpResponseRedirect('/create-task')
    return render(request, 'admin-template/data1.html', context)

# from .models import employperformance
# def employper(request):
#     s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)

#     projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
#     pr=AdminHod.objects.filter(admin=request.user.id).first()
#     pr1=pr.id
#     projects=Project.objects.filter(o_id=pr1)
#     employees = Employs.objects.all()
#     tsk1_data = employperformance.objects.all()

#     for task in tsk1_data:
#         try:
#             employee = Employs.objects.get(id=task.employ_name)
#             task.employee_name = employee.first_name
#         except Employs.DoesNotExist:
#             task.employee_name = "N/A"
        
#         try:
#             project = Project.objects.get(id=task.project_name)
#             task.project_name = project.p_name
#         except Project.DoesNotExist:
#             task.project_name = "N/A"

#     if request.method == "POST":
#         project_name=request.POST.get('project_name')
#         employ_name=request.POST.get('employ_name')
#         task_name=request.POST.get('task_name')
#         performance=request.POST.get('performance')
#         k=employperformance(project_name=project_name,employ_name=employ_name,task_name=task_name,performance=performance)
#         k.save()
#     return render(request,"admin-template/admin-employper.html",{'s':s,'projects_drops':projects_drops,'tsk1_data':tsk1_data,'projects':projects,'employees':employees})


from .models import employperformance, Project, employ_nav  # Add this line to import employ_nav

# The rest of your code remains unchanged...

def employper(request):
    s = employ_nav.objects.filter(is_name_exist=1, is_tl_option=0)
    projects_drops = project_drop.objects.filter(parent_category=None).order_by('id')
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = AdminHod.objects.filter(admin=request.user.id).first()
    da2 = da1.id
    admin = AdminHod.objects.get(id=1)
    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    data = AdminHod.objects.filter(id=request.user.id)
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')


    pr = AdminHod.objects.filter(admin=request.user.id).first()
    pr1 = pr.id
    projects = Project.objects.filter(o_id=pr1)
    employees = Employs.objects.all()
    tsk1_data = employperformance.objects.all()

    for task in tsk1_data:
        try:
            project = Project.objects.get(id=task.project_name.id)
            task.project_name = project  # Assign the Project instance directly
        except Project.DoesNotExist:
            task.project_name = None  # Handle the case when the project is not found

    if request.method == "POST":
        project_id = request.POST.get('project_name')  # Assuming the value is an ID
        employ_id = request.POST.get('employ_name')
        task_name = request.POST.get('task_name')
        performance = request.POST.get('performance')

        try:
            selected_project = Project.objects.get(id=project_id)
        except Project.DoesNotExist:
            selected_project = None  # Set selected_project as None if project is not found

        try:
            selected_employee = Employs.objects.get(id=employ_id)
        except Employs.DoesNotExist:
            selected_employee = None  # Set selected_employee as None if employee is not found
        try:
            project_name = projecttask.objects.get(id=task_name)
        except projecttask.DoesNotExist:
            project_name = None

        k = employperformance(
            project_name=selected_project,
            employ_name=selected_employee,
            project_task1=36,
            performance=performance
        )
        k.save()

    return render(request, "admin-template/admin-employper.html", {
        's': s,
        'projects_drops': projects_drops,
        'tsk1_data': tsk1_data,
        'projects': projects,
        'employees': employees,
        'user':user,
        'da2':da2,
        'admin':admin,
        'admin_drops':admin_drops,
        'h':h,
        'projectm':projectm,
        'projects_drops':projects_drops,
        'employs_all':employs_all,
        'data':data,
        'admin_home_drops':admin_home_drops

    })


from django.http import JsonResponse

def get_team_leads(request):
    project_id = request.GET.get('p_name')
    project = Project.objects.get(id=project_id)
    team_lead_info = project.get_team_leader()
    team_lead_name = f"{team_lead_info['team_first_name']}.{team_lead_info['team_last_name']}"
    return JsonResponse({'team_lead_name': team_lead_name})



from django.shortcuts import render, redirect
from django.shortcuts import get_object_or_404
from django.contrib import messages
from django.db.models import Q
from .models import Project, Employs, TeamMember, admin_drop  # Make sure to import your models

def update_data(request, pid):
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')

    # Get the project to be updated
    project = get_object_or_404(Project, id=pid)
    team_leader_info = project.get_team_leader()

    # Get available team members excluding team leads and those already selected for the project
    team_members = TeamMember.objects.filter(project=project)
    selected_team_member_ids = team_members.values_list('employee_id', flat=True)
    available_team_members = Employs.objects.filter(
        Q(teammember__isnull=True) | (Q(teammember__project=project) & ~Q(id__in=selected_team_member_ids))
    ).distinct()

    if request.method == 'POST':
        # Retrieve updated project details from the form
        p_name = request.POST['p_name']
        p_desc = request.POST['p_desc']
        pr_deadline = request.POST['project_deadline_date']
        project_manager = request.POST['manager_name']
        status = request.POST.get('status')

        # Update the project
        project.p_name = p_name
        project.p_desc = p_desc
        project.project_deadline = pr_deadline
        project.project_manager = project_manager
        project.status = status
        project.save()

        # Update team lead if changed
        selected_team_lead_id = request.POST.get('team_lead')
        if selected_team_lead_id:
            try:
                team_lead = Employs.objects.get(id=selected_team_lead_id)
                project.teammember_set.update(is_team_lead=False)  # Remove team leader status from existing team leader

                # Create a new TeamMember if it doesn't exist
                team_member, created = TeamMember.objects.get_or_create(project=project, employee=team_lead)
                team_member.is_team_lead = True
                team_member.save()
            except Employs.DoesNotExist:
                # Handle the case where the selected team leader doesn't exist
                messages.error(request, 'Selected team leader does not exist.')
                return redirect('/admin_home')

        # Update selected team members
        selected_employee_ids = request.POST.getlist('selected_employees[]')
        existing_team_member_ids = TeamMember.objects.filter(project=project).values_list('employee_id', flat=True)
        for emp_id in selected_employee_ids:
            emp_id = int(emp_id)
            if emp_id not in existing_team_member_ids:
                employee = get_object_or_404(Employs, id=emp_id)
                TeamMember.objects.create(project=project, employee=employee, is_team_lead=False)

        messages.success(request, 'Project updated successfully and employees assigned.')
        return redirect('/admin_home')

    # Render the template with the updated available team members
    return render(request, 'admin-template/editproject.html', {
        "admin_drops": admin_drops,
        "project": project,
        "team_members": team_members,
        "available_team_members": available_team_members,
        'team_leader_info': team_leader_info
    })




from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.shortcuts import render
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.shortcuts import render
from .models import employperformance
from .models import Employs, Project
from datetime import datetime
from django.db.models import Avg

from .models import employperformance, Employs, Project

def data_table(request):
    data = CustomUser.objects.filter(id=request.user.id,user_type=1)

    data1 = AdminHod.objects.get(id=1) 

    
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    # Number of items to display per page
    items_per_page = 5  # Two lines per page
    tsk1_data = employperformance.objects.values('employ_name').annotate(avg_performance=Avg('performance'))

    for task in tsk1_data:
        employ_id = task['employ_name']
        try:
            employee = Employs.objects.get(id=employ_id)
            task['employee_name'] = employee.first_name
            task['employee_id'] = employee.empid

        except Employs.DoesNotExist:
            task['employee_name'] = "N/A"
            task['employee_id'] = "N/A"


    # Fetch additional data for the detailed performance table
    performance_data = employperformance.objects.all()

    for task in performance_data:
        try:
            employee = Employs.objects.get(id=task.employ_name)
            task.employee_name = employee.first_name
            task.employee_id = employee.empid

        except Employs.DoesNotExist:
            task.employee_name = "N/A"

        try:
            project = Project.objects.get(id=task.project_name)
            task.project_name = project.p_name
        except Project.DoesNotExist:
            task.project_name = "N/A"

    # Perform pagination with two lines per page

    
    paginator = Paginator(performance_data, items_per_page)
    page = request.GET.get('page')

    try:
        performance_data = paginator.page(page)
    except PageNotAnInteger:
        performance_data = paginator.page(1)
    except EmptyPage:
        performance_data = paginator.page(paginator.num_pages)
    paginator = Paginator(tsk1_data, items_per_page)
    page = request.GET.get('page')

    try:
        tsk1_data = paginator.page(page)
    except PageNotAnInteger:
        tsk1_data = paginator.page(1)
    except EmptyPage:
        tsk1_data = paginator.page(paginator.num_pages)

    return render(request, 'admin-template/emptable.html', {'data':data,'data1':data1,'admin_drops':admin_drops, 'tsk1_data': tsk1_data,'admin_home_drops':admin_home_drops, 'performance_data': performance_data})



def get_team_lead(request, project_id):
    team_members = TeamMember.objects.filter(project_id=project_id)
    team_lead = TeamMember.objects.filter(project_id=project_id,is_team_lead=1).first()  # Assuming you have a TeamLead model
    data = [{'id': member.employee.id, 'name': f"{member.employee.first_name} {member.employee.last_name}"}
            for member in team_members]
    
    # Add team lead information to the data
    if team_lead:
        data.append({'id': team_lead.employee.id, 'name': f"Team Lead: {team_lead.employee.first_name} {team_lead.employee.last_name}"})
    
    return JsonResponse({'team_members': data})


import os
import zipfile
from io import BytesIO
from django.http import HttpResponse
from django.conf import settings
from .models import Employs

def download_all_employee_data(request):
    # Create a BytesIO stream to write the zip file.
    output = BytesIO()
    zip_file = zipfile.ZipFile(output, 'w')

    # Directory structure for the zip file.
    base_dir = os.path.join(settings.MEDIA_ROOT, 'employees')

    # Iterate over all employees.
    for employ in Employs.objects.all():
        employee_dir = os.path.join(base_dir, employ.first_name)

        # Add employee documents to the zip file.
        for empdoc in employ.empdocs_set.all():
            image_path = empdoc.imagefile.path
            image_name = os.path.basename(image_path)
            relative_path = os.path.join(employ.first_name, image_name)
            zip_file.write(image_path, relative_path)

    zip_file.close()

    # Set response headers to serve the zip file.
    response = HttpResponse(output.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename=all_employee_data.zip'
    return response

import os
import zipfile
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.conf import settings
from io import BytesIO
from .models import Employs  # Import your Employs model

def download_images_zip(request, employ_id):
    employ = get_object_or_404(Employs, id=employ_id)
      
    # Create a BytesIO stream to write the zip file.
    output = BytesIO()
    zip_file = zipfile.ZipFile(output, 'w')

    # Iterate over 'empdocs' related to the employee and add their image files to the zip.
    for empdoc in employ.empdocs_set.all():
        document_type = empdoc.documenttype1
        image_path = empdoc.imagefile.path
        image_name = os.path.basename(image_path)
        
        # Construct the relative path inside the zip file.
        relative_path = os.path.join(employ.first_name, image_name)
        zip_file.write(image_path, relative_path)

    zip_file.close()

    # Set response headers to serve the zip file.
    response = HttpResponse(output.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename={employ.first_name}_images.zip'
    return response


def performancetask(request):
    
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')


    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()
    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()

    projects = Project.objects.all()  
    items_per_page = 2
    performance_data = employperformance.objects.filter(employ_name_id__companyid=data1)
    paginator = Paginator(performance_data, items_per_page)
    page = request.GET.get('page')

    try:
        performance_data = paginator.page(page)
    except PageNotAnInteger:
        performance_data = paginator.page(1)
    except EmptyPage:
        performance_data = paginator.page(paginator.num_pages)
    return render(request, 'admin-template/emptaskper.html', {'a':a,'projects_drops':projects_drops,'projects':projects,'data':data,'data1':data1,'admin_drops':admin_drops,'admin_home_drops':admin_home_drops, 'performance_data': performance_data,'user':user,'s':s,'h':h,'projectm':projectm,'data':data,'employs_all':employs_all})



# def performanceproject(request):
#     data = AdminHod.objects.filter(admin=request.user.id,options=0)
#     data1 = AdminHod.objects.get(id=1) 
#     admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
#     admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
#     projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')

#     user = CustomUser.objects.filter(id=request.user.id).first()
#     userid1 = user.id
#     da1 = AdminHod.objects.filter(admin=request.user.id).first()
#     da2 = da1.id
#     admin = AdminHod.objects.get(id=1)
#     s = adminnav.objects.all()
#     projectm = admin_project_create.objects.filter(admin_id=userid1)
#     h = HR.objects.all()
#     employs_all = Employs.objects.all()
#     data = AdminHod.objects.filter(id=request.user.id)


#     projects = Project.objects.all()  

#     items_per_page = 10  # Adjust the number of items per page as needed

#     paginator = Paginator(employee_average_performances, items_per_page)
#     page = request.GET.get('page')

#     try:
#         employee_average_performances = paginator.page(page)
#     except PageNotAnInteger:
#         employee_average_performances = paginator.page(1)
#     except EmptyPage:
#         employee_average_performances = paginator.page(paginator.num_pages)

#     if request.method == 'POST':
#         selected_project_id = request.POST.get('project')  # Get the selected project ID

#         if selected_project_id:
#             selected_project = Project.objects.get(id=selected_project_id)
#             project_name = selected_project.p_name
#             # Fetch the employee average performances for the selected project
#             performances_for_project = employperformance.objects.filter(project_name=selected_project)
#             employee_average_performances = performances_for_project.values('employ_name').annotate(average_performance=Avg('performance'))
           

#             # Retrieve employee names and IDs
#             for emp in employee_average_performances:
#                 employ_id = emp['employ_name']
#                 try:
#                     employee = Employs.objects.get(id=employ_id)
#                     emp['employee_name'] = employee.first_name
#                     emp['employee_id'] = employee.empid
#                 except Employs.DoesNotExist:
#                     emp['employee_name'] = "N/A"
#                     emp['employee_id'] = "N/A"
#         else:
#             performances_for_project=None
#             project_name=None
#             employee_average_performances=None
#     else:
#             performances_for_project=None
#             project_name=None
#             employee_average_performances=None



#     return render(request, 'admin-template/empprojectper.html', {'projects_drops':projects_drops,'projects':projects,'project_name':project_name,'employee_average_performances':employee_average_performances,'data':data,'data1':data1,'admin_drops':admin_drops,'admin_home_drops':admin_home_drops, 'user':user,'da1':da1,'da2':da2,'admin':admin,'s':s,'h':h,'projectm':projectm,'data':data,'employs_all':employs_all})
def performanceproject(request):
    
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    admin_home_drops = admin_home_drop.objects.filter(parent_category=None).order_by('id')
    projects_drops = project_drop.objects.filter(parent_category=None).order_by('id')
    s=adminnav.objects.all()
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()
    projects = Project.objects.filter(o_id__companyid=data1)
    # items_per_page = 10  # Adjust the number of items per page as needed

    employee_average_performances = []  # Initialize as an empty list

    project_name = None  # Initialize project_name variable

    if request.method == 'POST':
        selected_project_id = request.POST.get('project')

        if selected_project_id:
            selected_project = Project.objects.get(id=selected_project_id)
            project_name = selected_project.p_name  # Set project_name here
            performances_for_project = employperformance.objects.filter(project_name=selected_project)
            employee_average_performances = performances_for_project.values('employ_name').annotate(
                average_performance=Avg('performance'))

            for emp in employee_average_performances:
                employ_id = emp['employ_name']
                try:
                    employee = Employs.objects.get(id=employ_id)
                    emp['employee_name'] = employee.first_name
                    emp['employee_id'] = employee.empid
                except Employs.DoesNotExist:
                    emp['employee_name'] = "N/A"
                    emp['employee_id'] = "N/A"

    # Check if employee_average_performances is None, then set it as an empty list
    if employee_average_performances is None:
        employee_average_performances = []
        
    items_per_page = 10  # Adjust the number of items per page as needed

    paginator = Paginator(employee_average_performances, items_per_page)
    page = request.GET.get('page')

    try:
        employee_average_performances = paginator.page(page)
    except PageNotAnInteger:
        employee_average_performances = paginator.page(1)
    except EmptyPage:
        employee_average_performances = paginator.page(paginator.num_pages)

    return render(request, 'admin-template/empprojectper.html', {
        'projects_drops': projects_drops,
        'projects': projects,
        'project_name': project_name,  # Pass project_name to the template
        'employee_average_performances': employee_average_performances,
        'data': data,
        'data1': data1,
        'admin_drops': admin_drops,
        'admin_home_drops': admin_home_drops,
        's':s,
        'a':a,
    })


# def performanceallproject(request):
    
#     data = AdminHod.objects.filter(admin=request.user.id,options=0)
#     data1 = AdminHod.objects.get(id=1) 
#     admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
#     admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
#     projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')

#     user = CustomUser.objects.filter(id=request.user.id).first()
#     userid1 = user.id
#     da1 = AdminHod.objects.filter(admin=request.user.id).first()
#     da2 = da1.id
#     admin = AdminHod.objects.get(id=1)
#     s = adminnav.objects.all()
#     projectm = admin_project_create.objects.filter(admin_id=userid1)
#     h = HR.objects.all()
#     employs_all = Employs.objects.all()
#     data = AdminHod.objects.filter(id=request.user.id)


#     projects = Project.objects.all()  
#     items_per_page = 2
#     tsk1_data = employperformance.objects.values('employ_name').annotate(avg_performance=Avg('performance'))
#     employees_data = []
#     for task in tsk1_data:
#         employ_id = task['employ_name']
#         try:
#             employee = Employs.objects.get(id=employ_id)
#             task['employee_name'] = employee.first_name
#             task['employee_id'] = employee.empid
#             project_names = employperformance.objects.filter(employ_name=employee).values('project_name__p_name')
#             task['project_names'] = project_names

#         except Employs.DoesNotExist:
#             task['employee_name'] = "N/A"
#             task['employee_id'] = "N/A"
#             task['project_names'] = "N/A"
#         employees_data.append(task)
#     items_per_page = 10  # Adjust the number of items per page to 10

#     paginator = Paginator(employees_data, items_per_page)
#     page = request.GET.get('page')

#     try:
#         tsk1_data = paginator.page(page)
#     except PageNotAnInteger:
#         tsk1_data = paginator.page(1)
#     except EmptyPage:
#         tsk1_data = paginator.page(paginator.num_pages)

#     projects = Project.objects.all()

#     return render(request, 'admin-template/empavgper.html', {'projects_drops':projects_drops,'projects':projects,'data':data,'data1':data1,'admin_drops':admin_drops, 'tsk1_data': employees_data,'admin_home_drops':admin_home_drops,'user':user,'da1':da1,'da2':da2,'admin':admin,'s':s,'h':h,'projectm':projectm,'data':data,'employs_all':employs_all})


def performanceallproject(request):
    
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    admin_home_drops = admin_home_drop.objects.filter(parent_category=None).order_by('id')
    projects_drops = project_drop.objects.filter(parent_category=None).order_by('id')
    s=adminnav.objects.all()
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()
    # Calculate the average performance for employees
    tsk1_data = employperformance.objects.filter(employ_name_id__companyid=data1).values('employ_name').annotate(avg_performance=Avg('performance'))

    employees_data = []
    for task in tsk1_data:
        employ_id = task['employ_name']
        try:
            employee = Employs.objects.get(id=employ_id)
            task['employee_name'] = employee.first_name
            task['employee_id'] = employee.empid
            project_names = employperformance.objects.filter(employ_name=employee).values('project_name__p_name')
            task['project_names'] = project_names

        except Employs.DoesNotExist:
            task['employee_name'] = "N/A"
            task['employee_id'] = "N/A"
            task['project_names'] = "N/A"
        employees_data.append(task)

    items_per_page = 10  # Adjust the number of items per page to 10

    paginator = Paginator(employees_data, items_per_page)
    page = request.GET.get('page')

    try:
        tsk1_data = paginator.page(page)
    except PageNotAnInteger:
        tsk1_data = paginator.page(1)
    except EmptyPage:
        tsk1_data = paginator.page(paginator.num_pages)

    projects = Project.objects.all()

    return render(request, 'admin-template/empavgper.html', {
        'projects_drops': projects_drops,
        'projects': projects,
        'data': data,
        'data1': data1,
        's':s,
        'admin_drops': admin_drops,
        'a':a,
        'admin_home_drops': admin_home_drops,
                'tsk1_data': tsk1_data,

    })


   
def data_tables(request):
    
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    s=adminnav.objects.all()

    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()

    today = date.today()
    tsk1_data = tlassigntask.objects.filter(task_date=today)  # Filter tasks for the current date only
    team_leads = TeamMember.objects.filter(is_team_lead=True,employee_id__companyid=data1)

    for task in tsk1_data:
        try:
            employee = Employs.objects.get(id=task.employid)
            task.employid = employee.first_name
        except Employs.DoesNotExist:
            task.employid = "N/A"
        
        try:
            project = Project.objects.get(id=task.project_id)
            task.project_name = project.p_name
        except Project.DoesNotExist:
            task.project_name = "N/A"

    items_per_page = 10  # Adjust the number of items per page to 10

    paginator = Paginator(tsk1_data, items_per_page)
    page = request.GET.get('page')

    try:
        tsk1_data = paginator.page(page)
    except PageNotAnInteger:
        tsk1_data = paginator.page(1)
    except EmptyPage:
        tsk1_data = paginator.page(paginator.num_pages)
    
    return render(request, 'admin-template/emptables.html', {'a':a,'tsk1_data': tsk1_data, 'team_leads': team_leads,'s':s,'data':data,'data1':data1,'admin_drops':admin_drops,'admin_home_drops':admin_home_drops,'projects_drops':projects_drops})




from django.shortcuts import render, redirect
from .models import admin_project_create  # Import your project model

def project_list(request):
    creater = CustomUser.objects.filter(adminhod__options=1)
    
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()
    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    # data = AdminHod.objects.filter(id=request.user.id)


    if request.method == 'POST':
        project_name = request.POST['project_name']
        project_dec = request.POST['project_dec']
        admin_id = request.POST['admin_id']
        # Create a new project instance and save it to the database
        project_manager = CustomUser.objects.get(id=admin_id)  # Get the project manager instance
        project = admin_project_create(project_name=project_name, project_dec=project_dec, admin_id=project_manager)
        project.save()

    # Query the list of projects and pass it to the template
    projects = admin_project_create.objects.filter(companyid=data1)
    items_per_page = 10  # You can adjust this value as needed

    # Pagination
    paginator = Paginator(projects, items_per_page)
    page = request.GET.get('page')

    try:
        projects = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver the first page.
        projects = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver the last page of results.
        projects = paginator.page(paginator.num_pages)
    return render(request, 'admin-template/admin-projectslist.html', {'a':a,'admin_drops': admin_drops, 'data': data, 'creater': creater, 'projects': projects,'user':user,'admin_home_drops':admin_home_drops,'s':s,'h':h,'projectm':projectm,'data':data,'employs_all':employs_all})


from django.shortcuts import render, redirect
from django.http import HttpResponse
from .models import admin_project_create

def projectstatus(request):
    
    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = AdminHod.objects.filter(admin=request.user.id).first()
    da2 = da1.id
    admin = AdminHod.objects.get(id=1)
    s = adminnav.objects.all()
    # projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')

    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    data = AdminHod.objects.filter(id=request.user.id)
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')

    if request.method == "POST":

       


        projectm = admin_project_create.objects.filter(admin_id=request.user.id)


        
        for project in projectm:
            project_id = project.id
            status = request.POST.get(f'status-{project_id}')
            
            if status in ('Complete', 'Incomplete','on going'):
                project.status = status
                project.save()

        return redirect('admin_home')  # You can replace 'admin_home' with the actual URL you want to redirect to

    # Only get the projects assigned to the manager
    projectm = admin_project_create.objects.filter(admin_id=request.user.id)

    return render(request, "admin-template/projectstatus.html", {'projectm': projectm,'user':user,'da1':da1,'da2':da2,'admin':admin,'admin_home_drops':admin_home_drops,'s':s,'h':h,'projectm':projectm,'projects_drops':projects_drops,'data':data,'employs_all':employs_all,'admin_drops':admin_drops})


    from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Alignment

def download_excel(request):
    tsk1_data = employperformance.objects.values('employ_name').annotate(avg_performance=Avg('performance'))

    # Create an Excel workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(["Employee ID", "Employee Name", "Project Names", "Average Performance"])

    for task in tsk1_data:
        employ_id = task['employ_name']
        try:
            employee = Employs.objects.get(id=employ_id)
            task['employee_name'] = employee.first_name
            task['employee_id'] = employee.empid

            # Aggregate unique project names for the employee
            project_names = set(employperformance.objects.filter(employ_name=employee).values_list('project_name__p_name', flat=True))
            # Join unique project names with line breaks
            task['project_names'] = "\n".join(project_names)

        except Employs.DoesNotExist:
            task['employee_name'] = "N/A"
            task['employee_id'] = "N/A"
            task['project_names'] = "N/A"

        # Append the data to the Excel sheet
        worksheet.append([task['employee_id'], task['employee_name'], task['project_names'], task['avg_performance']])

    # Apply alignment style to the cell containing project names for line breaks
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # Create an HttpResponse with the Excel file
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="tsk1_data.xlsx"'
    workbook.save(response)

    return response



import openpyxl
from django.http import HttpResponse
from django.db.models import Avg

def download_all_performance_data_excel(request):
    performance_data = employperformance.objects.all()

    # Create an Excel workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(["Employee ID", "Employee Name", "Project Name", "Date", "Performance"])  # Include "Date" in the header

    for performance in performance_data:
        try:
            employee = Employs.objects.get(id=performance.employ_name.id)
            project_name = performance.project_name.p_name
            worksheet.append([employee.empid, employee.first_name, project_name, performance.date, performance.performance])  # Include date in the row
        except Employs.DoesNotExist:
            pass  # Handle the case where the employee doesn't exist

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="all_performance_data.xlsx"'
    workbook.save(response)

    return response

def payslip_apply_view1(request):
    # staff_obj = Employs.objects.get(admin=request.user.id)
    # leave_data=payslip_request.objects.filter(student_id=staff_obj)
    # s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)

    # user = CustomUser.objects.filter(id=request.user.id).first()
    # userid1 = user.id
    # da1 = AdminHod.objects.filter(admin=request.user.id).first()
    # da2 = da1.id
    # admin = AdminHod.objects.get(id=1)
    # s = adminnav.objects.all()
    # projectm = admin_project_create.objects.filter(admin_id=userid1)
    # h = HR.objects.all()
    # employs_all = Employs.objects.all()
    # admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    # data = AdminHod.objects.filter(id=request.user.id)
    # projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')


    # return render(request,"admin-template/payslip_request_apply_view.html",{'s':s,'user':user,'da1':da1,'admin':admin,'projectm':projectm,'employs_all':employs_all,'data':data,'admin_home_drops':admin_home_drops,'projects_drops':projects_drops})
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = Employs.objects.filter(admin=request.user.id).first()
    if da1:
       da2 = da1.id
    else:
        da2=None
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')

    a=company_details.objects.filter(companyid=data1).first()

    # s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)
    s=adminnav.objects.all()
    employ = Employs.objects.all()
  
    leaves = payslip_request.objects.filter(student_id__companyid=data1)

    total_count = leaves.count()
    approved_count = leaves.filter(status=1).count()
    pending_count = leaves.filter(status=0).count()
    reject_count = leaves.filter(status=2).count()
    items_per_page = 10  # You can adjust this value as needed

    # Pagination
    paginator = Paginator(leaves, items_per_page)
    page = request.GET.get('page')

    try:
        leaves = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver the first page.
        leaves = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver the last page of results.
        leaves = paginator.page(paginator.num_pages)

    context = {
        'approved_count': approved_count,
        'pending_count': pending_count,
        'reject_count': reject_count,
        'total_count': total_count,
        'leaves': leaves,
        'admin_drops': admin_drops,
        'employ': employ,
        's':s,
        'a':a,
        'user':user,
        'da1':da1,
        'da2':da2,
        'data':data,
        'admin_home_drops':admin_home_drops,
        'employs_all':employs_all,
        'h':h,
      
        'projectm':projectm,
    }

    return render(request, "admin-template/payslip_request_status.html", context)

def trails1(request):
    # staff_obj = Employs.objects.get(admin=request.user.id)
    # leave_data=empdocs.objects.filter(employ_id=staff_obj)
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id

   
    # s = adminnav.objects.all()
    # projectm = admin_project_create.objects.filter(admin_id=userid1)
    # h = HR.objects.all()
    # employs_all = Employs.objects.all()
    # admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()

    s=adminnav.objects.filter(is_projectmanager = 1)
    projectm=admin_project_create.objects.filter(admin_id= userid1)
    h=HR.objects.all()
    employs_all=Employs.objects.all()
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')

    employ=empdocs.objects.filter(employ_id__companyid=data1)
    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id

    # s = adminnav.objects.all()
    # projectm = admin_project_create.objects.filter(admin_id=userid1)
    # h = HR.objects.all()
    # employs_all = Employs.objects.all()
    # admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    # admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    # data = AdminHod.objects.filter(id=request.user.id)
    r=documents_setup1.objects.all()
    items_per_page = 10  # Adjust the number of items per page as needed

    paginator = Paginator(employ, items_per_page)
    page = request.GET.get('page')

    try:
        employ = paginator.page(page)
    except PageNotAnInteger:
        employ = paginator.page(1)
    except EmptyPage:
        employ = paginator.page(paginator.num_pages)
    if request.method=="POST":
        documenttype1=request.POST["documenttype1"]
        imagefile=request.FILES.get('imagefile')
        description=request.POST["description"]
        employ_obj=Employs.objects.get(admin=request.user.id)
        k=empdocs(employ_id=employ_obj,documenttype1=documenttype1,imagefile=imagefile,description=description)
        k.save()
    leave_data = empdocs.objects.all()  # Update this to fetch your documents data

    items_per_page = 10  # Adjust the number of items per page as needed

    paginator = Paginator(leave_data, items_per_page)
    page = request.GET.get('page')

    try:
        leave_data = paginator.page(page)
    except PageNotAnInteger:
        leave_data = paginator.page(1)
    except EmptyPage:
        leave_data = paginator.page(paginator.num_pages)
        return redirect("/trails1")
    return render(request, "admin-template/documents_uploaded_view.html", {
        's': s,
        'r': r,
        'data': data,
        'user': user,
        'a':a,
        'projectm': projectm,
        'admin_home_drops': admin_home_drops,
        'h': h,
        'employs_all': employs_all,
        'leave_data': leave_data, 
    
        'employ':employ,'admin_drops':admin_drops,'user':user,'admin_home_drops':admin_home_drops,'s':s,'h':h,'projectm':projectm,'data':data,'employs_all':employs_all
          # Add the paginated data to the context
    })
      


def paysliprequest1(request):
    # staff_obj =Employs.objects.get(admin=request.user.id)
    s = employnav.objects.all()
    k = types.objects.all()

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = AdminHod.objects.filter(admin=request.user.id).first()
    da2 = da1.id
    admin = AdminHod.objects.get(id=1)
    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')

    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    data = AdminHod.objects.filter(id=request.user.id)
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')


    st4 = request.POST.get("ss")
    st3 = request.POST.get("vk")
    st = request.POST.get("d1")
    st1 = request.POST.get("d2")

    leave_data = payslip_request.objects.all()  # Initialize with all records

    # Apply filters based on user input
    if st4 and st4 != '----Select----':  # Check if a valid status is selected
        leave_data = leave_data.filter(payslip_request_status=st4)
    if st3:  # If "Select Type" is selected
        leave_data = leave_data.filter(typea__icontains=st3)
    if st and st1:
        leave_data = leave_data.filter(date__range=[st, st1])

    total_approved = leave_data.filter(status=1)
    total_pending = leave_data.filter(status=0)

    return render(request, "admin-template/payslipreq.html", {
        'leave_data': leave_data,
        's': s,
        'k': k,
        'total': total_approved, 
        'total1': total_pending,
        'user':user,
        'projectm':projectm,
        'projects_drops':projects_drops,
        'data':data,
        'h':h,
        'da2':da2,
        'admin':admin,
        'admin_home_drops':admin_home_drops,
        'employs_all':employs_all,
        'admin_drops':admin_drops,
    })


def advancesalary_request1(request):
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = AdminHod.objects.filter(admin=request.user.id).first()
    da2 = da1.id
    admin = AdminHod.objects.get(id=1)
    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')

    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    data = AdminHod.objects.filter(id=request.user.id)
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')



    # staff_obj = Employs.objects.get(admin=request.user.id)
    # leave_data=ad_salary.objects.filter(employ_id=staff_obj)
    return render(request,"admin-template/employ_advsalayr_request.html",{'s':s,'user':user,'da2':da2,'admin':admin,'admin_home_drops':admin_home_drops,'admin_drops':admin_drops,'h':h,'projectm':projectm,'projects_drops':projects_drops,'data':data,'employs_all':employs_all})
def data1(request):
    staff_obj = Employs.objects.all()
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')


    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id


    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()

    s=adminnav.objects.all()
    k = types.objects.all()



    redirect('/home')
    return render(request,"adminhelp/help.html",{        
        's': s,
        'k': k,
        'staff_obj':staff_obj,
        'admin_drops':admin_drops,
        'admin_home_drops':admin_home_drops,
        'user':user,
        'a':a,
        'data':data,
        'admin_home_drops':admin_home_drops,
        'employs_all':employs_all,
        'h':h,
      
})
def hom(request):
    
    return render(request,"adminhelp/getstarted.html")
def hom1(request):
    
    return render(request,"adminhelp/leave.html")
def hom2(request):
    
    return render(request,"adminhelp/taxreduction.html")
def hom3(request):
   
    return render(request,"adminhelp/reimbesement.html")
def hom4(request):
    
    return render(request,"adminhelp/investments.html")
def hom5(request):
    
    return render(request,"adminhelp/mytax.html")
def hom6(request):
    return render(request,"adminhelp/search.html")
def hom7(request):
    return render(request,"adminhelp/gstarted.html")
def hom8(request):
    return render(request,"adminhelp/people.html")
def hom9(request):
    return render(request,"adminhelp/payroll.html")
def hom10(request):
    return render(request,"adminhelp/statutory.html")
def hom11(request):
    return render(request,"adminhelp/payment.html")
def hom12(request):
    return render(request,"adminhelp/selfservice.html")
def hom13(request):
    return render(request,"adminhelp/insurance.html")
def hom14(request):
    return render(request,"adminhelp/accountintegration.html")
def hom15(request):
    return render(request,"adminhelp/integration.html")
def hom16(request):
    return render(request,"adminhelp/modules.html")
def hom17(request):
    return render(request,"adminhelp/updates.html")
def hom18(request):
    return render(request,"adminhelp/tax-regime.html")
def hom19(request):
    return render(request,"adminhelp/onetime.html")
def hom20(request):
    return render(request,"adminhelp/sal.html")
def inf1(request):
    return render(request,"adminhelp/contactsupport.html")
def inf2(request):
    return render(request,"adminhelp/guied1.html")
def inf3(request):
    return render(request,"adminhelp/guied2.html")
def inf4(request):
    return render(request,"adminhelp/guied3.html")
def dem1(request):
    return render(request,"adminhelp/demo1.html")
def dem2(request):
    return render(request,"adminhelp/demo2.html")
def dem3(request):
    return render(request,"adminhelp/demo3.html")
def dem4(request):
    return render(request,"adminhelp/demo4.html")
def dem5(request):
    return render(request,"adminhelp/demo5.html")
def dem6(request):
    return render(request,"adminhelp/demo6.html")
def dem7(request):
    return render(request,"adminhelp/demo7.html")
def dem8(request):
    return render(request,"adminhelp/demo8.html")
def peopl1(request):
    return render(request,"adminhelp/people1.html")
def peopl2(request):
    return render(request,"adminhelp/people2.html")
def peopl3(request):
    return render(request,"adminhelp/people3.html")
def peopl4(request):
    return render(request,"adminhelp/people4.html")
def peopl5(request):
    return render(request,"adminhelp/people5.html")
def payrol1(request):
    return render(request,"adminhelp/payroll1.html")
def payrol2(request):
    return render(request,"adminhelp/payroll2.html")
def payrol3(request):
    return render(request,"adminhelp/payroll3.html")
def payrol4(request):
    return render(request,"adminhelp/payroll4.html")
def payrol5(request):
    return render(request,"adminhelp/payroll5.html")
def payrol6(request):
    return render(request,"adminhelp/payroll6.html")
def statutor1(request):
    return render(request,"adminhelp/statutory1.html")
def statutor3(request):
    return render(request,"adminhelp/statutory3.html")
def sfun1(request):
    return render(request,"adminhelp/sfund1.html")
def sfun2(request):
    return render(request,"adminhelp/sfund2.html")
def sfun3(request):
    return render(request,"adminhelp/sfund3.html")
def es1(request):
    return render(request,"adminhelp/esi1.html")
def es2(request):
    return render(request,"adminhelp/esi2.html")
def es3(request):
    return render(request,"adminhelp/esi3.html")
def td1(request):
    return render(request,"adminhelp/tds1.html")
def td2(request):
    return render(request,"adminhelp/tds2.html")
def td3(request):
    return render(request,"adminhelp/tds3.html")
def td4(request):
    return render(request,"adminhelp/tds4.html")
def taxs1(request):
    return render(request,"adminhelp/invest.html")
def taxs2(request):
    return render(request,"adminhelp/tax2.html")
def paymen1(request):
    return render(request,"adminhelp/payment1.html")
def paymen2(request):
    return render(request,"adminhelp/payment2.html")
def paymen3(request):
    return render(request,"adminhelp/payment3.html")
def paymen4(request):
    return render(request,"adminhelp/payment4.html")
def paymen5(request):
    return render(request,"adminhelp/payment5.html")
def billin1(request):
    return render(request,"adminhelp/billing1.html")

def billin2(request):
    return render(request,"adminhelp/billing2.html")

def billin3(request):
    return render(request,"adminhelp/billing3.html")
def billin4(request):
    return render(request,"adminhelp/billing4.html")

def billin5(request):
    return render(request,"adminhelp/billing5.html")
def billin6(request):
    return render(request,"adminhelp/billing6.html")
def billin7(request):
    return render(request,"adminhelp/billing7.html")
def billin8(request):
    return render(request,"adminhelp/billing8.html")
def billin9(request):
    return render(request,"adminhelp/billing9.html")
def servic1(request):
    return render(request,"adminhelp/service1.html")
def servic2(request):
    return render(request,"adminhelp/service2.html")
def servic3(request):
    return render(request,"adminhelp/service3.html")
def servic5(request):
    return render(request,"adminhelp/service5.html")
def attendenc1(request):
    return render(request,"adminhelp/attendence1.html")
def attendenc2(request):
    return render(request,"adminhelp/attendence2.html")
def attendenc3(request):
    return render(request,"adminhelp/attendence3.html")
def attendenc4(request):
    return render(request,"adminhelp/attendence4.html")
def attendenc5(request):
    return render(request,"adminhelp/attendence5.html")
def attendenc6(request):
    return render(request,"adminhelp/attendence6.html")
def attendenc7(request):
    return render(request,"adminhelp/attendence7.html")
def attendenc8(request):
    return render(request,"adminhelp/attendence8.html")
def reimbu1(request):
    return render(request,"adminhelp/reimbus1.html")
def reimbu2(request):
    return render(request,"adminhelp/reimbus2.html")
def letter1(request):
    return render(request,"adminhelp/dletter1.html")
def letter2(request):
    return render(request,"adminhelp/dletter2.html")
def insuranc1(request):
    return render(request,"adminhelp/insurance1.html")
def insuranc2(request):
    return render(request,"adminhelp/insurance2.html")
def insuranc3(request):
    return render(request,"adminhelp/insurance3.html")
def insuranc4(request):
    return render(request,"adminhelp/insurance4.html")
def insuranc5(request):
    return render(request,"adminhelp/insurance5.html")
def insuranc6(request):
    return render(request,"adminhelp/insurance6.html")
def insuranc7(request):
    return render(request,"adminhelp/insurance7.html")
def insuranc8(request):
    return render(request,"adminhelp/insurance8.html")
def insuranc9(request):
    return render(request,"adminhelp/insurance9.html")
def insuranc10(request):
    return render(request,"adminhelp/insurance10.html")
def insuranc11(request):
    return render(request,"adminhelp/insurance11.html")
def insuranc12(request):
    return render(request,"adminhelp/insurance12.html")
def insuranc13(request):
    return render(request,"adminhelp/insurance13.html")
def insuranc14(request):
    return render(request,"adminhelp/insurance14.html")
def insuranc15(request):
    return render(request,"adminhelp/insurance15.html")
def insuranc16(request):
    return render(request,"adminhelp/insurance16.html")
def insuranc17(request):
    return render(request,"adminhelp/insurance17.html")
def insuranc18(request):
    return render(request,"adminhelp/insurance18.html")
def insuranc19(request):
    return render(request,"adminhelp/insurance19.html")
def slac(request):
    return render(request,"adminhelp/slack.html")
def book1(request):
    return render(request,"adminhelp/books1.html")
def book2(request):
    return render(request,"adminhelp/books2.html")
def book3(request):
    return render(request,"adminhelp/books3.html")
def book4(request):
    return render(request,"adminhelp/books4.html")
def book5(request):
    return render(request,"adminhelp/books5.html")
def book6(request):
    return render(request,"adminhelp/books6.html")
def book7(request):
    return render(request,"adminhelp/books7.html")
def book8(request):
    return render(request,"adminhelp/books8.html")