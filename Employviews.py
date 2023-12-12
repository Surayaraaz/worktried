import datetime

from django.contrib import messages
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render,redirect
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.core.mail import send_mail
from django.conf import settings
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.db.models import Sum
from ehrms .models import taskdata,list,HR, Employs,employ_drop,CustomUser,payslip_request,documents_setup,empdocs,ad_salary,duration_year,duration_months,Reimbursement,adminnav,employnav,types,typeofd,year,LeaveReportEmploy,employ_add_form,employ_payslip,NotificationEmploy,checkin,checkout,employ_tax_form,reimbursementsetup1,documents_setup1,company_details
from io import BytesIO
from django.template.loader import get_template
from django.core.paginator import Paginator

from xhtml2pdf import pisa  

from django import forms
from ehrms.forms import ScreenshotsForm

class HomeForm(forms.Form):
    name = forms.CharField(max_length=100)
    email = forms.EmailField()
    

from.models import project_drop
def Employ_home(request):
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    employ_obj=Employs.objects.get(admin=request.user.id)
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    compid=data.companyid
    a=company_details.objects.filter(companyid=compid).first()
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1)

    notification1=NotificationEmploy.objects.filter(employ_id=employ_obj.id)
    notifications1=NotificationEmploy.objects.filter(employ_id=employ_obj.id).count()

    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')
    user=Employs.objects.filter(admin=request.user.id).first()
    userid=user.id
    projectm=admin_project_create.objects.filter(admin_id = userid)
    da1=Employs.objects.filter(admin=request.user.id).first()
    da2=da1.id
    project_details = Project.objects.filter(o_id=da2).prefetch_related('teammember_set') # Adjust this query to filter projects as needed


    emp_details=Employs.objects.filter(admin=request.user.id) 
    emp_detail=Employs.objects.get(admin=request.user.id) 
    today = date.today()
    next_week = today + timedelta(days=6)
    tlop=TeamMember.objects.filter(employee=data1,is_team_lead=1)
    tloptions=employnav.objects.filter(is_name_exist=1,is_tl_option=1)

    # Query employees whose birthdays fall within the next week
    employees2 = employ_add_form.objects.filter(dob__day__gte=today.day, dob__day__lte=next_week.day, dob__month=today.month).order_by('dob')
    
    for employee in employees2:
        if employee.dob.day == today.day:
            subject =  'Heartfelt Birthday Greetings for a Valued Team Member 🎉'  
            message = f"Dear {employee.firstname1} {employee.lastname1},\n\n “Wishing you a day filled with joy, laughter, and unforgettable moments as you celebrate another year of wonderful experiences and accomplishments. 🎂 Happy Birthday!”\n\nYour presence and contributions have greatly enriched our team, bringing forth creativity, dedication, and a positive spirit. On this special day, we recognize and appreciate all that you do to make our workplace a better and more enjoyable environment for everyone.\n\nBest wishes,\n HR,\n DevelopTrees Refinding IT solutions\n"
            send_mail(subject, message, 'saipathivada1234@gmail.com', [employee.email2])

    team_members = TeamMember.objects.filter(is_team_lead=1,employee=data1).first()
    task = tlassigntask.objects.filter(employid=data1)
    k=taskdata.objects.filter(team_lead=data1).order_by('-date').first()
    k1=Meeting.objects.filter(team_member=data1).order_by('-date').first()

    if team_members:
       team1 = team_members.project
       employee_ids = TeamMember.objects.filter(project=team1).values_list('employee_id', flat=True)
       project_ids = TeamMember.objects.filter(project=team1).values_list('project_id', flat=True)
       employees = Employs.objects.filter(id__in=employee_ids)
       projects = Project.objects.filter(id__in=project_ids)
    else:
    # No team members found
      team1 = None
      employees = None
      projects = None

    missing=employ_add_form.objects.filter(student_id_id=employ_obj)
    
    missing_doc=empdocs.objects.filter(employ_id=employ_obj)
    
    try:
        user_data=employ_add_form.objects.get(student_id_id=employ_obj)
        is_form_complete=all(getattr(user_data,field) for field in ['pan','ifsecode','acno','beneficiaryname','phno','gender1','dob','address1','heq','aadharno','bloodgroup','profile_pic'])
    except employ_add_form.DoesNotExist:
        is_form_complete = False 
    try:
        user_data_doc = empdocs.objects.filter(employ_id=employ_obj).first()
        if user_data_doc:
         is_form_complete_doc = all(getattr(user_data_doc, field) for field in ['documenttype1', 'imagefile', 'description'])
        else:
         is_form_complete_doc = False
    except empdocs.DoesNotExist:
        is_form_complete_doc = False    
    return render(request,"employ-template/employ_home_template.html",{'project_details':project_details,'projectm':projectm,'projects_drops':projects_drops,'data2':data2,'tloptions':tloptions,'tlop':tlop,'k':k,'k1':k1,'task':task,'team_members':team_members,'projects':projects,'employees':employees,'data':data,'admin_drops':admin_drops,'employees2': employees2,'is_form_complete_doc':is_form_complete_doc,'user_data_doc':user_data_doc if is_form_complete_doc else None,'is_form_complete':is_form_complete,'user_data':user_data if is_form_complete else None,'emp_detail':emp_detail,'emp_details':emp_details,"employ_obj":employ_obj,'s':s,"notification1":notification1,'missing':missing,'a':a})

############### project views ################

from .models import employperformance, Project, employ_nav  # Add this line to import employ_nav

# The rest of your code remains unchanged...

def employper1(request):
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    compid=data.companyid
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1) 
    a=company_details.objects.filter(companyid=compid).first() 
    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')
    user=Employs.objects.filter(admin=request.user.id).first()
    userid=user.id
    projectm=admin_project_create.objects.filter(admin_id = userid)
    da1 = Employs.objects.filter(admin=request.user.id).first()
    da2 = da1.id
    # admin = AdminHod.objects.get(id=1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    # admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    pr = Employs.objects.filter(admin=request.user.id).first()
    pr1 = pr.id
    projects = Project.objects.filter(o_id=pr1)
    employees = Employs.objects.all()
    tsk1_data = employperformance.objects.filter(employ_name_id__companyid=em1,employ_name_id__hroptions=0)

    for task in tsk1_data:
        try:
            project = Project.objects.get(id=task.project_name.id)
            task.project_name = project  # Assign the Project instance directly
        except Project.DoesNotExist:
            task.project_name = None  # Handle the case when the project is not found

    if request.method == "POST":
        project_id = request.POST.get('project_name')  # Assuming the value is an ID
        employ_id = request.POST.get('employ_name')
        task_name = request.POST.get('task_name')
        performance = request.POST.get('performance')

        try:
            selected_project = Project.objects.get(id=project_id)
        except Project.DoesNotExist:
            selected_project = None  # Set selected_project as None if project is not found

        try:
            selected_employee = Employs.objects.get(id=employ_id)
        except Employs.DoesNotExist:
            selected_employee = None  # Set selected_employee as None if employee is not found

        try:
            project_name = projecttask.objects.get(id=task_name)
        except projecttask.DoesNotExist:
            project_name = None

        k = employperformance(
            project_name=selected_project,
            employ_name=selected_employee,
            project_task1=36,
            performance=performance
        )
        k.save()

    return render(request, "employ-template/employ-employper.html", {
        'admin_drops':admin_drops,
        's': s,
        'tsk1_data': tsk1_data,
        'projects': projects,
        'employees': employees,
        'user':user,
        'da2':da2,
        'h':h,
        'projects_drops':projects_drops,
        'projectm':projectm,
        'employs_all':employs_all,
        'data':data,
        'a':a
  

    })

def get_employee_tasks1(request, employee_id):
    try:
        # Query the Task model to retrieve tasks associated with the employee_id
        tasks = tlassigntask.objects.filter(employid=employee_id)

        # Prepare a list of task data
        task_data = [{'task_name': task.task, } for task in tasks]

        # Return the task data as a JSON response
        return JsonResponse({'tasks': task_data})

    except Exception as e:
        return JsonResponse({'error': str(e)})

from django.db.utils import IntegrityError

from django.db.models import Q

def create_proj1(request):
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    compid=data.companyid
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1) 
    a=company_details.objects.filter(companyid=compid).first() 
    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')
    em=Employs.objects.filter(admin=request.user.id).first()
    em1=em.id
    em2=em.companyid
    pname=admin_project_create.objects.filter(admin_id=em1)
    team_lead_employee = None
    selected_team_lead_id = None
    # s=adminnav.objects.filter(parent_category=None).order_by('id')
    team_members = Employs.objects.filter(companyid=em2,hroptions=0,projectmanagerop=0)

    if request.method == 'POST':
        p_name = request.POST['p_name']
        p_desc = request.POST['p_desc']
        pr_deadline = request.POST['project_deadline_date']
        # project_manager = request.POST['manager_name']
        # status = request.POST['status']
        o_id = Employs.objects.get(admin=request.user.id)

        selected_employee_ids = request.POST.getlist('selected_employees[]')
        project = Project.objects.create(
            p_name=p_name,
            p_desc=p_desc,
            o_id=o_id,
            # project_manager=project_manager,
            project_deadline=pr_deadline,
            # status=status
            
        )

        selected_team_lead_name = request.POST.get('team_lead')
        existing_team_member_ids = TeamMember.objects.filter(project=project).values_list('employee_id', flat=True)
        available_team_members = Employs.objects.exclude(id__in=existing_team_member_ids)
        team_lead_employee = Employs.objects.get(id=selected_team_lead_name)
        
        if team_lead_employee:
            selected_team_lead_id = team_lead_employee.id
        # Check if a TeamMember entry with the same project and employee already exists
        existing_team_member = TeamMember.objects.filter(project=project, employee=team_lead_employee.id, is_team_lead=True).first()
        if not existing_team_member:
            # Create the team lead only if it doesn't exist
            team_lead_member = TeamMember(
                project=project,
                employee=team_lead_employee,
                is_team_lead=True,
            )
            team_lead_member.save()

        # Mark the selected team lead as a team lead in the database
        

        for emp_id in selected_employee_ids:
            try:
                employee = Employs.objects.get(id=emp_id)
                TeamMember.objects.create(project=project, employee=employee, is_team_lead=False)
                # available_team_members = available_team_members.exclude(id=emp_id)
            except IntegrityError:
                existing_entry = TeamMember.objects.get(project=project, employee=employee, is_team_lead=False)
                existing_entry.save()
        
        # Fetch available team members (exclude team leads)
        # team_members = Employ.objects.filter(
        #         Q(teammember__isnull=True) | Q(teammember__project__status='completed')
        #         ).distinct()

        messages.success(request, 'Project created successfully and employees assigned.')
        return redirect('/Employ_home')
    
    

    return render(request, 'employ-template/OrgCreateProject.html', {'admin_drops':admin_drops,'a':a,'data':data,'s':s,'projects_drops':projects_drops,'pname':pname,"team_members": team_members, "selected_team_lead_id": selected_team_lead_id})

from .models import Project, projecttask
from django.shortcuts import render
from django.utils import timezone

def projecttask2(request):
    pr = Employs.objects.filter(admin=request.user.id).first()
    pr1 = pr.id
    projects = Project.objects.filter(o_id=pr1)


    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = Employs.objects.filter(admin=request.user.id).first()
    da2 = da1.id
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    compid=data.companyid
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1) 
    a=company_details.objects.filter(companyid=compid).first() 
    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    # admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')


    if request.method == "POST":
        p_id = request.POST['p_name']  # Assuming 'p_name' contains the ID of the project
        l_name = request.POST['l_name']
        task_names = request.POST.getlist('task_name[]')

        # Fetch the Project instance based on the ID obtained
        project_instance = Project.objects.get(pk=p_id)

        for task_name in task_names:
            # Assign the project instance (not just the ID) to the projecttask's p_name field
            task = projecttask(p_name=project_instance, tasks=task_name, l_name=l_name)
            task.save()

    return render(request, 'employ-template/form123.html', {'projects': projects,'user':user,'da1':da1,'da2':da2,'s':s,'h':h,'projectm':projectm,'projects_drops':projects_drops,'admin_drops':admin_drops,'data':data,'employs_all':employs_all,'a':a})

def get_team_members1(request, project_id):
    team_members = TeamMember.objects.filter(project_id=project_id)
    data = [{'id': member.employee.id, 'name': f"{member.employee.first_name} {member.employee.last_name}"}
            for member in team_members]
    return JsonResponse({'team_members': data})


def get_team_lead1(request, project_id):
    team_members = TeamMember.objects.filter(project_id=project_id)
    team_lead = TeamMember.objects.filter(project_id=project_id,is_team_lead=1).first()  # Assuming you have a TeamLead model
    data = [{'id': member.employee.id, 'name': f"{member.employee.first_name} {member.employee.last_name}"}
            for member in team_members]
    
    # Add team lead information to the data
    if team_lead:
        data.append({'id': team_lead.employee.id, 'name': f"Team Lead: {team_lead.employee.first_name} {team_lead.employee.last_name}"})
    
    return JsonResponse({'team_members': data})

from django.shortcuts import render, redirect
from .models import admin_project_create, taskdata, TeamMember

def display2(request):
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    compid=data.companyid
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1) 
    a=company_details.objects.filter(companyid=compid).first()
    pr=Employs.objects.filter(admin=request.user.id).first()
    pr1=pr.id
    projects=Project.objects.filter(o_id=pr1)
    tl=TeamMember.objects.filter(is_team_lead=1)

    if request.method == 'POST':
        p_name = request.POST.get('p_name')
        p_url = request.POST.get('p_url')
        team_lead = request.POST.get('team_lead')

        # Create a new project
        p1 = taskdata(
            p_name=p_name,
            p_url=p_url,
            team_lead=team_lead
        )
        p1.save()

        return redirect('/read-proj')

    return render(request, 'employ-template/data.html', {'tl':tl,'projects': projects,'a':a,'projects_drops':projects_drops,'s':s,'data':data})
  
def performancetask1(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    compid=data.companyid
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1) 
    a=company_details.objects.filter(companyid=compid).first() 
    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = Employs.objects.filter(admin=request.user.id).first()
    da2 = da1.id
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    projects = Project.objects.all()  
    items_per_page = 2
    performance_data = employperformance.objects.filter(employ_name_id__companyid=em1)
    paginator = Paginator(performance_data, items_per_page)
    page = request.GET.get('page')

    try:
        performance_data = paginator.page(page)
    except PageNotAnInteger:
        performance_data = paginator.page(1)
    except EmptyPage:
        performance_data = paginator.page(paginator.num_pages)
    return render(request, 'employ-template/emptaskper.html', {"projects_drops":projects_drops,'projects':projects,'data':data,'performance_data': performance_data,'a':a,'user':user,'da1':da1,'da2':da2,'s':s,'h':h,'projectm':projectm,'employs_all':employs_all})
import openpyxl
from django.http import HttpResponse
from django.db.models import Avg

def download_all_performance_data_excel1(request):
    performance_data = employperformance.objects.all()

    # Create an Excel workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(["Employee ID", "Employee Name", "Project Name", "Date", "Performance"])  # Include "Date" in the header

    for performance in performance_data:
        try:
            employee = Employs.objects.get(id=performance.employ_name.id)
            project_name = performance.project_name.p_name
            worksheet.append([employee.empid, employee.first_name, project_name, performance.date, performance.performance])  # Include date in the row
        except Employs.DoesNotExist:
            pass  # Handle the case where the employee doesn't exist

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="all_performance_data.xlsx"'
    workbook.save(response)

    return response

def performanceallproject1(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    compid=data.companyid
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1) 
    a=company_details.objects.filter(companyid=compid).first()
    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = Employs.objects.filter(admin=request.user.id).first()
    da2 = da1.id
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    projects = Project.objects.all()  
    items_per_page = 2
    tsk1_data = employperformance.objects.filter(employ_name_id__companyid=em1).values('employ_name').annotate(avg_performance=Avg('performance'))
    employees_data = []
    for task in tsk1_data:
        employ_id = task['employ_name']
        try:
            employee = Employs.objects.get(id=employ_id)
            task['employee_name'] = employee.first_name
            task['employee_id'] = employee.empid
            project_names = employperformance.objects.filter(employ_name=employee).values('project_name__p_name')
            task['project_names'] = project_names

        except Employs.DoesNotExist:
            task['employee_name'] = "N/A"
            task['employee_id'] = "N/A"
            task['project_names'] = "N/A"
        employees_data.append(task)


    try:
        paginator = Paginator(employees_data, items_per_page)
        page = request.GET.get('page')
        employees_data = paginator.page(page)
    except PageNotAnInteger:
        employees_data = paginator.page(1)
    except EmptyPage:
        employees_data = paginator.page(paginator.num_pages)

    return render(request, 'employ-template/empavgper.html', {'a':a,'projects_drops':projects_drops,'projects':projects,'data':data, 'tsk1_data': employees_data,'user':user,'da1':da1,'da2':da2,'s':s,'h':h,'projectm':projectm,'data':data,'employs_all':employs_all})

import openpyxl
from openpyxl.styles import Alignment

def download_excel1(request):
    tsk1_data = employperformance.objects.values('employ_name').annotate(avg_performance=Avg('performance'))

    # Create an Excel workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(["Employee ID", "Employee Name", "Project Names", "Average Performance"])

    for task in tsk1_data:
        employ_id = task['employ_name']
        try:
            employee = Employs.objects.get(id=employ_id)
            task['employee_name'] = employee.first_name
            task['employee_id'] = employee.empid

            # Aggregate unique project names for the employee
            project_names = set(employperformance.objects.filter(employ_name=employee).values_list('project_name__p_name', flat=True))
            # Join unique project names with line breaks
            task['project_names'] = "\n".join(project_names)

        except Employs.DoesNotExist:
            task['employee_name'] = "N/A"
            task['employee_id'] = "N/A"
            task['project_names'] = "N/A"

        # Append the data to the Excel sheet
        worksheet.append([task['employee_id'], task['employee_name'], task['project_names'], task['avg_performance']])

    # Apply alignment style to the cell containing project names for line breaks
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # Create an HttpResponse with the Excel file
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="tsk1_data.xlsx"'
    workbook.save(response)

    return response

from django.shortcuts import render, redirect
from django.http import HttpResponse
from .models import admin_project_create

def projectstatus1(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    compid=data.companyid
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1) 
    a=company_details.objects.filter(companyid=compid).first()
   
    da1 = Employs.objects.filter(admin=request.user.id).first()
    da2 = da1.id
    # projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()


    if request.method == "POST":

       


        projectm = admin_project_create.objects.filter(admin_id=request.user.id)


        
        for project in projectm:
            project_id = project.id
            status = request.POST.get(f'status-{project_id}')
            
            if status in ('Complete', 'Incomplete','on going'):
                project.status = status
                project.save()

        return redirect('admin_home')  # You can replace 'admin_home' with the actual URL you want to redirect to

    # Only get the projects assigned to the manager
    projectm = admin_project_create.objects.filter(companyid=em1)

    return render(request, "employ-template/projectstatus.html", {'a':a,'projectm': projectm,'user':user,'da1':da1,'da2':da2,'s':s,'h':h,'projectm':projectm,'projects_drops':projects_drops,'data':data,'employs_all':employs_all})


import pandas as pd
import plotly.express as px
from plotly.offline import plot
from datetime import timedelta
import json
from.models import AdminHod

def projectwise_task(request, pid):
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1)
    o_id = Employs.objects.get(admin=request.user.id)
    team_members = TeamMember.objects.filter(project=pid)
    team_member_scores = [(team_member, team_member.calculate_total_score_in_project()) for team_member in team_members]
    team_member_scores = sorted(team_member_scores, key=lambda x: x[1], reverse=True)
    project_details=Project.objects.filter(id=pid)
    task_details = Task.objects.filter(o_id_id=o_id, id=pid).all()
    # s = adminnav.objects.filter(parent_category=None).order_by('id')

    if request.method == 'GET':
        t_status = request.GET.get('t_status', 'total')  # Default to 'total' if not provided

        if t_status == 'total':
            task_details = Task.objects.filter(o_id_id=o_id, p_id=pid).all()
        elif t_status == 'completed':
            task_details = Task.objects.filter(o_id_id=o_id, p_id=pid, t_status='completed')
        elif t_status == 'in-progress':
            # task_details = Task.objects.filter(o_id_id=o_id, p_id=pid, t_status='in-progress').all()
            task_details = Task.objects.filter(o_id_id=o_id, p_id=pid).exclude(t_status='completed')

        

        tasks = task_details
        count_no_of_total_tasks = Task.objects.filter(o_id_id=o_id, p_id_id=pid).count()
        count_no_of_completed_tasks = Task.objects.filter(o_id_id=o_id, p_id_id=pid, t_status="completed").count()
        count_no_of_pending_tasks = count_no_of_total_tasks - count_no_of_completed_tasks
        task_data = []

        for task in tasks:
            now = timezone.now()
            task_score = task.calculate_task_score()
            remaining_time = task.calculate_remaining_time()
            progress_percentage = min(task_score / 10 * 100, 100)
            excess_points = max(task_score - 10, 0)  # Calculate the excess points (if any)
            neg_points = max(-task_score, 0)

            task_data.append({
                'task': task,
                'progress_percentage': progress_percentage,
                'excess_points': excess_points,
                'neg_points': neg_points,
                'task_score': task_score,
                'remaining_time': remaining_time,
            })
            
        
        context = {
        
        "project_details": project_details,
        'task_total': count_no_of_total_tasks,
        'task_completed': count_no_of_completed_tasks,
        'task_pending': count_no_of_pending_tasks,
        'team_members': team_members,
        's': s,
        't_status':t_status,
        "task_details": task_details,
        'task_data': task_data,
        'data':data,
        'team_member_scores': team_member_scores,

        }
    
    return render(request, 'employ-template/ViewProjectwiseTasks1.html', context)
from django.shortcuts import render, redirect
from django.shortcuts import get_object_or_404
from django.contrib import messages
from django.db.models import Q
from .models import Project, Employs, TeamMember, admin_drop  # Make sure to import your models

def update_data1(request, pid):
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    # Get the project to be updated
    project = get_object_or_404(Project, id=pid)
    team_leader_info = project.get_team_leader()

    # Get available team members excluding team leads and those already selected for the project
    team_members = TeamMember.objects.filter(project=project)
    selected_team_member_ids = team_members.values_list('employee_id', flat=True)
    available_team_members = Employs.objects.filter(companyid=em1,hroptions=0,projectmanagerop=0)
       

    if request.method == 'POST':
        # Retrieve updated project details from the form
        p_name = request.POST['p_name']
        p_desc = request.POST['p_desc']
        pr_deadline = request.POST['project_deadline_date']
        project_manager = request.POST['manager_name']
        status = request.POST.get('status')

        # Update the project
        project.p_name = p_name
        project.p_desc = p_desc
        project.project_deadline = pr_deadline
        project.project_manager = project_manager
        project.status = status
        project.save()

        # Update team lead if changed
        selected_team_lead_id = request.POST.get('team_lead')
        if selected_team_lead_id:
            try:
                team_lead = Employs.objects.get(id=selected_team_lead_id)
                project.teammember_set.update(is_team_lead=False)  # Remove team leader status from existing team leader

                # Create a new TeamMember if it doesn't exist
                team_member, created = TeamMember.objects.get_or_create(project=project, employee=team_lead)
                team_member.is_team_lead = True
                team_member.save()
            except Employs.DoesNotExist:
                # Handle the case where the selected team leader doesn't exist
                messages.error(request, 'Selected team leader does not exist.')
                return redirect('/admin_home')

        # Update selected team members
        selected_employee_ids = request.POST.getlist('selected_employees[]')
        existing_team_member_ids = TeamMember.objects.filter(project=project).values_list('employee_id', flat=True)
        for emp_id in selected_employee_ids:
            emp_id = int(emp_id)
            if emp_id not in existing_team_member_ids:
                employee = get_object_or_404(Employs, id=emp_id)
                TeamMember.objects.create(project=project, employee=employee, is_team_lead=False)

        messages.success(request, 'Project updated successfully and employees assigned.')
        return redirect('/admin_home')

    # Render the template with the updated available team members
    return render(request, 'employ-template/editproject.html', {
        "admin_drops": admin_drops,
        "project": project,
        "team_members": team_members,
        "available_team_members": available_team_members,
        'team_leader_info': team_leader_info,
    })

  
def delete_user1(request,employee_id , project_id):
        # Assuming your model has a field 'id' representing the user's ID
        employee = TeamMember.objects.get(employee_id=employee_id,project_id=project_id)
        employee.delete()
        referer = request.META.get('HTTP_REFERER')
        if referer:
          return redirect(referer)
        else:
          return redirect(reverse('home'))



############# project views end ##################
def employ_people_count(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1) 
    
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')
    employs = Employs.objects.filter(companyid=em1)
    hr=HR.objects.all()
    emp = CustomUser.objects.filter(user_type=2,is_active=True,employs__companyid=em1).count()  # Count active employees
    dis = CustomUser.objects.filter(user_type=2,is_active=False,employs__companyid=em1).count()  # Count inactive employees
    total_employees_count = emp + dis 
    
    paginator = Paginator(employs, per_page=10)  # Display 2 employees per page (you can adjust per_page as needed)

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, "employ-template/employpeoples.html", {
    
        'page_obj': page_obj,
        'admin_drops': admin_drops,
        'employs': employs,
        'emp': emp,
        'dis': dis,
        'total_employees_count': total_employees_count,
        'admin_drops':admin_drops,
        'hr':hr,
        's':s,
        'data':data,
        'admin_drops':admin_drops,
        
    })

def edit_people(request,std):
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1)


    request.session['employ_id']=std
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')
    # li=list.objects.all()
    employ=Employs.objects.get(admin=std)
    objss=employ.id 
    users=CustomUser.objects.get(id=std)
    datas=employ_add_form.objects.filter(student_id=objss)
    return render(request,"employ-template/people_employ.html",{'data':data,'admin_drops':admin_drops,'employ':employ,'admin_drops':admin_drops,'datas':datas,'objss':objss,'users':users,'s':s})

def enable_button(request, std):

    request.session['employ_id']=std
    instance = CustomUser.objects.get(id=std)
    instance.is_active = True
    instance.save()
    return redirect('edit_people', std=std)  # Redirect to the detail view

def disable_button(request, std):
    request.session['employ_id']=std
    instance = CustomUser.objects.get(id=std)
    instance.is_active = False
    instance.save()
    return redirect('edit_people',std=std)

def individual_tax_report(request,std):
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')


    data=Employs.objects.filter(admin=request.user.id).first()
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)

    
    request.session['employ_id']=std
    employ=Employs.objects.all()
    tax=employ_tax_form.objects.filter(employ=std)
    
    leave_date=LeaveReportEmploy.objects.filter()
    return render(request,"employ-template/individual_taxReport.html",{'employ':employ,'admin_drops':admin_drops,'tax':tax,'leave_date':leave_date,'s':s,'data':data})

def individual_employ_documentreport(request,std):
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')
    data=Employs.objects.filter(admin=request.user.id).first()
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)

    request.session['employ_id']=std
    employ_id=Employs.objects.all()
    doc=empdocs.objects.filter(employ_id=std)
 
    return render (request,"employ-template/individual_employ_docreport.html",{'doc':doc,'employ_id':employ_id,'admin_drops':admin_drops,'s':s,'data':data})

# def individual_employ_attendance(request,std):
#     admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
#     request.session['employ_id']=std
#     employ_id=Employs.objects.all()
#     attend=checkin.objects.filter(employ_id=std)
 
#     return render (request,"admin-template/individual_attendance_report.html",{'attend':attend,'employ_id':employ_id,'admin_drops':admin_drops})


def individual_employ_reimbursement_view(request,std):
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')

    data=Employs.objects.filter(admin=request.user.id).first()
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)

    
    request.session['employ_id']=std
    # s=adminnav.objects.all()
    employ_id=Employs.objects.all()
    leaves=Reimbursement.objects.filter(employ_id=std)
    return render(request,"employ-template/individual_employ_reimbursement.html",{"leaves":leaves,'s':s,"employ_id":employ_id,'admin_drops':admin_drops,'data':data})

def reimbursement_employ_approve_status(request,leave_id):
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')
    leave=Reimbursement.objects.get(id=leave_id)
    leave.reimbursement_status=1
    leave.save()
    user=CustomUser.objects.all()
    return HttpResponseRedirect(reverse("individual_reimbursement_view", args=[leave.employ_id_id]))

def reimbursement_employ_disapprove_status(request,leave_id):
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')
    leave=Reimbursement.objects.get(id=leave_id)
    leave.reimbursement_status=2
    leave.save()
    return HttpResponseRedirect(reverse("individual_reimbursement_view", args=[leave.employ_id_id]))

def individual_employ_leaveReport(request,std):
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')


    data=Employs.objects.filter(admin=request.user.id).first()
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)


    request.session['employ_id']=std
    employ_id=Employs.objects.all()
    leaves=LeaveReportEmploy.objects.filter(employ_id=std)
    return render(request,"employ-template/individual_leave_view.html",{"leaves":leaves,'admin_drops':admin_drops,'employ_id':employ_id,'data':data,'s':s})

def approve_leave(request,leave_id):
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')
    leave=LeaveReportEmploy.objects.get(id=leave_id)
    leave.leave_status=1
    leave.save()
    return HttpResponseRedirect(reverse("individual_leaveReport", args=[leave.employ_id]))
from datetime import date

def disapprove_leave(request,leave_id):
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')
    leave=LeaveReportEmploy.objects.get(id=leave_id)
    leave.leave_status=2
    leave.save()
    return HttpResponseRedirect(reverse("individual_leaveReport", args=[leave.employ_id_id]))
def data_tables1(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)

    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')
    today = date.today()
    tsk1_data = tlassigntask.objects.filter(task_date=today)  
    team_leads = TeamMember.objects.filter(is_team_lead=True,employee_id__companyid=em1)

    for task in tsk1_data:
        try:
            employee = Employs.objects.get(id=task.employid)
            task.employid = employee.first_name
        except Employs.DoesNotExist:
            task.employid = "N/A"
        
        try:
            project = Project.objects.get(id=task.project_id)
            task.project_name = project.p_name
        except Project.DoesNotExist:
            task.project_name = "N/A"
            
    items_per_page = 10  # Adjust the number of items per page as needed

    paginator = Paginator(tsk1_data, items_per_page)
    page = request.GET.get('page')

    try:
        tsk1_data = paginator.page(page)
    except PageNotAnInteger:
        tsk1_data = paginator.page(1)
    except EmptyPage:
        tsk1_data = paginator.page(paginator.num_pages)

    
    return render(request, 'employ-template/emptables.html', {'projects_drops':projects_drops,"data":data,'admin_drops':admin_drops,'s':s,'data':data,'tsk1_data': tsk1_data, 'team_leads': team_leads})
def individual_employ_advancesalary_Report(request,std):
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')


    data=Employs.objects.filter(admin=request.user.id).first()
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)

    
    request.session['employ_id']=std
    employ_id=Employs.objects.all()
    leaves=ad_salary.objects.filter(employ_id=std)
    return render(request,"employ-template/individual_employ_advancesalary_Report.html",{"leaves":leaves,'admin_drops':admin_drops,'employ_id':employ_id,'s':s,'data':data})

def approve_advancesalary(request,leave_id):
    # admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    leave=ad_salary.objects.get(id=leave_id)
    leave.leave_status=1
    leave.save()
    return HttpResponseRedirect(reverse("individual_advancesalary_Report", args=[leave.employ_id]))

def disapprove_advancesalary(request,leave_id):
    # admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    leave=ad_salary.objects.get(id=leave_id)
    leave.leave_status=2
    leave.save()
    return HttpResponseRedirect(reverse("individual_advancesalary_Report", args=[leave.employ_id_id]))


def employee_form(request):
    li=list.objects.all()
    if request.method == "POST":
        first_name = request.POST["first_name"]
        last_name = request.POST["last_name"]
        username = request.POST["username"]
        email = request.POST["email"]
        web_mail = request.POST["web_mail"]
        password = request.POST["password"]
        address = request.POST["address"]
        empid = request.POST["empid"]
        Manager = request.POST["Manager"]
        designation = request.POST["designation"]
        location = request.POST["location"]
        package = request.POST["package"]
        pincode = request.POST["pincode"]
        contactno = request.POST["contactno"]
        bloodgroup = request.POST["bloodgroup"]
        dateofjoining = request.POST['dateofjoining']
        sex = request.POST["gender"]
        role = request.POST.get('role') # Added field for role
        
        user = CustomUser.objects.create_user(username=username, password=password, email=email, user_type=2)
        user.employs.first_name = first_name
        user.employs.last_name = last_name
        user.employs.email = email
        user.employs.password = password
        user.employs.address = address
        user.employs.empid = empid
        user.employs.web_mail = web_mail
        user.employs.Manager = Manager
        user.employs.designation = designation
        user.employs.location = location
        user.employs.package = package
        user.employs.pincode = pincode
        user.employs.contactno = contactno
        user.employs.bloodgroup = bloodgroup
        user.employs.dateofjoining = dateofjoining
        user.employs.role = role  # Set the role field for HR

        user.save()
                
       
        html_content = render_to_string("email_template.html", {'title': 'test email', 'first_name': first_name,
                                                                'last_name': last_name, 'empid': empid})
        text_content = strip_tags(html_content)
        subject = "WELCOME TO DEVELOPTRESS"
        email = EmailMultiAlternatives(
            subject,
            text_content,
            settings.EMAIL_HOST_USER,
            [email],
        )
        email.attach_alternative(html_content, "text/html")
        email.fail_silently = True
        email.send()
        
        if user:
            messages.success(request, "Successfully Added employee")
            return redirect("/employee_form")
        else:
            messages.error(request,"Failed to add employ")
    # If the request method is not POST, render the form
    s = adminnav.objects.all()
    form = Employs.objects.all()
    hrform = HR.objects.all()
    role = list.objects.all()
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')
    return render(request, "employ-template/employ_form_template.html",
                  {'role': role, 'hrform': hrform, "form": form, 's': s, 'admin_drops': admin_drops,'li':li})

import pandas as pd

def employdetails_upload_excel(request,message=None, error_message=None):
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1)


    if request.method == 'POST':
        excel_file = request.FILES['excel_file']

        if excel_file.name.endswith('.xlsx'):
            try:
                df = pd.read_excel(excel_file)

                # Assuming the columns in the Excel file match the fields in Employs model
                for index, row in df.iterrows():
                    user=CustomUser.objects.create_user(username=row['username'],password=row['empid'],first_name=row['first_name'],last_name=row['last_name'],email=row['email'],user_type=2)
                    user.employs.first_name=row['first_name']
                    user.employs.last_name=row['last_name']
                    user.employs.email=row['email']
                    user.employs.address=row['address']
                    user.employs.empid=row['empid']
                    user.employs.role=row['role']
                    user.employs.location=row['location']
                    user.employs.package=row['package']
                    user.employs.pincode=row['pincode']
                    user.employs.contactno=row['contactno']
                    user.employs.gender=row['sex']
                    date_string = str(row['dateofjoining'])
                    date_obj = datetime.strptime(date_string, '%Y-%m-%d %H:%M:%S')  # Adjust the format as per your Excel file
                    user.employs.dateofjoining = date_obj.date()
                    user.save()
                    html_content = render_to_string("email_template.html",{'title':'test email','first_name':row['first_name'],'last_name':row['last_name'],'empid':row['empid']})
                    text_content=strip_tags(html_content)
                    subject="WELCOME TO DEVELOPTRESS"
                    email = EmailMultiAlternatives(
                            subject,
                            text_content,
                            settings.EMAIL_HOST_USER,
                            [row['email']],
                            )
                    email.attach_alternative(html_content,"text/html")
                    email.fail_silently = True
                    email.send()
                    
                message = f"{len(df)} records imported successfully."  
                return render(request, 'employ-template/employdetails_upload_bulk_data.html',{'message':message})
            except Exception as e:
                # Handle any exceptions that may occur during processing
                # return HttpResponseRedirect(reverse("upload_excel"))
                error_message = f"{str(e)} An error occurred during processing."
        else:
            # return HttpResponseRedirect(reverse("upload_excel"),{'messagea': 'Please upload a valid Excel file.'}) 
            error_message = 'Please upload a valid Excel file.'
            
    return render(request, 'employ-template/employdetails_upload_bulk_data.html',{'error_message':error_message,'s':s,'admin_drops':admin_drops,'data':data})

# def edit_role(request):
#     rl=Employs.objects.all()

# from .models import halfldayvreason    
# def employ_apply_leave(request):
#     staff_obj = Employs.objects.get(admin=request.user.id)
#     leave_data=LeaveReportEmploy.objects.filter(employ_id=staff_obj)
#     half=halfldayvreason.objects.first()
#     return render(request,"employ-template/employ_apply_leave.html",{"leave_data":leave_data,'half':half})

def employ_apply_leave_save(request):
    if request.method!="POST":
        return HttpResponseRedirect(reverse("employ_apply_leave"))
    else:
        leave_date=request.POST.get("leave_date")
        leave_type=request.POST.get("leave_type")
        to_date=request.POST.get("to_date")
        leave_msg=request.POST.get("leave_msg")
        email_leave=request.POST.get("email_leave")

        employ_obj=Employs.objects.get(admin=request.user.id)
        try:
            leave_report=LeaveReportEmploy(email_leave=email_leave,employ_id=employ_obj,leave_type=leave_type,leave_date=leave_date,to_date=to_date,leave_message=leave_msg,leave_status=0)
            leave_report.save()
            messages.success(request, "Successfully Applied for Leave")
            return HttpResponseRedirect(reverse("employ_apply_leave"))
        except:
            messages.error(request, "Failed To Apply for Leave")
            return HttpResponseRedirect(reverse("employ_apply_leave"))

def reimbursement_apply_view(request):
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')

    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')


    staff_obj = Employs.objects.get(admin=request.user.id)
    leave_data=Reimbursement.objects.filter(employ_id=staff_obj)
  
    k3=reimbursementsetup1.objects.all()
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    compid=data.companyid
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1)  
   
    a=company_details.objects.filter(companyid=compid).first()


    return render(request,"employ-template/reimbursement_apply_view.html",{'a':a,"data":data,"projects_drops": projects_drops,"data":data,"admin_drops":admin_drops,"leave_data":leave_data,'s':s,'k3':k3})

def reimbursement_apply_view_save(request):
    if request.method!="POST":
        return HttpResponseRedirect(reverse("reimbursement_apply_view"))
    else:
        typea=request.POST.get("typea")
        date=request.POST.get("date")
        detail=request.POST.get("detail")
        amount=request.POST.get("amount")
        image=request.POST.get("image")

        student_obj=Employs.objects.get(admin=request.user.id)
        try:
            leave_report=Reimbursement(employ_id=student_obj,typea=typea,date=date,detail=detail,amount=amount,image=image,reimbursement_status=0)
            leave_report.save()
            messages.success(request, "Successfully Applied for Reimbursement")
            return HttpResponseRedirect(reverse("reimbursement_apply_view"))
        except:
            messages.error(request, "Failed To Apply for Reimbursement")
            return HttpResponseRedirect(reverse("reimbursement_apply_view"))
from django.db.models import Sum
from .models import Employs, Reimbursement, employnav, types

def reg(request):
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    compid=data.companyid
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
        projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    else:
        projects_drops=None
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1)
    a=company_details.objects.filter(companyid=compid).first()

   
    staff_obj = Employs.objects.get(admin=request.user.id)
 
    k = types.objects.all()

    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')




    st4 = request.POST.get("ss")
    st3 = request.POST.get("vk")
    st = request.POST.get("d1")
    st1 = request.POST.get("d2")

    leave_data = Reimbursement.objects.filter(employ_id=staff_obj)

    # Apply filters based on user input
    if st4 and st4 != '----Select----':  # Check if a valid status is selected
        leave_data = leave_data.filter(reimbursement_status=st4)
    if st3:  # If "Select Type" is selected
        leave_data = leave_data.filter(typea__icontains=st3)
    if st and st1:
        leave_data = leave_data.filter(date__range=[st, st1])

    total_approved = leave_data.filter(reimbursement_status=1).aggregate(Sum('amount'))['amount__sum'] or 0
    total_pending = leave_data.filter(reimbursement_status=0).aggregate(Sum('amount'))['amount__sum'] or 0
    items_per_page = 10  # Adjust the number of items per page as needed

    paginator = Paginator(leave_data, items_per_page)
    page = request.GET.get('page')

    try:
        leave_data = paginator.page(page)
    except PageNotAnInteger:
        leave_data = paginator.page(1)
    except EmptyPage:
        leave_data = paginator.page(paginator.num_pages)


    return render(request, "employ-template/a.html", {
        'leave_data': leave_data,
        's': s,
        'k': k,
        'total': total_approved,
        'total1': total_pending,
        'data':data,
        'admin_drops':admin_drops,
        'projects_drops':projects_drops,
        'a':a
                })



    

def type(request):
    staff_obj = Employs.objects.get(admin=request.user.id)
    leave_data=Reimbursement.objects.filter(employ_id=staff_obj) 
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0).all()
    k=types.objects.all()
    

    if request.method=="POST":
        
        st3=request.POST.get("vk")
      
        if st3!=None:
            leave_data=Reimbursement.objects.filter(typea_icontains=st3).values() 

    return render(request,"employ-template/a.html",{'leave_data':leave_data,'s':s,'k':k})


def delete(request,id):
    k=Reimbursement.objects.get(id=id)
    k.delete()
    return redirect("/reg")

def edit(request,id):
    ks=Reimbursement.objects.get(id=id)
    k=types.objects.all()
    return render(request,"employ-template/update.html",{'ks':ks,'k':k})
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage

def update(request,id):
    if request.method=="POST":
        
        date=request.POST["date"]
        detail=request.POST["detail"]
        typea=request.POST["typea"]
        amount=request.POST["amount"]
        image=request.FILES.get('image')
        k=Reimbursement.objects.get(id=id);
        
        k.date=date
        k.detail=detail
        k.typea=typea
        k.amount=amount
        k.image=image
        k.save();
        return redirect("/reg")
    return render(request,"employ-template/update.html")

def documents_uploaded_view(request):
    staff_obj = Employs.objects.get(admin=request.user.id)
    leave_data=empdocs.objects.filter(employ_id=staff_obj)
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)   
    r=documents_setup.objects.all()
    return render(request,"employ-template/documents_uploaded_view.html",{"leave_data":leave_data,'s':s,'r':r})


def trails(request):
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    compid=data.companyid
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1) 
    a=company_details.objects.filter(companyid=compid).first() 
    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')
    staff_obj = Employs.objects.get(admin=request.user.id)
    leave_data=empdocs.objects.filter(employ_id=staff_obj)
    r=documents_setup1.objects.all()
    
   
    items_per_page = 10  # Adjust the number of items per page as needed

    paginator = Paginator( leave_data, items_per_page)
    page = request.GET.get('page')

    try:
        leave_data = paginator.page(page)
    except PageNotAnInteger:
        leave_data = paginator.page(1)
    except EmptyPage:
        leave_data = paginator.page(paginator.num_pages)  
    if request.method=="POST":
        documenttype1=request.POST["documenttype1"]
        imagefile=request.FILES.get('imagefile')
        description=request.POST["description"]
        employ_obj=Employs.objects.get(admin=request.user.id)
        k=empdocs(employ_id=employ_obj,documenttype1=documenttype1,imagefile=imagefile,description=description)
        k.save()
        return redirect("/trails")
    return render(request,"employ-template/documents_uploaded_view.html",{"a":a,"data":data,'projects_drops':projects_drops,'data':data,'admin_drops':admin_drops,'leave_data':leave_data,'s':s,'r':r})
from datetime import datetime, time
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

def task_data(request):
   
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    compid=data.companyid
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1) 
    a=company_details.objects.filter(companyid=compid).first() 
    em1=data.companyid
    employ_obj=Employs.objects.get(admin=request.user.id)

    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')
    today = date.today()
    tsk1_data = tlassigntask.objects.filter(task_date=today)  

   
    employ_obj = Employs.objects.get(admin=request.user.id)
   

    # Query tasks assigned to the employee
    tasks = projecttask.objects.filter(p_name__o_id__companyid=em1)

    # Calculate status for each task based on performance
    tasks_with_status = []
    for task in tasks:
        performance_data = employperformance.objects.filter(
            project_task1=task,
            employ_name=data1,
        ).first()
        if performance_data:
            performance_value = performance_data.performance
            if performance_value < 30:
                status = "Work Started"
            elif performance_value < 50:
                status = "Partial Completion"
            elif performance_value < 70:
                status = "In Progress"
            elif performance_value < 90:
                status = "In Progress"
            else:
                status = "Completed"
            tasks_with_status.append({'task': task, 'status': status})
        else:
            tasks_with_status.append({'task': task, 'status': "Not Started"})

    # Pagination
    items_per_page = 10
    paginator = Paginator(tasks_with_status, items_per_page)
    page = request.GET.get('page', 1)

    try:
        tasks_with_status = paginator.page(page)
    except PageNotAnInteger:
        tasks_with_status = paginator.page(1)
    except EmptyPage:
        tasks_with_status = paginator.page(paginator.num_pages)

    return render(request, "employ-template/task.html", {
        'tasks_with_status': tasks_with_status,
        'employ_obj': employ_obj,
        's': s,
         'admin_drops':admin_drops,
         'data':data,
         'data1':data1,
         'data2':data2,
        'projects_drops':projects_drops,
        'a':a
        
    })

def upload_photo(request):
    user_obj = Employs.objects.get(admin=request.user.id)
    
    try:
            data = employ_add_form.objects.get(student_id=user_obj)
    except employ_add_form.DoesNotExist:
        return redirect('employ_add')

    

    if request.method == "POST":
        profile_pic = request.FILES.get("profile_pic")
        if profile_pic:
           
            data.profile_pic = profile_pic
            
                # Create a new instance if it doesn't exist
            data,_ = employ_add_form.objects.get_or_create(student_id=user_obj)
            data.profile_pic=profile_pic

            data.save()
            messages.success(request, "Profile picture uploaded successfully.")
        else:
            messages.error(request, "Profile picture upload failed.")

    return render(request, "employ-template/employprofilepic.html", {'data': data})

def delete_profile_pic(request):
    user_obj = Employs.objects.get(admin=request.user.id)
    try:
        data = employ_add_form.objects.get(student_id=user_obj.id)
        data.profile_pic.delete()  # Delete the profile picture file
        data.profile_pic = None  # Clear the profilepic field
        data.save()
        messages.success(request, "Profile picture removed successfully.")
    except employ_add_form.DoesNotExist:
        messages.error(request, "User data not found. Please add user data first.")
    
    return redirect("upload_photo")


from .models import task1
def employ_profile(request):
    half=task1.objects.first()
    user=CustomUser.objects.get(id=request.user.id)
    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
        projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    else:
        projects_drops=None
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1) 



    employ=Employs.objects.get(admin=user)
    profile=Employs.objects.all()
    email2=request.session.get('email_1');
    q=employ_add_form.objects.filter(email2=email2).values();
    datas=employ_add_form.objects.filter(student_id=employ)
    return render(request,"employ-template/employ_profile.html",{'data':data,"profile":profile,"half":half,"datas":datas,"user":user,"employ":employ,'s':s,"q":q,'data':data,'admin_drops':admin_drops,'projects_drops':projects_drops})



import cv2

# Check OpenCV version



# screenshot_project/screenshot_app/views.py

from django.shortcuts import render
from django.views import View
import time
import pyautogui as pg
import os
from django.http import  JsonResponse
from .models import Screenshots , SystemStatus
import threading
import subprocess
import json
from django.views.decorators.csrf import csrf_exempt


class ScreenshotView(View):
    template_name = 'employ-template/calendar.html'
    is_capturing = False
    capture_thread = None

    def get(self, request, *args, **kwargs):
        screenshotform = ScreenshotsForm()
        return render(request, self.template_name, {'screenshotform': screenshotform})

    def post(self, request, *args, **kwargs):
        screenshotform = ScreenshotsForm(request.POST, request.FILES)
        if screenshotform.is_valid():
            employee_id =request.user.employs.id
            screenshot = screenshotform.save(commit=False)
            screenshot.employee_id = employee_id  # Assign the employee field
            print(f"Employee ID: {screenshot.employee_id}")
            screenshot.image = self.capture_screenshot(employee_id)
            screenshot.save()                
            return JsonResponse({'message': 'Image inserted in the database'})
        else:
            print(screenshotform.errors)
            

        return JsonResponse({'error': 'Form is not valid'})

    def capture_screenshot(self, employee_id):
        random = int(time.time())
        filename = f"screenshot_{random}.png"
        full_path = os.path.join(r"C:\Users\91944\AppData\Local\Programs\Python\Python310\Scripts\Developtreeshrms(12)\media\screenshots", filename)
        directory = os.path.dirname(full_path)
        os.makedirs(directory, exist_ok=True)
        # Create the directory if it doesn't exist
        screenshot = Screenshots(image=filename , employee_id=employee_id)
        screenshot.save()
        pg.screenshot(full_path)
        return full_path

    @classmethod
    def start_capture(cls, request):
        cls.is_capturing = True

        # Check if the capture thread is not already running
        if not cls.capture_thread or not cls.capture_thread.is_alive():
            # Create a thread for capturing screenshots at regular intervals
            employee_id = request.user.employs.id
            cls.capture_thread = threading.Thread(target=cls.capture_screenshots_thread, args=(employee_id,))
            cls.capture_thread.start()

        return JsonResponse({'message': 'Capture started'})

    @classmethod
    def stop_capture(cls, request):
        cls.is_capturing = False

        # Wait for the capture thread to finish
        if cls.capture_thread and cls.capture_thread.is_alive():
            cls.capture_thread.join()

        return JsonResponse({'message': 'Capture stopped'})

    @classmethod
    def capture_screenshots_thread(cls, employee_id):
        while cls.is_capturing:
            screenshot = cls().capture_screenshot(employee_id)  # Instantiate the class to call the method
            time.sleep(5)  # Wait for 5 seconds between captures




# Shared variable to track the monitoring status
monitoring_status = False
init_app_process = None  # Store the subprocess object for initApp.py

@csrf_exempt
def update_status(request, status_type, status_message):
    # Save the status information to the database
    employee_id = request.user.employs.id
    SystemStatus.objects.create(status_type=status_type, status_message=status_message , employee_id=employee_id)

    return JsonResponse({'status': 'success'})

def powercheck(request):
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'start':
            return start_activity_tracking(request)
        elif action == 'stop':
            return stop_and_save_activities(request)
        elif action == 'update_status':
            status_type = request.POST.get('status_type')
            status_message = request.POST.get('status_message')
            return update_status(request, status_type, status_message)

    return render(request, "employ-template/calendar.html")

def start_activity_tracking(request):
    global monitoring_status, init_app_process
    employee_id = request.user.employs.id
    json_data = {
         "employee_id": employee_id,
    }
    json_string = json.dumps(json_data)

    # Stop the existing initApp.py subprocess if it's running
    if init_app_process and init_app_process.poll() is None:
        init_app_process.terminate()

    # Start a new initApp.py subprocess
    init_app_process = subprocess.Popen(['python', 'engine/initApp.py', json_string])

    monitoring_status = True
    return JsonResponse({'status': 'success', 'message': 'Activities started.'})

def stop_and_save_activities(request):
    global monitoring_status, init_app_process
    employee_id = request.user.employs.id


    json_data = {
         "employee_id": employee_id,
    }
    json_string = json.dumps(json_data)

    # Stop the initApp.py subprocess if it's running
    if init_app_process and init_app_process.poll() is None:
        init_app_process.terminate()
        init_app_process.wait()

    # Start the destroyApp.py subprocess
    subprocess.run(['python', 'engine/destroyApp.py', json_string])
    
    monitoring_status = False
    return JsonResponse({'status': 'success', 'message': 'Activities stopped.'})





















def employ_profile_save(request):
    if request.method!="POST":
        return HttpResponseRedirect(reverse("employ_profile"))
    else:
        first_name=request.POST.get("first_name")
        last_name=request.POST.get("last_name")
        password=request.POST.get("password")
        location=request.POST.get("location")

        
        try:
            customuser=CustomUser.objects.get(id=request.user.id)
            customuser.first_name=first_name
            customuser.last_name=last_name
            
            if password!=None and password!="":
                customuser.set_password(password)
            customuser.save()

            employ=Employs.objects.get(admin=customuser)
            employ.first_name=first_name
            employ.last_name=last_name
            employ.location=location
            employ
            employ.save()
            messages.success(request, "Successfully Updated Profile")
            return HttpResponseRedirect(reverse("employ_profile"))
        except:
            messages.error(request, "Failed to Update Profile")
            return HttpResponseRedirect(reverse("employ_profile"))


def employ_add(request):
    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
        projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    else:
        projects_drops=None
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    if request.method=="POST":
        firstname1=request.POST["firstname1"]
        lastname1=request.POST["lastname1"]
        
        pan=request.POST["pan"]
        ifsecode=request.POST["ifsecode"]
        acno=request.POST["acno"]
        beneficiaryname=request.POST["beneficiaryname"]
        phno=request.POST["phno"]  
        gender1=request.POST.get('gender1')
        dob=request.POST["dob"]
        address1=request.POST["address1"]
        heq=request.POST["heq"]
        aadharno=request.POST["aadharno"] 
        bloodgroup=request.POST["bloodgroup"] 
        # fathername=request.POST.get("fathername")
        # fathersdob=request.POST.get("fathersdob")
        # mothername=request.POST.get("mothername")
        # mothersdob=request.POST.get("mothersdob")
        # Childdetails1=request.POST.get("Childdetails1")
        # Childdetails2=request.POST.get("Childdetails2")
        # maritalstatus=request.POST.get("maritalstatus")
        # workexperiance=request.POST.get("workexperiance")
        # previousemploye=request.POST.get("previousemploye")
        # previousdesignation=request.POST.get("previousdesignation")
        # Marriageannivarsary=request.POST.get("Marriageannivarsary")
        # emergencycontactname=request.POST.get("emergencycontactname")
        # emergencycontactnumber=request.POST.get("emergencycontactnumber")
        # emergencycontactrelation=request.POST.get("emergencycontactrelation")
        # nationality=request.POST.get('nationality')
        # qualification=request.POST.get("qualification")
        profile_pic=request.FILES.get('profile_pic')
        student_obj=Employs.objects.get(admin=request.user.id)
        try:
            emp_instance=employ_add_form.objects.get(student_id=student_obj)
            emp_instance.firstname1=firstname1
            emp_instance.lastname1=lastname1
            emp_instance.pan=pan
            emp_instance.ifsecode=ifsecode
            emp_instance.acno=acno
            emp_instance.beneficiaryname=beneficiaryname
            emp_instance.phno=phno
            emp_instance.gender1=gender1
            emp_instance.dob=dob
            emp_instance.address1=address1
            emp_instance.heq=heq
            emp_instance.aadharno=aadharno
            emp_instance.bloodgroup=bloodgroup
            # emp_instance.fathername=fathername
            # emp_instance.fathersdob=fathersdob
            # emp_instance.mothername=mothername
            # emp_instance.mothersdob=mothersdob
            # emp_instance.Childdetails1=Childdetails1
            # emp_instance.Childdetails2=Childdetails2
            # emp_instance.maritalstatus=maritalstatus
            # emp_instance.workexperiance=workexperiance
            # emp_instance.previousdesignation=previousdesignation
            # emp_instance.Marriageannivarsary=Marriageannivarsary
            # emp_instance.emergencycontactname=emergencycontactname
            # emp_instance.emergencycontactnumber=emergencycontactnumber
            # emp_instance.emergencycontactrelation=emergencycontactrelation
            # emp_instance.nationality=nationality
            # emp_instance.qualification=qualification
            emp_instance.profile_pic=profile_pic
            # emp_instance.previousemploye=previousemploye

            emp_instance.save()
        except employ_add_form.DoesNotExist:
            k=employ_add_form(profile_pic=profile_pic,student_id=student_obj,firstname1=firstname1,lastname1=lastname1,email2=student_obj.email,pan=pan,ifsecode=ifsecode,acno=acno,beneficiaryname=beneficiaryname,phno=phno,gender1=gender1,dob=dob,address1=address1,heq=heq,aadharno=aadharno,bloodgroup=bloodgroup)
            k.save();
   
        return HttpResponseRedirect(reverse("employ_profile"))
    return render(request,"employ-template/emp1form.html",{'s':s,'admin_drops':admin_drops,'data':data})


# def employ_add(request):
#     s=employnav.objects.all()
#     if request.method=="POST":
#         firstname1=request.POST["firstname1"]
#         lastname1=request.POST["lastname1"]
        
#         pan=request.POST["pan"]
#         ifsecode=request.POST["ifsecode"]
#         acno=request.POST["acno"]
#         beneficiaryname=request.POST["beneficiaryname"]
#         phno=request.POST["phno"]  
#         gender1=request.POST['gender1']
#         dob=request.POST["dob"]
#         address1=request.POST["address1"]
#         heq=request.POST["heq"]
#         aadharno=request.POST["aadharno"] 
#         bloodgroup=request.POST["bloodgroup"] 
#         profile_pic=request.FILES.get('profile_pic')
#         student_obj=Employs.objects.get(admin=request.user.id)
#         try:
#             emp_instance=employ_add_form.objects.get(student_id=student_obj)
#             emp_instance.firstname1=firstname1
#             emp_instance.lastname1=lastname1
#             emp_instance.pan=pan
#             emp_instance.ifsecode=ifsecode
#             emp_instance.acno=acno
#             emp_instance.beneficiaryname=beneficiaryname
#             emp_instance.phno=phno
#             emp_instance.gender1=gender1
#             emp_instance.dob=dob
#             emp_instance.address1=address1
#             emp_instance.heq=heq
#             emp_instance.aadharno=aadharno
#             emp_instance.bloodgroup=bloodgroup
#             emp_instance.profile_pic=profile_pic
            

#             emp_instance.save()
#         except employ_add_form.DoesNotExist:
#             k=employ_add_form(profile_pic=profile_pic,student_id=student_obj,firstname1=firstname1,lastname1=lastname1,email2=student_obj.email,pan=pan,ifsecode=ifsecode,acno=acno,beneficiaryname=beneficiaryname,phno=phno,gender1=gender1,dob=dob,address1=address1,heq=heq,aadharno=aadharno,bloodgroup=bloodgroup)
#             k.save();
   
#         return HttpResponseRedirect(reverse("employ_profile"))
#     return render(request,"employ-template/emp1form.html",{'s':s})


def edit_basic_info(request,id):
  k = Employs.objects.get(id=id);
  return render(request,"employ-template/update_basic_info.html",{'k':k})

def update_basic_info(request,id):
    if (request.method=="POST"):
        first_name=request.POST.get("first_name")
        last_name=request.POST.get("last_name")
        email=request.POST.get("email")
        k=Employs.objects.get(id=id);
        
        k.first_name=first_name
        k.last_name=last_name
        k.email=email
        k.save();
        return HttpResponseRedirect(reverse("employ_profile"))
    return render(request,"employ-template/update_basic_info.html")

def edit_pay_info(request,id):
  k = employ_add_form.objects.get(id=id);
  return render(request,"employ-template/update_pay_info.html",{'k':k})

def update_pay_info(request,id):
    if (request.method=="POST"):
        pan=request.POST.get("pan")
        ifsecode=request.POST.get("ifsecode")
        acno=request.POST.get("acno")
        beneficiaryname=request.POST.get("beneficiaryname")
       
        k=employ_add_form.objects.get(id=id);
        
        k.pan=pan
        k.ifsecode=ifsecode
        k.acno=acno
        k.beneficiaryname=beneficiaryname
        k.save();
        return HttpResponseRedirect(reverse("employ_profile"))
    return render(request,"employ-template/update_pay_info.html")

def edit_other_info(request,id):
  k = employ_add_form.objects.get(id=id);
  return render(request,"employ-template/update_other_info.html",{'k':k})

# def update_other_info(request,id):
#     if (request.method=="POST"):
#         phno=request.POST.get("phno")
#         gender1=request.POST.get("gender1")
#         dob=request.POST.get("dob")
#         address1=request.POST.get("address1")
#         heq=request.POST.get("heq")
#         aadharno=request.POST.get("aadharno")
#         bloodgroup=request.POST.get("bloodgroup")

#         k=employ_add_form.objects.get(id=id);
        
#         k.phno=phno
#         k.gender1=gender1
#         k.dob=dob
#         k.address1=address1
#         k.heq=heq
#         k.aadharno=aadharno
#         k.bloodgroup=bloodgroup
#         k.save();
#         return HttpResponseRedirect(reverse("employ_profile"))
#     return render(request,"employ-template/update_other_info.html")



def update_other_info(request,id):
    if (request.method=="POST"):
        phno=request.POST.get("phno")
        gender1=request.POST.get("gender1")
        dob=request.POST.get("dob")
        address1=request.POST.get("address1")
        heq=request.POST.get("heq")
        aadharno=request.POST.get("aadharno")
        fathername=request.POST.get("fathername")
        fathersdob=request.POST.get("fathersdob")
        mothername=request.POST.get("mothername")
        mothersdob=request.POST.get("mothersdob")
        Childdetails1=request.POST.get("Childdetails1")
        Childdetails2=request.POST.get("Childdetails2")
        maritalstatus=request.POST.get("maritalstatus")
        workexperiance=request.POST.get("workexperiance")
        previousemploye=request.POST.get("previousemploye")
        previousdesignation=request.POST.get("previousdesignation")
        Marriageannivarsary=request.POST.get("Marriageannivarsary")
        emergencycontactname=request.POST.get("emergencycontactname")
        emergencycontactnumber=request.POST.get("emergencycontactnumber")
        emergencycontactrelation=request.POST.get("emergencycontactrelation")
        nationality=request.POST.get('nationality')
        qualification=request.POST.get("qualification")
        bloodgroup=request.POST.get("bloodgroup")

        k=employ_add_form.objects.get(id=id);
        
        k.phno=phno
        k.gender1=gender1
        k.dob=dob
        k.address1=address1
        k.heq=heq
        k.aadharno=aadharno
        k.fathername=fathername
        k.fathersdob=fathersdob
        k.mothername=mothername
        k.mothersdob=mothersdob
        k.Childdetails1=Childdetails1
        k.Childdetails2=Childdetails2
        k.maritalstatus=maritalstatus
        k.workexperiance=workexperiance
        k.previousemploye=previousemploye
        k.previousdesignation=previousdesignation
        k.Marriageannivarsary=Marriageannivarsary
        k.emergencycontactname=emergencycontactname
        k.emergencycontactnumber=emergencycontactnumber
        k.emergencycontactrelation=emergencycontactrelation
        k.nationality=nationality
        k.qualification=qualification
        k.bloodgroup=bloodgroup
        k.save();
        return HttpResponseRedirect(reverse("employ_profile"))
    return render(request,"employ-template/update_other_info.html")

def payslip_table(request):
    user=CustomUser.objects.get(id=request.user.id)
    employ=Employs.objects.get(admin=user)
    email_id=request.session.get('e_mail')
    p=employ_payslip.objects.filter(email=email_id).values();
    
    return render(request,"employ-template/paysliptable.html",{'p':p,'employ':employ})


def send_payslip_mail(request):
    # Get the selected year and month from the request
    selected_year = int(request.GET.get('year', datetime.now().year))
    selected_month = int(request.GET.get('month', datetime.now().month))

    # Check if December of the current year is completed
    today = datetime.now()
    if today.month == 12 and today.day > 31:
        selected_month = 1  # Reset to January of the next year
        selected_year += 1  # Increment the year

    # Calculate the start date for filtering past months
    start_date = datetime(selected_year, selected_month, 1) - timedelta(days=1)

    # Calculate the end date based on the selected time frame (3 or 6 months)
    time_frame = request.GET.get('time_frame', '3')  # Default to 3 months if not specified
    if time_frame == '6':
        end_date = start_date - timedelta(days=181)
    else:
        end_date = start_date - timedelta(days=91)

    # Filter data from the `pay` model for the selected time frame
    # filtered_data = MonthlyTotal.objects.filter(date_field__range=[end_date, start_date])

    # Prepare data for the dropdowns
    year_range = range(2021, datetime.now().year + 1)
    month_range = range(1, 13)  # 1 to 12 for months

    # Retrieve data from the `MonthlyTotal` model for the selected time frame
    if time_frame == '6':
        # If the time frame is 6 months, fetch data for the past 6 months
        start_date_month = selected_month - 5 if selected_month >= 6 else 12 - (5 - selected_month)
        start_date_year = selected_year if selected_month >= 6 else selected_year - 1
    else:
        # Default to fetching data for the past 3 months
        start_date_month = selected_month - 2 if selected_month >= 3 else 12 - (2 - selected_month)
        start_date_year = selected_year if selected_month >= 3 else selected_year - 1

    monthly_total_data = MonthlyTotal.objects.filter(
        year__gte=start_date_year,
        year__lte=selected_year,
        month__gte=start_date_month,
        month__lte=selected_month
    )
   

    return render(request, 'employ-template/search_payslip.html', {
        'selected_year': selected_year,
        'selected_month': selected_month,
        # 'filtered_data': filtered_data,
        'year_range': year_range,
        'month_range': month_range,
        'time_frame': time_frame,
        'monthly_total_data': monthly_total_data,})


def send_payslip_to_employeemail(request):
    from_email = settings.EMAIL_HOST_USER
    email=request.session.get('email_2')
    e=Employs.objects.filter(email=email).values();
    recipient_list =[email]
    html_content = render_to_string("employ-template/employ_email_template.html",{'title':'test email','e':e})
    text_content=strip_tags(html_content)
    subject="YOUR PAYSLIP"
    email = EmailMultiAlternatives(
        subject,
        text_content,
        from_email,
        recipient_list,
        )
    email.attach_alternative(html_content,"text/html")
    email.fail_silently = True
    email.send() 
    return redirect('/send_payslip_mail',{'e':e})

def payslip_table(request):
    user=CustomUser.objects.get(id=request.user.id)
    employ=Employs.objects.get(admin=user)
    email=request.session.get('email_2')
    e=Employs.objects.filter(email=email).values()
    return render(request,"employ-template/paysliptable.html",{'e':e,'employ':employ})

def pdf_report_create(request):
    user=CustomUser.objects.get(id=request.user.id)
    employ=Employs.objects.get(admin=user)
    email=request.session.get('email_2')
    c=company_details.objects.all()
    e=Employs.objects.filter(email=email).values()
    # email_id=request.session.get('e_mail')
    # p=employ_payslip.objects.filter(email=email_id).values();
    template_path = 'employ-template/pay.html'

    context = {'employ':employ,'e':e,'c':c,}

    response = HttpResponse(content_type='application/pdf')

    response['Content-Disposition'] = 'filename="products_report.pdf"'

    template = get_template(template_path)

    html = template.render(context)

    # create a pdf
    pisa_status = pisa.CreatePDF(
       html, dest=response)
    # if error then show some funy view
    if pisa_status.err:
       return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response


def pay(request):
    user=CustomUser.objects.get(id=request.user.id)
    employ=Employs.objects.get(admin=user)
    email_id=request.session.get('e_mail')
    p=employ_payslip.objects.filter(email=email_id).values();
    
    return render(request,"employ-template/pay.html",{'p':p,'employ':employ})

def employ_payroll_table(request):
    user=CustomUser.objects.get(id=request.user.id)
    employ=Employs.objects.get(admin=user)
    email_id=request.session.get('e_mail')
    p=employ_payslip.objects.filter(email=email_id).values()
    return render(request,"employ-template/employ_table.html",{'p':p,'employ':employ})


def advancesalary_request(request):
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
        projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    else:
        projects_drops=None
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1)
    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')

    staff_obj = Employs.objects.get(admin=request.user.id)
    leave_data=ad_salary.objects.filter(employ_id=staff_obj)
    items_per_page = 10  # Adjust the number of items per page as needed
    paginator = Paginator(leave_data, items_per_page)
    page = request.GET.get('page')

    try:
        leave_data = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        leave_data = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        leave_data = paginator.page(paginator.num_pages)
    return render(request,"employ-template/employ_advsalayr_request.html",{"leave_data":leave_data,'s':s,'data':data,'admin_drops':admin_drops,'projects_drops':projects_drops})

def advancesalary_request_save(request):
      if request.method!="POST":
        return HttpResponseRedirect(reverse("advancesalary_request"))
      else:
        amount=request.POST.get("amount")
        emi=request.POST.get("emi")
        reason=request.POST.get("reason")
        employ_obj=Employs.objects.get(admin=request.user.id)
        try:
            leave_report=ad_salary(employ=employ_obj,amount=amount,emi=emi,reason=reason,request_status=0)
            leave_report.save()
            messages.success(request, "Successfully Applied for Advancesalary")
            return HttpResponseRedirect(reverse("advancesalary_request"))
        except:
            messages.error(request, "Failed To Apply for  Advancesalary")
            return HttpResponseRedirect(reverse("advancesalary_request"))



def tax_slip(request):
    user = CustomUser.objects.get(id=request.user.id)
    emp=Employs.objects.get(admin=user)
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)
    email=Employs.objects.get(admin=request.user.id)
    tax=employ_tax_form.objects.filter(employ=email)
    return render(request,"employ-template/tax_table.html",{"tax":tax,'s':s,'emp':emp})

def edit_home_rent(request,id):
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)
    k = employ_tax_form.objects.get(id=id);
    return render(request,"employ-template/update_home_rent.html",{'k':k,'s':s})

def update_home_rent(request,id):
    if (request.method=="POST"):
        Current_Monthly_Rent=request.POST.get("Current_Monthly_Rent")
        Name_of_landlord=request.POST.get("Name_of_landlord")
        PAN_of_landlord=request.POST.get("PAN_of_landlord")
        Address_of_landlord=request.POST.get("Address_of_landlord")
        home_rent_proof=request.FILES.get("home_rent_proof")
        from_month=request.POST.get("from_month")
        to_month=request.POST.get("to_month")
        
        k=employ_tax_form.objects.get(id=id);
        
        k.Current_Monthly_Rent=Current_Monthly_Rent
        k.Name_of_landlord=Name_of_landlord
        k.PAN_of_landlord=PAN_of_landlord
        k.Address_of_landlord=Address_of_landlord
        k.home_rent_proof=home_rent_proof
        k.from_month=from_month
        k.to_month=to_month
        k.save();
        return HttpResponseRedirect(reverse("tax_slip"))
    return render(request,"employ-template/update_home_rent.html")

def edit_tax_sections(request,id):
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)
    k = employ_tax_form.objects.get(id=id);
    return render(request,"employ-template/update_tax_sections.html",{'k':k,'s':s})

def update_tax_sections(request,id):
    if (request.method=="POST"):
        Section_80C=request.POST.get("Section_80C")
        Section_80CCD=request.POST.get("Section_80CCD")
        Section_80D=request.POST.get("Section_80D")
        Section_80DD=request.POST.get("Section_80DD")
        Section_80E=request.POST.get("Section_80E")
        Section_80EEB=request.POST.get("Section_80EEB")
        Section_80G=request.POST.get("Section_80G")
        Section_80U=request.POST.get("Section_80U")
        Section_80DDB=request.POST.get("Section_80DDB")
        Section_80TTA=request.POST.get("Section_80TTA")
        Section_80TTB=request.POST.get("Section_80TTB")
        Section_80C_proof=request.FILES.get("Section_80C_proof")
        Section_80D_proof=request.FILES.get("Section_80D_proof")
        Section_80CCD1BS_proof=request.FILES.get("Section_80CCD1BS_proof")
        Section_80DD_proof=request.FILES.get("Section_80DD_proof")
        Section_80E_proof=request.FILES.get("Section_80E_proof")
        Section_80G_proof=request.FILES.get("Section_80G_proof")
        Section_80U_proof=request.FILES.get("Section_80U_proof")
        Section_80EEB_proof=request.FILES.get("Section_80EEB_proof")
        Section_80DDB_proof=request.FILES.get("Section_80DDB_proof")
        Section_80TTA_proof=request.FILES.get("Section_80TTA_proof")
        Section_80TTB_proof=request.FILES.get("Section_80TTB_proof")

        k=employ_tax_form.objects.get(id=id);
        k.Section_80C=Section_80C
        k.Section_80CCD=Section_80CCD
        k.Section_80D=Section_80D
        k.Section_80DD=Section_80DD
        k.Section_80E=Section_80E
        k.Section_80EEB=Section_80EEB
        k.Section_80G=Section_80G
        k.Section_80U=Section_80U
        k.Section_80DDB=Section_80DDB
        k.Section_80TTA=Section_80TTA
        k.Section_80TTB=Section_80TTB
        k.Section_80C_proof=Section_80C_proof
        k.Section_80D_proof=Section_80D_proof
        k.Section_80CCD1BS_proof=Section_80CCD1BS_proof
        k.Section_80DD_proof=Section_80DD_proof
        k.Section_80E_proof=Section_80E_proof
        k.Section_80G_proof=Section_80G_proof
        k.Section_80U_proof=Section_80U_proof
        k.Section_80EEB_proof=Section_80EEB_proof
        k.Section_80DDB_proof=Section_80DDB_proof
        k.Section_80TTA_proof=Section_80TTA_proof
        k.Section_80TTB_proof=Section_80TTB_proof
        k.save();
        return HttpResponseRedirect(reverse("tax_slip"))
    return render(request,"employ-template/update_tax_sections.html")

def edit_home_loan(request,id):
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)
    k = employ_tax_form.objects.get(id=id);
    return render(request,"employ-template/update_home_loan.html",{'k':k,'s':s})

def update_home_loan(request,id):
    if (request.method=="POST"):
        Annual_interest_payable=request.POST.get("Annual_interest_payable")
        Additional_benefit_under_Section=request.POST.get("Additional_benefit_under_Section")
        Name_of_lender=request.POST.get("Name_of_lender")
        PAN_of_lender=request.POST.get("PAN_of_lender")
        Address_of_lender=request.POST.get("Address_of_lender")
        Section_80EEA=request.POST.get("Section_80EEA")
        House_Property_proof=request.FILES.get("House_Property_proof")
        Section_80EE_proof=request.FILES.get("Section_80EE_proof")
        Section_80EEA_proof=request.FILES.get("Section_80EEA_proof")

        k=employ_tax_form.objects.get(id=id);
        k.Annual_interest_payable=Annual_interest_payable
        k.Additional_benefit_under_Section=Additional_benefit_under_Section
        k.Name_of_lender=Name_of_lender
        k.PAN_of_lender=PAN_of_lender
        k.Address_of_lender=Address_of_lender
        k.Section_80EEA=Section_80EEA
        k.House_Property_proof=House_Property_proof
        k.Section_80EE_proof=Section_80EE_proof
        k.Section_80EEA_proof=Section_80EEA_proof
        k.save();
        return HttpResponseRedirect(reverse("tax_slip"))
    return render(request,"employ-template/update_home_loan.html")


def edit_travel_allowance(request,id):
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)
    k = employ_tax_form.objects.get(id=id);
    return render(request,"employ-template/update_travel_allowance.html",{'k':k,'s':s})

def update_travel_allowance(request,id):
    if (request.method=="POST"):
        Amount=request.POST.get("Amount")
        Origin=request.POST.get("Origin")
        Destination=request.POST.get("Destination")
        TravelStartDate=request.POST.get("TravelStartDate")
        k=employ_tax_form.objects.get(id=id);
        k.Amount=Amount
        k.Origin=Origin
        k.Destination=Destination
        k.TravelStartDate=TravelStartDate
        k.save();
        return HttpResponseRedirect(reverse("tax_slip"))
    return render(request,"employ-template/update_travel_allowance.html")

@csrf_exempt
def employ_fcmtoken_save(request):
    token=request.POST.get("token")
    try:
        employ=Employs.objects.get(admin=request.user.id)
        employ.fcm_token=token
        employ.save()
        return HttpResponse("True")
    except:
        return HttpResponse("False")


def pay(request):
    user=CustomUser.objects.get(id=request.user.id)
    employ=Employs.objects.get(admin=user)
    email_id=request.session.get('e_mail')
    p=employ_payslip.objects.filter(email=email_id).values();
    
    return render(request,"employ-template/pay.html",{'p':p,'employ':employ})

def employ_payroll_table(request):
    today = datetime.now()
    current_year = today.year
    user=CustomUser.objects.get(id=request.user.id)
    email_id=request.session.get('e_mail')
    user_obj=Employs.objects.get(admin=request.user.id)
    package = int(user_obj.package)
    # p=employ_payslip.objects.filter(email=email_id).values()
    return render(request,"employ-template/employ_table.html",{'user':user,'user_obj':user_obj})

def employ_all_notification(request):
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)
    employ=Employs.objects.get(admin=request.user.id)
    notifications=NotificationEmploy.objects.filter(employ_id=employ.id)
    notifications1=NotificationEmploy.objects.filter(employ_id=employ.id).count()

    return render(request,"employ-template/all_notification.html",{"notifications":notifications,"notifications1":notifications1,'s':s})


def payslip_apply_view(request):
    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')
    
    employs = Employs.objects.all()
    hr=HR.objects.all()
    emp = CustomUser.objects.filter(user_type=2,is_active=True).count()  # Count active employees
    dis = CustomUser.objects.filter(user_type=2,is_active=False).count()  # Count inactive employees


    staff_obj = Employs.objects.get(admin=request.user.id)
    leave_data=payslip_request.objects.filter(student_id=staff_obj)
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    compid=data.companyid
    a=company_details.objects.filter(companyid=compid).first()
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1)   
    return render(request,"employ-template/payslip_request_apply_view.html",{"data":data,"projects_drops": projects_drops,"a":a,"data":data,"admin_drops":admin_drops,"leave_data":leave_data,'s':s,'employs':employs,'hr':hr,'emp':emp,'dis':dis})

def payslip_apply_view_save(request):
    if request.method != "POST":
        return HttpResponseRedirect(reverse("payslip_apply_view"))
    else:
        # Get the month and year as input, e.g., "2023-10"
        month_year = request.POST.get("monthyear")
        duration = request.POST.get("duration")
        reason = request.POST.get("reason")

        # Convert the month and year string to a Python datetime object
        try:
            date = datetime.strptime(month_year, "%Y-%m")
        except ValueError:
            messages.error(request, "Invalid month and year format")
            return HttpResponseRedirect(reverse("payslip_apply_view"))

        student_obj = Employs.objects.get(admin=request.user.id)
        try:
            leave_report = payslip_request(student_id=student_obj, date=date, duration=duration, reason=reason, status=0)
            leave_report.save()
            messages.success(request, "Successfully Applied for payslip")
            return HttpResponseRedirect(reverse("payslip_apply_view"))
        except:
            messages.error(request, "Failed To Apply for payslip")
            return HttpResponseRedirect(reverse("payslip_apply_view"))


def paysliprequest(request):
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')

    staff_obj =Employs.objects.get(admin=request.user.id)

    k = types.objects.all()
    
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1)
    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')




    st4 = request.POST.get("ss")
    st3 = request.POST.get("vk")
    st = request.POST.get("d1")
    st1 = request.POST.get("d2")

    leave_data=payslip_request.objects.filter(student_id=staff_obj)

    # Apply filters based on user input
    if st4 and st4 != '----Select----':  # Check if a valid status is selected
        leave_data = leave_data.filter(payslip_request_status=st4)
    if st3:  # If "Select Type" is selected
        leave_data = leave_data.filter(typea__icontains=st3)
    if st and st1:
        leave_data = leave_data.filter(date__range=[st, st1])

    total_approved = leave_data.filter(status=1)
    total_pending = leave_data.filter(status=0)
    items_per_page = 20  # Adjust the number of items per page as needed

    paginator = Paginator( leave_data, items_per_page)

    page = request.GET.get('page')

    try:

        leave_data = paginator.page(page)

    except PageNotAnInteger:

        leave_data = paginator.page(1)

    except EmptyPage:

        leave_data = paginator.page(paginator.num_pages)

    return render(request, "employ-template/payslipreq.html", {
        'leave_data': leave_data,
        'k': k,
        's':s,'data1':data1,'data':data,'admin_drops':admin_drops,
'projects_drops':projects_drops,

        'total': total_approved,
        'total1': total_pending,
    })



def help(request):
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)
    return render(request,"employ-template/help.html",{'s':s})
    from django.shortcuts import render,redirect



from .models import employhelp
# Create your views here.
def data(request):
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')

    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    compid=data.companyid
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1)  
    a=company_details.objects.filter(companyid=compid).first() 
    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')
    redirect('/home')
    return render(request,"employhelp/help.html",{'projects_drops':projects_drops,'s':s,'data':data,'admin_drops':admin_drops,'data':data,'a':a})
def home(request):
    
    return render(request,"employhelp/getstarted.html")
def home1(request):
    
    return render(request,"employhelp/leave.html")
def home2(request):
    
    return render(request,"employhelp/taxreduction.html")
def home3(request):
   
    return render(request,"employhelp/reimbesement.html")
def home4(request):
    
    return render(request,"employhelp/investments.html")
def home5(request):
    
    return render(request,"employhelp/mytax.html")
def home6(request):
    return render(request,"employhelp/search.html")
def home7(request):
    return render(request,"employhelp/gstarted.html")
def home8(request):
    return render(request,"employhelp/people.html")
def home9(request):
    return render(request,"employhelp/payroll.html")
def home10(request):
    return render(request,"employhelp/statutory.html")
def home11(request):
    return render(request,"employhelp/payment.html")
def home12(request):
    return render(request,"employhelp/selfservice.html")
def home13(request):
    return render(request,"employhelp/insurance.html")
def home14(request):
    return render(request,"employhelp/accountintegration.html")
def home15(request):
    return render(request,"employhelp/integration.html")
def home16(request):
    return render(request,"employhelp/modules.html")
def home17(request):
    return render(request,"employhelp/updates.html")
def info1(request):
    return render(request,"employhelp/contactsupport.html")
def info2(request):
    return render(request,"employhelp/guied1.html")
def info3(request):
    return render(request,"employhelp/guied2.html")
def info4(request):
    return render(request,"employhelp/guied3.html")
def demo1(request):
    return render(request,"employhelp/demo1.html")
def demo2(request):
    return render(request,"employhelp/demo2.html")
def demo3(request):
    return render(request,"employhelp/demo3.html")
def demo4(request):
    return render(request,"employhelp/demo4.html")
def demo5(request):
    return render(request,"employhelp/demo5.html")
def demo6(request):
    return render(request,"employhelp/demo6.html")
def demo7(request):
    return render(request,"employhelp/demo7.html")
def demo8(request):
    return render(request,"employhelp/demo8.html")
def people1(request):
    return render(request,"employhelp/people1.html")
def people2(request):
    return render(request,"employhelp/people2.html")
def people3(request):
    return render(request,"employhelp/people3.html")
def people4(request):
    return render(request,"employhelp/people4.html")
def people5(request):
    return render(request,"employhelp/people5.html")
def payroll1(request):
    return render(request,"employhelp/payroll1.html")
def payroll2(request):
    return render(request,"employhelp/payroll2.html")
def payroll3(request):
    return render(request,"employhelp/payroll3.html")
def payroll4(request):
    return render(request,"employhelp/payroll4.html")
def payroll5(request):
    return render(request,"employhelp/payroll5.html")
def payroll6(request):
    return render(request,"employhelp/payroll6.html")
def statutory1(request):
    return render(request,"employhelp/statutory1.html")
def statutory3(request):
    return render(request,"employhelp/statutory3.html")
def sfund1(request):
    return render(request,"employhelp/sfund1.html")
def sfund2(request):
    return render(request,"employhelp/sfund2.html")
def sfund3(request):
    return render(request,"employhelp/sfund3.html")
def esi1(request):
    return render(request,"employhelp/esi1.html")
def esi2(request):
    return render(request,"employhelp/esi2.html")
def esi3(request):
    return render(request,"employhelp/esi3.html")
def tds1(request):
    return render(request,"employhelp/tds1.html")
def tds2(request):
    return render(request,"employhelp/tds2.html")
def tds3(request):
    return render(request,"employhelp/tds3.html")
def tds4(request):
    return render(request,"employhelp/tds4.html")
def tax1(request):
    return render(request,"employhelp/tax1.html")
def tax2(request):
    return render(request,"employhelp/tax2.html")
def payment1(request):
    return render(request,"employhelp/payment1.html")
def payment2(request):
    return render(request,"employhelp/payment2.html")
def payment3(request):
    return render(request,"employhelp/payment3.html")
def payment4(request):
    return render(request,"employhelp/payment4.html")
def payment5(request):
    return render(request,"employhelp/payment5.html")
def billing1(request):
    return render(request,"employhelp/billing1.html")

def billing2(request):
    return render(request,"employhelp/billing2.html")

def billing3(request):
    return render(request,"employhelp/billing3.html")
def billing4(request):
    return render(request,"employhelp/billing4.html")

def billing5(request):
    return render(request,"employhelp/billing5.html")
def billing6(request):
    return render(request,"employhelp/billing6.html")
def billing7(request):
    return render(request,"employhelp/billing7.html")
def billing8(request):
    return render(request,"employhelp/billing8.html")
def billing9(request):
    return render(request,"employhelp/billing9.html")
def service1(request):
    return render(request,"employhelp/service1.html")
def service2(request):
    return render(request,"employhelp/service2.html")
def service3(request):
    return render(request,"employhelp/service3.html")
def service5(request):
    return render(request,"employhelp/service5.html")
def attendence1(request):
    return render(request,"employhelp/attendence1.html")
def attendence2(request):
    return render(request,"employhelp/attendence2.html")
def attendence3(request):
    return render(request,"employhelp/attendence3.html")
def attendence4(request):
    return render(request,"employhelp/attendence4.html")
def attendence5(request):
    return render(request,"employhelp/attendence5.html")
def attendence6(request):
    return render(request,"employhelp/attendence6.html")
def attendence7(request):
    return render(request,"employhelp/attendence7.html")
def attendence8(request):
    return render(request,"employhelp/attendence8.html")
def reimbus1(request):
    return render(request,"employhelp/reimbus1.html")
def reimbus2(request):
    return render(request,"employhelp/reimbus2.html")
def dletter1(request):
    return render(request,"employhelp/dletter1.html")
def dletter2(request):
    return render(request,"employhelp/dletter2.html")
def insurance1(request):
    return render(request,"employhelp/insurance1.html")
def insurance2(request):
    return render(request,"employhelp/insurance2.html")
def insurance3(request):
    return render(request,"employhelp/insurance3.html")
def insurance4(request):
    return render(request,"employhelp/insurance4.html")
def insurance5(request):
    return render(request,"employhelp/insurance5.html")
def insurance6(request):
    return render(request,"employhelp/insurance6.html")
def insurance7(request):
    return render(request,"employhelp/insurance7.html")
def insurance8(request):
    return render(request,"employhelp/insurance8.html")
def insurance9(request):
    return render(request,"employhelp/insurance9.html")
def insurance10(request):
    return render(request,"employhelp/insurance10.html")
def insurance11(request):
    return render(request,"employhelp/insurance11.html")
def insurance12(request):
    return render(request,"employhelp/insurance12.html")
def insurance13(request):
    return render(request,"employhelp/insurance13.html")
def insurance14(request):
    return render(request,"employhelp/insurance14.html")
def insurance15(request):
    return render(request,"employhelp/insurance15.html")
def insurance16(request):
    return render(request,"employhelp/insurance16.html")
def insurance17(request):
    return render(request,"employhelp/insurance17.html")
def insurance18(request):
    return render(request,"employhelp/insurance18.html")
def insurance19(request):
    return render(request,"employhelp/insurance19.html")
def slack(request):
    return render(request,"employhelp/slack.html")
def books1(request):
    return render(request,"employhelp/books1.html")
def books2(request):
    return render(request,"employhelp/books2.html")
def books3(request):
    return render(request,"employhelp/books3.html")
def books4(request):
    return render(request,"employhelp/books4.html")
def books5(request):
    return render(request,"employhelp/books5.html")
def books6(request):
    return render(request,"employhelp/books6.html")
def books7(request):
    return render(request,"employhelp/books7.html")
def books8(request):
    return render(request,"employhelp/books8.html")
def home18(request):
    return render(request,"employhelp/tax-regime.html")

def setup_guide(request):
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)

    return render(request,"employ-template/setup_guide.html",{'s':s})




from datetime import date, timedelta
from django.core.mail import send_mail
from django.shortcuts import render
from .models import employ_add_form  # Make sure to import the correct model

def send_birthday_wishes(request):
    today = date.today()
    next_week = today + timedelta(days=6)

    # Query employees whose birthdays fall within the next week
    employees2 = employ_add_form.objects.filter(dateofbirth__day__gte=today.day, dateofbirth__day__lte=next_week.day, dateofbirth__month=today.month).order_by('dob')

    for employee in employees2:
        # Check if email has been sent today
        if not employee.email_sent_today:
            # Check if today is the birthday
            if employee.dob.day == today.day:
                # Send email for today's birthday
                message = f"Dear {employee.firstname},\n\nToday is your birthday! Happy birthday!\n\nBest regards,\nThe HRMS Team"
                send_mail('Happy Birthday', message, 'saipathivada1234@gmail.com', [employee.email])
            else:
                # Send email for upcoming birthday (within 1 or 2 days)
                days_to_birthday = (employee.dateofbirth - today).days
                if 0 < days_to_birthday <= 2:
                    message = f"Dear {employee.firstname},\n\nYour birthday is coming up in {days_to_birthday} days! We wish you an advanced happy birthday!\n\nBest regards,\nThe HRMS Team"
                    send_mail('Upcoming Birthday', message, 'saipathivada1234@gmail.com', [employee.email])

            # Mark email as sent for the current day
            employee.email_sent_today = True
            employee.save()

    s = employnav.objects.all()  # Make sure to replace 'employnav' with the correct model
    return render(request, 'employ-template/birthday.html', {'employees2': employees2, 's': s})


# from django.shortcuts import render
# from datetime import date, timedelta, datetime
# from.models import Employee1
# from django.core.mail import send_mail
# from django.db.models import Q
# from django.utils import timezone
# from django_crontab.crontab import Crontab
# from django.http import HttpResponse, HttpResponseRedirect

# def upcoming_birthdays(request):
#     s=employnav.objects.all()
#     today = date.today()
#     next_week = today + timedelta(days=6)
#     employees = Employee1.objects.filter(Q(birthday_date__day__gte=today.day) & Q(birthday_date__day__lte=next_week.day))
#     upcoming_birthdays = Employee1.objects.filter(birthday_date__day=today.day)
#     for employee in upcoming_birthdays:
#         # Check if an email has already been sent to this employee
#         if Employee1.objects.filter(email=employee.email, sent_att__isnull=False).exists():
#             continue
#         # If an email has not been sent, send it now
#         html_content = render_to_string("email_template.html", {'employee': employee})
#         subject = "Happy birthday to our Best employee!"
#         from_email = settings.DEFAULT_FROM_EMAIL
#         recipient_list = [employee.email]
#         message = EmailMultiAlternatives(subject, html_content, from_email, recipient_list)
#         message.content_subtype = "html"
#         message.fail_silently = False
#         message.send()
#         # Set the sent_birthday_email field for the employee
#         employee.sent_att = timezone.now()
#         employee.save()
#         # Send a separate email to the second employee
#         second_employee = Employee1.objects.filter(birthday_date__day=today.day, id__in=[employee.id for employee in upcoming_birthdays]).exclude(id=employee.id).first()
#         if second_employee:
#             html_content = render_to_string("email_template.html", {'employee': second_employee})
#             subject = "Happy birthday to our Best employee!"
#             from_email = settings.DEFAULT_FROM_EMAIL
#             recipient_list = [second_employee.email]
#             message = EmailMultiAlternatives(subject, html_content, from_email, recipient_list)
#             message.content_subtype = "html"
#             message.fail_silently = False
#             message.send()
#             # Set the sent_birthday_email field for the second employee
#             second_employee.sent_att = timezone.now()
#             second_employee.save()
#     return render(request, 'birthday.html', {'employees': employees,'upcoming_birthdays':upcoming_birthdays,'s':s})



from django.http import HttpResponse


from datetime import datetime, timedelta, date
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, HttpResponseRedirect
from django.views import generic
from django.urls import reverse
from django.utils.safestring import mark_safe
import calendar
import requests
from django.views.generic import TemplateView
from django.template import Context, Template

from django.db.models import Count,Q
from .utils import Calendar,caltable
from django import forms
from .models import employlevsheet,editholiday12,customholidays,publicholidays
def index(request):
    return HttpResponse('hello')

class HomeForm(forms.Form):
    name=forms.CharField(max_length=100)
    email=forms.EmailField()

class CalendarView(TemplateView):
    template_name = "employ-template/calendar.html"
    
    def get_context_data(self, **kwargs,):
        data=Employs.objects.filter(admin=self.request.user.id).first()
        data1=data.id
        data2=data.hroptions
        project_manager=data.projectmanagerop
        compid=data.companyid
        teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
        if data2:
            s=employnav.objects.filter(is_name_exist=1,hr_options=1)
        if project_manager:
            s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
        if teamleadop:
            s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
        else:
            s=employnav.objects.filter(is_name_exist=1,employ_options=1)
        a=company_details.objects.filter(companyid=compid).first()


        admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')
        projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')

        tlop=TeamMember.objects.filter(employee=data1,is_team_lead=1)
        tloptions=employnav.objects.filter(is_name_exist=1,is_tl_option=1)



        context = super(CalendarView,self).get_context_data(**kwargs)
        d = get_date(self.request.GET.get('month', None))
        cal = Calendar(d.year, d.month,d.day)
        html_cal = cal.formatmonth(self.request,withyear=True)
        days_list,weekoff_days = caltable(self,d.year,d.month,self.request)
        employlevsheetcausel = employlev.objects.filter(leave_id=1).first()
        employlevsheetmedical=employlev.objects.filter(leave_id=2).first()
        employlevsheetearned=employlev.objects.filter(leave_id=3).first()
        employs = Employs.objects.get(admin=self.request.user)
        current_month = datetime.today().date().replace(day=1)
        next_month1 = (current_month.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
        weekholiday=editholiday12.objects.first()
        leave_report_datamedical = LeaveReportEmploy.objects.filter(Q(leave_date__range=[current_month, next_month1]) | Q(to_date__range=[current_month, next_month1]),employ_id=employs.id, leave_status=1,leave_type=2)
        leave_report_datacausal = LeaveReportEmploy.objects.filter(Q(leave_date__range=[current_month, next_month1]) | Q(to_date__range=[current_month, next_month1]),employ_id=employs.id, leave_status=1,leave_type=1)
        leave_report_dataearned = LeaveReportEmploy.objects.filter(Q(leave_date__range=[current_month, next_month1]) | Q(to_date__range=[current_month, next_month1]),employ_id=employs.id, leave_status=1,leave_type=3)
        total_leave_duration = sum([leave.get_leave_duration() for leave in leave_report_datamedical])
        total_causel=sum([leave.get_leave_duration() for leave in leave_report_datacausal])
        total_earned=sum([leave.get_leave_duration() for leave in leave_report_dataearned])
        formatted_data =  f"Medical Leave Total Duration (Days): {total_leave_duration}"
        usedmedical=employlevsheetmedical.defaultleave-total_leave_duration
        usedcausel=employlevsheetcausel.defaultleave-total_causel
        usedearned=employlevsheetearned.defaultleave-total_earned
        today = date.today()
        is_special_weekend = customholidays.objects.filter(date=today).exists()
        is_public_holiday = publicholidays.objects.filter(publicholiday_date=today).exists()
        is_weekoff_today =today.strftime('%A') in weekoff_days or is_special_weekend or is_public_holiday
        context['leave_report_data'] = formatted_data
        context['calendar'] = mark_safe(html_cal)
        context['prev_month'] = prev_month(d,self.request)
        context['next_month'] = next_month(d,self.request)
        context['employlevsheetcausel'] = employlevsheetcausel
        context['employlevsheetmedical']=employlevsheetmedical
        context['employlevsheetearned']=employlevsheetearned
        context['weekholiday']=weekholiday
        context['employs']=employs
        context['formatted_data']=formatted_data
        context['usedmedical']=usedmedical
        context['usedcausel']=usedcausel
        context['usedearned']=usedearned
        context['weekoff_days']=weekoff_days
        context['is_weekoff_today']=is_weekoff_today
        context['is_public_holiday']=is_public_holiday
        context['is_special_weekend']=is_special_weekend
        context['s']=s
        context['tlop']=tlop
        context['tloptions']=tloptions
        context['data']=data
        context['admin_drops']=admin_drops
        context['projects_drops']=projects_drops
        context['a']=a



        inline_context={
            "days_list":"days_list",
            
            
        }
        inline_html_template=Template(f"{days_list}")
        inline_view=inline_html_template.render(Context(inline_context))
        context['inline_view']=inline_view
        return  context
    def post(self, request, *args, **kwargs):
        form = HomeForm(request.POST)
        if form.is_valid():
            name = form.cleaned_data['name']
            email = form.cleaned_data['email']
            # Do something with the name and email
            return HttpResponseRedirect(reverse('calendar'))
        else:
            context = self.get_context_data(**kwargs)
            context['form'] = form
           
            context['check_in'] = check_in(self.request)
           
            return redirect('/calendar/')
    

from.models import employlev,halfldayvreason
def employ_apply_leave(request):
    half=halfldayvreason.objects.first()
    half1=employlev.objects.all()
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    compid=data.companyid
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
        projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    else:
        projects_drops=None
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1)
    a=company_details.objects.filter(companyid=compid).first()

    staff_obj = Employs.objects.get(admin=request.user.id)
    leave_data=LeaveReportEmploy.objects.filter(employ_id=staff_obj)
    
    items_per_page = 10

    page = request.GET.get('page', 1)

    paginator = Paginator(leave_data, items_per_page)

    try:
        leave_data_page = paginator.page(page)
    except PageNotAnInteger:
        leave_data_page = paginator.page(1)
    except EmptyPage:
        leave_data_page = paginator.page(paginator.num_pages)


    return render(request,"employ-template/employ_apply_leave.html",{'a':a,"leave_data":leave_data,'s':s,'half':half,'half1':half1,"leave_data": leave_data_page,'data':data,'projects_drops':projects_drops,
})



    
def get_date(req_month):
    if req_month:
        year, month = (int(x) for x in req_month.split('-'))
        curr_day = datetime.now().day
        curr_year = datetime.now().year
        curr_month = datetime.now().month
        return date(year, month, day=1)

    return datetime.today()

def prev_month(d,request):
    first = d.replace(day=1)
    prev_month = first - timedelta(days=1)
    month = 'month=' + str(prev_month.year) + '-' + str(prev_month.month)
    request.session['year']=str(prev_month.year);
    return month

def next_month(d,request):
    days_in_month = calendar.monthrange(d.year, d.month)[1]
    last = d.replace(day=days_in_month)
    next_month = last + timedelta(days=1)
    month = 'month=' + str(next_month.year) + '-' + str(next_month.month)
    request.session['month']=str(next_month.month);
    request.session['year']=str(next_month.year);
    return month



def sidenav(request):
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)
    return render(request,'employ-template/calendar.html',{'s':s})

        

# def check_in(request):
#     today = datetime.now().date()
#     now = datetime.now().time()
#     stat = "present"
#     email = request.session.get('email_2')

#     if request.method == 'POST' and 'vk' in request.POST:
#         check = checkin.objects.filter(date=today, empid=email).first()
#         if not check:
#             checkin.objects.create(date=today, time=now, status=stat, empid=email)
#             formatted_time = now.strftime("%I:%M %p")  # Format time as HH:MM AM/PM
#             messages.success(request,f'You have successfully checked in at {formatted_time}.')
#             return redirect('/calendar/')
#         else:
#             messages.warning(request,'You have already checked in today.')
#             return redirect('/calendar/')


#     if request.method == 'POST' and 'vk1' in request.POST:
#         checkt = checkout.objects.filter(date=today, empid=email).first()
#         if not checkt:
#             checkout.objects.create(date=today, time=now, empid=email,date_value=1)
#             formatted_time = now.strftime("%I:%M %p")  # Format time as HH:MM AM/PM
#             messages.success(request, f'You have successfully checked out at  {formatted_time}.')
#             return redirect('/calendar/')

#         else:
#             messages.warning(request, 'You have already checked out today.')
#             return redirect('/calendar/')

#     return render(request, 'employ-template/calendar.html')


from datetime import datetime, timedelta
from django.contrib import messages
from django.shortcuts import redirect, render
from .models import working_shifts, checkin, checkout, Employs

def check_in(request):
    user = request.user
    employs = Employs.objects.get(admin=user)
    today = datetime.now().date()
    now = datetime.now().time()
    stat="present"
    email = request.session.get('email_2')

    # Fetch the working shift data based on 'employs.working12' value
    workings = working_shifts.objects.filter(shift_name=employs.working12).first()

    if workings:
        if request.method == 'POST':
            if 'vk' in request.POST:
                # Check if the user has already checked in
                check_in = checkin.objects.filter(date=today, empid=email).first()
                if not check_in:
                    # Calculate the check-in and cutoff times
                    starting_time = workings.starting_time
                    before_time = workings.befor_time  # Use the correct field name
                    starting_datetime = datetime.combine(today, starting_time)
                    check_in_time = starting_datetime - timedelta(minutes=before_time)
                    check_in_cutoff = starting_datetime + timedelta(minutes=workings.cutoff_time)

                    if check_in_time <= datetime.now() <= check_in_cutoff:
                        checkin.objects.create(date=today, time=now,status=stat, empid=email, shift_name=workings.shift_name,companyid=employs.companyid,is_employee="1")
                        formatted_time = now.strftime("%I:%M %p")  # Format time as HH:MM AM/PM
                        messages.success(request, f'You have successfully checked in at {formatted_time}.')
                    else:
                        messages.warning(request, 'It is too early/late to check in.')
                else:
                    messages.warning(request, 'You have already checked in today.')

            if 'vk1' in request.POST:
                # Check if the user has already checked out
                check_out = checkout.objects.filter(date=today, empid=email).first()
                if not check_out:
                    checkout.objects.create(date=today, time=now, empid=email, shift_name=workings.shift_name, date_value=1,companyid=employs.companyid,is_employee="1")
                    formatted_time = now.strftime("%I:%M %p")  # Format time as HH:MM AM/PM
                    messages.success(request, f'You have successfully checked out at {formatted_time}.')
                else:
                    messages.warning(request, 'You have already checked out today.')

            return redirect('/calendar/')
    else :
        messages.warning(request, 'You are not allowed to check in or out because your shift is not valid.')
        return redirect('/calendar/')
    return render (request,'employ-template/calendar.html')



from django.shortcuts import render
from datetime import datetime
from django.db.models import Sum

from .models import  MonthlyTotal,company_details_first
from .utils import get_current_month_holidays_count

def count_business_days(year, month):
    total_days = calendar.monthrange(year, month)[1]
    business_days = 0
    
    for day in range(1, total_days + 1):
        if datetime(year, month, day).weekday() < 5:  # Monday to Friday (0 to 4)
            business_days += 1
    
    return business_days

def get_month_name(month_number):
    return calendar.month_abbr[month_number]

def emp_payslip(request):
    today = datetime.now()
    
    current_year = today.year
    current_month = today.month
    company_detail=company_details_first.objects.all()
    objs=Employs.objects.get(admin=request.user.id)
    details=employ_add_form.objects.filter(student_id=objs)
    # Filter data for the current month
    # data_for_month = CheckData.objects.filter(
    #     created_at__year=current_year,
    #     created_at__month=current_month
    # )
    data_for_months = checkout.objects.filter(
        date__year=current_year,
        date__month=current_month,
        empid=objs.email
    )
    india_holidays_count = get_current_month_holidays_count('india')
    user_obj=Employs.objects.get(admin=request.user.id)
    package = int(user_obj.package)
    tax= int(package * 0.66667)
    
    b_tax=int(tax * 0.5)
    h_tax=int(b_tax * 0.5)
    s_tax=int(b_tax * 0.3)
    l_tax=int(b_tax * 0.2)
    standard_deductions= 50000
    net_taxable_income=tax - standard_deductions
    quotient = int(package) / 12
    main_salary=quotient
    # Multiply the remainder by 0.5%
    month_salary = main_salary * 0.5
    value=(main_salary * 0.25) +  (main_salary * 0.15) + (main_salary * 0.10)
    rent=int(main_salary * 0.25)
    travel=int(main_salary * 0.15)
    leave_travel=int(main_salary * 0.10)
    # Get the number of days in the current month
    business_days_in_current_month = count_business_days(current_year, current_month)

    # Adjust the amount based on the number of days in the current month
    quotients = month_salary / business_days_in_current_month 
    try:
        monthly_total_result = data_for_months.aggregate(total=Sum('date_value'))
        monthly_total = monthly_total_result['total'] + india_holidays_count
    except TypeError:
        monthly_total = 0
    totalss = int(monthly_total * quotients)
    
    total= int(totalss + value )
    month_and_year = f"{get_month_name(current_month)} {current_year}"
    if monthly_total is None:
        monthly_total = 0
    monthly_total_obj, created = MonthlyTotal.objects.update_or_create(
        year=current_year,
        month=current_month,
        student_id=user_obj,
        month_and_year = month_and_year,
        defaults={'total': totalss}
    )
    data=MonthlyTotal.objects.get(student_id=user_obj,month=current_month)
    return render(request, 'employ-template/emp_payslip.html', {'standard_deductions':standard_deductions,'net_taxable_income':net_taxable_income,'tax':tax,'b_tax':b_tax,'h_tax':h_tax,'s_tax':s_tax,'l_tax':l_tax,'travel':travel,'leave_travel':leave_travel,'rent':rent,'details':details,'user_obj':user_obj,'data':data,'company_detail':company_detail,'total':total,'monthly_total': monthly_total_obj.total,'totalss':totalss})


from .models import TeamMember , Task , Project
def employee_record_emp(request, employee_id):
    employee = Employs.objects.get(id=employee_id)
    is_team_lead = TeamMember.objects.filter(employee=employee, is_team_lead=True).exists()
    # Get employee's overall performance data
    performance_data = employee.calculate_overall_performance()

    # Get the list of projects associated with the employee
    projects = Project.objects.filter(teammember__employee=employee)
    team_members = TeamMember.objects.filter(employee=employee)
    total_score_across_all_projects = sum(member.calculate_total_score_across_all_projects() for member in team_members)
    # Get the list of tasks associated with the employee
    tasks = Task.objects.filter(e_id=employee)

    context = {
        'employee': employee,
        'performance_data': performance_data,
        'projects': projects,
        'tasks': tasks,
        'total_score_across_all_projects':total_score_across_all_projects,
        'is_team_lead':is_team_lead
    }


    return render(request,"employ-template/employee_record.html" , context)



def team_lead_read_proj(request, teamlead_id):
    projects = Project.objects.filter(teammember__employee=teamlead_id)
    completed_count = projects.filter(status='completed').count()
    ongoing_count = projects.filter(status='ongoing').count()
    featured_count = projects.filter(status='featured').count()
    if request.method == 'GET':
        status = request.GET.get('status')

        if status == 'completed':
            projects = projects.filter(status='completed')
        elif status == 'ongoing':
            projects = projects.filter(status='ongoing')
        elif status == 'featured':
            projects = projects.filter(status='featured')
        else:
            # If the status is not specified or is invalid, show all projects
            projects = Project.objects.filter(teammember__employee=teamlead_id)

        # Count the number of projects for each status
        
    return render(request, 'employ-template/TeamLeadViewProjects.html', {'projects': projects,"status":status,"completed_count":completed_count,"ongoing_count":ongoing_count,"featured_count":featured_count})

from django.utils import timezone

def team_lead_projectwise_task(request, pid):
    
    
    
    team_members = TeamMember.objects.filter(project=pid)
    team_member_scores = [(team_member, team_member.calculate_total_score_in_project()) for team_member in team_members]
    team_member_scores = sorted(team_member_scores, key=lambda x: x[1], reverse=True)
    project_details=Project.objects.filter(id=pid)
    task_details = Task.objects.filter( id=pid).all()

    if request.method == 'GET':
        t_status = request.GET.get('t_status', 'total')  # Default to 'total' if not provided

        if t_status == 'total':
            task_details = Task.objects.filter( p_id=pid).all()
        elif t_status == 'completed':
            task_details = Task.objects.filter( p_id=pid, t_status='completed')
        elif t_status == 'in-progress':
            # task_details = Task.objects.filter( p_id=pid, t_status='in-progress').all()
            task_details = Task.objects.filter( p_id=pid).exclude(t_status='completed')

        

        tasks = task_details
        count_no_of_total_tasks = Task.objects.filter( p_id_id=pid).count()
        count_no_of_completed_tasks = Task.objects.filter( p_id_id=pid, t_status="completed").count()
        count_no_of_pending_tasks = count_no_of_total_tasks - count_no_of_completed_tasks
        task_data = []

        for task in tasks:
            now = timezone.now()
            task_score = task.calculate_task_score()
            remaining_time = task.calculate_remaining_time()
            progress_percentage = min(task_score / 10 * 100, 100)
            excess_points = max(task_score - 10, 0)  # Calculate the excess points (if any)
            neg_points = max(-task_score, 0)

            task_data.append({
                'task': task,
                'progress_percentage': progress_percentage,
                'excess_points': excess_points,
                'neg_points': neg_points,
                'task_score': task_score,
                'remaining_time': remaining_time,
            })
            
        
        context = {
        
        "project_details": project_details,
        'task_total': count_no_of_total_tasks,
        'task_completed': count_no_of_completed_tasks,
        'task_pending': count_no_of_pending_tasks,
        'team_members': team_members,
        
        't_status':t_status,
        "task_details": task_details,
        'task_data': task_data,
        
        'team_member_scores': team_member_scores,

        }
    
    return render(request, 'employ-template/TeamLeadViewProjectwiseTasks.html', context)

def team_lead_delete_user(request,employee_id,project_id):
        # Assuming your model has a field 'id' representing the user's ID
        employee = TeamMember.objects.get(employee_id=employee_id,project_id=project_id)
        employee.delete()
        referer = request.META.get('HTTP_REFERER')
        if referer:
          return redirect(referer)
        else:
          return redirect(reverse('home'))

from django.http import HttpResponse, HttpResponseRedirect, JsonResponse

def tl_get_team_members(request, project_id):
    team_members = TeamMember.objects.filter(project_id=project_id)
    data = [{'id': member.employee.id, 'name': f"{member.employee.first_name} {member.employee.last_name}"}
            for member in team_members]
    return JsonResponse({'team_members': data})

from django.core.mail import send_mail
from django.conf import settings

def team_lead_create_task(request,pid):
    
    boards = HR.objects.all()
    projects = Project.objects.filter(id=pid ).values()
    employees = Employs.objects.all()
    context = {"boards": boards, "projects": projects, "employees": employees}
    if request.method == 'POST':
        t_name = request.POST['t_name']
        t_desc = request.POST['t_desc']
        t_deadline_date = request.POST['t_deadline_date']
        t_status = "todo"
        t_priority = request.POST['t_priority']
        b_id = request.POST['b_id']
        p_id = pid
        e_id = request.POST['e_id']
        taskObj = Task.objects.create(t_name=t_name, t_desc=t_desc, t_deadline_date=t_deadline_date,
                                      t_status=t_status, t_priority=t_priority,  b_id_id=b_id, p_id_id=p_id, e_id_id=e_id)
        if taskObj:
            empDetails = Employs.objects.filter(id=e_id).values()
            subject = 'DevelopTrees - New Task Created for you'
            message = f'Hi {empDetails[0]["first_name"]} , Your organization as created a new task : {t_name} , description : {t_desc}, priority : {t_priority} and deadline for task is : {t_deadline_date}, Login in your account to get more information. From: DevelopTrees. '
            email_from = settings.EMAIL_HOST_USER
            recipient_list = [empDetails[0]["email"], ]
            send_mail(subject, message, email_from, recipient_list)
            messages.success(request, "Task was created successfully!")
            dynamic_url = reverse ('team_lead_projectwise_tasks' , args=[pid])
            return HttpResponseRedirect(dynamic_url)
        else:
            messages.error(request, "Some Error was occurred!")
            return HttpResponseRedirect('/team_lead_create_task')
    return render(request, 'employ-template/TeamLeadCreateTask.html', context)

def team_lead_delete_task(request, pk):
    try:
        # Fetch the task and its project ID
        task = Task.objects.get(id=pk)
        pid = task.p_id.id  # Assuming Task has a foreign key to Project model

        # Check if the user has the necessary permissions (you might want to customize this logic)
        
        task.delete()
            

    except Task.DoesNotExist:
        messages.error(request, "Task does not exist.")
    except Exception as e:
        messages.error(request, f"An error occurred: {str(e)}")

    dynamic_url = reverse('team_lead_projectwise_tasks', args=[pid])
    return HttpResponseRedirect(dynamic_url)

def team_lead_update_task(request, pk):
    # try:
    tasks = Task.objects.get(id=pk)
    p_id = tasks.p_id
    pid=tasks.p_id.id
    projects_emp_link = TeamMember.objects.filter(project_id=p_id)
     
    if request.method == 'POST':
        t_name = request.POST['t_name']
        t_desc = request.POST['t_desc']
        t_deadline_date = request.POST['t_deadline_date']
        t_priority = request.POST['t_priority']
        e_id = request.POST['e_id']
        employ_instance = Employs.objects.get(id=e_id)
        t_status = request.POST['t_status']
        tasks.t_deadline_date = datetime.strptime(t_deadline_date, '%Y-%m-%dT%H:%M')
        tasks.t_name = t_name
        tasks.t_desc = t_desc
        
        tasks.t_priority = t_priority
        tasks.t_status = t_status
        tasks.e_id = employ_instance
        tasks.save()
        if tasks:
            empDetails = Employs.objects.filter(id=e_id).values()
            subject = 'DevelopTrees - Task Updated for you'
            message = f'Hi {empDetails[0]["first_name"]} , Your organization as updated a task : {t_name} , description : {t_desc}, priority : {t_priority} and deadline for task is : {t_deadline_date}, Login in your account to get more information. From: DevelopTrees. '
            email_from = settings.EMAIL_HOST_USER
            recipient_list = [empDetails[0]["email"], ]
            send_mail(subject, message, email_from, recipient_list)
            messages.success(request, "Task was updated successfully!")
            dynamic_url = reverse('team_lead_projectwise_tasks', args=[pid])
            return HttpResponseRedirect(dynamic_url)

        else:
            messages.error(request, "Some Error was occurred!2")
            
    else:
        if tasks:
            return render(request, 'employ-template/TeamLeadUpdateTask.html', {"tasks": tasks , "projects_emp_link": projects_emp_link } )
        else:
            messages.error(request, "Some Error was occurred!1")
            
    dynamic_url = reverse('team_lead_projectwise_tasks', args=[pid])
    return HttpResponseRedirect(dynamic_url)

def employ_upload_photo(request):
    # data = Employs.objects.get(admin=request.user.id)
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
        projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    else:
        projects_drops=None

    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1) 


    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')
    
    if request.method == "POST":
        profile_pic = request.FILES.get("profile_pic")
        if profile_pic:
           
            data.profile_pic = profile_pic
            
                # Create a new instance if it doesn't exist
            data,_ = Employs.objects.get_or_create(admin=request.user.id)
            data.profile_pic=profile_pic

            data.save()
            messages.success(request, "Profile picture uploaded successfully.")
        else:
            messages.error(request, "Profile picture upload failed.")
        return redirect('/Employ_home')
    return render(request, "employ-template/employprofile_pic.html", {'data':data,'s':s,'data1':data1,'admin_drops':admin_drops,'projects_drops':projects_drops})
def employdelete_profile_pic(request):
    user_obj = Employs.objects.get(admin=request.user.id)
    try:
       
        user_obj.profile_pic.delete()  # Delete the profile picture file
        user_obj.save()
        messages.success(request, "Profile picture removed successfully.")
    except Employs.DoesNotExist:
        messages.error(request, "User data not found. Please add user data first.")
    
    return redirect("employ_upload_photo")



def team_lead_update_project(request, pid):
    # Get the project to be updated
    project = get_object_or_404(Project, id=pid)
    team_leader_info=project.get_team_leader()
    
    # Get available team members excluding team leads
    teams=Employs.objects.filter(Q(teammember__isnull=True) | Q(teammember__project=project)).distinct()
    team_members = TeamMember.objects.filter(project=pid)

    if request.method == 'POST':

       
        # Retrieve updated project details from the form
        p_name = request.POST['p_name']
        p_desc = request.POST['p_desc']
        pr_deadline = request.POST['project_deadline_date']
        project_manager = request.POST['manager_name']
        status = request.POST.get('status')

        # Update the project
        project.p_name = p_name
        project.p_desc = p_desc
        project.project_deadline = pr_deadline
        project.project_manager = project_manager
        project.status = status
        project.save()  
            # Update team lead if changed
        selected_team_lead_id = request.POST.get('team_lead')
        if selected_team_lead_id:
            try:
                team_lead = Employs.objects.get(id=selected_team_lead_id)
                project.teammember_set.update(is_team_lead=False)  # Remove team leader status from existing team leader

                # Create a new TeamMember if it doesn't exist
                team_member, created = TeamMember.objects.get_or_create(project=project, employee=team_lead)
                team_member.is_team_lead = True
                team_member.save()
            except Employs.DoesNotExist:
                # Handle the case where the selected team leader doesn't exist
                messages.error(request, 'Selected team leader does not exist.')
                return redirect('/read-proj')

        # Update selected team members
        selected_employee_ids = request.POST.getlist('selected_employees[]')
        existing_team_member_ids = TeamMember.objects.filter(project=project).values_list('employee_id', flat=True)
        for emp_id in selected_employee_ids:
            emp_id = int(emp_id)
            if emp_id not in existing_team_member_ids:
                employee = get_object_or_404(Employs, id=emp_id)
                TeamMember.objects.create(project=project, employee=employee, is_team_lead=False)

        messages.success(request, 'Project updated successfully and employees assigned.')
        return redirect('/read-proj')

    # Fetch available team members (exclude team leads)
    # selected_team_lead_id = project.team_lead.id if project.team_lead else None
    return render(request, 'employ-template/team_lead_update_project.html', {"project": project,"teams":teams, "team_members": team_members,'team_leader_info':team_leader_info})

from .models import Project,TeamMember
def read_proj_hr(request):
    projects = Project.objects.all()
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')


    return render(request, 'employ-template/ViewProjects1.html',{'projects':projects,'admin_drops':admin_drops})

def reports1(request):
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1)

    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')

    return render(request,"employ-template/reports.html",{'s':s,'admin_drops':admin_drops,'data':data})

def projectwise_task_1(request, pid):
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')

    project_details=Project.objects.filter(id=pid)
    team_members = TeamMember.objects.filter(project=pid)
    team_member_scores = [(i, i.calculate_total_score_in_project()) for i in team_members]
    team_member_scores = sorted(team_member_scores, key=lambda x: x[1], reverse=True)
    return render(request, 'employ-template/ViewProjectwiseTasks1.html', {'project_details':project_details,'team_members':team_members,'team_member_scores':team_member_scores,'admin_drops':admin_drops})


from .models import projecttask,tlassigntask

from datetime import datetime


def tlassigntask1(request):
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    compid=data.companyid
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    a=company_details.objects.filter(companyid=compid).first()
    
    team_members = TeamMember.objects.filter(is_team_lead=1, employee=data1).first()
    if team_members:
        team1 = team_members.project
        project_ids = TeamMember.objects.filter(project=team1).values_list('project_id', flat=True)
        employee_ids = TeamMember.objects.filter(project=team1).values_list('employee_id', flat=True)
        tlemp_id = TeamMember.objects.filter(project=team1, is_team_lead=1).values_list('employee_id', flat=True)
        employees = Employs.objects.filter(id__in=employee_ids)
        projects = Project.objects.filter(id__in=project_ids).first()
        tlid = Employs.objects.filter(id__in=tlemp_id).first()
        tlidc = tlid.id
        project_id1 = projects.id
        # Filter projecttask by date
        today = datetime.now()
        # task = projecttask.objects.filter(l_name=tlidc, p_name=project_id1, date__date=today.date()).values()
        task = projecttask.objects.filter(l_name=tlidc, p_name=project_id1, task_date=today).values()

    else:
        # No team members found
        team1 = None
        employees = None
        projects = None
        project_id1 = None
        task = None
    items_per_page = 10  # Adjust the number of items per page as needed
    paginator = Paginator(task, items_per_page)
    page = request.GET.get('page')

    try:
        task = paginator.page(page)
    except PageNotAnInteger:
        task = paginator.page(1)
    except EmptyPage:
        task = paginator.page(paginator.num_pages)        

    if request.method == "POST":
        task = request.POST.getlist('task')
        employid = request.POST.getlist('employid')
        project_id = request.POST.getlist('project_id')
        for task, employid, project_id in zip(task, employid, project_id):
            if task:
                assign = tlassigntask(task=task, employid=employid, project_id=project_id)
                assign.save()
        return redirect('/Employ_home')

    return render(request, "employ-template/tltaskassign.html", {'a':a,'tlidc': tlidc, 'task': task, 'team1': team1, 'project_id1': project_id1, 'employees': employees,'s':s,'data':data})



from.models import employperformance
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger



from.models import employperformance

def table1(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    tsk1_data = employperformance.objects.filter(employ_name_id__companyid=em1)
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    items_per_page = 10  # Adjust the number of items per page as needed
    paginator = Paginator(tsk1_data, items_per_page)
    page = request.GET.get('page')

    try:
        tsk1_data = paginator.page(page)
    except PageNotAnInteger:
        tsk1_data = paginator.page(1)
    except EmptyPage:
        tsk1_data = paginator.page(paginator.num_pages) 
    
    context = {
        'tsk1_data': tsk1_data,'s':s,"data":data,'data1':data1,'projects_drops':projects_drops
    }
    return render(request, 'employ-template/emptable.html', context)

from .models import admin_project_create,Meeting

def displayed(request):
    
    pmanager = admin_project_create.objects.filter(admin_id=request.user.id)

    data1=Employs.objects.get(admin=request.user.id)
    if request.method == 'POST':
           meeting_url = request.POST.get('meeting_url')
           team_member = request.POST.get('team_member')

        # Create a new project
           p1 = Meeting(
               meeting_url=meeting_url,
               team_member=team_member
            )
           p1.save()

    team_members = TeamMember.objects.filter(is_team_lead=1,employee=data1).first()
    if team_members:
        team1=team_members.project
        employee_ids=TeamMember.objects.filter(project=team1).values_list('employee_id',flat=True)
        employees=Employs.objects.filter(id__in=employee_ids)
    else:
        employees=None

    
        
        return redirect('/Employ_home')

    return render(request, 'employ-template/data3.html', {'data1':data1,'employees':employees, 'pmanager': pmanager,  'team_members': TeamMember.objects.all()})


from django.db.models import Avg

def performancetask(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    employ_obj=Employs.objects.get(admin=request.user.id)
    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')
    team=TeamMember.objects.filter(employee=data1,is_team_lead=1)  
    projects = Project.objects.all()  
    items_per_page = 2
    performance_data = employperformance.objects.filter(employ_name_id__companyid=em1)
    paginator = Paginator(performance_data, items_per_page)
    page = request.GET.get('page')

    try:
        performance_data = paginator.page(page)
    except PageNotAnInteger:
        performance_data = paginator.page(1)
    except EmptyPage:
        performance_data = paginator.page(paginator.num_pages)
    return render(request, 'employ-template/emptaskper.html', {'s':s,'team':team,'employ_obj':employ_obj,'projects':projects,'data':data,'data1':data1,'admin_drops':admin_drops, 'performance_data': performance_data})


def performanceproject(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    compid=data.companyid
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1) 
    a=company_details.objects.filter(companyid=compid).first()
    employ_obj=Employs.objects.get(admin=request.user.id)
    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')
    team=TeamMember.objects.filter(employee=data1,is_team_lead=1)
    projects = Project.objects.filter(o_id__companyid=em1) 
    if request.method == 'POST':
        selected_project_id = request.POST.get('project')  # Get the selected project ID

        if selected_project_id:
            selected_project = Project.objects.get(id=selected_project_id)
            project_name = selected_project.p_name
            # Fetch the employee average performances for the selected project
            performances_for_project = employperformance.objects.filter(project_name=selected_project)
            employee_average_performances = performances_for_project.values('employ_name').annotate(average_performance=Avg('performance'))
           

            # Retrieve employee names and IDs
            for emp in employee_average_performances:
                employ_id = emp['employ_name']
                try:
                    employee = Employs.objects.get(id=employ_id)
                    emp['employee_name'] = employee.first_name
                    emp['employee_id'] = employee.empid
                except Employs.DoesNotExist:
                    emp['employee_name'] = "N/A"
                    emp['employee_id'] = "N/A"
        else:
            performances_for_project=None
            project_name=None
            employee_average_performances=None
    else:
            performances_for_project=None
            project_name=None
            employee_average_performances=None
            
    items_per_page = 2
    employee_average_performances = employperformance.objects.all()
    paginator = Paginator(employee_average_performances, items_per_page)
    page = request.GET.get('page')

    try:
        employee_average_performances = paginator.page(page)
    except PageNotAnInteger:
        employee_average_performances = paginator.page(1)
    except EmptyPage:
        employee_average_performances = paginator.page(paginator.num_pages)


    return render(request, 'employ-template/empprojectper.html', {'a':a,'s':s,'admin_drops':admin_drops,'team':team,'employ_obj':employ_obj,'projects':projects,'project_name':project_name,'employee_average_performances':employee_average_performances,'data':data,'data1':data1,'projects_drops':projects_drops})

def performanceallproject(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')
    employ_obj=Employs.objects.get(admin=request.user.id)
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1)
    team=TeamMember.objects.filter(employee=data1,is_team_lead=1)
  
    projects = Project.objects.all()  
    items_per_page = 2
    tsk1_data = employperformance.objects.filter(employ_name_id__companyid=em1).values('employ_name').annotate(avg_performance=Avg('performance'))
    employees_data = []
    for task in tsk1_data:
        employ_id = task['employ_name']
        try:
            employee = Employs.objects.get(id=employ_id)
            task['employee_name'] = employee.first_name
            task['employee_id'] = employee.empid
            project_names = employperformance.objects.filter(employ_name=employee).values('project_name__p_name')
            task['project_names'] = project_names

        except Employs.DoesNotExist:
            task['employee_name'] = "N/A"
            task['employee_id'] = "N/A"
            task['project_names'] = "N/A"
        employees_data.append(task)
      
    items_per_page = 2
    paginator = Paginator(employees_data, items_per_page)
    page = request.GET.get('page')

    try:
        employees_data = paginator.page(page)
    except PageNotAnInteger:
        employees_data = paginator.page(1)
    except EmptyPage:
        employees_data = paginator.page(paginator.num_pages)
    return render(request, 'employ-template/empavgper.html', {'s':s,'employ_obj':employ_obj,'projects':projects,'data':data,'data1':data1,'admin_drops':admin_drops, 'tsk1_data': employees_data,'team':team})

def reports1(request):
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1)

    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')

    return render(request,"employ-template/reports.html",{'s':s,'admin_drops':admin_drops,'data':data})

# from django.shortcuts import render
from django.db import models
# from.models import Employee,Department
from datetime import date, timedelta,timezone
from datetime import datetime, timedelta
from django.shortcuts import render
from django.db.models import Sum
from .models import adminnav, admin_drop, Department, Employs

def search1(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1)

    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')
    departments = Department.objects.all()
    employees = Employs.objects.filter(companyid=em1)
    
    search_term = request.GET.get('search_term')
    designation = request.GET.get('department')
    distinct_department_designations = employees.values('designation').distinct()
    selected_date_range = request.GET.get('dateofjoining')
    
    employees_by_designation = {}

    if designation:
        designation_employees = employees.filter(designation=designation)
        
        if selected_date_range:
            today = datetime.now().date()
            start_date = today - timedelta(days=int(selected_date_range) * 30)
            designation_employees = designation_employees.filter(dateofjoining__gte=start_date)
        
        if search_term:
            designation_employees = designation_employees.filter(first_name__icontains=search_term)
        
        total_count = designation_employees.count()
        total_package = designation_employees.aggregate(Sum('package'))['package__sum']
        
        employees_by_designation[designation] = designation_employees
    else:
        total_count = 0
        total_package = 0
    
    labels = []
    salaries = []

    context = {
        'employees': employees_by_designation,
        'total_count': total_count,
        'total_package': total_package,
        'labels': labels,
        'salaries': salaries,
        's': s,
        'admin_drops': admin_drops,
        'departments': distinct_department_designations,
        'selected_date_range': selected_date_range, 
        'data':data, 
    }

    return render(request, 'employ-template/variance_report.html', context)

def documentreport1(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1)

    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')
    employ=empdocs.objects.filter(employ_id__companyid=em1)
    return render (request,"employ-template/docreport.html",{'employ':employ,'s':s,'data':data,'admin_drops':admin_drops})
import os
import zipfile
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.conf import settings
from io import BytesIO
from .models import Employs  # Import your Employs model

def download_images_zip1(request, employ_id):
    employ = get_object_or_404(Employs, id=employ_id)
      
    # Create a BytesIO stream to write the zip file.
    output = BytesIO()
    zip_file = zipfile.ZipFile(output, 'w')

    # Iterate over 'empdocs' related to the employee and add their image files to the zip.
    for empdoc in employ.empdocs_set.all():
        document_type = empdoc.documenttype1
        image_path = empdoc.imagefile.path
        image_name = os.path.basename(image_path)
        
        # Construct the relative path inside the zip file.
        relative_path = os.path.join(employ.first_name, image_name)
        zip_file.write(image_path, relative_path)

    zip_file.close()

    # Set response headers to serve the zip file.
    response = HttpResponse(output.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename={employ.first_name}_images.zip'
    return response

import os
import zipfile
from io import BytesIO
from django.http import HttpResponse
from django.conf import settings
from .models import Employs

def download_all_employee_data1(request):
    # Create a BytesIO stream to write the zip file.
    output = BytesIO()
    zip_file = zipfile.ZipFile(output, 'w')

    # Directory structure for the zip file.
    base_dir = os.path.join(settings.MEDIA_ROOT, 'employees')

    # Iterate over all employees.
    for employ in Employs.objects.all():
        employee_dir = os.path.join(base_dir, employ.first_name)

        # Add employee documents to the zip file.
        for empdoc in employ.empdocs_set.all():
            image_path = empdoc.imagefile.path
            image_name = os.path.basename(image_path)
            relative_path = os.path.join(employ.first_name, image_name)
            zip_file.write(image_path, relative_path)

    zip_file.close()

    # Set response headers to serve the zip file.
    response = HttpResponse(output.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename=all_employee_data.zip'
    return response
from django.shortcuts import render
from datetime import date, timedelta
from .models import Employs, checkin, LeaveReportEmploy, editholiday12, customholidays, publicholidays



def paid1(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1)
    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')
    results = []
   

    # Query your data with date filtering
    employees = Employs.objects.filter(companyid=em1)
    
    # current_day=datetime.now().day
    current_month = datetime.now().month
    current_year = datetime.now().year

    _, num_days = calendar.monthrange(current_year, current_month)
    start_date = datetime(current_year, current_month, 1)
    
    days_of_month = [start_date + timedelta(days=i) for i in range(num_days)]

    checkin_statuses = {}  # A dictionary to store check-in status for each day

    # Iterate through each employee
    for employee in employees:
        empid = employee.empid
        first_name = employee.first_name

        # Calculate the current month
        today = date.today()
        current_month = today.month
        current_year = today.year

        last_day_of_month = date(current_year, current_month, 1)
        last_day_of_month = last_day_of_month.replace(day=28)
        while last_day_of_month.month == current_month:
            last_day_of_month += timedelta(days=1)
        last_day_of_month -= timedelta(days=1)

        # Calculate the number of days in the current month
        days_in_month = (last_day_of_month - date(current_year, current_month, 1)).days + 1

        # Query your database for admin-configured holidays
        editholidays = editholiday12.objects.first()
        admin_holidays = sum(getattr(editholidays, day, False) or 0 for day in ['sun', 'sat1', 'sat2', 'sat3', 'sat4', 'sat5', 'alsat', 'mon', 'tue', 'web', 'thu', 'fri'])

        # Query your database for custom holidays
        custom_holidays = customholidays.objects.filter(date__month=current_month).count()

        # Query your database for public holidays
        public_holidays = publicholidays.objects.filter(publicholiday_date__month=current_month).count()

        # Calculate the remaining days
        remaining_days = days_in_month - admin_holidays - custom_holidays - public_holidays

        # Calculate the total check-ins for this employee (based on email in the checkin table)
        total_checkins = checkin.objects.filter(empid=employee.email, date__month=current_month).count()

        # Query for leaves for this employee in the current month
        half_leave = LeaveReportEmploy.objects.filter(employ_id=employee, leave_status=1, leave_date__month=current_month, leave_type="Half-Day").count()
        unpaid_leave = LeaveReportEmploy.objects.filter(employ_id=employee, leave_status=1, leave_date__month=current_month, leave_type="Unpaid_leave").count()
        leave = LeaveReportEmploy.objects.filter(employ_id=employee, leave_status=1, leave_date__month=current_month, leave_type="Leave").count()
        open_request = LeaveReportEmploy.objects.filter(employ_id=employee, leave_status=0, leave_date__month=current_month).count()

        checkin_status_for_employee = []  # Store check-in status for this employee

        # Iterate through each day in the month
        for day in days_of_month:
            # Check if there's a check-in record for this employee on this day
            checkin_exists = checkin.objects.filter(
                empid=employee.email,  # Match email
                date=day
            ).exists()

            checkin_status_for_employee.append('Present' if checkin_exists else '-NA-')

        checkin_statuses[employee] = checkin_status_for_employee

        results.append({
            'empid': empid,
            'first_name': first_name,
            'total_checkins': total_checkins,
            'remaining_days': remaining_days,
            'original_working_days': remaining_days - total_checkins,
            'half_leave': half_leave,
            'unpaid_leave': unpaid_leave,
            'leave': leave,
            'open_request': open_request,
           
        })

    return render(request, "employ-template/paid.html", {
        'days_of_month': days_of_month,
        'employ_id': employees,
        'checkin_statuses': checkin_statuses,
        'results': results,
        'admin_drops':admin_drops,
        's':s,
        'data':data,
    })
from datetime import datetime, timedelta
from calendar import monthrange
from django.shortcuts import render
from .models import Employs, checkin

def paid3(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1)
    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')
    employees = Employs.objects.filter(companyid=em1)
   
    # Get start_date and end_date from query parameters (if provided)
    start_date_param = request.GET.get('start_date')
    end_date_param = request.GET.get('end_date')
    checkin_statuses = {}
    
    # Determine today's date
    today = datetime.now()

    # Calculate start_date and end_date based on provided parameters or default to the current month
    if start_date_param and end_date_param:
        start_date = datetime.strptime(start_date_param, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_param, '%Y-%m-%d')
    else:
        current_month = today.month
        current_year = today.year
        _, num_days = monthrange(current_year, current_month)
        start_date = datetime(current_year, current_month, 1)
        end_date = datetime(current_year, current_month, num_days)
    
    # Create a list of dates in the date range
    date_range = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]

    # Categorize dates into "previous," "present," or "future"
    previous_dates = [date for date in date_range if date < today]
    present_dates = [date for date in date_range if date == today]
    future_dates = [date for date in date_range if date > today]

    for employee in employees:
        checkin_status_for_employee = []

        for day in date_range:
            checkin_exists = checkin.objects.filter(
                empid=employee.email,
                date=day
            ).exists()
            checkin_status_for_employee.append('Present' if checkin_exists else '-NA-')

        checkin_statuses[employee] = checkin_status_for_employee

    # Determine the category of the date range
    if start_date < today and end_date < today:
        date_category = "Previous"
    elif start_date > today and end_date > today:
        date_category = "Future"
    else:
        date_category = "Present"

    return render(request, "employ-template/paid2.html", {
        'date_category': date_category,
        'date_range': date_range,
        'employ_id': employees,
        'checkin_statuses': checkin_statuses,
        'previous_dates':previous_dates,
        'present_dates':present_dates,
        'future_dates':future_dates,
        'admin_drops':admin_drops,
        's':s,
        'data':data,

    })



def balance1(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1)
    results = []
   

    # Query your data with date filtering
    employees = Employs.objects.filter(companyid=em1)
    
    # current_day=datetime.now().day
    current_month = datetime.now().month
    current_year = datetime.now().year

    _, num_days = calendar.monthrange(current_year, current_month)
    start_date = datetime(current_year, current_month, 1)
    
    days_of_month = [start_date + timedelta(days=i) for i in range(num_days)]
    for employee in employees:
        empid = employee.empid
        first_name = employee.first_name

        # Calculate the current month
        today = date.today()
        current_month = today.month
        current_year = today.year

        last_day_of_month = date(current_year, current_month, 1)
        last_day_of_month = last_day_of_month.replace(day=28)
        while last_day_of_month.month == current_month:
            last_day_of_month += timedelta(days=1)
        last_day_of_month -= timedelta(days=1)

        # Calculate the number of days in the current month
        days_in_month = (last_day_of_month - date(current_year, current_month, 1)).days + 1
   

        # Calculate the number of days in the current month
        half_leave = LeaveReportEmploy.objects.filter(employ_id=employee, leave_status=1, leave_date__month=current_month, leave_type="Half-Day").count()
        unpaid_leave = LeaveReportEmploy.objects.filter(employ_id=employee, leave_status=1, leave_date__month=current_month, leave_type="Unpaid_leave").count()
        leave = LeaveReportEmploy.objects.filter(employ_id=employee, leave_status=1, leave_date__month=current_month, leave_type="Leave").count()
        open_request = LeaveReportEmploy.objects.filter(employ_id=employee, leave_status=0, leave_date__month=current_month).count()


        # Add information to the results
        results.append({
            'empid': empid,
            'first_name': first_name,
            'half_leave': half_leave,
            'unpaid_leave': unpaid_leave,
            'leave': leave,
            'open_request': open_request,
            's':s,
            'data':data,
            'admin_drops':admin_drops
            
        })

    return render(request, 'employ-template/balance.html', {'results': results})


from datetime import datetime, timedelta
from calendar import monthrange
from django.shortcuts import render
from .models import Employs, checkin

def opeanrequest1(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1)
    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')
    employees = Employs.objects.filter(companyid=em1)
   
    # Get start_date and end_date from query parameters (if provided)
    start_date_param = request.GET.get('start_date')
    end_date_param = request.GET.get('end_date')
    checkin_statuses = {}
    
    # Determine today's date
    today = datetime.now()

    # Calculate start_date and end_date based on provided parameters or default to the current month
    if start_date_param and end_date_param:
        start_date = datetime.strptime(start_date_param, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_param, '%Y-%m-%d')
    else:
        current_month = today.month
        current_year = today.year
        _, num_days = monthrange(current_year, current_month)
        start_date = datetime(current_year, current_month, 1)
        end_date = datetime(current_year, current_month, num_days)
    
    # Create a list of dates in the date range
    date_range = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]

    # Categorize dates into "previous," "present," or "future"
    previous_dates = [date for date in date_range if date < today]
    present_dates = [date for date in date_range if date == today]
    future_dates = [date for date in date_range if date > today]

    for employee in employees:
        checkin_status_for_employee = []

        for day in date_range:
            checkin_exists = checkin.objects.filter(
                empid=employee.email,
                date=day
            ).exists()
            checkin_status_for_employee.append('Present' if checkin_exists else '-NA-')

        checkin_statuses[employee] = checkin_status_for_employee

    # Determine the category of the date range
    if start_date < today and end_date < today:
        date_category = "Previous"
    elif start_date > today and end_date > today:
        date_category = "Future"
    else:
        date_category = "Present"

    return render(request, "employ-template/openanrequest.html", {
        'date_category': date_category,
        'date_range': date_range,
        'employ_id': employees,
        'checkin_statuses': checkin_statuses,
        'previous_dates':previous_dates,
        'present_dates':present_dates,
        'future_dates':future_dates,
        'admin_drops':admin_drops,
        's':s,
        'data':data,

    })    

from django.core.mail import send_mail
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Employs, empdocs, employ_add_form
from django.http import Http404

from django.core.exceptions import ObjectDoesNotExist

def all_employees_missing_info1(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1)
    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')
    a=companylogo.objects.all()
    if request.method == 'POST':
        selected_employee_ids = request.POST.getlist('selected_employees')

        # Loop through the selected employee IDs and send emails
        for employee_id in selected_employee_ids:
            try:
                employee = Employs.objects.get(id=employee_id)
                employee_email = employee.email
                try:
                    employ_add_form_info = employ_add_form.objects.get(student_id=employee)
                except employ_add_form.DoesNotExist:
                    employ_add_form_info = None

                empdocs_info = empdocs.objects.filter(employ_id=employee)

                # Compose the email message with missing information
                missing_info = []
                if employee.missing_info():
                    missing_info.append(", ".join(employee.missing_info().keys()))
                if employ_add_form_info and employ_add_form_info.missing_info():
                    missing_info.append(", ".join(employ_add_form_info.missing_info().keys()))
                for empdoc in empdocs_info:
                    if empdoc.missing_info():
                        missing_info.append(", ".join(empdoc.missing_info().keys()))
                        
                if missing_info:
                    subject = f"Mr/Miss {employee.first_name}.{employee.last_name} Please Update Your Missing Information."
                    message = f"Dear {employee.first_name} .{employee.last_name},\n\n"
                    message += f"The following information is missing or incomplete:\n"
                    message += "\n".join(missing_info)
                    # message += "\n\nPlease Login to DevelopTrees HRMS to update your information as soon as possible."
                    message += "\n\nPlease click the following link to update your information as soon as possible.\n"
                    # Generate the profile URL using the employee's ID
                    # message += "\n"
                    # profile_url = reverse('employ_profile', kwargs={'employee_id': employee_id})
                    # message += request.build_absolute_uri(profile_url)

                    from_email = settings.EMAIL_HOST_USER  # Replace with your email address
                    recipient_list = [employee.email]  # Send email to the employee's email address

                    send_mail(subject, message, from_email, recipient_list)

            except Employs.DoesNotExist:
                pass  # Handle the case where an employee does not exist

        return redirect('all_employees_missing_info1')

    employees = Employs.objects.filter(companyid=em1)
    employee_data = []

    for employee in employees:
        try:
            employ_add_form_info = employ_add_form.objects.get(student_id=employee)
        except ObjectDoesNotExist:
            employ_add_form_info = None

        employee_info = {
            'employs': employee,
            'empdocs_info': empdocs.objects.filter(employ_id=employee),
            'employ_add_form_info': employ_add_form_info,
        }
        employee_data.append(employee_info)

    context = {
        'employee_data': employee_data,
        'admin_drops':admin_drops,
        'a':a,
        's':s,
        'data':data,
    }

    return render(request, 'employ-template/all_employees_missing_info1.html', context)

def all_employees_reimbursement_report1(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1)
    
    staff_obj = Employs.objects.all()

    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')   
    k = types.objects.all()

    st4 = request.POST.get("ss")
    st3 = request.POST.get("vk")
    st = request.POST.get("d1")
    st1 = request.POST.get("d2")

    leave_data = Reimbursement.objects.filter(employ_id__companyid=em1)

    # Apply filters based on user input
    if st4 and st4 != '----Select----':  # Check if a valid status is selected
        leave_data = leave_data.filter(reimbursement_status=st4)
    if st3:  # If "Select Type" is selected
        leave_data = leave_data.filter(typea__icontains=st3)
    if st and st1:
        leave_data = leave_data.filter(date__range=[st, st1])

    total_approved = leave_data.filter(reimbursement_status=1).aggregate(Sum('amount'))['amount__sum'] or 0
    total_pending = leave_data.filter(reimbursement_status=0).aggregate(Sum('amount'))['amount__sum'] or 0

    return render(request, "employ-template/all_employees_reimbursement_report.html", {
        'leave_data': leave_data,
        's': s,
        'k': k,
        'total': total_approved,
        'total1': total_pending,
        'staff_obj':staff_obj,
        'admin_drops':admin_drops,
        'data':data,
    })
from django.db import connection
from.models import companylogo
# def yearmonth_uploaded1(request):
#     data=Employs.objects.filter(admin=request.user.id).first()
#     admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')
#     s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)


#     a=companylogo.objects.all()
#     ad=ad_salary.objects.first()
#     b=Reimbursement.objects.first()
#     x=connection.cursor()
#     x.execute("SELECT ehrms_employs.empid,ehrms_employs.first_name,ehrms_employs.package,ehrms_employs.dateofjoining,ehrms_checkin.status,ehrms_reimbursement.reimbursement_status,ehrms_employs.id FROM ehrms_employs INNER JOIN ehrms_checkin on ehrms_employs.email=ehrms_checkin.empid INNER JOIN ehrms_reimbursement on ehrms_employs.id=ehrms_reimbursement.employ_id_id;")
#     rs=x.fetchall()  
#     return render(request,"employ-template/salaryreport.html",{'data':data,'s':s,'rs':rs,'a':a,'b':b,'admin_drops':admin_drops,'ad':ad})
def yearmonth_uploaded1(request):
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1)

    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')   
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid

    a=companylogo.objects.all()
    ad=ad_salary.objects.first()
    b=Reimbursement.objects.first()
   
   
   
    rst = Employs.objects.filter(companyid=em1)

   
    page = request.GET.get('page', 1)
    items_per_page = 5  # Adjust the number of items per page as needed
    paginator = Paginator(rst, items_per_page)

    try:
        rst = paginator.page(page)
    except PageNotAnInteger:
        rst = paginator.page(1)
    except EmptyPage:
        rst = paginator.page(paginator.num_pages)
 
    return render(request,"employ-template/salaryreport.html",{'rst':rst,'a':a,'b':b,'admin_drops':admin_drops,'ad':ad,'s':s,'data':data,'paginator': paginator
})
# def masterctcs1(request):
#     admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')
#     s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)
#     data=Employs.objects.filter(admin=request.user.id).first()

#     a=companylogo.objects.all()
#     k=connection.cursor()
#     k.execute("SELECT ehrms_employs.empid,ehrms_customuser.first_name,ehrms_customuser.email,ehrms_employs.role,ehrms_employs.location,ehrms_employs.status,ehrms_employs.dateofjoining,ehrms_employs.dateofbirth,ehrms_employs.package FROM ehrms_employs INNER JOIN ehrms_customuser ON ehrms_employs.id=ehrms_customuser.id;")

#     rs=k.fetchall()
#     return render(request,'employ-template/masterctc.html',{'rs':rs,'s':s,'admin_drops':admin_drops,'a':a,'data':data})

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

def masterctcs1(request):
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1)

    a = companylogo.objects.all()
    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    rst = Employs.objects.filter(companyid=em1)

    # Pagination
    page = request.GET.get('page', 1)
    items_per_page = 1
    paginator = Paginator(rst, items_per_page)

    try:
        rst = paginator.page(page)
    except PageNotAnInteger:
        rst = paginator.page(1)
    except EmptyPage:
        rst = paginator.page(paginator.num_pages)

    return render(request, 'employ-template/masterctc.html', {'data': data, 'rst': rst, 's': s, 'admin_drops': admin_drops, 'a': a})
def register2(request):
    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1)

    query = request.GET.get('query')
    x=connection.cursor()
    x.execute("SELECT ehrms_employs.id,ehrms_employs.empid,ehrms_employs.first_name,ehrms_employs.last_name,ehrms_employs.email,ehrms_employ_add_form.phno,ehrms_employs.gender,ehrms_employ_add_form.pan,ehrms_employs.dateofjoining,ehrms_employs.role FROM ehrms_employs INNER JOIN ehrms_employ_add_form ON ehrms_employs.email=ehrms_employ_add_form.email2;")
    rs=x.fetchall()
    paginator = Paginator(rs, 10)  # Show 10 items per page
    page_number = request.GET.get('page')
    res = paginator.get_page(page_number)
    return render(request, 'employ-template/register_data.html', {'data':data,'res': res,'s':s,'admin_drops':admin_drops,'query':query})





from .models import  CustomUser, Employs,Progress
def add_employ(request):
    li=list.objects.all()
    admin_drops=employ_drop.objects.filter(parent_category=None).order_by('id')
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    if project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    if teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1)


    existing_user_email = None
    existing_user_username = None

    if request.method == "POST":
        email = request.POST.get("email")
        username = request.POST.get("username")

        if email:
            try:
                existing_user_email = CustomUser.objects.get(email=email)
                messages.info(request, 'Email already exists')
                return redirect("/add_employ")  # Redirect here to avoid processing the rest of the view
            except ObjectDoesNotExist:
                existing_user_email = None
            except MultipleObjectsReturned:
                messages.error(request, ' Email already exists.')
                return redirect("/add_employ")  # Redirect here to avoid processing the rest of the view

        if username:
            try:
                existing_user_username = CustomUser.objects.get(username=username)
                messages.error(request, 'Username already exists')
                return redirect("/add_employ")  # Redirect here to avoid processing the rest of the view
            except ObjectDoesNotExist:
                existing_user_username = None
            except MultipleObjectsReturned:
                messages.error(request, 'Both username and email already exist. Contact the administrator.')
                return redirect("/add_employ")  # Redirect here to avoid processing the rest of the view

    if request.method == "POST":
        first_name = request.POST["first_name"]
        last_name = request.POST["last_name"]
        username = request.POST["username"]
        email = request.POST["email"]
        web_mail = request.POST["web_mail"]
        password = request.POST["password"]
        address = request.POST["address"]
        empid = request.POST["empid"]
        designation = request.POST.get("designation")
        location = request.POST["location"]
        package = request.POST["package"]
        pincode = request.POST["pincode"]
        contactno = request.POST["contactno"]
        bloodgroup = request.POST.get("bloodgroup")
        dateofjoining = request.POST['dateofjoining']
        sex = request.POST["gender"]
        role = request.POST.get('role') # Added field for role
    
        if role == "Employ":
            user = CustomUser.objects.create_user(username=username, password=password, email=email, user_type=2)
            user.employs.first_name = first_name
            user.employs.last_name = last_name
            user.employs.email = email
            user.employs.password = password
            user.employs.address = address
            user.employs.empid = empid
            user.employs.web_mail = web_mail
            user.employs.designation = designation
            user.employs.location = location
            user.employs.package = package
            user.employs.pincode = pincode
            user.employs.contactno = contactno
            user.employs.bloodgroup = bloodgroup
            user.employs.dateofjoining = dateofjoining
            user.employs.role = role  # Set the role field for HR
            user.employs.hroptions=0
            user.save()
            messages.success(request, "Successfully Added Employee")
                
        elif role == "HR":
            user = CustomUser.objects.create_user(username=username, password=password, email=email, user_type=2)
            user.employs.first_name = first_name
            user.employs.last_name = last_name
            user.employs.email = email
            user.employs.password = password
            user.employs.address = address
            user.employs.empid = empid
            user.employs.web_mail = web_mail
            user.employs.designation = designation
            user.employs.location = location
            user.employs.package = package
            user.employs.pincode = pincode
            user.employs.contactno = contactno
            user.employs.bloodgroup = bloodgroup
            user.employs.dateofjoining = dateofjoining
            user.employs.role = role  # Set the role field for HR
            user.employs.hroptions=1
            user.save()
            messages.success(request, "Successfully Added HR")
        else:
            user = CustomUser.objects.create_user(username=username, password=password, email=email, user_type=1)
            user.adminhod.firstname = first_name
            user.adminhod.lastname = last_name
            user.adminhod.email = email
            user.adminhod.password = password
            user.adminhod.address = address
            user.adminhod.empid = empid
            user.adminhod.web_mail = web_mail
            user.adminhod.designation = designation
            user.adminhod.location = location
            user.adminhod.package = package
            user.adminhod.pincode = pincode
            user.adminhod.contactno = contactno
            user.adminhod.bloodgroup = bloodgroup
            user.adminhod.dateofjoining = dateofjoining
            user.adminhod.role = role  # Set the role field for HR
            user.adminhod.options=1
            user.save()
            messages.success(request, "Successfully Added Project Manager")
            

       
        html_content = render_to_string("email_template.html", {'title': 'test email', 'first_name': first_name,
                                                                'last_name': last_name, 'empid': empid})
        text_content = strip_tags(html_content)
        subject = "WELCOME TO DEVELOPTRESS"
        email = EmailMultiAlternatives(
            subject,
            text_content,
            settings.EMAIL_HOST_USER,
            [email],
        )
        email.attach_alternative(html_content, "text/html")
        email.fail_silently = True
        email.send()
        progress = 20
        val = 4
        value_message = 1
        progress_obj, _ = Progress.objects.get_or_create(pk=1, defaults={'value': 0})

        progress_obj.value3 = value_message
        progress_obj.progress_form4 = progress

        progress_obj.save()
        instance = admin_home_drop.objects.get(id=5)
        instance.progress_value = val
        instance.save()
        
        return redirect("/add_employ")

    # If the request method is not POST, render the form
    # s = adminnav.objects.all()
    form = Employs.objects.all()
    a1=working_shifts.objects.all()
    hrform = HR.objects.all()
    role = list.objects.all()
    # admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    return render(request, "employ-template/add_employ.html",
                  {'role': role, 'hrform': hrform, "form": form, 's': s, 'admin_drops': admin_drops,'li':li,'a1':a1,'admin_drops':admin_drops,'data':data})

